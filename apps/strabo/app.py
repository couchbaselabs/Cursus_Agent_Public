#!/usr/bin/env python3
"""
Strabo — Couchbase Supportal Explorer
======================================
Multi-tab dashboard for scraping, analysing, and chatting with support data.

Auth modes:
  1. Cookie paste  — copy the Cookie header from browser DevTools, paste here.
                     Uses requests (no browser required).
  2. Browser login — opens a real Chromium window so you can complete SSO,
                     then scrapes headlessly with the saved session.

Usage:
  ./venv/bin/python run_strabo.py
  # then open http://localhost:8765 in your browser
"""

__version__ = "2.7.70"

import asyncio
import threading
import csv
import datetime
import hashlib
import io
import json
import os
import re
import time
import urllib.parse
import uuid
from pathlib import Path
from typing import Callable, Optional

import requests
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
from bs4 import BeautifulSoup
from nicegui import run, ui
import subprocess

# When imported as a library (e.g. by the unified app), skip NiceGUI page
# registration so the importing app's routes are not overwritten.
_LIBRARY_MODE = os.environ.get("STRABO_LIBRARY_MODE") == "1"

# ── Extracted modules ─────────────────────────────────────────────────────────
from supportal.prompts import (
    SYSTEM_PROMPT_TEMPLATE,
    CLASSIFY_PROMPT,
    EXTRACT_PROMPT,
    RERANK_PROMPT,
    CRITIQUE_PROMPT,
    SCORING_SYSTEM_PROMPT,
    _SUMMARY_SYSTEM,
    _SUMMARY_PROMPT_TMPL,
)
from supportal.llm_providers import (
    fetch_ollama_models,
    fetch_openai_compat_models,
    _THINKING_MODEL_PATTERNS,
    _model_has_thinking_by_name,
    fetch_ollama_model_info,
    _parse_num_ctx_from_params,
    poll_ollama_ps,
    poll_lmstudio_model_info,
    lmstudio_load_model,
    lmstudio_ensure_model_loaded,
)
from supportal.constants import BASE_URL, UA, TICKET_HREF_RE, SETTINGS_FILE, COOKIES_FILE, PROFILE_DIR
from supportal.ticket_parser import (
    _find_label_value,
    _DATE_RE,
    _guess_author_from_text,
    _extract_comments,
    _extract_all_dl_fields,
    _extract_named_section,
    _normalize_field_key,
    parse_ticket_detail,
    _is_deleted_api_ticket,
    parse_ticket_from_api,
    _STATUS_MAP,
    _extract_ticket_links,
    _extract_ticket_rows,
    _resolve_customer_input,
    _normalize_customer_url,
    _find_customer_url_in_search,
)
from supportal.api_client import (
    _make_api_session,
    query_supportal_analytics,
    fetch_snapshots_via_analytics,
    fetch_ticket_api,
    _get_customer_ticket_listing_api,
    _get_customer_snapshot_listing_api,
    search_customers_on_supportal,
    search_customers_via_analytics,
    resolve_customer_name,
    _find_tickets_tab_url,
)
from supportal.snapshot_parser import (
    _UUID_RE,
    _SNAP_ID_RE,
    _SNAP_HREF_RE,
    _SNAP_DEBUG_DIR,
    _topo_str,
    _highest_snap_id,
    extract_cluster_snapshot_info,
    _normalize_checker_name,
    _parse_snapshot_checker_text,
    _write_snap_debug,
    _parse_structured_api_json,
    fetch_snapshot_topology_api,
    fetch_snapshot_topology,
    enrich_tickets_with_snapshots,
    _find_snapshots_tab_url,
    _extract_snapshot_rows,
    scrape_snapshots_from_stubs,
    scrape_snapshots_for_customer,
)
from supportal.ticket_parser import _extract_ticket_ids, _parse_ticket_fields
from supportal.scoring import (
    _openai_base_url,
    _tls_openai,
    _get_openai_client,
    _APP_CLUSTER_ALIASES_SEED,
    _cluster_app_dynamic,
    _app_cluster_dynamic,
    _get_cluster_to_app,
    _get_app_cluster_aliases,
    _FOLLOWUP_TRIGGERS,
    call_llm,
    rewrite_query_for_retrieval,
    _ticket_date,
    _parse_ticket_date,
    _ticket_cluster_ids,
    build_dataset_stats,
    prefilter_for_query,
    compute_aggregations,
    build_rag_context,
    classify_query,
    extract_ticket_fields,
    rerank_tickets,
    self_critique_answer,
    run_deep_reasoning,
    _build_memory_section,
    contextualize_question,
    chat_batch_map_reduce,
)
from supportal.cb_helpers import (
    _cb_conn_str,
    _cb_kv_get_multi,
    search_orgs_from_cb,
    load_tickets_for_orgs_from_cb,
    load_tickets_from_cb,
    fetch_tickets_by_keys,
    _make_snap_col,
    load_to_couchbase,
    build_embed_text,
    build_snapshot_embed_text,
    embed_text_ollama,
    embed_text,
    embed_all_snapshots,
    embed_all_tickets,
    migrate_ticket_fields_in_cb,
    create_vector_index,
    vector_search_cb,
    _snap_keys_to_ticket_keys,
    fts_keyword_search_cb,
    _load_cluster_app_map,
    _KEYWORD_STOPWORDS,
    build_structured_query,
    structured_search_cb,
    tool_query_tickets,
    search_tickets_retrieve_rerank,
    snapshot_topology_search,
    fetch_snapshots_for_clusters,
    reciprocal_rank_fusion,
    hybrid_retrieval,
    _mlx_emb_cache,
)
from supportal.agent_tools import (
    _AGENT_TOOLS,
    _SUPPORTAL_TICKET_URL,
    _SUPPORTAL_CUSTOMER_URL,
    _ARTIFACT_RE,
    _CODE_ASSET_RE,
    _ASSET_MIME,
    _ASSET_ICONS,
    _build_agent_echart_option,
    _agent_filters_from_args,
    _extract_text_tool_calls,
    _normalise_tool_args,
    call_llm_with_tools,
    _classify_agent_error,
    _generate_followup_suggestions,
    _canonical_priority,
    _compute_health_score,
    _compute_sla_compliance,
    _get_digest,
    _save_query_to_cb,
    _list_saved_queries,
    _tag_ticket_in_cb,
    _generate_customer_report,
    _make_asset_thumbnail,
    _ensure_assets_collection,
    _save_asset_to_cb,
    _list_assets_from_cb,
    _get_asset_content_from_cb,
    _delete_asset_from_cb,
    _fleet_query,
    _query_fleet_tickets,
    _list_at_risk_clusters,
    _fleet_version_distribution,
    _fleet_cbse_impact,
    _compute_health_score_with_cluster,
)
from supportal.prompts import build_agent_system_prompt


def _safe_notify(client, message: str, type: str = "info") -> None:  # noqa: A002
    """Notify the client only if it is still connected; silently no-ops on disconnect."""
    if client is None:
        ui.notify(message, type=type)
        return
    try:
        from nicegui.client import Client as _NClient
        if client.id not in _NClient.instances:
            return
        with client:
            ui.notify(message, type=type)
    except Exception:
        pass




# Optional — Couchbase SDK (Phase 1 + 2).  Import lazily so the app starts
# even if the package is not installed.
try:
    from couchbase.auth import PasswordAuthenticator
    from couchbase.cluster import Cluster
    from couchbase.options import ClusterOptions, UpsertOptions, SearchOptions, QueryOptions
    from couchbase.exceptions import CouchbaseException
    from couchbase.search import SearchRequest, MatchQuery, DisjunctionQuery
    from couchbase.vector_search import VectorQuery, VectorSearch
    import couchbase.subdocument as _SD
    from datetime import timedelta
    _CB_AVAILABLE = True
except ImportError:
    _CB_AVAILABLE = False

# Optional — LLM providers (Phase 2).
try:
    import anthropic as _anthropic_mod
    _ANTHROPIC_AVAILABLE = True
except ImportError:
    _ANTHROPIC_AVAILABLE = False

try:
    from google import genai as _genai_mod
    _GEMINI_AVAILABLE = True
except ImportError:
    _GEMINI_AVAILABLE = False

try:
    import openai as _openai_mod   # also covers Ollama chat + LMStudio
    _OPENAI_AVAILABLE = True
except ImportError:
    _OPENAI_AVAILABLE = False

try:
    import boto3 as _boto3_mod
    _BOTO3_AVAILABLE = True
except ImportError:
    _boto3_mod = None
    _BOTO3_AVAILABLE = False

try:
    from mlx_embeddings import load as _mlx_emb_load
    import mlx.core as _mx
    try:
        from mlx_embeddings import pool_embeddings as _mlx_pool
    except ImportError:
        _mlx_pool = None
    _MLX_EMB_AVAILABLE = True
    _MLX_EMB_IMPORT_ERROR = None
except Exception as _mlx_err:
    _MLX_EMB_AVAILABLE = False
    _MLX_EMB_IMPORT_ERROR = str(_mlx_err)

try:
    import cairosvg as _cairosvg
    _CAIROSVG_AVAILABLE = True
except ImportError:
    _CAIROSVG_AVAILABLE = False

import dataclasses


@dataclasses.dataclass
class CbConfig:
    """Couchbase connection parameters — pass as a unit to pipeline functions."""
    url: str
    bucket: str
    username: str
    password: str
    use_tls: bool = False
    scope: str = "_default"
    ticket_collection: str = "tickets"
    snap_collection: str = "snapshots"


# (moved to supportal/ package, imported at top of file)
# ── RAG chat cache helpers ────────────────────────────────────────────────────

def _chat_cache_key(prefix: str, *parts: str) -> str:
    """Build a Couchbase doc key for a chat cache entry."""
    digest = hashlib.sha256("|".join(parts).encode()).hexdigest()[:24]
    return f"chat_cache::{prefix}::{digest}"


def chat_cache_get(
    cache_key: str,
    cb_url: str, bucket: str, username: str, password: str, use_tls: bool,
    scope: str, collection: str,
) -> dict | None:
    """Fetch a cached doc from Couchbase; returns None on miss or any error."""
    if not _CB_AVAILABLE:
        return None
    try:
        conn_str = _cb_conn_str(cb_url, use_tls)
        cluster  = Cluster(conn_str, ClusterOptions(PasswordAuthenticator(username, password)))
        cluster.wait_until_ready(timedelta(seconds=10))
        col      = cluster.bucket(bucket).scope(scope).collection(collection)
        result   = col.get(cache_key)
        cluster.close()
        return result.content_as[dict]
    except Exception:
        return None


def chat_cache_set(
    cache_key: str,
    doc: dict,
    ttl_seconds: int,
    cb_url: str, bucket: str, username: str, password: str, use_tls: bool,
    scope: str, collection: str,
) -> None:
    """Upsert a cache doc into Couchbase.
    ttl_seconds=0 → permanent (no expiry).  Silently ignores errors."""
    if not _CB_AVAILABLE:
        return
    try:
        conn_str = _cb_conn_str(cb_url, use_tls)
        cluster  = Cluster(conn_str, ClusterOptions(PasswordAuthenticator(username, password)))
        cluster.wait_until_ready(timedelta(seconds=10))
        col = cluster.bucket(bucket).scope(scope).collection(collection)
        opts = UpsertOptions(expiry=timedelta(seconds=ttl_seconds)) if ttl_seconds > 0 else UpsertOptions()
        col.upsert(cache_key, doc, opts)
        cluster.close()
    except Exception:
        pass


def _chat_cache_delete_by_prefix(
    prefix: str,
    cb_url: str, bucket: str, username: str, password: str, use_tls: bool,
    scope: str, collection: str,
) -> int:
    """Delete all docs whose key starts with prefix. Returns count deleted."""
    if not _CB_AVAILABLE:
        return 0
    try:
        conn_str = _cb_conn_str(cb_url, use_tls)
        cluster  = Cluster(conn_str, ClusterOptions(PasswordAuthenticator(username, password)))
        cluster.wait_until_ready(timedelta(seconds=10))
        keyspace = f"`{bucket}`.`{scope}`.`{collection}`"
        rows = list(cluster.query(
            f"SELECT META(t).id AS k FROM {keyspace} AS t "
            f"WHERE META(t).id LIKE $1",
            QueryOptions(
                positional_parameters=[f"{prefix}%"],
                timeout=timedelta(seconds=30),
            ),
        ))
        col = cluster.bucket(bucket).scope(scope).collection(collection)
        deleted = 0
        for row in rows:
            try:
                col.remove(row["k"])
                deleted += 1
            except Exception:
                pass
        cluster.close()
        return deleted
    except Exception:
        return 0


def chat_cache_clear(
    cb_url: str, bucket: str, username: str, password: str, use_tls: bool,
    scope: str, collection: str,
) -> int:
    """Delete embed + search cache entries (chat_cache::embed::* and chat_cache::search::*)."""
    n  = _chat_cache_delete_by_prefix("chat_cache::embed::", cb_url, bucket, username, password, use_tls, scope, collection)
    n += _chat_cache_delete_by_prefix("chat_cache::search::", cb_url, bucket, username, password, use_tls, scope, collection)
    return n


def chat_memory_clear(
    cb_url: str, bucket: str, username: str, password: str, use_tls: bool,
    scope: str, collection: str,
) -> int:
    """Delete permanent memory summary entries (chat_cache::memory::*)."""
    return _chat_cache_delete_by_prefix("chat_cache::memory::", cb_url, bucket, username, password, use_tls, scope, collection)


def _ensure_chat_history_col(bkt) -> None:
    """Create chat scope + history collection if missing. Ignores errors."""
    try:
        from couchbase.management.collections import CollectionSpec
        cm = bkt.collections()
        existing = {s.name: {c.name for c in s.collections} for s in cm.get_all_scopes()}
        if "chat" not in existing:
            cm.create_scope("chat")
        if "history" not in existing.get("chat", set()):
            cm.create_collection(CollectionSpec("history", scope_name="chat"))
    except Exception:
        pass


def _history_key(customer: str) -> str:
    return "history::" + (customer or "__all__").strip().lower().replace(" ", "_")


def save_customer_chat_history(
    customer: str,
    history: list[dict],
    cb_url: str, bucket: str, username: str, password: str, use_tls: bool,
) -> None:
    """Persist per-customer NiceGUI/shared chat history in the `chat.history` CB collection."""
    if not _CB_AVAILABLE or not history:
        return
    try:
        conn_str = _cb_conn_str(cb_url, use_tls)
        cluster = Cluster(conn_str, ClusterOptions(PasswordAuthenticator(username, password)))
        cluster.wait_until_ready(timedelta(seconds=10))
        bkt = cluster.bucket(bucket)
        _ensure_chat_history_col(bkt)
        bkt.scope("chat").collection("history").upsert(_history_key(customer), {
            "customer": customer or "",
            "updated": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
            "messages": history,
        })
        cluster.close()
    except Exception:
        pass


def load_customer_chat_history(
    customer: str,
    cb_url: str, bucket: str, username: str, password: str, use_tls: bool,
) -> list[dict]:
    """Load per-customer shared chat history from the `chat.history` CB collection."""
    if not _CB_AVAILABLE:
        return []
    try:
        conn_str = _cb_conn_str(cb_url, use_tls)
        cluster = Cluster(conn_str, ClusterOptions(PasswordAuthenticator(username, password)))
        cluster.wait_until_ready(timedelta(seconds=10))
        bkt = cluster.bucket(bucket)
        try:
            doc = bkt.scope("chat").collection("history").get(_history_key(customer)).content_as[dict]
            cluster.close()
            return doc.get("messages", [])
        except Exception:
            cluster.close()
            return []
    except Exception:
        return []


def save_chat_session(
    session_id: str,
    turns: list[tuple],
    organization: str,
    cb_url: str, bucket: str, username: str, password: str, use_tls: bool,
    scope: str, collection: str,
) -> None:
    """Persist a conversation session with ticket IDs and topic tags for staleness-aware recall.

    Each turn is (question, answer, timestamp_iso, ticket_ids) where ticket_ids is a
    list[str] of ticket IDs that were in context for that turn (empty list for turns
    with no retrieval, e.g. Batch mode).  Old 3-tuple turns are accepted for backward
    compatibility.  Stored under ``chat_cache::session::<session_id>``.
    """
    if not _CB_AVAILABLE or not turns:
        return
    try:
        # Collect all ticket IDs touched across the session (union, order-preserved)
        _all_ticket_ids: list[str] = []
        _seen_ids: set[str] = set()
        interleaved: list[dict] = []
        for turn in turns:
            q, a, ts = turn[0], turn[1], turn[2]
            t_ids: list[str] = list(turn[3]) if len(turn) > 3 else []
            interleaved.append({"role": "user",      "content": q,  "timestamp": ts})
            interleaved.append({"role": "assistant", "content": a,  "timestamp": ts,
                                 "ticket_ids": t_ids})
            for tid in t_ids:
                if tid and tid not in _seen_ids:
                    _all_ticket_ids.append(tid)
                    _seen_ids.add(tid)

        # Build topic tags: lower-case non-stopword words from questions (≥5 chars)
        _stop = {"what", "which", "where", "there", "their", "these", "those", "about",
                 "would", "could", "should", "please", "many", "have", "that", "with",
                 "from", "this", "were", "show", "list", "give", "tell", "across"}
        _tags: list[str] = []
        for turn in turns:
            for w in re.findall(r"[a-z][a-z0-9]{4,}", turn[0].lower()):
                if w not in _stop and w not in _tags:
                    _tags.append(w)
        topic_tags = _tags[:30]

        _now_epoch = int(time.time())
        doc = {
            "type":           "chat_session",
            "session_id":     session_id,
            "organization":   organization,
            "turn_count":     len(turns),
            "turns":          interleaved,
            "ticket_ids":     _all_ticket_ids,    # union of all tickets touched
            "topic_tags":     topic_tags,
            "started_at":     turns[0][2] if turns else "",
            "ended_at":       turns[-1][2] if turns else "",
            "last_active_at": _now_epoch,          # epoch seconds — staleness anchor
            "created_at":     time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
        }
        chat_cache_set(
            f"chat_cache::session::{session_id}",
            doc,
            0,   # permanent
            cb_url, bucket, username, password, use_tls, scope, collection,
        )
    except Exception:
        pass


def check_freshness(
    ticket_ids: list[str],
    since_epoch: int,
    cb_url: str, bucket: str, username: str, password: str, use_tls: bool,
    scope: str, collection: str,
) -> list[str]:
    """Return ticket IDs where last_scraped_at > since_epoch.

    Uses a single N1QL query so it scales to hundreds of IDs without per-doc
    round-trips.  Returns an empty list when CB is unavailable or on error.
    This is the staleness signal that prevents the LLM from reasoning about
    stale ticket state carried over from a prior session.
    """
    if not _CB_AVAILABLE or not ticket_ids:
        return []
    try:
        conn_str = _cb_conn_str(cb_url, use_tls)
        cluster  = Cluster(conn_str, ClusterOptions(PasswordAuthenticator(username, password)))
        cluster.wait_until_ready(timedelta(seconds=10))
        keyspace = f"`{bucket}`.`{scope}`.`{collection}`"
        id_list  = ", ".join(json.dumps(str(tid)) for tid in ticket_ids[:500])
        rows = list(cluster.query(
            f"SELECT META(t).id AS doc_key, t.ticket_id, t.last_scraped_at "
            f"FROM {keyspace} AS t "
            f"WHERE t.ticket_id IN [{id_list}] "
            f"AND t.last_scraped_at > $1",
            QueryOptions(positional_parameters=[since_epoch], timeout=timedelta(seconds=15)),
        ))
        cluster.close()
        stale: list[str] = []
        for row in rows:
            tid = str(row.get("ticket_id") or "").strip()
            if tid and tid not in stale:
                stale.append(tid)
        return stale
    except Exception:
        return []


def fetch_prior_session_context(
    organization: str,
    cb_url: str, bucket: str, username: str, password: str, use_tls: bool,
    scope: str, collection: str,
    limit: int = 3,
) -> str:
    """Build a brief prior-session context block for system prompt injection.

    Fetches the most recent sessions for the customer, runs check_freshness on
    each session's ticket IDs, and flags which tickets have been updated since
    that session ended.  Returns an empty string when there are no prior sessions
    or CB is unavailable.

    The block is intentionally short — it gives the model topic continuity and
    staleness signals without re-injecting stale ticket content.
    """
    sessions = fetch_recent_sessions(
        cb_url, bucket, username, password, use_tls, scope, collection,
        limit=limit, organization=organization,
    )
    if not sessions:
        return ""

    lines: list[str] = ["## Prior Session Context"]
    for sess in sessions:
        _dt    = (sess.get("ended_at") or sess.get("created_at") or "")[:10]
        _tc    = sess.get("turn_count", 0)
        _tags  = ", ".join((sess.get("topic_tags") or [])[:8])
        _tids  = sess.get("ticket_ids") or []
        _since = sess.get("last_active_at") or 0

        _stale: list[str] = []
        if _tids and _since:
            _stale = check_freshness(
                _tids, _since,
                cb_url, bucket, username, password, use_tls, scope, collection,
            )

        _line = f"Session {_dt} ({_tc} turns)"
        if _tags:
            _line += f" — topics: {_tags}"
        if _tids:
            _line += f" — {len(_tids)} tickets in scope"
        if _stale:
            _line += f" — ⚠ {len(_stale)} ticket(s) updated since session: {', '.join(_stale[:10])}"
        lines.append(_line)

        # Include a one-sentence distillation of the last assistant turn if available
        _turns = sess.get("turns") or []
        _last_a = next(
            (t["content"][:200] for t in reversed(_turns) if t.get("role") == "assistant"),
            None,
        )
        if _last_a:
            lines.append(f"  Last answer: {_last_a}{'…' if len(_last_a) == 200 else ''}")

    return "\n".join(lines)


def fetch_recent_sessions(
    cb_url: str, bucket: str, username: str, password: str, use_tls: bool,
    scope: str, collection: str,
    limit: int = 5,
    organization: str = "",
) -> list[dict]:
    """Return the most recent conversation session docs for context injection."""
    if not _CB_AVAILABLE:
        return []
    try:
        conn_str = _cb_conn_str(cb_url, use_tls)
        cluster  = Cluster(conn_str, ClusterOptions(PasswordAuthenticator(username, password)))
        cluster.wait_until_ready(timedelta(seconds=10))
        keyspace = f"`{bucket}`.`{scope}`.`{collection}`"
        org_filter = "AND (t.organization = $1 OR t.organization IS MISSING)" if organization else ""
        params = [organization] if organization else []
        rows = list(cluster.query(
            f"SELECT t.* FROM {keyspace} AS t "
            f"WHERE META(t).id LIKE 'chat_cache::session::%' {org_filter} "
            f"ORDER BY t.created_at DESC LIMIT {limit}",
            QueryOptions(positional_parameters=params, timeout=timedelta(seconds=15)),
        ))
        cluster.close()
        return rows
    except Exception:
        return []


def fetch_chat_memories(
    cb_url: str, bucket: str, username: str, password: str, use_tls: bool,
    scope: str, collection: str,
    limit: int = 50,
    organization: str = "",
) -> list[dict]:
    """
    Return permanent chat memory docs, optionally filtered by organization.
    Fetches up to `limit` most recent docs; caller scores by relevance.
    Returns all fields (including question_vector for semantic ranking).
    """
    if not _CB_AVAILABLE:
        return []
    try:
        conn_str = _cb_conn_str(cb_url, use_tls)
        cluster  = Cluster(conn_str, ClusterOptions(PasswordAuthenticator(username, password)))
        cluster.wait_until_ready(timedelta(seconds=10))
        keyspace = f"`{bucket}`.`{scope}`.`{collection}`"
        if organization:
            rows = list(cluster.query(
                f"SELECT t.* FROM {keyspace} AS t "
                f"WHERE META(t).id LIKE 'chat_cache::memory::%' "
                f"AND (t.organization = $1 OR t.organization IS MISSING) "
                f"ORDER BY t.created_at DESC LIMIT {limit}",
                QueryOptions(positional_parameters=[organization], timeout=timedelta(seconds=15)),
            ))
        else:
            rows = list(cluster.query(
                f"SELECT t.* FROM {keyspace} AS t "
                f"WHERE META(t).id LIKE 'chat_cache::memory::%' "
                f"ORDER BY t.created_at DESC LIMIT {limit}",
                QueryOptions(timeout=timedelta(seconds=15)),
            ))
        cluster.close()
        return rows   # newest-first; fetch_relevant_memories will re-rank
    except Exception:
        return []


def fetch_relevant_memories(
    query_vec: list[float],
    cb_url: str, bucket: str, username: str, password: str, use_tls: bool,
    scope: str, collection: str,
    top_k: int = 5,
    organization: str = "",
) -> list[dict]:
    """
    Retrieve the top-k most semantically relevant chat memories for the current question.
    Fetches all memories from CB then cosine-scores in Python (memory counts are small).
    Falls back to recency order when memories have no question_vector stored.
    Returns oldest-first for chronological prompt ordering.
    """
    all_memories = fetch_chat_memories(
        cb_url, bucket, username, password, use_tls, scope, collection,
        limit=200, organization=organization,
    )
    if not all_memories:
        return []

    def _cosine(a: list[float], b: list[float]) -> float:
        dot = sum(x * y for x, y in zip(a, b))
        na  = sum(x * x for x in a) ** 0.5
        nb  = sum(x * x for x in b) ** 0.5
        return dot / (na * nb) if na and nb else 0.0

    has_vectors = any(m.get("question_vector") for m in all_memories)
    if has_vectors and query_vec:
        scored = sorted(
            all_memories,
            key=lambda m: _cosine(query_vec, m["question_vector"]) if m.get("question_vector") else -1.0,
            reverse=True,
        )
    else:
        # Fallback: recency order (already newest-first from query)
        scored = all_memories

    top = scored[:top_k]
    # Return chronological order so the prompt reads naturally
    return list(reversed(top))


def fetch_existing_ticket_ids_from_cb(
    cb_url: str, bucket: str, username: str, password: str, use_tls: bool,
    scope: str, collection: str,
) -> set[str]:
    """Return the set of ticket IDs (numeric string) already stored in Couchbase."""
    if not _CB_AVAILABLE:
        return set()
    try:
        conn_str = _cb_conn_str(cb_url, use_tls)
        cluster  = Cluster(conn_str, ClusterOptions(PasswordAuthenticator(username, password)))
        cluster.wait_until_ready(timedelta(seconds=15))
        keyspace = f"`{bucket}`.`{scope}`.`{collection}`"
        rows = list(cluster.query(
            f"SELECT RAW META(t).id FROM {keyspace} AS t WHERE META(t).id LIKE 'ticket::%'",
            QueryOptions(timeout=timedelta(seconds=30)),
        ))
        cluster.close()
        # Keys are "ticket::12345" — extract the numeric part
        return {str(k).split("::")[-1] for k in rows if k}
    except Exception:
        return set()


def fetch_ticket_signals_from_cb(
    cb_url: str, bucket: str, username: str, password: str, use_tls: bool,
    scope: str, collection: str,
) -> dict[str, dict]:
    """Return {ticket_id: {status, solved, priority, last_scraped_at, is_stub,
    has_score, has_embedding, sfdc_matched, _deleted}} for all stored tickets.

    Extended signals power the smart_refresh diff:
      - priority:        detect priority escalations from the listing
      - last_scraped_at: proxy for updated_at (Supportal listing has no updated_at)
      - has_score:       enrichment gap — ticket needs LLM scoring
      - has_embedding:   enrichment gap — ticket needs vector embedding
      - sfdc_matched:    enrichment gap — SFDC correlation never run
    """
    if not _CB_AVAILABLE:
        return {}
    try:
        conn_str = _cb_conn_str(cb_url, use_tls)
        cluster  = Cluster(conn_str, ClusterOptions(PasswordAuthenticator(username, password)))
        cluster.wait_until_ready(timedelta(seconds=15))
        keyspace = f"`{bucket}`.`{scope}`.`{collection}`"
        rows = list(cluster.query(
            f"SELECT META(t).id AS doc_id, t.status, t.solved, t.priority, "
            f"t.last_scraped_at, t.requester, t.`_deleted`, "
            f"(t.score IS NOT MISSING AND t.score IS NOT NULL) AS has_score, "
            f"(t.embedding IS NOT MISSING AND t.embedding IS NOT NULL) AS has_embedding, "
            f"t.sfdc_matched "
            f"FROM {keyspace} AS t WHERE META(t).id LIKE 'ticket::%'",
            QueryOptions(timeout=timedelta(seconds=30)),
        ))
        cluster.close()
        _now = time.time()
        signals: dict[str, dict] = {}
        for row in rows:
            tid = str(row.get("doc_id", "")).split("::")[-1]
            if tid:
                lsa = row.get("last_scraped_at") or 0
                signals[tid] = {
                    "status":          row.get("status"),
                    "solved":          row.get("solved"),
                    "priority":        (row.get("priority") or "").lower().strip(),
                    "last_scraped_at": lsa,
                    "age_hours":       (_now - lsa) / 3600 if lsa else None,
                    "is_stub":         not row.get("requester") and not row.get("_deleted"),
                    "has_score":       bool(row.get("has_score")),
                    "has_embedding":   bool(row.get("has_embedding")),
                    "sfdc_matched":    row.get("sfdc_matched"),  # None = never run
                    "_deleted":        bool(row.get("_deleted")),
                }
        return signals
    except Exception:
        return {}


def _reconcile_deleted_tickets(
    scraped_ids: set[str],
    customer: str,
    cb_url: str, bucket: str, username: str, password: str,
    use_tls: bool, scope: str, collection: str,
    progress_cb=None,
) -> tuple[int, int]:
    """Delete CB ticket docs for the given org that are absent from scraped_ids.

    Returns (deleted_count, error_count).
    Only touches docs whose LOWER(organization) matches customer — never
    deletes tickets belonging to other orgs.
    """
    if not _CB_AVAILABLE:
        return 0, 0
    try:
        from couchbase.cluster import Cluster as _Cl
        from couchbase.options import ClusterOptions as _CO
        from couchbase.auth import PasswordAuthenticator as _PA
        from couchbase.options import QueryOptions as _QO
        from datetime import timedelta as _td
        conn   = _cb_conn_str(cb_url, use_tls)
        _clust = _Cl(conn, _CO(_PA(username, password)))
        _clust.wait_until_ready(_td(seconds=15))
        keyspace = f"`{bucket}`.`{scope}`.`{collection}`"
        cust_pat = customer.lower().replace("'", "''")
        rows = list(_clust.query(
            f"SELECT RAW META(t).id FROM {keyspace} AS t "
            f"WHERE META(t).id LIKE 'ticket::%' "
            f"AND LOWER(t.organization) LIKE '%{cust_pat}%'",
            _QO(timeout=_td(seconds=60)),
        ))
        cb_ids = {str(k).split("::")[-1] for k in rows if k}
        orphans = cb_ids - scraped_ids
        if not orphans:
            return 0, 0
        col     = _clust.bucket(bucket).scope(scope).collection(collection)
        deleted = 0
        errors  = 0
        total   = len(orphans)
        for i, tid in enumerate(sorted(orphans), 1):
            try:
                col.remove(f"ticket::{tid}")
                deleted += 1
            except Exception:
                errors += 1
            if progress_cb and i % 10 == 0:
                progress_cb(f"Removed {deleted}/{total}…", i / total)
        _clust.close()
        return deleted, errors
    except Exception:
        return 0, 1


def _filter_changed_tickets(
    listing_summaries: list[dict],
    stored_signals: dict[str, dict],
    max_tickets: int = 0,
    stale_open_hours: float = 4.0,
) -> tuple[list[dict], int, int, int]:
    """Partition listing summaries into scrape vs skip based on change signals.

    Change reasons detected (in priority order):
      1. New — ticket ID not in CB at all
      2. Status changed — listing status ≠ CB status
      3. Solved date changed — ticket was closed/reopened
      4. Priority escalated — listing priority ≠ CB priority (e.g. P3→P1)
      5. Stub — CB record has no requester (detail page never fetched successfully)
      6. Stale-open — open/pending ticket scraped > stale_open_hours ago (proxy for
         updated_at since the listing API does not expose updated_at)

    Priority order for the returned list: new → status/solved/priority changed →
    stub → stale-open. Newest ticket IDs break ties within each group.

    Returns:
        (to_scrape, n_new, n_changed, n_skipped)
        to_scrape is truncated to max_tickets if max_tickets > 0.
    """
    _OPEN_STATUSES = {"open", "pending", "on-hold", "hold", "new"}

    new_tickets:      list[dict] = []
    changed_tickets:  list[dict] = []
    stale_open:       list[dict] = []
    skipped = 0

    for s in listing_summaries:
        tid = str(s.get("ticket_id", ""))
        if tid not in stored_signals:
            new_tickets.append(s)
            continue
        stored = stored_signals[tid]
        # Never re-scrape permanently deleted tickets
        if stored.get("_deleted"):
            skipped += 1
            continue

        listing_status   = (s.get("status") or "").strip().lower()
        stored_status    = (stored.get("status") or "").strip().lower()
        listing_solved   = (s.get("solved") or "").strip()
        stored_solved    = (stored.get("solved") or "").strip()
        listing_priority = (s.get("priority") or "").strip().lower()
        stored_priority  = (stored.get("priority") or "").strip().lower()
        is_stub          = stored.get("is_stub", False)
        age_h            = stored.get("age_hours")

        if (listing_status != stored_status
                or listing_solved != stored_solved
                or (listing_priority and listing_priority != stored_priority)
                or is_stub):
            # Tag the stub with what changed for reporting
            s["_change_reason"] = (
                "stub" if is_stub else
                f"status {stored_status}→{listing_status}" if listing_status != stored_status else
                f"priority {stored_priority}→{listing_priority}" if listing_priority != stored_priority else
                "solved date changed"
            )
            changed_tickets.append(s)
        elif listing_status in _OPEN_STATUSES and age_h is not None and age_h >= stale_open_hours:
            s["_change_reason"] = f"stale open ({age_h:.1f}h since last scrape)"
            stale_open.append(s)
        else:
            skipped += 1

    def _tid_num(s: dict) -> int:
        tid = str(s.get("ticket_id") or "")
        return int(tid) if tid.isdigit() else 0

    for group in (new_tickets, changed_tickets, stale_open):
        group.sort(key=_tid_num, reverse=True)

    to_scrape = new_tickets + changed_tickets + stale_open
    if max_tickets > 0:
        to_scrape = to_scrape[:max_tickets]

    n_changed_total = len(changed_tickets) + len(stale_open)
    return to_scrape, len(new_tickets), n_changed_total, skipped


def fetch_snapshot_signals_from_cb(
    cb_url: str, bucket: str, username: str, password: str, use_tls: bool,
    scope: str, snap_collection: str,
) -> dict[str, dict]:
    """Return {snap_id: {complete: bool}} for all stored snapshots.

    A snapshot is considered complete if cb_version is present and non-empty.
    ticket_ids is not used as a completeness signal — snapshots without tickets
    are valid and should not be re-scraped.
    """
    if not _CB_AVAILABLE:
        return {}
    try:
        conn_str = _cb_conn_str(cb_url, use_tls)
        cluster  = Cluster(conn_str, ClusterOptions(PasswordAuthenticator(username, password)))
        cluster.wait_until_ready(timedelta(seconds=15))
        keyspace = f"`{bucket}`.`{scope}`.`{snap_collection}`"
        rows = list(cluster.query(
            f"SELECT META(s).id AS doc_id, s.cb_version "
            f"FROM {keyspace} AS s WHERE META(s).id LIKE 'snapshot::%'",
            QueryOptions(timeout=timedelta(seconds=30)),
        ))
        cluster.close()
        signals: dict[str, dict] = {}
        for row in rows:
            snap_id = str(row.get("doc_id", "")).split("snapshot::")[-1]
            if snap_id:
                signals[snap_id] = {
                    "complete": bool(row.get("cb_version")),
                }
        return signals
    except Exception:
        return {}


def _filter_incomplete_snapshots(
    stubs: list[dict],
    stored_signals: dict[str, dict],
    max_snapshots: int = 0,
) -> tuple[list[dict], int, int, int]:
    """Partition snapshot stubs into scrape vs skip based on CB completeness.

    - New (snap_id not in CB): always scrape
    - Incomplete (in CB but cb_version missing): re-scrape
    - Complete (in CB with cb_version): skip

    Returns:
        (to_scrape, n_new, n_incomplete, n_skipped)
        to_scrape is truncated to max_snapshots if max_snapshots > 0.
    """
    new_snaps:        list[dict] = []
    incomplete_snaps: list[dict] = []
    skipped = 0

    for s in stubs:
        snap_id = s.get("snap_id", "")
        if snap_id not in stored_signals:
            new_snaps.append(s)
        elif not stored_signals[snap_id].get("complete"):
            incomplete_snaps.append(s)
        else:
            skipped += 1

    to_scrape = new_snaps + incomplete_snaps
    if max_snapshots > 0:
        to_scrape = to_scrape[:max_snapshots]

    return to_scrape, len(new_snaps), len(incomplete_snaps), skipped


def query_cluster_topology(cluster_uuid: str, snapshot_idx: int | None = None) -> dict:
    """Fetch cluster topology from Supportal nutshellresults for a given cluster UUID.

    Assembles node layout (services, RAM, disk, CB version) via three split queries
    to stay within the analytics API's 2-distinct-r-*-keys-per-query limit.
    CPU count is omitted — r-cpus-available causes a 500 in the analytics index
    (tracked with Supportal team; re-add when fixed).

    Args:
        cluster_uuid:  Cluster UUID (matches snapshot.uuid / cluster.uuid).
        snapshot_idx:  Snapshot sequence index. Defaults to the latest available.

    Returns dict with keys:
        cluster_uuid, snapshot_idx, snapshot_timestamp, node_count,
        nodes (list of per-node dicts), missing_fields (list of unavailable signals).
    """
    from supportal.api_client import query_supportal_analytics as _qsa

    uuid_clean = cluster_uuid.strip()

    # ── Step 1: resolve snapshot index and detect format ─────────────────────
    # New format: Nutshell::uuid::idx  / Nutshell::uuid::idx::Node::hostname
    # Old format: Nutshell::uuid::Cluster / Nutshell::uuid::Node::hostname
    old_format = False
    if snapshot_idx is None:
        id_rows = _qsa(
            f"SELECT META(n).id FROM nutshellresults n "
            f"WHERE META(n).id LIKE 'Nutshell::{uuid_clean}::%' "
            f"AND META(n).id NOT LIKE '%::Node::%' "
            f"ORDER BY META(n).id DESC LIMIT 1"
        )
        if not id_rows:
            return {"error": f"No nutshell results found for cluster {uuid_clean}"}
        latest_id = id_rows[0].get("id", "")
        parts = latest_id.split("::")
        if len(parts) >= 3 and parts[2] == "Cluster":
            old_format = True
            snapshot_idx = 0  # sentinel; not used in prefix below
        else:
            try:
                snapshot_idx = int(parts[2]) if len(parts) >= 3 else 0
            except (ValueError, IndexError):
                snapshot_idx = 0

    if old_format:
        snap_prefix  = f"Nutshell::{uuid_clean}::Cluster"
        node_pattern = f"Nutshell::{uuid_clean}::Node::%"
    else:
        snap_prefix  = f"Nutshell::{uuid_clean}::{snapshot_idx}"
        node_pattern = f"{snap_prefix}::Node::%"

    # ── Step 2: node hostname map from cluster-level doc ──────────────────────
    cluster_rows = _qsa(
        f"SELECT META(n).id, n.nodes FROM nutshellresults n "
        f"WHERE META(n).id = '{snap_prefix}' LIMIT 1"
    )
    node_map: dict = cluster_rows[0].get("nodes") or {} if cluster_rows else {}

    # ── Step 3: three split queries (≤2 distinct r-* keys each) ──────────────
    try:
        rows_a = _qsa(                              # services + RAM
            f"SELECT META(n).id, "
            f"n.results.`r-mds-services-node`.values AS services, "
            f"n.results.`r-memory-limit`.bytes AS ram "
            f"FROM nutshellresults n WHERE META(n).id LIKE '{node_pattern}' LIMIT 50"
        )
    except Exception:
        rows_a = []

    try:
        rows_b = _qsa(                              # disk (two sub-fields, one r-* key)
            f"SELECT META(n).id, "
            f"n.results.`r-data-directory`.directories[0].size AS disk_bytes, "
            f"n.results.`r-data-directory`.directories[0].used_percent AS disk_pct "
            f"FROM nutshellresults n WHERE META(n).id LIKE '{node_pattern}' LIMIT 50"
        )
    except Exception:
        rows_b = []

    try:
        rows_c = _qsa(                              # version
            f"SELECT META(n).id, n.results.`r-cbs-version`.version AS version "
            f"FROM nutshellresults n WHERE META(n).id LIKE '{node_pattern}' LIMIT 50"
        )
    except Exception:
        rows_c = []

    # ── Step 4: snapshot timestamp ────────────────────────────────────────────
    try:
        ts_rows = _qsa(
            f"SELECT s.timestamp FROM snapshot s "
            f"WHERE META(s).id = 'Snapshot::{uuid_clean}::{snapshot_idx}' LIMIT 1"
        )
        snapshot_timestamp = ts_rows[0].get("timestamp") if ts_rows else None
    except Exception:
        snapshot_timestamp = None

    # ── Step 5: merge by node hostname ───────────────────────────────────────
    def _hostname(doc_id: str) -> str:
        parts = doc_id.split("::Node::")
        return parts[1] if len(parts) == 2 else ""

    merged: dict[str, dict] = {}

    for row in rows_a:
        h = _hostname(row.get("id", ""))
        if not h:
            continue
        merged.setdefault(h, {"hostname": h})
        if row.get("services") is not None:
            merged[h]["services"] = sorted(row["services"])
        if row.get("ram") is not None:
            merged[h]["ram_gb"] = round(row["ram"] / 1_073_741_824, 1)

    for row in rows_b:
        h = _hostname(row.get("id", ""))
        if not h:
            continue
        merged.setdefault(h, {"hostname": h})
        if row.get("disk_bytes") is not None:
            merged[h]["disk_gb"] = round(row["disk_bytes"] / 1_073_741_824, 0)
        if row.get("disk_pct") is not None:
            merged[h]["disk_used_pct"] = row["disk_pct"]

    for row in rows_c:
        h = _hostname(row.get("id", ""))
        if not h:
            continue
        merged.setdefault(h, {"hostname": h})
        if row.get("version") is not None:
            merged[h]["version"] = row["version"]

    # seed any hostnames from the cluster-level map not seen in query results
    for h in node_map:
        merged.setdefault(h, {"hostname": h})

    nodes = sorted(merged.values(), key=lambda n: n["hostname"])

    # ── Step 6: report which fields have zero coverage ────────────────────────
    missing = ["cpus"]  # always missing — analytics API bug, tracked with Supportal
    for field in ("services", "ram_gb", "disk_gb", "version"):
        if not any(field in n for n in nodes):
            missing.append(field)

    return {
        "cluster_uuid":       uuid_clean,
        "snapshot_idx":       snapshot_idx,
        "snapshot_timestamp": snapshot_timestamp,
        "node_count":         len(nodes),
        "nodes":              nodes,
        "missing_fields":     missing,
    }


# ─────────────────────────── Constants ────────────────────────────────────────

# (constants moved to supportal/constants.py)


# ── Settings persistence ───────────────────────────────────────────────────────

def _load_settings_file() -> dict:
    """Return all saved profiles as a dict keyed by profile name."""
    if SETTINGS_FILE.exists():
        try:
            return json.loads(SETTINGS_FILE.read_text())
        except Exception:
            pass
    return {}


def _save_settings_file(profiles: dict) -> None:
    SETTINGS_FILE.write_text(json.dumps(profiles, indent=2))


def _get_profile_cookie() -> str:
    """Return the session cookie from the active saved profile, or empty string."""
    try:
        s = _load_settings_file()
        active = s.get("__last__", "")
        return s.get(active, {}).get("cookie", "") if active else ""
    except Exception:
        return ""


# Module-level state — safe because NiceGUI runs as a single process.
# ── Server-level persistent state ────────────────────────────────────────────
# Survives browser refreshes and WebSocket reconnects.  Written by operations
# (scrape/score/embed) so a new page load can restore results instead of starting
# blank.  Also tracks in-flight operation progress so a refreshed page can
# reconnect to a running scrape/score without orphaning threads.
_SERVER_STATE: dict = {
    "results":       [],    # last loaded/scraped tickets
    "scores":        {},    # last scored results
    "customer_name": "",    # last customer name
}

# Active operation progress — updated by background threads, polled by UI timers.
# A refreshed page creates a new timer that picks this up immediately.
_OP_STATUS: dict = {
    "op":       None,    # "scrape" | "score" | "embed" | None
    "status":   "",
    "progress": 0.0,
    "done":     True,
}

# Background scrape job registry — keyed by short job_id (6-char hex).
# Kept in insertion order; trimmed to MAX_SCRAPE_JOBS entries.
_SCRAPE_JOBS: dict[str, dict] = {}
_MAX_SCRAPE_JOBS = 20

# Cancel signals — set the event to request clean cancellation of a running job.
_JOB_CANCEL_EVENTS: dict[str, threading.Event] = {}

# Browser-login state — populated by login_browser.py subprocess.
_browser_state: dict = {
    "logged_in":     False,
    "cookie_string": "",
}

# Coordination events for the browser-login flow.
_browser_closed_event: threading.Event = threading.Event()
_browser_ready_event:  threading.Event = threading.Event()  # set when subprocess has started

# Shared results — written by worker thread, read by UI download handlers.
_results: list[dict] = []


# (ticket_parser block moved to supportal/ticket_parser.py)

# (api_client block moved to supportal/api_client.py)



def _parse_all_page_urls(html: str, base_url: str) -> list[str]:
    """
    Parse the numbered page list at the bottom of the ticket listing.
    Returns a sorted list of all distinct page URLs found (including any
    not yet visited), derived from page-number links.
    Handles both ?page=N query-param style and path-based /page/N style.
    """
    soup = BeautifulSoup(html, "html.parser")
    page_map: dict[int, str] = {}

    for a in soup.find_all("a", href=True):
        href = a["href"]
        text = a.get_text(strip=True)

        # Query-param style: ?page=N or &page=N
        m = re.search(r"[?&]page=(\d+)", href)
        if m:
            n = int(m.group(1))
            url = href if href.startswith("http") else urllib.parse.urljoin(base_url, href)
            page_map[n] = url
            continue

        # Path style: /page/N
        m = re.search(r"/page/(\d+)", href)
        if m:
            n = int(m.group(1))
            url = href if href.startswith("http") else urllib.parse.urljoin(base_url, href)
            page_map[n] = url
            continue

        # Link whose visible text is a plain number (e.g. "2", "3", ... "581")
        if text.isdigit():
            n = int(text)
            url = href if href.startswith("http") else urllib.parse.urljoin(base_url, href)
            page_map.setdefault(n, url)

    return [url for _, url in sorted(page_map.items())]


def _find_next_page(html: str, current_url: str) -> Optional[str]:
    """
    Fallback single-step next-page detector used when no full page list is found.
    """
    soup = BeautifulSoup(html, "html.parser")

    # rel="next"
    a = soup.find("a", attrs={"rel": "next"})
    if a and a.get("href"):
        return urllib.parse.urljoin(current_url, a["href"])

    # Text-based
    for a in soup.find_all("a", href=True):
        txt = (a.get_text() or "").strip().lower()
        if txt in {"next", "›", "»", "→", "older", "more", "next »", "next ›"}:
            return urllib.parse.urljoin(current_url, a["href"])

    # Increment page= param
    parsed = urllib.parse.urlparse(current_url)
    qs = urllib.parse.parse_qs(parsed.query)
    if "page" in qs:
        try:
            page = int(qs["page"][0])
            candidate = re.sub(r"([?&]page=)\d+", rf"\g<1>{page + 1}", current_url)
            if candidate != current_url:
                return candidate
        except (ValueError, IndexError):
            pass

    return None


# ─────────────────────────── Scraper workers (blocking — run in thread) ───────

DEBUG_DIR = os.path.join(os.path.dirname(__file__), "debug_html")


def _save_debug(name: str, url: str, html: str) -> str:
    """Save HTML to debug_html/<name>.html and return the file path."""
    os.makedirs(DEBUG_DIR, exist_ok=True)
    safe = re.sub(r"[^\w\-]", "_", name)
    path = os.path.join(DEBUG_DIR, f"{safe}.html")
    with open(path, "w", encoding="utf-8") as f:
        f.write(f"<!-- URL: {url} -->\n{html}")
    return path


def _scrape_listing(get_html: Callable[[str], str], customer: str, max_pages: int,
                    progress_cb: Callable[[str, float], None],
                    debug: bool = False) -> list[tuple[str, str]]:
    """
    Full navigation flow:
      1. Search Supportal for the customer name
      2. Find the correct Customer result (by alias / name match)
      3. Navigate to the customer page
      4. Find and follow the Tickets tab
      5. Paginate through the full numbered page list
      6. Collect every ticket link

    Set debug=True to save HTML at every step to ./debug_html/ for inspection.
    """
    all_tickets: list[tuple[str, str]] = []
    seen_urls: set[str] = set()
    visited_pages: set[str] = set()

    def log(msg: str, pct: float):
        print(f"[SCRAPE] {msg}")
        progress_cb(msg, pct)

    # ── Step 1: search ────────────────────────────────────────────────────────
    _cust_dec = urllib.parse.unquote(customer)
    search_url = f"{BASE_URL}/search/{urllib.parse.quote(_cust_dec)}"
    log(f"GET {search_url}", 0.01)
    search_html = get_html(search_url)
    if debug:
        p = _save_debug("01_search", search_url, search_html)
        log(f"Saved search HTML → {p}", 0.01)

    # Report what the page looks like
    soup_s = BeautifulSoup(search_html, "html.parser")
    page_title = (soup_s.find("title") or soup_s.find("h1") or "")
    page_title = page_title.get_text(strip=True) if hasattr(page_title, "get_text") else str(page_title)
    customer_links = [a["href"] for a in soup_s.find_all("a", href=True) if "/customer/" in a["href"]]
    log(f"Search page title: '{page_title}' | /customer/ links found: {len(customer_links)}", 0.015)
    if customer_links:
        log(f"  First few: {customer_links[:5]}", 0.015)

    # ── Step 2: find customer URL ─────────────────────────────────────────────
    customer_url = _find_customer_url_in_search(search_html, customer)
    if customer_url:
        log(f"Matched customer URL: {customer_url}", 0.02)
    else:
        slug = customer.lower().replace(" ", "-")
        customer_url = f"{BASE_URL}/customer/{urllib.parse.quote(slug, safe='')}"
        log(f"No search match — falling back to slug URL: {customer_url}", 0.02)

    # ── Step 3: load customer page ────────────────────────────────────────────
    log(f"GET customer page: {customer_url}", 0.03)
    customer_html = get_html(customer_url)
    if debug:
        p = _save_debug("02_customer", customer_url, customer_html)
        log(f"Saved customer HTML → {p}", 0.03)

    soup_c = BeautifulSoup(customer_html, "html.parser")
    ctitle = (soup_c.find("title") or soup_c.find("h1") or "")
    ctitle = ctitle.get_text(strip=True) if hasattr(ctitle, "get_text") else ""
    all_links = [(a.get_text(strip=True), a["href"]) for a in soup_c.find_all("a", href=True)]
    log(f"Customer page title: '{ctitle}' | total links: {len(all_links)}", 0.035)
    tab_like = [(t, h) for t, h in all_links if t.lower() in
                {"tickets", "entitlements", "snapshots", "clusters", "users", "analytics"}]
    log(f"  Tab-like links: {tab_like}", 0.035)

    # ── Step 4: find Tickets tab ──────────────────────────────────────────────
    tickets_url = _find_tickets_tab_url(customer_html, customer_url)
    if tickets_url:
        log(f"Tickets tab URL: {tickets_url}", 0.04)
    else:
        tickets_url = customer_url.rstrip("/")
        log(f"Tickets tab not found — using customer page directly: {tickets_url}", 0.04)

    # ── Step 5: load first tickets page ──────────────────────────────────────
    log(f"GET tickets page 1: {tickets_url}", 0.05)
    first_html = get_html(tickets_url)
    visited_pages.add(tickets_url)
    if debug:
        p = _save_debug("03_tickets_page1", tickets_url, first_html)
        log(f"Saved tickets page 1 HTML → {p}", 0.05)

    soup_t = BeautifulSoup(first_html, "html.parser")
    ttitle = (soup_t.find("title") or soup_t.find("h1") or "")
    ttitle = ttitle.get_text(strip=True) if hasattr(ttitle, "get_text") else ""
    ticket_links_found = _extract_ticket_links(first_html)
    all_page_hrefs = [(a.get_text(strip=True), a["href"])
                      for a in soup_t.find_all("a", href=True)
                      if re.search(r"[?&]page=\d+|/page/\d+", a["href"])]
    log(f"Tickets page title: '{ttitle}' | ticket links: {len(ticket_links_found)} | page links: {len(all_page_hrefs)}", 0.06)
    if all_page_hrefs:
        log(f"  Pagination sample: {all_page_hrefs[:6]}", 0.06)

    for item in ticket_links_found:
        if item[1] not in seen_urls:
            all_tickets.append(item)
            seen_urls.add(item[1])

    log(f"Page 1 collected: {len(all_tickets)} tickets", 0.07)

    # Parse full page list
    all_page_urls = _parse_all_page_urls(first_html, tickets_url)
    total_pages = len(all_page_urls) + 1

    if total_pages > 1:
        log(f"Pagination: {total_pages} total pages", 0.08)
    else:
        all_page_urls = []
        log("No numbered pagination found — will use next-page fallback", 0.08)

    # ── Step 6: iterate remaining pages ──────────────────────────────────────
    if all_page_urls:
        for i, page_url in enumerate(all_page_urls, start=2):
            if max_pages and i > max_pages:
                break
            if page_url in visited_pages:
                continue
            visited_pages.add(page_url)

            pct = 0.08 + 0.87 * (i / max(total_pages, 1))
            log(f"Listing page {i}/{total_pages}…", pct)
            html = get_html(page_url)

            for item in _extract_ticket_links(html):
                if item[1] not in seen_urls:
                    all_tickets.append(item)
                    seen_urls.add(item[1])

            log(f"Page {i}/{total_pages}: {len(all_tickets)} tickets total", pct)
            time.sleep(0.3)

    else:
        current_url = tickets_url
        current_html = first_html
        page_num = 1
        while True:
            if max_pages and page_num >= max_pages:
                break
            nxt = _find_next_page(current_html, current_url)
            if not nxt or nxt in visited_pages:
                break
            visited_pages.add(nxt)
            page_num += 1
            log(f"Next-page {page_num}: {nxt}", 0.08 + 0.02 * page_num)
            current_html = get_html(nxt)
            for item in _extract_ticket_links(current_html):
                if item[1] not in seen_urls:
                    all_tickets.append(item)
                    seen_urls.add(item[1])
            current_url = nxt
            time.sleep(0.3)

    return all_tickets, customer_url


def diagnose(get_html: Callable[[str], str], customer: str,
             progress_cb: Callable[[str, float], None]) -> str:
    """
    Run just the discovery steps (no ticket detail scraping) with debug=True.
    Returns a text report summarising what was found at each step.
    Saves HTML files to ./debug_html/ for manual inspection.
    """
    report_lines: list[str] = []

    def log_cb(msg: str, pct: float):
        report_lines.append(msg)
        progress_cb(msg, pct)

    _scrape_listing(get_html, customer, max_pages=1, progress_cb=log_cb, debug=True)  # returns (tickets, url); ignored here
    report_lines.append(f"\nDebug HTML saved to: {DEBUG_DIR}")
    return "\n".join(report_lines)


def scrape_with_cookie(
    cookie: str,
    customer: str,
    max_pages: int,
    progress_cb: Callable[[str, float], None],
    skip_ids: set | None = None,
    workers: int = 4,
) -> list[dict]:
    """Auth mode A: plain HTTP requests with a Cookie header.

    Ticket detail pages are fetched in parallel (``workers`` threads) using
    direct URLs (BASE_URL/zendesk/ticket/<id>).  A per-worker delay of 0.4 s
    keeps the request rate well below DoS thresholds.
    """
    import concurrent.futures
    import threading

    _headers = {
        "User-Agent": UA,
        "Accept": "text/html,application/xhtml+xml,*/*;q=0.9",
        "Cookie": cookie,
    }
    # Shared session for the listing phase (sequential, thread-safe for reads)
    listing_session = requests.Session()
    listing_session.headers.update(_headers)

    def get_html(url: str) -> str:
        resp = listing_session.get(url, timeout=30, allow_redirects=True)
        resp.raise_for_status()
        return resp.text

    all_tickets, customer_url = _scrape_listing(get_html, customer, max_pages, progress_cb)

    if skip_ids:
        to_scrape = [(tid, turl) for tid, turl in all_tickets if str(tid) not in skip_ids]
        skipped = len(all_tickets) - len(to_scrape)
        progress_cb(
            f"Incremental mode: {len(to_scrape)} new tickets, {skipped} skipped.",
            0.1,
        )
    else:
        to_scrape = all_tickets

    total = len(to_scrape)
    progress_cb(f"Fetching {total} ticket detail pages ({workers} parallel workers)…", 0.1)

    results: list[dict | None] = [None] * total
    completed = 0
    lock = threading.Lock()

    # Each worker thread gets its own Session to avoid sharing internal state
    _tls = threading.local()

    def _worker_session() -> requests.Session:
        if not getattr(_tls, "session", None):
            s = requests.Session()
            s.headers.update(_headers)
            _tls.session = s
        return _tls.session

    def _fetch_one(idx: int, tid: str, turl: str) -> None:
        nonlocal completed
        url = f"{BASE_URL}/zendesk/ticket/{tid}"  # always use canonical URL
        try:
            resp = _worker_session().get(url, timeout=30, allow_redirects=True, verify=False)
            if _is_deleted_ticket_page(resp.status_code, resp.text):
                rec = {"ticket_id": tid, "url": url, "_deleted": True}
            else:
                resp.raise_for_status()
                rec = parse_ticket_detail(resp.text, url)
        except Exception as exc:
            rec = {"ticket_id": tid, "url": url, "error": str(exc)}
        results[idx] = rec
        time.sleep(0.4)  # polite rate-limit per worker
        with lock:
            completed += 1
            pct = 0.1 + 0.89 * (completed / max(total, 1))
            progress_cb(f"[{completed}/{total}] Ticket #{tid}", pct)

    with concurrent.futures.ThreadPoolExecutor(max_workers=workers) as pool:
        futs = [pool.submit(_fetch_one, i, tid, turl)
                for i, (tid, turl) in enumerate(to_scrape)]
        concurrent.futures.wait(futs)

    progress_cb("Done", 1.0)
    out = [r for r in results if r is not None]
    for r in out:
        r.setdefault("customer_url", customer_url)
    return out


def scrape_with_cookie(
    cookie: str,
    customer: str,
    max_pages: int,
    progress_cb: Callable[[str, float], None],
    skip_ids: set | None = None,
    change_signals: dict | None = None,
    max_tickets: int = 0,
) -> list[dict]:
    """Scrape all tickets for a customer using REST APIs only (no browser required)."""
    customer = customer.strip().strip('"\'')
    session  = _make_api_session(cookie)
    customer_url = f"{BASE_URL}/customer/{urllib.parse.quote(customer, safe='')}"

    # ── Step 1: get ticket listing ─────────────────────────────────────────
    progress_cb("Fetching ticket listing via REST API…", 0.02)
    raw_listing = _get_customer_ticket_listing_api(customer, session, progress_cb)

    listing_summaries: list[dict] = []
    _skipped_deleted = 0
    for item in raw_listing:
        tid = str(item.get("id") or "").strip()
        if not tid:
            continue
        item_status = str(item.get("status") or "").lower()
        if item_status == "deleted":
            _skipped_deleted += 1
            continue
        listing_summaries.append({
            "ticket_id": tid,
            "url":       f"{BASE_URL}/zendesk/ticket/{tid}",
            "status":    item.get("status", ""),
            "priority":  item.get("Priority", ""),
            "subject":   item.get("subject", ""),
            "created":   item.get("created_at", ""),
            "solved":    item.get("solved_at", ""),
        })

    _deleted_note = f" ({_skipped_deleted} deleted skipped)" if _skipped_deleted else ""
    progress_cb(f"Listing complete — {len(listing_summaries)} tickets found{_deleted_note}. Fetching details…", 0.15)

    # ── Step 2: change detection / incremental filtering ──────────────────
    if change_signals is not None:
        listing_summaries, n_new, n_changed, n_skipped = _filter_changed_tickets(
            listing_summaries, change_signals, max_tickets
        )
        progress_cb(
            f"Change detection: {n_new} new, {n_changed} changed, {n_skipped} unchanged (skipped)"
            + (f", capped at {max_tickets}" if max_tickets > 0 else "") + ".",
            0.15,
        )
    elif skip_ids:
        all_count = len(listing_summaries)
        listing_summaries = [s for s in listing_summaries if str(s.get("ticket_id", "")) not in skip_ids]
        if max_tickets > 0:
            listing_summaries = listing_summaries[:max_tickets]
        skipped = all_count - len(listing_summaries)
        progress_cb(
            f"Incremental mode: {len(listing_summaries)} new tickets, {skipped} skipped.",
            0.15,
        )
    elif max_tickets > 0:
        listing_summaries = listing_summaries[:max_tickets]
        progress_cb(f"Capped at {max_tickets} tickets.", 0.15)

    # ── Step 3: fetch full ticket details via /status API ─────────────────
    total = len(listing_summaries)
    results: list[dict] = []
    for i, summary in enumerate(listing_summaries):
        tid = str(summary.get("ticket_id", ""))
        if not tid:
            continue
        pct = 0.15 + 0.84 * (i / max(total, 1))
        progress_cb(f"Detail {i + 1}/{total}  ticket #{tid}", pct)
        rec = fetch_ticket_api(tid, session)
        for field in ("status", "priority", "subject", "created", "solved"):
            if not rec.get(field) and summary.get(field):
                rec[field] = summary[field]
        rec.setdefault("customer_url", customer_url)
        results.append(rec)
        time.sleep(0.05)

    progress_cb("Done", 1.0)
    return results


def scrape_with_cookie_playwright(
    cookie: str,
    customer: str,
    max_pages: int,
    progress_cb: Callable[[str, float], None],
    skip_ids: set | None = None,
    change_signals: dict | None = None,
    max_tickets: int = 0,
) -> list[dict]:
    """Alias for scrape_with_cookie (Playwright no longer used)."""
    return scrape_with_cookie(cookie, customer, max_pages, progress_cb, skip_ids, change_signals, max_tickets)


def _scrape_listing_playwright(
    page,
    customer: str,
    max_pages: int,
    progress_cb: Callable[[str, float], None],
    debug: bool = False,
) -> tuple[list[dict], str]:
    """Stub — Playwright listing removed; callers should use scrape_with_cookie instead."""
    progress_cb("Playwright listing not available — use cookie auth.", 0.0)
    return [], f"{BASE_URL}/customer/{urllib.parse.quote(customer.strip(), safe='')}"


def scrape_with_playwright(
    customer: str,
    max_pages: int,
    progress_cb: Callable[[str, float], None],
    skip_ids: set | None = None,
    change_signals: dict | None = None,
    max_tickets: int = 0,
) -> list[dict]:
    """Auth mode B: uses saved browser session cookie for REST-based scraping."""
    cookie = _browser_state.get("cookie_string", "") or _get_profile_cookie()
    if not cookie:
        raise RuntimeError("No session cookie available — complete browser login first.")
    return scrape_with_cookie(cookie, customer, max_pages, progress_cb, skip_ids, change_signals, max_tickets)


_DELETED_TICKET_PHRASES = (
    "ticket has been deleted",
    "ticket was deleted",
    "this ticket no longer exists",
    "has been permanently deleted",
    "this record has been deleted",
    "record not found",
    "ticket not found",
    "no longer available",
)


def _is_deleted_ticket_page(status_code: int, html: str) -> bool:
    """Return True if the page indicates the ticket was deleted or does not exist."""
    if not html:
        return False
    if status_code not in (200, 404):
        return False
    text = BeautifulSoup(html, "html.parser").get_text(" ", strip=True).lower()
    if any(phrase in text for phrase in _DELETED_TICKET_PHRASES):
        return True
    # 404 pages with generic "not found" copy also count as deleted
    if status_code == 404 and "not found" in text:
        return True
    return False


def scrape_single_ticket_cookie(cookie: str, ticket_id: str) -> dict:
    """Fetch and parse a single ticket detail page using cookie auth."""
    session = requests.Session()
    session.headers.update({
        "User-Agent": UA,
        "Accept": "text/html,application/xhtml+xml,*/*;q=0.9",
        "Cookie": cookie,
    })
    url = f"{BASE_URL}/zendesk/ticket/{ticket_id}"
    resp = session.get(url, timeout=30, allow_redirects=True, verify=False)
    if _is_deleted_ticket_page(resp.status_code, resp.text):
        return {"ticket_id": ticket_id, "url": url, "_deleted": True}
    resp.raise_for_status()
    return parse_ticket_detail(resp.text, url)


def scrape_single_ticket_playwright(ticket_id: str) -> dict:
    """Fetch a single ticket via REST API using the saved browser session cookie."""
    cookie = _browser_state.get("cookie_string", "") or _get_profile_cookie()
    if not cookie:
        raise RuntimeError("No session cookie available — complete browser login first.")
    session = _make_api_session(cookie)
    return fetch_ticket_api(ticket_id, session)


def validate_and_recover_pipeline(
    scores: dict[str, dict],
    cb_url: str,
    bucket: str,
    username: str,
    password: str,
    use_tls: bool,
    scope: str,
    collection: str,
    emb_provider: str,
    emb_model: str,
    emb_api_key: str,
    emb_base_url: str,
    emb_dims: int,
    llm_provider: str,
    llm_model: str,
    llm_api_key: str,
    llm_base_url: str,
    score_batch_size: int,
    cookie: str,
    use_playwright: bool = False,  # kept for backwards compat; ignored (always uses REST)
    progress_cb: Callable[[str, float], None] = lambda m, p: None,
    cancel: threading.Event | None = None,
    raw_tickets: list[dict] | None = None,
) -> tuple[int, int]:
    """
    Validate that every scored ticket exists in CB. For any that are missing
    (DocumentNotFoundException / _stub), re-scrape, re-embed, re-score, and save.
    Returns (recovered, errors).
    """
    if not _CB_AVAILABLE:
        raise RuntimeError("couchbase SDK not installed")

    import datetime

    # ── Report scraping failures from the raw ticket batch ───────────────
    # Tickets whose subject is an HTTP error title were stored as-is in CB;
    # log them here so the operator knows which URLs need a re-scrape.
    _HTTP_ERR_SUBJECTS = frozenset({
        "404 page not found", "403 forbidden", "401 unauthorized",
        "500 internal server error", "502 bad gateway",
        "503 service unavailable", "access denied",
    })
    if raw_tickets:
        scrape_failures = [
            t for t in raw_tickets
            if (t.get("subject") or "").strip().lower() in _HTTP_ERR_SUBJECTS
        ]
        if scrape_failures:
            progress_cb(
                f"Scraping failures ({len(scrape_failures)} doc(s) returned HTTP error pages "
                f"— stored with error title, need re-scrape):", 0.0,
            )
            for t in scrape_failures:
                # Determine document type from whichever ID field is present
                if t.get("ticket_id"):
                    doc_type = "ticket"
                    doc_id   = t["ticket_id"]
                    url = f"{BASE_URL}/zendesk/ticket/{doc_id}"
                elif t.get("snap_id") or t.get("snapshot_id"):
                    doc_type = "snapshot"
                    doc_id   = t.get("snap_id") or t.get("snapshot_id", "?")
                    url = t.get("url") or f"{BASE_URL}/snapshot/{doc_id}"
                else:
                    doc_type = "unknown"
                    doc_id   = "?"
                    url = t.get("url", "")
                subj = (t.get("subject") or "").strip()
                progress_cb(f"  ✗ [{doc_type}] #{doc_id}  {subj}  →  {url}", 0.0)

    conn_str = _cb_conn_str(cb_url, use_tls)
    cluster  = Cluster(conn_str, ClusterOptions(PasswordAuthenticator(username, password)))
    cluster.wait_until_ready(timedelta(seconds=15))
    col = cluster.bucket(bucket).scope(scope).collection(collection)

    # 1. Find tickets that are missing or are stubs
    missing_ids: list[str] = []
    for tid in scores:
        if cancel and cancel.is_set():
            break
        doc_key = f"ticket::{tid}"
        try:
            result = col.get(doc_key)
            doc = result.content_as[dict]
            if doc.get("_stub"):
                missing_ids.append(tid)
        except CouchbaseException as exc:
            if "document_not_found" in str(exc) or "KEY_ENOENT" in str(exc):
                missing_ids.append(tid)

    if not missing_ids:
        progress_cb("Validation passed — all ticket documents present in Couchbase.", 1.0)
        cluster.close()
        return 0, 0

    progress_cb(f"Validation found {len(missing_ids)} missing/stub ticket(s) — recovering…", 0.0)
    recovered = errors = 0
    total = len(missing_ids)
    scored_at = datetime.datetime.now(datetime.UTC).strftime("%Y-%m-%dT%H:%M:%SZ")

    for i, tid in enumerate(missing_ids):
        if cancel and cancel.is_set():
            progress_cb(f"Recovery cancelled after {recovered}/{total}.", i / total)
            break
        pct = i / total
        progress_cb(f"[{i+1}/{total}] Re-scraping ticket #{tid}…", pct)
        doc_key = f"ticket::{tid}"

        # 2. Re-scrape via REST API
        try:
            _ck = cookie or _browser_state.get("cookie_string", "") or _get_profile_cookie()
            if _ck:
                sess = _make_api_session(_ck)
                ticket = fetch_ticket_api(tid, sess)
            else:
                ticket = scrape_single_ticket_cookie(cookie, tid)
        except Exception as exc:
            errors += 1
            progress_cb(f"[{i+1}/{total}] Scrape error for {tid}: {exc}", pct)
            continue

        if ticket.get("error"):
            errors += 1
            progress_cb(f"[{i+1}/{total}] Scrape returned error for {tid}: {ticket['error']}", pct)
            continue

        # 3. Re-embed
        try:
            progress_cb(f"[{i+1}/{total}] Embedding ticket #{tid}…", pct)
            vec = embed_text(
                build_embed_text(ticket),
                emb_provider, emb_model, emb_api_key, emb_base_url, emb_dims,
            )
            if emb_dims and len(vec) > emb_dims:
                vec = vec[:emb_dims]
                norm = sum(x * x for x in vec) ** 0.5
                if norm > 0:
                    vec = [x / norm for x in vec]
            ticket["embedding"] = vec
            ticket["embedding_model"] = emb_model
        except Exception as exc:
            errors += 1
            progress_cb(f"[{i+1}/{total}] Embed error for {tid}: {exc}", pct)
            # Save without embedding rather than drop the ticket entirely
            ticket.pop("embedding", None)

        # 4. Re-score (use existing score if available, otherwise re-score)
        score_data = scores.get(str(tid))
        if not score_data:
            try:
                progress_cb(f"[{i+1}/{total}] Scoring ticket #{tid}…", pct)
                batch_scores = score_all_tickets(
                    [ticket], llm_provider, llm_model, llm_api_key, llm_base_url,
                    score_batch_size,
                    lambda msg, p: None,  # silence inner progress
                )
                score_data = batch_scores.get(str(tid)) or batch_scores.get(tid)
            except Exception as exc:
                errors += 1
                progress_cb(f"[{i+1}/{total}] Score error for {tid}: {exc}", pct)

        if score_data:
            ticket["score"] = {**score_data, "scored_at": scored_at}

        # 5. Save to CB
        try:
            col.upsert(doc_key, ticket)
            recovered += 1
            progress_cb(f"[{i+1}/{total}] Ticket #{tid} recovered and saved.", (i + 1) / total)
        except Exception as exc:
            errors += 1
            progress_cb(f"[{i+1}/{total}] CB save error for {tid}: {exc}", pct)

    cluster.close()
    return recovered, errors


def open_browser_thread() -> None:
    """
    Launch login_browser.py as a subprocess for interactive SSO login.
    The subprocess opens a headed browser, detects successful login,
    saves cookies to ~/.supportal_cookies.json, then exits.
    """
    import sys as _sys
    _browser_closed_event.clear()
    _browser_ready_event.clear()
    _browser_state["logged_in"]     = False
    _browser_state["cookie_string"] = ""

    script = Path(__file__).parent.parent.parent / "tools" / "login_browser.py"
    venv_python = Path(__file__).parent.parent.parent / "venv" / "bin" / "python"
    python = str(venv_python) if venv_python.exists() else _sys.executable

    # Signal "subprocess starting" so the UI unblocks from _browser_ready_event.wait()
    _browser_ready_event.set()

    try:
        subprocess.run([python, str(script)], check=False, timeout=300)
    except subprocess.TimeoutExpired:
        print("[LOGIN] login_browser.py timed out after 5 min")
    except Exception as exc:
        print(f"[LOGIN] subprocess error: {exc}")

    # Read cookies saved by login_browser.py
    try:
        if COOKIES_FILE.exists():
            data = json.loads(COOKIES_FILE.read_text())
            ck = data.get("cookie", "")
            if ck:
                _browser_state["cookie_string"] = ck
                _browser_state["logged_in"]     = True
    except Exception as exc:
        print(f"[LOGIN] cookie read error: {exc}")

    _browser_closed_event.set()


def confirm_login_thread() -> None:
    """Read saved cookie from COOKIES_FILE (login_browser.py already saved it)."""
    try:
        if COOKIES_FILE.exists():
            data = json.loads(COOKIES_FILE.read_text())
            ck = data.get("cookie", "")
            if ck:
                _browser_state["cookie_string"] = ck
                _browser_state["logged_in"]     = True
    except Exception as exc:
        print(f"[LOGIN] confirm_login_thread read error: {exc}")


# ─────────────────────────── Export helpers ───────────────────────────────────

_FLAT_FIELDS = [
    "ticket_id", "url", "subject", "status", "priority",
    "requester", "assignee", "organization",
    "created", "updated", "solved", "tags",
    "description", "ticket_information", "ticket_timeline",
    "escalations", "cbses", "jira_issues", "snapshots", "comment_count", "error",
    # enriched snapshot topology — serialised as JSON in CSV
    "snapshot_topology",
]


def to_csv_bytes(data: list[dict]) -> bytes:
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=_FLAT_FIELDS, extrasaction="ignore")
    writer.writeheader()
    for row in data:
        flat = {k: row.get(k, "") for k in _FLAT_FIELDS}
        # Serialise nested dicts/lists so they land in a single CSV cell
        topo = flat.get("snapshot_topology")
        if isinstance(topo, dict):
            flat["snapshot_topology"] = json.dumps(topo, ensure_ascii=False)
        writer.writerow(flat)
    return buf.getvalue().encode()


def to_json_bytes(data: list[dict]) -> bytes:
    return json.dumps(data, ensure_ascii=False, indent=2).encode()


def to_xls_bytes(data: list[dict]) -> bytes:
    """Build an .xlsx workbook from tickets. Falls back to CSV bytes if openpyxl missing."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Tickets"
        # Header
        ws.append(list(_FLAT_FIELDS))
        hdr_fill = PatternFill("solid", fgColor="1A237E")
        hdr_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 18
        # Rows
        for ticket in data:
            row = []
            for f in _FLAT_FIELDS:
                val = ticket.get(f, "")
                if isinstance(val, dict):
                    val = json.dumps(val, ensure_ascii=False)
                elif isinstance(val, list):
                    val = json.dumps(val, ensure_ascii=False)
                row.append("" if val is None else val)
            ws.append(row)
        # Column widths (capped at 60)
        for col in ws.columns:
            max_w = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max_w + 2, 60)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
    except ImportError:
        return to_csv_bytes(data)


def native_save_dialog(default_name: str, content: bytes) -> str | None:
    """
    Show a native OS "Save As" dialog and write content to the chosen path.
    Returns the saved path, or None if the user cancelled.
    Raises on unexpected errors so callers can surface them.

    macOS  → AppleScript via osascript, activated to foreground
    Other  → tkinter filedialog
    """
    import platform, subprocess, os

    ext  = os.path.splitext(default_name)[1].lstrip(".")
    path: str | None = None

    if platform.system() == "Darwin":
        # activate brings the dialog in front of whatever window is focused
        script = "\n".join([
            'tell application "System Events" to set frontmost of every process whose bundle identifier is "com.apple.finder" to true',
            f'set f to choose file name default name "{default_name}" with prompt "Save as:"',
            'return POSIX path of f',
        ])
        result = subprocess.run(
            ["osascript", "-e", script],
            capture_output=True, text=True, timeout=120,
        )
        raw = result.stdout.strip()
        if result.returncode != 0 and not raw:
            # User cancelled — osascript exits non-zero on cancel
            return None
        if raw:
            path = raw
    else:
        import tkinter as tk
        from tkinter import filedialog
        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        path = filedialog.asksaveasfilename(
            initialfile=default_name,
            defaultextension=f".{ext}",
            filetypes=[(ext.upper(), f"*.{ext}"), ("All files", "*.*")],
        ) or None
        root.destroy()

    if path:
        with open(path, "wb") as fh:
            fh.write(content)

    return path


# ── UI helpers ────────────────────────────────────────────────────────────────

class _PctBar:
    """Drop-in for ui.linear_progress that shows the value as a percentage label.

    Usage mirrors ui.linear_progress — supports .set_value(), .set_visibility(),
    .classes(), .props(), and .value.  The percentage label is kept in sync
    automatically: 0.0454… → "4.55%".
    """

    def __init__(self, value: float = 0.0, *, show_value: bool = True,  # noqa: ARG002
                 color: str = "primary", size: str | None = None):
        with ui.row().classes("w-full items-center gap-2") as self._row:
            self._bar = ui.linear_progress(
                value=value, show_value=False, color=color, size=size,
            ).classes("flex-1")
            self._lbl = ui.label(self._fmt(value)).classes(
                "text-xs font-mono w-14 text-right shrink-0 opacity-70"
            )

    # ── forwarded Element-like API ──────────────────────────────────────────
    def classes(self, add: str = "", **kw):
        self._row.classes(add, **kw)
        return self

    def props(self, add: str = "", **kw):
        self._bar.props(add, **kw)
        return self

    def set_visibility(self, visible: bool):
        self._row.set_visibility(visible)

    # ── value API ───────────────────────────────────────────────────────────
    @staticmethod
    def _fmt(v: float) -> str:
        if v <= 0:
            return "0%"
        if v >= 1:
            return "100%"
        pct = v * 100
        # up to 3 significant digits, no trailing zeros
        return f"{pct:.3g}%"

    def set_value(self, v: float):
        self._bar.set_value(v)
        self._lbl.set_text(self._fmt(v))

    @property
    def value(self) -> float:
        return self._bar.value


# ─────────────────────────── UI ───────────────────────────────────────────────

def _ui_page(path: str):
    """Conditional page decorator — skips registration in library mode."""
    if _LIBRARY_MODE:
        return lambda f: f
    return ui.page(path)

@_ui_page("/")
def main_page():
    # Per-page state (NiceGUI re-creates this for each browser tab)
    state = {
        "results":          [],
        "auth_mode":        "cookie",   # "cookie" | "browser"
        "chat_history":     [],         # list of {"role": "user"|"assistant", "content": str}
        "scores":           {},         # ticket_id -> score dict from Phase 3 LLM scoring
        "customer_name":    "",         # resolved customer name from last scrape/load
        "chat_session_id":  str(uuid.uuid4()),  # unique ID for current conversation session
        "chat_session_turns": [],       # [(question, answer, ts, ticket_ids)] for session storage
        "prior_session_block": "",      # fetched once on first turn; injected into every system prompt
        # AFTER v1.5.0: chat flow
        "last_suggestions": [],         # follow-up question chips rendered after last assistant message
        "_session_log":     {},         # tool call tally carried across turns
    }

    # ── Restore from server-level state (survives page refresh) ─────────────
    if _SERVER_STATE["results"]:
        state["results"]       = _SERVER_STATE["results"]
        state["scores"]        = _SERVER_STATE["scores"]
        state["customer_name"] = _SERVER_STATE["customer_name"]

    _cancel = threading.Event()   # set() to request cancellation of the active operation
    _agent_cancel = threading.Event()  # AFTER v1.5.0: cancel for in-flight agent runs

    # ── Header ──────────────────────────────────────────────────────────────
    with ui.header().classes("bg-blue-900 text-white items-center px-6 py-3 shadow-md gap-3"):
        ui.label("Strabo").classes("text-xl font-bold tracking-tight")
        ui.label(f"v{__version__}").classes(
            "text-xs font-mono bg-blue-700 text-blue-200 px-2 py-0.5 rounded"
        )
        ui.space()
        # Active customer chip — updated whenever customer changes
        _hdr_cust = ui.label("").classes(
            "text-sm font-medium bg-blue-800 text-blue-100 px-3 py-0.5 rounded-full hidden"
        )
        # CB connectivity dot — green when URL is set, grey when not
        with ui.row().classes("items-center gap-1"):
            _cb_dot = ui.element("div").classes(
                "w-2.5 h-2.5 rounded-full bg-gray-400"
            ).tooltip("Couchbase: not configured")
            ui.label("CB").classes("text-xs text-blue-300 font-mono")

    # ── Reconnect banner — shown when page loads while an op is running ─────
    _reconnect_banner = ui.notify(
        f"Reconnected — {_OP_STATUS['op'] or 'operation'} still in progress: "
        f"{_OP_STATUS['status']}",
        type="warning",
        timeout=0,       # stays until dismissed
        close_button=True,
    ) if (not _OP_STATUS["done"] and _OP_STATUS["op"]) else None

    # Timer polls _OP_STATUS every 2 s so this page shows live progress
    # even after a WebSocket reconnect.
    _op_banner_label = ui.label("").classes(
        "text-xs text-amber-700 bg-amber-50 border border-amber-200 "
        "rounded px-3 py-1 w-full"
    )
    _op_banner_label.set_visibility(not _OP_STATUS["done"] and bool(_OP_STATUS["op"]))

    with ui.column().classes("w-full px-4 pt-2 gap-0"):

        # ── Settings profile bar ─────────────────────────────────────────────
        with ui.card().classes("w-full mb-2 bg-blue-50 border border-blue-200"):
            with ui.row().classes("items-center gap-3 flex-wrap"):
                ui.icon("manage_accounts", size="sm").classes("text-blue-700")
                ui.label("Profile:").classes("text-sm font-medium text-blue-800 whitespace-nowrap")
                _all_profiles   = _load_settings_file()
                _profile_names  = sorted(_all_profiles.keys()) or ["default"]
                _last_used      = _all_profiles.get("__last__", _profile_names[0])
                profile_select  = ui.select(
                    _profile_names,
                    value=_last_used if _last_used in _profile_names else _profile_names[0],
                    label="",
                ).classes("w-44")
                profile_name_input = ui.input(placeholder="new profile name…").classes("w-44")
                btn_save_profile   = ui.button("Save", icon="save",   on_click=lambda: asyncio.ensure_future(_save_profile())).props("outline color=blue-8 size=sm")
                btn_load_profile   = ui.button("Load", icon="download", on_click=lambda: asyncio.ensure_future(_load_profile())).props("outline color=blue-8 size=sm")
                btn_delete_profile = ui.button("Delete", icon="delete", on_click=lambda: asyncio.ensure_future(_delete_profile())).props("outline color=red size=sm")
                profile_status     = ui.label("").classes("text-xs text-gray-500 ml-2")

        with ui.tabs().classes("w-full") as main_tabs:
            tab_config   = ui.tab("Configuration",      icon="settings")
            tab_scrape   = ui.tab("Scraping",           icon="search")
            tab_results  = ui.tab("Results",            icon="table_view")
            tab_chat     = ui.tab("Chat",               icon="chat")
            tab_scoring  = ui.tab("Scoring & Analysis", icon="analytics")
            tab_custs    = ui.tab("Customers",          icon="people")
            tab_assets   = ui.tab("Assets",             icon="folder")
        with ui.tab_panels(main_tabs, value=tab_config).classes("w-full pt-4"):

            with ui.tab_panel(tab_scrape):
                with ui.column().classes("w-full gap-6"):
                    # ── Settings card ────────────────────────────────────────────────────
                    with ui.card().classes("w-full"):
                        ui.label("Scraper Settings").classes("text-base font-semibold mb-1")
                        with ui.row().classes("gap-4 w-full flex-wrap items-start"):
                            with ui.row().classes("flex-1 min-w-64 items-center gap-2"):
                                customer_input = (
                                    ui.input(
                                        label="Customer URL or name",
                                        placeholder="https://supportal.couchbase.com/customer/American%20Express%20AZ",
                                    )
                                    .classes("flex-1")
                                    .props("outlined clearable")
                                )
                                btn_cust_search = ui.button(
                                    icon="search", color="blue-grey",
                                ).props("flat round").tooltip("Search Supportal for customer")
                            ui.label(
                                "Tip: type a name and click 🔍 to search Supportal, or paste the full URL directly."
                            ).classes("text-xs text-gray-400 w-full -mt-2")
                            max_pages_input = (
                                ui.number(label="Max listing pages  (0 = all)", value=0, min=0, step=1)
                                .classes("w-56")
                                .props("outlined")
                            )
                            scrape_mode_select = (
                                ui.select(
                                    ["All tickets", "Changed only (skip existing)"],
                                    label="Scrape mode",
                                    value="All tickets",
                                )
                                .classes("w-72")
                                .props("outlined")
                                .tooltip(
                                    "All tickets: re-scrape everything.\n"
                                    "Changed only: compares listing status/solved against Couchbase — "
                                    "re-scrapes new tickets and tickets whose status or solved date changed; "
                                    "skips unchanged tickets."
                                )
                            )
                            max_tickets_input = (
                                ui.number(label="Max tickets  (0 = all)", value=0, min=0, step=25)
                                .classes("w-48")
                                .props("outlined")
                                .tooltip(
                                    "Limit how many tickets are detail-fetched per run. "
                                    "Applied after change detection — new/changed tickets are prioritised. "
                                    "0 = no limit."
                                )
                            )

                        # Customer search dialog (scrape tab)
                        cust_search_dlg = ui.dialog()
                        with cust_search_dlg, ui.card().classes("w-[560px]"):
                            ui.label("Search Supportal Customers").classes("text-base font-semibold mb-2")
                            with ui.row().classes("w-full gap-2 items-center"):
                                cust_search_q = ui.input(
                                    label="Customer name", placeholder="e.g. Royal Caribbean"
                                ).classes("flex-1").props("outlined clearable")
                                btn_cust_search_go = ui.button("Search", icon="search").props("color=teal")
                            cust_search_status = ui.label("").classes("text-xs text-gray-500")
                            cust_search_table = ui.table(
                                columns=[
                                    {"name": "display_name", "label": "Display Name",  "field": "display_name", "align": "left"},
                                    {"name": "slug",         "label": "Slug / URL key", "field": "slug",         "align": "left"},
                                    {"name": "source",       "label": "Source",         "field": "source",       "align": "left"},
                                ],
                                rows=[],
                                row_key="slug",
                            ).classes("w-full").props("flat bordered dense")
                            cust_search_table.add_slot("body-row", """
                                <q-tr :props="props" class="cursor-pointer hover:bg-blue-50"
                                      @click="$emit('rowclick', props.row)">
                                  <q-td v-for="col in props.cols" :key="col.name" :props="props">
                                    {{ col.value }}
                                  </q-td>
                                </q-tr>
                            """)
                            with ui.row().classes("w-full justify-end mt-2"):
                                ui.button("Cancel", on_click=cust_search_dlg.close).props("flat")

                        async def _do_cust_search():
                            q = cust_search_q.value.strip()
                            if not q:
                                return
                            cust_search_status.set_text("Searching…")
                            cust_search_table.rows = []
                            _ck = (cookie_input.value or "").strip() or os.environ.get("SUPPORTAL_COOKIE", "")
                            try:
                                all_hits: list[dict] = []
                                seen_slugs: set[str] = set()

                                # 1 — Supportal UI search API
                                try:
                                    for h in await run.io_bound(search_customers_on_supportal, q, _ck or None):
                                        _k = (h.get("slug") or "").lower()
                                        if _k not in seen_slugs:
                                            seen_slugs.add(_k)
                                            h.setdefault("source", "Supportal")
                                            all_hits.append(h)
                                except Exception:
                                    pass

                                # 2 — Analytics LIKE '%q%' (open endpoint, no cookie needed)
                                if True:
                                    cust_search_status.set_text("Searching analytics…")
                                    try:
                                        for h in await run.io_bound(search_customers_via_analytics, q, _ck):
                                            _k = (h.get("slug") or "").lower()
                                            if _k not in seen_slugs:
                                                seen_slugs.add(_k)
                                                all_hits.append(h)
                                    except Exception:
                                        pass

                                # 3 — Local Couchbase LIKE '%q%' (already-scraped customers)
                                _cb_url  = cb_url_input.value.strip()
                                _cb_user = cb_user_input.value.strip()
                                if _CB_AVAILABLE and _cb_url and _cb_user:
                                    cust_search_status.set_text("Checking local database…")
                                    try:
                                        _local_orgs = await run.io_bound(
                                            search_orgs_from_cb,
                                            _cb_url,
                                            cb_bucket_input.value.strip() or "supportal",
                                            _cb_user,
                                            cb_pass_input.value,
                                            cb_tls_toggle.value,
                                            cb_scope_input.value.strip() or "_default",
                                            cb_collection_input.value.strip() or "tickets",
                                            q,
                                        )
                                        for _org in _local_orgs:
                                            _k = _org.lower()
                                            if _k not in seen_slugs:
                                                seen_slugs.add(_k)
                                                _url = f"{BASE_URL}/customer/{urllib.parse.quote(_org, safe='')}"
                                                all_hits.append({"slug": _org, "display_name": _org, "url": _url, "source": "Local DB"})
                                    except Exception:
                                        pass

                                cust_search_table.rows = all_hits
                                cust_search_table.update()
                                cust_search_status.set_text(
                                    f"{len(all_hits)} result(s)" if all_hits else "No customers found."
                                )
                            except Exception as exc:
                                cust_search_status.set_text(f"Search error: {exc}")

                        def _cust_search_pick(e):
                            row = e.args
                            if isinstance(row, dict):
                                customer_input.set_value(row.get("url") or row.get("slug", ""))
                            cust_search_dlg.close()

                        btn_cust_search_go.on_click(lambda: asyncio.ensure_future(_do_cust_search()))
                        cust_search_q.on("keydown.enter", lambda: asyncio.ensure_future(_do_cust_search()))
                        cust_search_table.on("rowclick", _cust_search_pick)
                        btn_cust_search.on_click(lambda: (
                            cust_search_q.set_value(
                                (customer_input.value or "").strip()
                                if not (customer_input.value or "").startswith("http") else ""
                            ),
                            cust_search_dlg.open(),
                        ))

                    # ── Run card ─────────────────────────────────────────────────────────
                    with ui.card().classes("w-full"):
                        progress_bar   = _PctBar(value=0).props("stripe color=blue-9 rounded")
                        progress_label = ui.label("").classes("text-sm text-gray-500 mt-1 min-h-5")

                        async def do_scrape():
                            if not customer_input.value.strip():
                                ui.notify("Enter a customer URL or name first.", type="warning")
                                return
                            customer, _resolved_url = _resolve_customer_input(customer_input.value)
                            progress_label.set_text(f"Resolved → {_resolved_url}")
                            if not customer:
                                ui.notify("Could not resolve a customer name from the input.", type="warning")
                                return
                            max_pages = int(max_pages_input.value or 0)

                            # Resolve auth — always use cookie (from paste or browser login)
                            cookie = (cookie_input.value or "").strip() or os.environ.get("SUPPORTAL_COOKIE", "")
                            if not cookie:
                                cookie = _browser_state.get("cookie_string", "") or _get_profile_cookie()

                            # Auto-resolve plain name to exact Supportal slug
                            if not customer_input.value.strip().startswith("http"):
                                _best_hit: dict | None = None

                                # 1 — Supportal UI search
                                if cookie:
                                    progress_label.set_text(f"Searching Supportal for '{customer}'…")
                                    try:
                                        _ui_hits = await run.io_bound(search_customers_on_supportal, customer, cookie, 10)
                                        if _ui_hits:
                                            _best_hit = _ui_hits[0]
                                    except Exception:
                                        pass

                                # 2 — Analytics LIKE fallback
                                if not _best_hit and cookie:
                                    progress_label.set_text(f"UI search empty — trying analytics for '{customer}'…")
                                    try:
                                        _an_hits = await run.io_bound(search_customers_via_analytics, customer, cookie, 10)
                                        if _an_hits:
                                            _best_hit = _an_hits[0]
                                    except Exception:
                                        pass

                                # 3 — Local Couchbase LIKE fallback
                                if not _best_hit and _CB_AVAILABLE and cb_url_input.value.strip() and cb_user_input.value.strip():
                                    progress_label.set_text(f"Checking local database for '{customer}'…")
                                    try:
                                        _local_orgs = await run.io_bound(
                                            search_orgs_from_cb,
                                            cb_url_input.value.strip(),
                                            cb_bucket_input.value.strip() or "supportal",
                                            cb_user_input.value.strip(),
                                            cb_pass_input.value,
                                            cb_tls_toggle.value,
                                            cb_scope_input.value.strip() or "_default",
                                            cb_collection_input.value.strip() or "tickets",
                                            customer,
                                        )
                                        if _local_orgs:
                                            _local_url = f"{BASE_URL}/customer/{urllib.parse.quote(_local_orgs[0], safe='')}"
                                            _best_hit = {"slug": _local_orgs[0], "display_name": _local_orgs[0], "url": _local_url}
                                    except Exception:
                                        pass

                                if _best_hit:
                                    customer      = _best_hit.get("slug") or customer
                                    _resolved_url = _best_hit.get("url")  or _resolved_url
                                    customer_input.set_value(_resolved_url)
                                    progress_label.set_text(
                                        f"Resolved '{_best_hit.get('display_name', customer)}' → {_resolved_url}"
                                    )
                                    print(f"[SCRAPE] auto-resolved to slug={customer!r} url={_resolved_url}")
                            if not cookie:
                                ui.notify(
                                    "Paste a cookie string in the Cookie tab, or use Browser Login first.",
                                    type="warning",
                                )
                                return

                            loop = asyncio.get_event_loop()

                            def progress_cb(msg: str, pct: float):
                                _OP_STATUS["op"] = "scrape"
                                _OP_STATUS["status"] = msg
                                _OP_STATUS["progress"] = pct
                                _OP_STATUS["done"] = (pct >= 1.0)
                                async def _update():
                                    progress_bar.set_value(pct)
                                    progress_label.set_text(msg)
                                asyncio.run_coroutine_threadsafe(_update(), loop)

                            btn_scrape.props("loading disabled")
                            progress_bar.set_value(0)
                            progress_label.set_text("Starting…")

                            # Incremental mode: fetch change signals from Couchbase
                            skip_ids: set | None = None
                            change_signals: dict | None = None
                            max_tickets = int(max_tickets_input.value or 0)
                            if scrape_mode_select.value == "Changed only (skip existing)" and _CB_AVAILABLE:
                                progress_label.set_text("Change detection: fetching stored ticket signals from Couchbase…")
                                change_signals = await run.io_bound(
                                    fetch_ticket_signals_from_cb,
                                    cb_url_input.value.strip(),
                                    cb_bucket_input.value.strip() or "supportal",
                                    cb_user_input.value.strip(),
                                    cb_pass_input.value,
                                    cb_tls_toggle.value,
                                    cb_scope_input.value.strip() or "_default",
                                    cb_collection_input.value.strip() or "tickets",
                                )
                                progress_label.set_text(
                                    f"Fetched signals for {len(change_signals)} stored tickets. Enumerating listing…"
                                )
                            elif scrape_mode_select.value == "Changed only (skip existing)" and not _CB_AVAILABLE:
                                ui.notify("Couchbase not available — falling back to full scrape.", type="warning")

                            try:
                                data = await run.io_bound(
                                    scrape_with_cookie, cookie, customer, max_pages, progress_cb,
                                    skip_ids, change_signals, max_tickets,
                                )

                                state["results"] = data
                                _results_empty.set_visibility(False)
                                _results_card.set_visibility(True)
                                # Save old customer's history, then load new customer's
                                _old_cust = state.get("customer_name", "")
                                if _CB_AVAILABLE and state.get("chat_history") and _old_cust != customer:
                                    await run.io_bound(
                                        save_customer_chat_history, _old_cust, list(state["chat_history"]),
                                        cb_url_input.value.strip(), cb_bucket_input.value.strip(),
                                        cb_user_input.value.strip(), cb_pass_input.value, cb_tls_toggle.value,
                                    )
                                state["customer_name"] = customer
                                _SERVER_STATE["results"] = data
                                _SERVER_STATE["customer_name"] = customer
                                _OP_STATUS["op"] = None
                                _OP_STATUS["done"] = True
                                _results.clear()
                                _results.extend(data)
                                _set_customer_banner(customer or "All Customers")
                                if _CB_AVAILABLE and _old_cust != customer:
                                    _loaded_hist = await run.io_bound(
                                        load_customer_chat_history, customer,
                                        cb_url_input.value.strip(), cb_bucket_input.value.strip(),
                                        cb_user_input.value.strip(), cb_pass_input.value, cb_tls_toggle.value,
                                    )
                                    state["chat_history"] = _loaded_hist
                                    state["chat_session_turns"] = []
                                    state["chat_session_id"] = str(uuid.uuid4())
                                    state["prior_session_block"] = ""
                                    _render_chat()

                                _refresh_table(data)
                                btn_dl_json.set_enabled(True)
                                btn_dl_csv.set_enabled(True)
                                btn_dl_xls.set_enabled(True)
                                btn_cb_load.set_enabled(_CB_AVAILABLE)
                                btn_embed.set_enabled(_CB_AVAILABLE)
                                btn_score.set_enabled(True)
                                btn_load_scores.set_enabled(True)
                                btn_rescore_all.set_enabled(True)
                                btn_render_charts.set_enabled(True)
                                progress_label.set_text(f"Done — {len(data)} tickets scraped.")
                                try:
                                    ui.notify(f"Done — {len(data)} tickets scraped.", type="positive")
                                except RuntimeError:
                                    pass  # client context gone after long scrape; element updates above are sufficient

                                # ── Pipeline continuation ──────────────────────────────────
                                if _CB_AVAILABLE and pipeline_save_toggle.value and data:
                                    import time as _time
                                    _pipe_step_start: dict = {}
                                    _is_full_scrape = scrape_mode_select.value != "Changed only (skip existing)"
                                    _steps_enabled = (
                                        ["save"]
                                        + (["enrich"]    if pipeline_enrich_toggle.value    else [])
                                        + (["embed"]     if pipeline_embed_toggle.value     else [])
                                        + (["score"]     if pipeline_score_toggle.value     else [])
                                        + (["validate"]  if pipeline_validate_toggle.value  else [])
                                        + (["reconcile"] if pipeline_reconcile_toggle.value and _is_full_scrape else [])
                                    )

                                    _cancel.clear()
                                    pipe_obs_card.set_visibility(True)
                                    btn_stop_pipeline.set_enabled(True)
                                    _inv: dict = {}  # inventory step results, written to CB at end
                                    pipe_overall_label.set_text(f"0 / {len(_steps_enabled)} steps complete")
                                    _pipe_log_ts = lambda: _time.strftime("%H:%M:%S")

                                    # Reset all step rows to grey/waiting state for this run
                                    for _rk in ("save", "enrich", "embed", "score", "validate"):
                                        _rr = _pipe_step_rows[_rk]
                                        _rr["icon"].props("color=grey")
                                        _rr["label"].set_text("Waiting")
                                        _rr["progress"].set_value(0)
                                        _rr["progress"].set_visibility(False)
                                        _rr["timing"].set_text("")
                                        _rr["detail"].set_text("")

                                    async def _step_activate(key: str):
                                        _pipe_step_start[key] = _time.time()
                                        r = _pipe_step_rows[key]
                                        r["icon"].props("color=blue")
                                        r["label"].set_text("Running…")
                                        r["progress"].set_value(0)
                                        r["progress"].set_visibility(True)
                                        r["timing"].set_text(f"Started {_time.strftime('%H:%M:%S')}")
                                        r["detail"].set_text("")
                                        pipe_log.push(f"{_pipe_log_ts()}  [{key.upper():5}]  Started — {len(data)} tickets")

                                    async def _step_finish(key: str, summary: str, ok: bool = True):
                                        elapsed = _time.time() - _pipe_step_start.get(key, _time.time())
                                        r = _pipe_step_rows[key]
                                        # Treat cancellation as failure regardless of partial ok
                                        if _cancel.is_set() and ok:
                                            ok = False
                                            summary = f"Cancelled — {summary}"
                                        r["icon"].props(f"color={'positive' if ok else 'negative'}")
                                        r["label"].set_text(summary)
                                        r["progress"].set_value(1.0 if ok else r["progress"].value)
                                        r["timing"].set_text(f"Done in {elapsed:.1f}s")
                                        r["detail"].set_text("")
                                        done_count = sum(
                                            1 for k in _steps_enabled
                                            if _pipe_step_rows[k]["timing"].text.startswith("Done")
                                        )
                                        pipe_overall_label.set_text(f"{done_count} / {len(_steps_enabled)} steps complete")
                                        pipe_log.push(f"{_pipe_log_ts()}  [{key.upper():5}]  {'Done' if ok else 'ERROR'} in {elapsed:.1f}s — {summary}")

                                    def _make_step_prog(key: str):
                                        _start = [_time.time()]
                                        _last  = [0.0, _time.time()]  # [last_pct, last_time]

                                        def _prog(msg: str, pct: float):
                                            now     = _time.time()
                                            elapsed = now - _start[0]
                                            dpct = pct - _last[0]
                                            dt   = now  - _last[1]
                                            eta_str = ""
                                            rate_str = ""
                                            if dpct > 0 and dt > 0 and pct > 0:
                                                pct_per_sec = dpct / dt
                                                if pct_per_sec > 0:
                                                    eta_sec = (1.0 - pct) / pct_per_sec
                                                    eta_str = f" · ETA {eta_sec:.0f}s" if eta_sec > 2 else " · almost done"
                                                # tickets/sec estimate from pct * total / elapsed
                                                if elapsed > 1:
                                                    tps = (pct * len(data)) / elapsed
                                                    rate_str = f" · {tps:.1f} t/s"
                                            _last[0] = pct
                                            _last[1] = now

                                            detail = f"Elapsed {elapsed:.0f}s{rate_str}{eta_str}"

                                            async def _upd():
                                                r = _pipe_step_rows[key]
                                                r["label"].set_text(msg)
                                                r["progress"].set_value(pct)
                                                r["detail"].set_text(detail)
                                                # Only push noteworthy lines to the log to avoid
                                                # errors scrolling off with max_lines cap
                                                _msg_lo = msg.lower()
                                                _is_notable = (
                                                    "error" in _msg_lo
                                                    or "fail" in _msg_lo
                                                    or "warn" in _msg_lo
                                                    or "missing" in _msg_lo
                                                    or "empty" in _msg_lo
                                                    or "cancelled" in _msg_lo
                                                    or pct in (0.0, 1.0)
                                                    or pct >= 0.99
                                                )
                                                if _is_notable:
                                                    pipe_log.push(f"{_pipe_log_ts()}  [{key.upper():5}]  {msg}  ({detail})")

                                            asyncio.run_coroutine_threadsafe(_upd(), loop)

                                        _start[0] = _time.time()
                                        return _prog

                                    pipeline_status.set_text("Pipeline running…")

                                    # ── Step 1: Save ───────────────────────────────────────
                                    await _step_activate("save")
                                    _save_ok = False
                                    try:
                                        saved, errs = await run.io_bound(
                                            load_to_couchbase,
                                            data,
                                            cb_url_input.value.strip(),
                                            cb_bucket_input.value.strip(),
                                            cb_user_input.value.strip(),
                                            cb_pass_input.value,
                                            cb_tls_toggle.value,
                                            cb_scope_input.value.strip() or "_default",
                                            cb_collection_input.value.strip() or "tickets",
                                            _make_step_prog("save"),
                                            _cancel,
                                        )
                                        await _step_finish(
                                            "save",
                                            f"{saved} saved" + (f", {errs} errors" if errs else ""),
                                            ok=errs == 0,
                                        )
                                        _save_ok = True
                                        _inv["save"] = {"at": datetime.datetime.now(datetime.timezone.utc).isoformat().replace("+00:00", "Z"), "done": saved, "total": len(data), "errors": errs}
                                    except Exception as exc:
                                        await _step_finish("save", f"Error: {exc}", ok=False)
                                        ui.notify(f"Pipeline save error: {exc}", type="negative")

                                    # ── Step 1.5: Enrich with Snapshot Topology ────────────
                                    if pipeline_enrich_toggle.value and _save_ok and not _cancel.is_set():
                                        await _step_activate("enrich")
                                        _enrich_cookie = (cookie_input.value or "").strip() or os.environ.get("SUPPORTAL_COOKIE", "")
                                        try:
                                            # Build a snap upsert function when CB is configured so
                                            # snapshots found via ticket enrichment are also persisted
                                            # to the snapshots collection in the same pass.
                                            _snap_upsert_fn = None
                                            if _CB_AVAILABLE and cb_url_input.value.strip():
                                                _enrich_snap_col = _make_snap_col(
                                                    cb_url_input.value.strip(),
                                                    cb_bucket_input.value.strip(),
                                                    cb_user_input.value.strip(),
                                                    cb_pass_input.value,
                                                    cb_tls_toggle.value,
                                                    cb_scope_input.value.strip() or "_default",
                                                    ch_snap_coll.value.strip() or "snapshots",
                                                )
                                                if _enrich_snap_col is not None:
                                                    def _snap_upsert_fn(doc, _col=_enrich_snap_col):
                                                        try:
                                                            _d = doc.copy()
                                                            _d["last_scraped_at"] = int(time.time())
                                                            _col.upsert(f"snapshot::{doc['snap_id']}", _d)
                                                        except Exception:
                                                            pass
                                            enriched_n, enrich_errs = await run.io_bound(
                                                enrich_tickets_with_snapshots,
                                                data,
                                                _enrich_cookie or None,
                                                _make_step_prog("enrich"),
                                                _cancel,
                                                int(enrich_workers_input.value or 4),
                                                _snap_upsert_fn,
                                            )
                                            _snap_count = len([t for t in data if t.get("snapshots")])
                                            # ok = skipped (no snap tickets), partial success, or full success
                                            _enrich_ok = _snap_count == 0 or enriched_n > 0 or enrich_errs < _snap_count

                                            # Re-save enriched tickets to Couchbase so that
                                            # snapshot_topology + snap_ids + snapshot_summary are
                                            # persisted.  The initial save (Step 1) ran before
                                            # enrichment, so the docs on disk don't have them yet.
                                            if enriched_n > 0 and _CB_AVAILABLE:
                                                enriched_tickets = [t for t in data if t.get("snapshot_topology")]
                                                await run.io_bound(
                                                    load_to_couchbase,
                                                    enriched_tickets,
                                                    cb_url_input.value.strip(),
                                                    cb_bucket_input.value.strip(),
                                                    cb_user_input.value.strip(),
                                                    cb_pass_input.value,
                                                    cb_tls_toggle.value,
                                                    cb_scope_input.value.strip() or "_default",
                                                    cb_collection_input.value.strip() or "tickets",
                                                    lambda msg, pct: None,
                                                )

                                            await _step_finish(
                                                "enrich",
                                                f"Enriched {enriched_n} tickets; {enrich_errs} errors",
                                                ok=_enrich_ok,
                                            )
                                            _inv["enrich"] = {"at": datetime.datetime.now(datetime.timezone.utc).isoformat().replace("+00:00", "Z"), "done": enriched_n, "total": len(data), "errors": enrich_errs}
                                        except Exception as exc:
                                            await _step_finish("enrich", f"Error: {exc}", ok=False)
                                            ui.notify(f"Enrich error: {exc}", type="negative")

                                    # ── Step 2: Embed ──────────────────────────────────────
                                    if pipeline_embed_toggle.value and _save_ok:
                                        await _step_activate("embed")
                                        try:
                                            emb_provider, emb_model, emb_api_key, emb_base_url, emb_dims, emb_num_ctx = _get_embed_config()
                                            # Ensure LMStudio embedding model is loaded before embedding
                                            if emb_provider == "lmstudio":
                                                _emb_lms_base = (emb_base_url or "http://localhost:1234").rstrip("/v1").rstrip("/")
                                                _loaded = await run.io_bound(lmstudio_ensure_model_loaded, _emb_lms_base, emb_model, 45, "embeddings")
                                                if _loaded:
                                                    if _loaded != emb_model:
                                                        progress_label.set_text(f"LMStudio: using embedding model '{_loaded}'")
                                                    emb_model = _loaded
                                            done_emb, errs_emb = await run.io_bound(
                                                embed_all_tickets,
                                                data,
                                                cb_url_input.value.strip(),
                                                cb_bucket_input.value.strip(),
                                                cb_user_input.value.strip(),
                                                cb_pass_input.value,
                                                cb_tls_toggle.value,
                                                cb_scope_input.value.strip() or "_default",
                                                cb_collection_input.value.strip() or "tickets",
                                                emb_provider,
                                                emb_model,
                                                emb_api_key,
                                                emb_base_url,
                                                emb_dims,
                                                _make_step_prog("embed"),
                                                _cancel,
                                                emb_num_ctx,
                                                int(embed_parallel_input.value or 1),
                                            )
                                            await _step_finish(
                                                "embed",
                                                f"{done_emb} embedded" + (f", {errs_emb} errors" if errs_emb else ""),
                                                ok=errs_emb == 0,
                                            )
                                            _inv["embed"] = {"at": datetime.datetime.now(datetime.timezone.utc).isoformat().replace("+00:00", "Z"), "done": done_emb, "total": len(data), "errors": errs_emb}
                                        except Exception as exc:
                                            await _step_finish("embed", f"Error: {exc}", ok=False)
                                            ui.notify(f"Pipeline embed error: {exc}", type="negative")

                                    # ── Step 3: Score & Save ───────────────────────────────
                                    if pipeline_score_toggle.value and _save_ok:
                                        await _step_activate("score")
                                        try:
                                            provider, model, api_key, base_url = _get_llm_config()
                                            _warn_if_small_model(model)
                                            scores = await run.io_bound(
                                                score_all_tickets,
                                                data,
                                                provider,
                                                model,
                                                api_key,
                                                base_url,
                                                int(score_batch_input.value or 20),
                                                _make_step_prog("score"),
                                                _cancel,
                                                int(score_ctx_input.value or 131) * 1024,
                                                score_no_think_toggle.value,
                                                int(score_parallel_input.value or 1),
                                            )
                                            state["scores"] = scores
                                            _SERVER_STATE["scores"] = scores
                                            _OP_STATUS["op"] = None
                                            _OP_STATUS["done"] = True
                                            saved_sc, errs_sc = await run.io_bound(
                                                persist_scores_to_cb,
                                                scores,
                                                cb_url_input.value.strip(),
                                                cb_bucket_input.value.strip(),
                                                cb_user_input.value.strip(),
                                                cb_pass_input.value,
                                                cb_tls_toggle.value,
                                                cb_scope_input.value.strip() or "_default",
                                                cb_collection_input.value.strip() or "tickets",
                                                _make_step_prog("score"),
                                                data,
                                            )
                                            await _step_finish(
                                                "score",
                                                f"{len(scores)} scored, {saved_sc} saved" + (f", {errs_sc} errors" if errs_sc else ""),
                                                ok=errs_sc == 0,
                                            )
                                            _inv["score"] = {"at": datetime.datetime.now(datetime.timezone.utc).isoformat().replace("+00:00", "Z"), "done": len(scores), "total": len(data), "errors": errs_sc}
                                            btn_render_charts.set_enabled(True)
                                        except Exception as exc:
                                            await _step_finish("score", f"Error: {exc}", ok=False)
                                            ui.notify(f"Pipeline score error: {exc}", type="negative")

                                    # ── Step 4: Validate & Recover ─────────────────────────
                                    if pipeline_validate_toggle.value and _save_ok and not _cancel.is_set():
                                        await _step_activate("validate")
                                        try:
                                            emb_provider, emb_model, emb_api_key, emb_base_url, emb_dims, emb_num_ctx = _get_embed_config()
                                            llm_provider, llm_model, llm_api_key, llm_base_url = _get_llm_config()
                                            recovered_v, errs_v = await run.io_bound(
                                                validate_and_recover_pipeline,
                                                state.get("scores", {}),
                                                cb_url_input.value.strip(),
                                                cb_bucket_input.value.strip(),
                                                cb_user_input.value.strip(),
                                                cb_pass_input.value,
                                                cb_tls_toggle.value,
                                                cb_scope_input.value.strip() or "_default",
                                                cb_collection_input.value.strip() or "tickets",
                                                emb_provider, emb_model, emb_api_key, emb_base_url, emb_dims,
                                                llm_provider, llm_model, llm_api_key, llm_base_url,
                                                int(score_batch_input.value or 20),
                                                (cookie_input.value or "").strip() or os.environ.get("SUPPORTAL_COOKIE", ""),
                                                False,
                                                _make_step_prog("validate"),
                                                _cancel,
                                                data,
                                            )
                                            await _step_finish(
                                                "validate",
                                                f"{recovered_v} recovered" + (f", {errs_v} errors" if errs_v else " — all clear"),
                                                ok=errs_v == 0,
                                            )
                                            _inv["validate"] = {"at": datetime.datetime.now(datetime.timezone.utc).isoformat().replace("+00:00", "Z"), "done": recovered_v, "total": len(data), "errors": errs_v}
                                        except Exception as exc:
                                            await _step_finish("validate", f"Error: {exc}", ok=False)
                                            ui.notify(f"Pipeline validate error: {exc}", type="negative")

                                    # ── Step 5: Reconcile (delete removed tickets) ──────────
                                    if pipeline_reconcile_toggle.value and _is_full_scrape and _save_ok and not _cancel.is_set():
                                        await _step_activate("reconcile")
                                        try:
                                            _scraped_ids = {str(t["ticket_id"]) for t in data if t.get("ticket_id")}
                                            _rec_deleted, _rec_errors = await run.io_bound(
                                                _reconcile_deleted_tickets,
                                                _scraped_ids,
                                                customer,
                                                cb_url_input.value.strip(),
                                                cb_bucket_input.value.strip(),
                                                cb_user_input.value.strip(),
                                                cb_pass_input.value,
                                                cb_tls_toggle.value,
                                                cb_scope_input.value.strip() or "_default",
                                                cb_collection_input.value.strip() or "tickets",
                                                _make_step_prog("reconcile"),
                                            )
                                            await _step_finish(
                                                "reconcile",
                                                f"{_rec_deleted} removed" + (f", {_rec_errors} errors" if _rec_errors else "") if _rec_deleted else "Nothing to remove",
                                                ok=_rec_errors == 0,
                                            )
                                        except Exception as exc:
                                            await _step_finish("reconcile", f"Error: {exc}", ok=False)
                                            ui.notify(f"Pipeline reconcile error: {exc}", type="negative")

                                    # ── Write customer inventory doc ────────────────────────
                                    if _inv and state.get("customer_name"):
                                        await run.io_bound(
                                            upsert_inventory_doc,
                                            state["customer_name"],
                                            _inv,
                                            cb_url_input.value.strip(),
                                            cb_bucket_input.value.strip(),
                                            cb_user_input.value.strip(),
                                            cb_pass_input.value,
                                            cb_tls_toggle.value,
                                            cb_scope_input.value.strip() or "_default",
                                            cb_collection_input.value.strip() or "tickets",
                                        )

                                    btn_stop_pipeline.set_enabled(False)

                                    # Mark any enabled steps that never ran as cancelled (red)
                                    if _cancel.is_set():
                                        for _rk in _steps_enabled:
                                            _rr = _pipe_step_rows[_rk]
                                            if _rr["label"].text in ("Waiting", ""):
                                                _rr["icon"].props("color=negative")
                                                _rr["label"].set_text("Cancelled — did not run")
                                                _rr["timing"].set_text("")
                                        pipeline_status.set_text("Pipeline cancelled.")
                                    else:
                                        pipeline_status.set_text(
                                            "Pipeline complete — see progress card above for details."
                                        )

                            except Exception as exc:
                                progress_label.set_text(f"Error: {exc}")
                                try:
                                    ui.notify(f"Scrape error: {exc}", type="negative", timeout=10000)
                                except RuntimeError:
                                    pass
                            finally:
                                btn_scrape.props(remove="loading disabled")

                        async def do_diagnose():
                            if not customer_input.value.strip():
                                ui.notify("Enter a customer URL or name first.", type="warning")
                                return
                            customer, _resolved_url = _resolve_customer_input(customer_input.value)
                            if not customer:
                                ui.notify("Could not resolve a customer name from the input.", type="warning")
                                return

                            _ck_diag = (cookie_input.value or "").strip() or os.environ.get("SUPPORTAL_COOKIE", "")
                            if not _ck_diag:
                                _ck_diag = _browser_state.get("cookie_string", "") or _get_profile_cookie()
                            if not _ck_diag:
                                ui.notify(
                                    "Paste a cookie string or complete Browser Login first.",
                                    type="warning",
                                )
                                return

                            loop = asyncio.get_event_loop()

                            def progress_cb(msg: str, pct: float):
                                async def _update():
                                    progress_bar.set_value(pct)
                                    progress_label.set_text(msg)
                                asyncio.run_coroutine_threadsafe(_update(), loop)

                            btn_diagnose.props("loading disabled")
                            progress_bar.set_value(0)
                            progress_label.set_text("Running diagnostics (REST API)…")

                            def run_diag():
                                report_lines: list[str] = []
                                def log_cb(msg: str, pct: float):
                                    report_lines.append(msg)
                                    progress_cb(msg, pct)

                                log_cb(f"Testing listing API for: {customer}", 0.1)
                                try:
                                    sess = _make_api_session(_ck_diag)
                                    tickets = _get_customer_ticket_listing_api(customer, sess, log_cb)
                                    log_cb(f"Listing OK — {len(tickets)} tickets found", 0.5)
                                    if tickets:
                                        t0 = tickets[0]
                                        log_cb(f"First ticket: id={t0.get('id')} status={t0.get('status')} subject={t0.get('subject','')[:60]}", 0.6)
                                except Exception as exc:
                                    log_cb(f"Listing error: {exc}", 0.5)

                                return "\n".join(report_lines)

                            try:
                                report = await run.io_bound(run_diag)
                                diag_output.set_text(report)
                                diag_output.set_visibility(True)
                                progress_label.set_text("Diagnostics complete")
                            except Exception as exc:
                                diag_output.set_text(str(exc))
                                diag_output.set_visibility(True)
                                progress_label.set_text(f"Diagnostics error: {exc}")
                            finally:
                                btn_diagnose.props(remove="loading disabled")

                        with ui.row().classes("gap-3 items-center"):
                            btn_scrape = ui.button("Scrape Tickets", on_click=do_scrape, icon="search").props(
                                "color=primary size=lg"
                            )
                            btn_diagnose = ui.button("Diagnose", on_click=do_diagnose, icon="bug_report").props(
                                "color=orange outline size=lg"
                            )
                            ui.label("(Diagnose runs only the navigation steps and saves HTML to debug_html/)").classes(
                                "text-xs text-gray-400"
                            )

                        with ui.expansion("Pipeline Options (auto-run after scrape)", icon="account_tree").classes("w-full mt-2 border rounded"):
                            ui.label(
                                "Enable steps to run automatically after a successful scrape. "
                                "Uses the Couchbase, Embedding, and LLM configurations from the Analysis tab."
                            ).classes("text-xs text-gray-500 mb-2")
                            _step_label_cls = "text-sm min-w-[13rem]"
                            _ctrl_col_cls   = "gap-0 items-start"
                            _ctrl_lbl_cls   = "text-xs text-gray-400 leading-none mb-0.5"

                            # ── Save ──────────────────────────────────────────────────────
                            with ui.row().classes("items-center gap-4 py-1"):
                                pipeline_save_toggle = ui.checkbox("").tooltip("Save scraped tickets to Couchbase")
                                ui.label("Save to Couchbase").classes(_step_label_cls)

                            # ── Enrich ────────────────────────────────────────────────────
                            with ui.row().classes("items-center gap-4 py-1"):
                                pipeline_enrich_toggle = ui.checkbox("").tooltip("Fetch first snapshot per ticket; requires Save")
                                ui.label("Enrich with Snapshots").classes(_step_label_cls)
                                with ui.column().classes(_ctrl_col_cls):
                                    ui.label("Workers").classes(_ctrl_lbl_cls)
                                    enrich_workers_input = ui.number(value=4, min=1, max=16).classes("w-20").tooltip(
                                        "Concurrent HTTP requests to Supportal for snapshot data"
                                    )

                            # ── Embed ─────────────────────────────────────────────────────
                            with ui.row().classes("items-center gap-4 py-1"):
                                pipeline_embed_toggle = ui.checkbox("").tooltip("Generate embeddings; requires Save")
                                ui.label("Embed Tickets").classes(_step_label_cls)
                                with ui.column().classes(_ctrl_col_cls):
                                    ui.label("Parallel").classes(_ctrl_lbl_cls)
                                    embed_parallel_input = ui.number(value=1, min=1, max=32).classes("w-20").tooltip(
                                        "Concurrent embedding requests. Match to LMStudio 'Parallel requests' "
                                        "or OLLAMA_NUM_PARALLEL. Use 1 for MLX (not thread-safe)."
                                    )
                                async def _probe_lmstudio_concurrency():
                                    _, _, _, emb_base_url, _, _ = _get_embed_config()
                                    url = emb_base_url.strip() or "http://localhost:1234"
                                    info = await run.io_bound(poll_lmstudio_model_info, url)
                                    if info.get("n_parallel"):
                                        embed_parallel_input.set_value(info["n_parallel"])
                                        ui.notify(
                                            f"LMStudio reports {info['n_parallel']} parallel slots"
                                            + (f", context {info['context_length']}" if info.get("context_length") else ""),
                                            type="positive",
                                        )
                                    elif info.get("api_version"):
                                        ui.notify(
                                            f"LMStudio API v{info['api_version']} reachable but did not "
                                            "expose parallel count — set manually.",
                                            type="info",
                                        )
                                    else:
                                        ui.notify("Could not reach LMStudio API", type="warning")
                                ui.button("Probe LMStudio", icon="network_ping",
                                          on_click=_probe_lmstudio_concurrency).props("outline color=teal dense")

                            # ── Score ─────────────────────────────────────────────────────
                            with ui.row().classes("items-center gap-4 py-1"):
                                pipeline_score_toggle = ui.checkbox("").tooltip("Score tickets with LLM and save scores; requires Save")
                                ui.label("Score & Save").classes(_step_label_cls)
                                with ui.column().classes(_ctrl_col_cls):
                                    ui.label("Batch").classes(_ctrl_lbl_cls)
                                    score_batch_input = ui.number(value=5, min=1, max=50).classes("w-20")
                                with ui.column().classes(_ctrl_col_cls):
                                    ui.label("Parallel").classes(_ctrl_lbl_cls)
                                    score_parallel_input = ui.number(value=1, min=1, max=8).classes("w-20").tooltip(
                                        "Batches sent to the LLM concurrently. Match to your model server's "
                                        "parallel request capacity."
                                    )
                                score_no_think_toggle = ui.checkbox("No-think").classes("ml-2").tooltip(
                                    "Suppresses Qwen3/QwQ reasoning traces (Ollama think=false). "
                                    "Auto-enabled when a thinking-capable model is detected. "
                                    "Ignored for Claude, Gemini, and LMStudio."
                                )

                            # ── Validate ──────────────────────────────────────────────────
                            with ui.row().classes("items-center gap-4 py-1"):
                                pipeline_validate_toggle = ui.checkbox("").tooltip("Re-scrape, re-embed, re-score any tickets missing from Couchbase")
                                ui.label("Validate & Recover").classes(_step_label_cls)

                            # ── Reconcile ─────────────────────────────────────────────────
                            with ui.row().classes("items-center gap-4 py-1"):
                                pipeline_reconcile_toggle = ui.checkbox("").tooltip(
                                    "Delete tickets from Couchbase that no longer exist in Zendesk. "
                                    "Only runs on full scrapes — skipped for 'Changed only' mode."
                                )
                                ui.label("Reconcile (delete removed)").classes(_step_label_cls)
                        pipeline_status = ui.label("").classes("text-sm text-gray-500 mt-1")

                        # ── Pipeline observability card ───────────────────────────────────
                        pipe_obs_card = ui.card().classes("w-full mt-2 bg-gray-50")
                        pipe_obs_card.set_visibility(False)
                        with pipe_obs_card:
                            with ui.row().classes("items-center justify-between w-full mb-2"):
                                ui.label("Pipeline Progress").classes("text-sm font-semibold")
                                pipe_overall_label = ui.label("").classes("text-xs text-gray-400")
                                btn_stop_pipeline = ui.button("Stop Pipeline", icon="stop_circle", on_click=lambda: (_cancel.set(), btn_stop_pipeline.set_enabled(False))).props("outline color=red size=sm")
                                btn_stop_pipeline.set_enabled(False)
                            _pipe_step_rows: dict = {}
                            for _sk, _sicon, _sname in [
                                ("save",      "save",            "Save to Couchbase"),
                                ("enrich",    "biotech",         "Enrich with Snapshot Topology"),
                                ("embed",     "model_training",  "Embed Tickets"),
                                ("score",     "psychology",      "Score & Save"),
                                ("validate",  "verified_user",   "Validate & Recover"),
                                ("reconcile", "delete_sweep",    "Reconcile (delete removed)"),
                            ]:
                                with ui.row().classes("items-start gap-3 w-full py-2 border-b last:border-b-0"):
                                    _ico = ui.icon(_sicon, size="xs").classes("text-gray-300 mt-1")
                                    with ui.column().classes("flex-1 gap-0"):
                                        with ui.row().classes("items-center justify-between w-full"):
                                            ui.label(_sname).classes("text-sm font-medium")
                                            _tim = ui.label("").classes("text-xs text-gray-400 font-mono")
                                        _lbl = ui.label("Waiting").classes("text-xs text-gray-400")
                                        _prg = _PctBar(value=0, color="blue").classes("mt-1")
                                        _prg.set_visibility(False)
                                        _det = ui.label("").classes("text-xs text-gray-500 font-mono")
                                _pipe_step_rows[_sk] = {
                                    "icon": _ico, "label": _lbl,
                                    "progress": _prg, "timing": _tim, "detail": _det,
                                }
                            ui.separator().classes("my-2")
                            ui.label("Activity Log").classes("text-xs font-semibold text-gray-500 mb-1")
                            pipe_log = ui.log(max_lines=500).classes("w-full h-40 text-xs font-mono border rounded")

                    # ── Background Jobs ───────────────────────────────────────────────────
                    _jobs_card = ui.card().classes("w-full mt-2")
                    with _jobs_card:
                        with ui.row().classes("items-center justify-between w-full mb-2"):
                            ui.label("Background Jobs").classes("text-base font-semibold")
                            ui.label("Auto-refreshes every 3s").classes("text-xs text-gray-400")

                        @ui.refreshable
                        def _render_jobs():
                            if not _SCRAPE_JOBS:
                                ui.label("No jobs started yet.").classes("text-xs text-gray-400 italic")
                                return
                            import time as _jt
                            _now = _jt.time()
                            for job in reversed(list(_SCRAPE_JOBS.values())):
                                proc  = job.get("processed") or 0
                                total = job.get("total")
                                elap  = int(_now - job["started_at"])
                                is_run = job["status"] == "running"
                                color = "blue" if is_run else ("positive" if job["status"] == "done" else "negative")
                                icon  = "sync" if is_run else ("check_circle" if job["status"] == "done" else "error")
                                with ui.row().classes("items-start gap-3 w-full py-2 border-b last:border-b-0"):
                                    ui.icon(icon, size="xs").classes(f"mt-1 text-{color}-500")
                                    with ui.column().classes("flex-1 gap-0.5"):
                                        with ui.row().classes("items-center gap-2"):
                                            ui.label(f"{job['org']}").classes("text-sm font-medium")
                                            ui.badge(job["mode"], color="teal").classes("text-xs")
                                            ui.badge(job["phase"], color=color).classes("text-xs")
                                            ui.label(f"#{job['job_id']}").classes("text-xs text-gray-400 font-mono")
                                        if total:
                                            ui.linear_progress(
                                                value=proc / total,
                                                color=color,
                                            ).classes("w-full mt-0.5").props("stripe" if is_run else "")
                                        ui.label(job.get("last_message") or "").classes("text-xs text-gray-500")
                                        if is_run:
                                            ui.label(f"Elapsed: {elap}s").classes("text-xs text-gray-400 font-mono")
                                        elif job.get("finished_at"):
                                            dur = int(job["finished_at"] - job["started_at"])
                                            ui.label(
                                                f"Done in {dur}s — {proc} scraped, "
                                                f"{job.get('saved',0)} saved, "
                                                f"{job.get('embedded',0)} embedded, "
                                                f"{job.get('scored',0)} scored"
                                                + (f", {job.get('errors',0)} errors" if job.get("errors") else "")
                                            ).classes("text-xs text-gray-500 font-mono")

                        _render_jobs()
                        ui.timer(3.0, _render_jobs.refresh)

                    # ── Diagnostics output ────────────────────────────────────────────────
                    with ui.card().classes("w-full"):
                        ui.label("Diagnostics Output").classes("text-base font-semibold mb-1")
                        diag_output = ui.label("").classes("font-mono text-xs whitespace-pre-wrap text-gray-700")
                        diag_output.set_visibility(False)

                    # ── Snapshot scraping ─────────────────────────────────────────────────
                    ui.separator().classes("my-2")
                    with ui.card().classes("w-full"):
                        ui.label("Snapshot Scraping").classes("text-base font-semibold mb-2")
                        ch_snap_state: dict = {
                            "snapshots": [], "cluster_index": {},
                            "health_data": {}, "last_customer": "",
                        }
                        with ui.row().classes("w-full items-end gap-3 flex-wrap"):
                            with ui.row().classes("flex-1 min-w-64 items-center gap-1"):
                                ch_cust_input = ui.input(
                                    label="Customer name or URL",
                                    placeholder="e.g. Convera or https://supportal…/customer/convera",
                                ).classes("flex-1")
                                btn_ch_cust_search = ui.button(
                                    icon="search", color="blue-grey",
                                ).props("flat round").tooltip("Search Supportal for customer")
                            ch_max_pages = ui.number(
                                label="Max listing pages (0=all)", value=0, min=0, step=1,
                            ).classes("w-36")
                            ch_workers = ui.number(
                                label="Detail fetch workers", value=4, min=1, max=16, step=1,
                            ).classes("w-36")
                            ch_max_snapshots = ui.number(
                                label="Max snapshots (0=all)", value=0, min=0, step=25,
                            ).classes("w-40").tooltip(
                                "Limit topology scrapes per run. "
                                "New and incomplete snapshots are prioritised. 0 = no limit."
                            )
                        with ui.row().classes("gap-3 flex-wrap mt-2 items-center"):
                            btn_ch_scrape = ui.button(
                                "Scrape Snapshots", icon="download",
                                on_click=lambda: asyncio.ensure_future(_ch_scrape(ui.context.client)),
                            ).props("color=teal")
                            def _click_fetch_analytics():
                                _c = ui.context.client
                                asyncio.ensure_future(_ch_fetch_analytics(_c))
                            btn_ch_fetch_analytics = ui.button(
                                "Fetch via Analytics API", icon="bolt",
                                on_click=_click_fetch_analytics,
                            ).props("color=cyan-8 outline").tooltip(
                                "Fast: fetches snapshot list + ticket IDs in one SQL++ query. "
                                "No topology detail — use Scrape Snapshots for full cluster data."
                            )
                            ch_analytics_limit = ui.number(
                                "Limit", value=200, min=10, max=50000, step=100,
                            ).classes("w-24").tooltip("Max snapshots to fetch from analytics API")
                            def _click_scrape_stubs():
                                _c = ui.context.client
                                asyncio.ensure_future(_ch_scrape_stubs(_c))
                            btn_ch_scrape_stubs = ui.button(
                                "Scrape Analytics Snaps", icon="travel_explore",
                                on_click=_click_scrape_stubs,
                            ).props("color=deep-orange outline").tooltip(
                                "Scrape full topology for snapshots already fetched via Analytics API. "
                                "Skips listing-page enumeration — uses known snap_ids directly."
                            )
                            btn_ch_clear = ui.button(
                                "Clear Cache", icon="delete_sweep",
                                on_click=lambda: _ch_clear_cache(),
                            ).props("color=red-4 outline")
                            btn_ch_load_cb = ui.button(
                                "Load from Couchbase", icon="storage",
                                on_click=lambda: asyncio.ensure_future(_ch_load_cb()),
                            ).props("color=blue-grey outline")
                            btn_ch_save_cb = ui.button(
                                "Save to Couchbase", icon="save",
                                on_click=lambda: asyncio.ensure_future(_ch_save_cb()),
                            ).props("color=indigo outline")
                            btn_ch_save_cb.set_enabled(False)
                            btn_ch_embed_cb = ui.button(
                                "Embed Snapshots", icon="hub",
                                on_click=lambda: asyncio.ensure_future(_ch_embed_cb()),
                            ).props("color=teal outline").tooltip(
                                "Generate vector embeddings for loaded snapshots and upsert them to Couchbase. "
                                "Required for vector search / topology queries."
                            )
                            btn_ch_embed_cb.set_enabled(False)
                            ch_auto_save_cb = ui.checkbox(
                                "Auto-save to Couchbase after scrape", value=False,
                            ).tooltip("Automatically persist scraped snapshots to Couchbase when the scrape completes")
                        ch_status = ui.label("Ready.").classes("text-sm text-gray-500 mt-1")
                        ch_progress = ui.linear_progress(value=0).classes("w-full mt-1")
                        ch_progress.set_visibility(False)

                    # Customer search dialog for snapshot scraping
                    ch_cust_search_dlg = ui.dialog()
                    with ch_cust_search_dlg, ui.card().classes("w-[560px]"):
                        ui.label("Search Supportal Customers").classes("text-base font-semibold mb-2")
                        with ui.row().classes("w-full gap-2 items-center"):
                            ch_cust_search_q = ui.input(
                                label="Customer name", placeholder="e.g. Royal Caribbean"
                            ).classes("flex-1").props("outlined clearable")
                            btn_ch_cust_search_go = ui.button("Search", icon="search").props("color=teal")
                        ch_cust_search_status = ui.label("").classes("text-xs text-gray-500")
                        ch_cust_search_table = ui.table(
                            columns=[
                                {"name": "display_name", "label": "Display Name",   "field": "display_name", "align": "left"},
                                {"name": "slug",         "label": "Slug / URL key", "field": "slug",         "align": "left"},
                            ],
                            rows=[],
                            row_key="slug",
                        ).classes("w-full").props("flat bordered dense")
                        ch_cust_search_table.add_slot("body-row", """
                            <q-tr :props="props" class="cursor-pointer hover:bg-blue-50"
                                  @click="$emit('rowclick', props.row)">
                              <q-td v-for="col in props.cols" :key="col.name" :props="props">
                                {{ col.value }}
                              </q-td>
                            </q-tr>
                        """)
                        with ui.row().classes("w-full justify-end mt-2"):
                            ui.button("Cancel", on_click=ch_cust_search_dlg.close).props("flat")

                    async def _do_ch_cust_search():
                        q = ch_cust_search_q.value.strip()
                        if not q:
                            return
                        ch_cust_search_status.set_text("Searching…")
                        ch_cust_search_table.rows = []
                        _ck = (cookie_input.value or "").strip() or os.environ.get("SUPPORTAL_COOKIE", "")
                        try:
                            hits = await run.io_bound(
                                search_customers_on_supportal, q, _ck or None
                            )
                            ch_cust_search_table.rows = hits
                            ch_cust_search_table.update()
                            ch_cust_search_status.set_text(
                                f"{len(hits)} result(s)" if hits else "No customers found."
                            )
                        except Exception as exc:
                            ch_cust_search_status.set_text(f"Search error: {exc}")

                    def _ch_cust_search_pick(e):
                        row = e.args
                        if isinstance(row, dict):
                            ch_cust_input.set_value(row.get("url") or row.get("slug", ""))
                        ch_cust_search_dlg.close()

                    btn_ch_cust_search_go.on_click(lambda: asyncio.ensure_future(_do_ch_cust_search()))
                    ch_cust_search_q.on("keydown.enter", lambda: asyncio.ensure_future(_do_ch_cust_search()))
                    ch_cust_search_table.on("rowclick", _ch_cust_search_pick)
                    btn_ch_cust_search.on_click(lambda: (
                        ch_cust_search_q.set_value(
                            (ch_cust_input.value or "").strip()
                            if not (ch_cust_input.value or "").startswith("http") else ""
                        ),
                        ch_cust_search_dlg.open(),
                    ))

            with ui.tab_panel(tab_results):
                with ui.column().classes("w-full gap-6"):
                    # ── Customer banner ───────────────────────────────────────────────────
                    with ui.row().classes("w-full items-center gap-3 px-4 py-2 rounded-lg bg-indigo-50 border border-indigo-200"):
                        ui.icon("business").classes("text-indigo-400")
                        results_banner = ui.label("No customer loaded").classes("text-sm font-semibold text-indigo-700 flex-1")

                    # ── Results empty state ───────────────────────────────────────────────
                    with ui.column().classes("w-full items-center gap-3 py-12 text-center") as _results_empty:
                        ui.icon("search_off", size="4rem").classes("text-gray-200")
                        ui.label("No tickets loaded").classes("text-lg font-medium text-gray-400")
                        ui.label("Scrape a customer or load tickets from Couchbase to get started.").classes("text-xs text-gray-400")
                        ui.button("Go to Scraping", icon="search", on_click=lambda: main_tabs.set_value(tab_scrape)).props("outline color=primary size=sm")
                    _has_results = bool(state["results"])
                    _results_empty.set_visibility(not _has_results)

                    # ── Results card ──────────────────────────────────────────────────────
                    with ui.card().classes("w-full") as _results_card:
                        _results_card.set_visibility(_has_results)
                        ui.label("Results").classes("text-base font-semibold mb-1")

                        columns = [
                            {"name": "ticket_id",    "label": "Ticket",       "field": "ticket_id",    "sortable": True, "align": "left"},
                            {"name": "status",       "label": "Status",       "field": "status",       "sortable": True},
                            {"name": "priority",     "label": "Priority",     "field": "priority",     "sortable": True},
                            {"name": "subject",      "label": "Subject",      "field": "subject",      "sortable": True, "align": "left"},
                            {"name": "requester",    "label": "Requester",    "field": "requester",    "sortable": True, "align": "left"},
                            {"name": "assignee",     "label": "Assignee",     "field": "assignee",     "sortable": True, "align": "left"},
                            {"name": "organization", "label": "Org",          "field": "organization", "sortable": True, "align": "left"},
                            {"name": "created",      "label": "Created",      "field": "created",      "sortable": True},
                            {"name": "updated",      "label": "Updated",      "field": "updated",      "sortable": True},
                            {"name": "solved",       "label": "Solved",       "field": "solved",       "sortable": True},
                            {"name": "comment_count","label": "Comments",     "field": "comment_count","sortable": True},
                            {"name": "tags",         "label": "Tags",         "field": "tags",         "sortable": True, "align": "left"},
                        ]

                        # Pagination state
                        _page_state = {"current": 1, "per_page": 25, "all_rows": []}

                        with ui.row().classes("items-center gap-4 mb-1"):
                            page_info_label = ui.label("").classes("text-xs text-gray-500 flex-1")
                            ui.label("Per page:").classes("text-xs text-gray-500")
                            per_page_select = ui.select(
                                options=[10, 25, 50, 100],
                                value=25,
                            ).classes("w-20 text-xs")

                        result_table = ui.table(columns=columns, rows=[], row_key="ticket_id").classes("w-full")
                        result_table.props("flat dense")

                        pager = ui.pagination(1, 1, direction_links=True).classes("mt-1 self-center")

                        # Row click → detail dialog
                        detail_dialog = ui.dialog().props("maximized")
                        with detail_dialog:
                            with ui.card().classes("w-full h-full overflow-auto p-6"):
                                dialog_title   = ui.label("").classes("text-lg font-bold mb-2")
                                dialog_content = ui.column().classes("w-full gap-2")

                        def _show_detail(ticket: dict):
                            dialog_title.set_text(f"Ticket #{ticket.get('ticket_id')} — {ticket.get('subject', '')}")
                            dialog_content.clear()
                            with dialog_content:
                                # Well-known fields
                                meta_fields = [
                                    ("Status",       ticket.get("status")),
                                    ("Priority",     ticket.get("priority")),
                                    ("Requester",    ticket.get("requester")),
                                    ("Assignee",     ticket.get("assignee")),
                                    ("Organization", ticket.get("organization")),
                                    ("Created",      ticket.get("created")),
                                    ("Updated",      ticket.get("updated")),
                                    ("Solved",       ticket.get("solved")),
                                    ("Tags",         ticket.get("tags")),
                                ]
                                with ui.grid(columns=2).classes("w-full gap-1 text-sm"):
                                    for label, val in meta_fields:
                                        if val:
                                            ui.label(label).classes("font-semibold text-gray-500")
                                            ui.label(val).classes("text-gray-800")

                                # Extra fields found on the page
                                all_fields = ticket.get("all_fields", {})
                                known_keys = {
                                    "status", "priority", "requester", "assignee",
                                    "organization", "created", "updated", "solved", "tags",
                                }
                                extra = {k: v for k, v in all_fields.items()
                                         if k.lower() not in known_keys and v}
                                if extra:
                                    ui.separator()
                                    ui.label("Additional fields").classes("font-semibold mt-1")
                                    with ui.grid(columns=2).classes("w-full gap-1 text-sm"):
                                        for k, v in extra.items():
                                            ui.label(k).classes("font-semibold text-gray-500")
                                            ui.label(str(v)).classes("text-gray-800")

                                # Named sections
                                def _section(label: str, content: Optional[str]):
                                    if not content:
                                        return
                                    ui.separator()
                                    ui.label(label).classes("font-semibold mt-1 text-blue-800")
                                    ui.label(content).classes("text-sm whitespace-pre-wrap text-gray-700 bg-gray-50 p-2 rounded")

                                _section("Description",  ticket.get("description"))

                                # ── Ticket Fields (native dict or legacy JSON string) ──────
                                tf_raw = ticket.get("ticket_fields")
                                if tf_raw:
                                    tf = _parse_ticket_fields(ticket)
                                    if tf:
                                        ui.separator()
                                        ui.label("Ticket Fields").classes("font-semibold mt-1 text-blue-800")
                                        with ui.element("table").classes("w-full text-sm border-collapse"):
                                            for k, v in sorted(tf.items()):
                                                if not v or str(v).strip() in ("", "None"):
                                                    continue
                                                display_k = k.replace("_", " ")
                                                with ui.element("tr").classes("border-b border-gray-100 hover:bg-gray-50"):
                                                    with ui.element("td").classes("py-1 pr-3 text-xs text-gray-500 font-medium w-48 align-top"):
                                                        ui.label(display_k)
                                                    with ui.element("td").classes("py-1 text-gray-800 whitespace-pre-wrap align-top"):
                                                        ui.label(str(v))

                                _section("Tags",          ticket.get("tags"))
                                _section("Escalations",   ticket.get("escalations"))
                                _section("Snapshots",     ticket.get("snapshots"))

                                # Comments / conversation (stored as JSON string or list)
                                comments = ticket.get("comments", [])
                                if isinstance(comments, str):
                                    try:
                                        comments = json.loads(comments)
                                    except Exception:
                                        comments = []
                                if comments:
                                    ui.separator()
                                    ui.label(f"Conversation  ({len(comments)} entries)").classes("font-semibold mt-1 text-blue-800")
                                    for c in comments:
                                        with ui.card().classes("w-full bg-gray-50"):
                                            with ui.row().classes("justify-between text-xs text-gray-400"):
                                                ui.label(c.get("author") or "—")
                                                ui.label(c.get("timestamp") or "")
                                            ui.label(c.get("body", "")).classes("text-sm whitespace-pre-wrap mt-1")

                                ui.button("Close", on_click=detail_dialog.close).classes("mt-4").props("flat color=grey")
                            detail_dialog.open()

                        def _render_page():
                            pp   = _page_state["per_page"]
                            pg   = _page_state["current"]
                            all_rows = _page_state["all_rows"]
                            total = len(all_rows)
                            start = (pg - 1) * pp
                            end   = start + pp
                            result_table.rows = all_rows[start:end]
                            result_table.update()
                            pages = max(1, -(-total // pp))  # ceiling division
                            pager.max = pages
                            if pager.value > pages:
                                pager.value = pages
                            showing_end = min(end, total)
                            page_info_label.set_text(
                                f"Showing {start + 1}–{showing_end} of {total} tickets" if total else "No tickets"
                            )

                        def _refresh_table(data: list[dict]):
                            _page_state["all_rows"] = [
                                {c["field"]: (rec.get(c["field"]) or "") for c in columns}
                                for rec in data
                            ]
                            _page_state["current"] = 1
                            pager.value = 1
                            _render_page()

                        def _on_page_change(e):
                            _page_state["current"] = int(pager.value)
                            _render_page()

                        def _on_per_page_change(e):
                            _page_state["per_page"] = int(per_page_select.value)
                            _page_state["current"] = 1
                            pager.value = 1
                            _render_page()

                        pager.on("update:model-value", _on_page_change)
                        per_page_select.on("update:model-value", _on_per_page_change)

                        def _on_row_click(e):
                            # NiceGUI 3.x / Quasar QTable fires row-click as (evt, row, index)
                            # so e.args is a list: [quasar_evt_dict, row_dict, row_index]
                            args = e.args
                            row_data = args[1] if isinstance(args, list) and len(args) > 1 else {}
                            if isinstance(row_data, dict):
                                ticket_id = str(row_data.get("ticket_id", ""))
                            else:
                                return
                            ticket = next(
                                (r for r in state["results"] if str(r.get("ticket_id")) == ticket_id), {}
                            )
                            _show_detail(ticket)

                        result_table.on("rowClick", _on_row_click)

                        # Downloads
                        with ui.row().classes("gap-3 mt-3"):
                            async def dl_json():
                                if state["results"]:
                                    try:
                                        path = await run.io_bound(
                                            native_save_dialog, "tickets.json", to_json_bytes(state["results"])
                                        )
                                        if path:
                                            ui.notify(f"Saved → {path}", type="positive", timeout=8000)
                                    except Exception as exc:
                                        ui.notify(f"Export error: {exc}", type="negative", timeout=10000)

                            async def dl_csv():
                                if state["results"]:
                                    try:
                                        path = await run.io_bound(
                                            native_save_dialog, "tickets.csv", to_csv_bytes(state["results"])
                                        )
                                        if path:
                                            ui.notify(f"Saved → {path}", type="positive", timeout=8000)
                                    except Exception as exc:
                                        ui.notify(f"Export error: {exc}", type="negative", timeout=10000)

                            async def dl_xls():
                                if state["results"]:
                                    try:
                                        path = await run.io_bound(
                                            native_save_dialog, "tickets.xlsx", to_xls_bytes(state["results"])
                                        )
                                        if path:
                                            ui.notify(f"Saved → {path}", type="positive", timeout=8000)
                                    except Exception as exc:
                                        ui.notify(f"Export error: {exc}", type="negative", timeout=10000)

                            btn_dl_json = ui.button("Download JSON", on_click=dl_json, icon="download").props("outline color=primary")
                            btn_dl_csv  = ui.button("Download CSV",  on_click=dl_csv,  icon="download").props("outline color=secondary")
                            btn_dl_xls  = ui.button("Download XLS",  on_click=dl_xls,  icon="table_chart").props("outline color=green-8")
                            btn_dl_json.set_enabled(False)
                            btn_dl_csv.set_enabled(False)
                            btn_dl_xls.set_enabled(False)

            with ui.tab_panel(tab_config):
                with ui.column().classes("w-full gap-0"):
                    with ui.tabs().classes("w-full bg-white border-b border-gray-200") as config_sub_tabs:
                        cfg_auth      = ui.tab("Authentication",       icon="lock")
                        cfg_cb        = ui.tab("Couchbase",            icon="storage")
                        cfg_embed     = ui.tab("Data Operations", icon="model_training")
                        cfg_chat_mem  = ui.tab("Chat & Memory",        icon="memory")
                        cfg_analytics = ui.tab("Analytics",            icon="tune")
                        cfg_ai        = ui.tab("AI Models",            icon="smart_toy")
                        cfg_preflight = ui.tab("Preflight",            icon="checklist")
                with ui.tab_panels(config_sub_tabs, value=cfg_cb).classes("w-full"):
                    with ui.tab_panel(cfg_auth):
                        ui.label("Authentication").classes("text-base font-semibold mb-1")

                        with ui.tabs().classes("w-full") as auth_tabs:
                            tab_cookie  = ui.tab("Cookie / Header")
                            tab_browser = ui.tab("Browser Login (SSO)")

                        with ui.tab_panels(auth_tabs, value=tab_cookie).classes("w-full border-t"):

                            # ── Cookie tab ──────────────────────────────────────────────
                            with ui.tab_panel(tab_cookie):
                                ui.label(
                                    "Open DevTools → Network → any request → Headers → Cookie. "
                                    "Copy the full Cookie header value and paste it below. "
                                    "Alternatively set the SUPPORTAL_COOKIE environment variable before launching."
                                ).classes("text-sm text-gray-500 mb-2")
                                cookie_input = (
                                    ui.textarea(label="Cookie string", placeholder="session=abc123; other=xyz …")
                                    .classes("w-full font-mono text-sm")
                                    .props('rows=3 outlined clearable')
                                )
                                # Pre-fill from env if available
                                env_cookie = os.environ.get("SUPPORTAL_COOKIE", "")
                                if env_cookie:
                                    cookie_input.value = env_cookie
                                    ui.label("(pre-filled from SUPPORTAL_COOKIE env var)").classes("text-xs text-green-600")

                            # ── Browser login tab ───────────────────────────────────────
                            with ui.tab_panel(tab_browser):
                                ui.label(
                                    "Click Open Browser to launch login_browser.py. "
                                    "A Chromium window will open — complete the Okta SSO login. "
                                    "The browser auto-closes when login is detected and the cookie is saved automatically."
                                ).classes("text-sm text-gray-500 mb-3")

                                with ui.row().classes("items-center gap-2 mb-2"):
                                    browser_dot    = ui.icon("circle").classes("text-red-500 text-sm")
                                    browser_status = ui.label("Not logged in").classes("text-sm font-semibold text-gray-600")

                                async def do_open_browser():
                                    btn_open.props("loading disabled")
                                    browser_dot.classes(replace="text-orange-500 text-sm")
                                    browser_status.set_text("Starting browser (login_browser.py)…")
                                    try:
                                        threading.Thread(target=open_browser_thread, daemon=True).start()
                                        # Wait until subprocess has started
                                        await asyncio.get_event_loop().run_in_executor(
                                            None, _browser_ready_event.wait
                                        )
                                        browser_dot.classes(replace="text-blue-500 text-sm")
                                        browser_status.set_text("Browser open — log in (will auto-close when complete)")
                                        # Wait for login to complete in background
                                        asyncio.ensure_future(_poll_login_complete())
                                    except Exception as exc:
                                        browser_dot.classes(replace="text-red-600 text-sm")
                                        browser_status.set_text(f"Error: {exc}")
                                    finally:
                                        btn_open.props(remove="loading disabled")

                                async def _poll_login_complete():
                                    await asyncio.get_event_loop().run_in_executor(
                                        None, _browser_closed_event.wait
                                    )
                                    captured = _browser_state.get("cookie_string", "")
                                    if captured:
                                        cookie_input.set_value(captured)
                                        browser_dot.classes(replace="text-green-500 text-sm")
                                        browser_status.set_text("Logged in ✓ — cookie captured and ready")
                                    else:
                                        browser_dot.classes(replace="text-red-600 text-sm")
                                        browser_status.set_text(
                                            "Login subprocess finished but no cookie found — "
                                            "check that login_browser.py saved ~/.supportal_cookies.json"
                                        )

                                async def do_load_saved_cookie():
                                    """Read cookie from ~/.supportal_cookies.json if it exists."""
                                    try:
                                        if COOKIES_FILE.exists():
                                            data = json.loads(COOKIES_FILE.read_text())
                                            ck = data.get("cookie", "")
                                            if ck:
                                                cookie_input.set_value(ck)
                                                _browser_state["cookie_string"] = ck
                                                _browser_state["logged_in"]     = True
                                                browser_dot.classes(replace="text-green-500 text-sm")
                                                browser_status.set_text("Loaded saved cookie ✓")
                                                ui.notify("Loaded saved cookie.", type="positive")
                                                return
                                        ui.notify("No saved cookie found — run Open Browser first.", type="warning")
                                    except Exception as exc:
                                        ui.notify(f"Load error: {exc}", type="negative")

                                with ui.row().classes("gap-3"):
                                    btn_open = ui.button("Open Browser & Login", on_click=do_open_browser, icon="open_in_browser")
                                    ui.button("Load Saved Cookie", on_click=do_load_saved_cookie, icon="key").props("outline color=teal")

                    with ui.tab_panel(cfg_cb):
                        with ui.row().classes("items-center justify-between w-full"):
                            ui.label("Load to Couchbase").classes("text-base font-semibold")
                            if not _CB_AVAILABLE:
                                ui.badge("SDK not installed", color="red").props("outline")

                        with ui.grid(columns=2).classes("w-full gap-x-4 gap-y-2 mt-2"):
                            cb_url_input      = ui.input("Cluster URL", placeholder="127.0.0.1").classes("w-full")
                            cb_bucket_input   = ui.input("Bucket",      placeholder="supportal").classes("w-full")
                            cb_user_input     = ui.input("Username",    placeholder="Administrator").classes("w-full")
                            cb_pass_input     = ui.input("Password").props("type=password").classes("w-full")
                            cb_scope_input      = ui.input("Scope",              placeholder="_default").classes("w-full")
                            cb_collection_input = ui.input("Collection",         placeholder="tickets").classes("w-full")
                            ch_snap_coll        = ui.input("Snapshot collection", placeholder="snapshots", value="snapshots").classes("w-full")
                            cb_summary_coll     = ui.input("Summary collection",  placeholder="summary",   value="summary").classes("w-full")

                        with ui.row().classes("items-center gap-4 mt-2"):
                            cb_tls_toggle = ui.switch("TLS (couchbases://)", value=False)

                        cb_status = ui.label("").classes("text-sm text-gray-500 mt-1")
                        cb_progress = _PctBar(value=0).classes("mt-1")
                        cb_progress.set_visibility(False)

                        async def _do_cb_load():
                            if not state["results"]:
                                ui.notify("Run a scrape first — no tickets loaded.", type="warning")
                                return
                            if not _CB_AVAILABLE:
                                ui.notify("couchbase SDK not installed. Run: venv/bin/pip install couchbase", type="negative")
                                return

                            btn_cb_load.set_enabled(False)
                            cb_progress.set_visibility(True)
                            cb_progress.set_value(0)
                            cb_status.set_text("Starting…")

                            loop = asyncio.get_event_loop()

                            def _cb_progress(msg: str, pct: float):
                                asyncio.run_coroutine_threadsafe(
                                    _update_cb_progress(msg, pct), loop
                                )

                            async def _update_cb_progress(msg: str, pct: float):
                                cb_status.set_text(msg)
                                cb_progress.set_value(pct)

                            try:
                                upserted, errors = await run.io_bound(
                                    load_to_couchbase,
                                    state["results"],
                                    cb_url_input.value.strip(),
                                    cb_bucket_input.value.strip(),
                                    cb_user_input.value.strip(),
                                    cb_pass_input.value,
                                    cb_tls_toggle.value,
                                    cb_scope_input.value.strip() or "_default",
                                    cb_collection_input.value.strip() or "tickets",
                                    _cb_progress,
                                )
                                msg = f"Done — {upserted} upserted, {errors} errors."
                                cb_status.set_text(msg)
                                cb_progress.set_value(1.0)
                                ui.notify(msg, type="positive" if errors == 0 else "warning")
                            except Exception as exc:
                                cb_status.set_text(f"Error: {exc}")
                                ui.notify(str(exc), type="negative")
                            finally:
                                btn_cb_load.set_enabled(True)

                        btn_cb_load = ui.button(
                            "Load to Couchbase",
                            on_click=_do_cb_load,
                            icon="upload",
                        ).props("color=red-8").classes("mt-2")
                        btn_cb_load.set_enabled(False)

                        # ── One-time migration: normalize ticket_fields + comments ─────────
                        ui.separator().classes("mt-4")
                        with ui.row().classes("items-center gap-2 mt-1"):
                            ui.icon("build_circle").classes("text-orange-500")
                            ui.label("Document Field Migration").classes("text-sm font-semibold text-gray-600")
                        ui.label(
                            "Normalizes existing Couchbase documents: converts ticket_fields from "
                            "escaped JSON strings to native sub-documents (keys: spaces→_, (EOL)→_EOL), "
                            "and converts comments from escaped JSON strings to native arrays. "
                            "Safe to run multiple times — already-normalized docs are skipped."
                        ).classes("text-xs text-gray-400 mt-1")

                        async def _do_migrate_fields():
                            if not _CB_AVAILABLE:
                                ui.notify("Couchbase SDK not installed.", type="negative")
                                return
                            btn_migrate_fields.set_enabled(False)
                            cb_status.set_text("Migrating documents…")
                            loop = asyncio.get_event_loop()
                            def _prog(msg, pct):
                                async def _upd():
                                    cb_status.set_text(msg)
                                    cb_progress.set_value(pct)
                                asyncio.run_coroutine_threadsafe(_upd(), loop)
                            try:
                                migrated, skipped = await run.io_bound(
                                    migrate_ticket_fields_in_cb,
                                    cb_url_input.value.strip(),
                                    cb_bucket_input.value.strip(),
                                    cb_user_input.value.strip(),
                                    cb_pass_input.value,
                                    cb_tls_toggle.value,
                                    cb_scope_input.value.strip() or "_default",
                                    cb_collection_input.value.strip() or "tickets",
                                    _prog,
                                )
                                cb_status.set_text(
                                    f"Migration complete: {migrated} updated, {skipped} skipped (already normalized or no fields)."
                                )
                                ui.notify(f"Migrated {migrated} documents.", type="positive")
                            except Exception as exc:
                                cb_status.set_text(f"Migration error: {exc}")
                                ui.notify(str(exc), type="negative")
                            finally:
                                btn_migrate_fields.set_enabled(True)

                        btn_migrate_fields = ui.button(
                            "Normalize ticket_fields & comments",
                            on_click=_do_migrate_fields,
                            icon="build",
                        ).props("color=orange outline").classes("mt-1")

                        # ── GSI index management ───────────────────────────────────────────
                        ui.separator().classes("mt-4")
                        with ui.row().classes("items-center gap-2 mt-1"):
                            ui.icon("schema").classes("text-indigo-500")
                            ui.label("Couchbase GSI Indexes").classes("text-sm font-semibold text-gray-600")
                        ui.label(
                            "Create or verify query indexes on both the tickets and snapshots collections. "
                            "Required for efficient customer directory queries, cluster health aggregations, "
                            "and ticket↔snapshot cross-reference lookups. Safe to run multiple times "
                            "(uses IF NOT EXISTS)."
                        ).classes("text-xs text-gray-400 mt-1")

                        async def _do_create_gsi():
                            if not _CB_AVAILABLE:
                                ui.notify("Couchbase SDK not installed.", type="negative")
                                return
                            btn_create_gsi.set_enabled(False)
                            cb_progress.set_visibility(True)
                            cb_progress.set_value(0)
                            loop = asyncio.get_event_loop()
                            def _prog(msg, pct):
                                async def _upd():
                                    cb_status.set_text(msg)
                                    cb_progress.set_value(pct)
                                asyncio.run_coroutine_threadsafe(_upd(), loop)
                            try:
                                results = await run.io_bound(
                                    ensure_cb_indexes,
                                    cb_url_input.value.strip(),
                                    cb_bucket_input.value.strip(),
                                    cb_user_input.value.strip(),
                                    cb_pass_input.value,
                                    cb_tls_toggle.value,
                                    cb_scope_input.value.strip() or "_default",
                                    ch_snap_coll.value.strip() or "snapshots",
                                    cb_collection_input.value.strip() or "tickets",
                                    _prog,
                                )
                                errors_gsi = [r for r in results if r.startswith("ERROR")]
                                summary = f"{len(results)} indexes processed; {len(errors_gsi)} errors."
                                if errors_gsi:
                                    cb_status.set_text(summary + " " + " | ".join(errors_gsi[:3]))
                                    ui.notify(summary, type="warning")
                                else:
                                    cb_status.set_text(summary)
                                    ui.notify(summary, type="positive")
                            except Exception as exc:
                                cb_status.set_text(f"Index error: {exc}")
                                ui.notify(str(exc), type="negative")
                            finally:
                                btn_create_gsi.set_enabled(True)

                        btn_create_gsi = ui.button(
                            "Create / Verify GSI Indexes",
                            on_click=_do_create_gsi,
                            icon="schema",
                        ).props("color=indigo outline").classes("mt-1")

                        # ── Ticket ↔ Snapshot link migration ──────────────────────────────
                        ui.separator().classes("mt-4")
                        with ui.row().classes("items-center gap-2 mt-1"):
                            ui.icon("link").classes("text-teal-500")
                            ui.label("Ticket ↔ Snapshot Link Migration").classes("text-sm font-semibold text-gray-600")
                        ui.label(
                            "Backfills snap_ids / snapshot_summary on existing ticket docs and populates "
                            "ticket_ids on snapshot docs — without re-scraping the site. "
                            "Run once after upgrading, then new scrapes maintain the links automatically."
                        ).classes("text-xs text-gray-400 mt-1")

                        async def _do_migrate_links():
                            if not _CB_AVAILABLE:
                                ui.notify("Couchbase SDK not installed.", type="negative")
                                return
                            btn_migrate_links.set_enabled(False)
                            cb_progress.set_visibility(True)
                            cb_progress.set_value(0)
                            loop = asyncio.get_event_loop()
                            def _prog(msg, pct):
                                async def _upd():
                                    cb_status.set_text(msg)
                                    cb_progress.set_value(pct)
                                asyncio.run_coroutine_threadsafe(_upd(), loop)
                            try:
                                t_upd, s_upd, errs = await run.io_bound(
                                    migrate_ticket_snapshot_links,
                                    cb_url_input.value.strip(),
                                    cb_bucket_input.value.strip(),
                                    cb_user_input.value.strip(),
                                    cb_pass_input.value,
                                    cb_tls_toggle.value,
                                    cb_scope_input.value.strip() or "_default",
                                    cb_collection_input.value.strip() or "tickets",
                                    ch_snap_coll.value.strip() or "snapshots",
                                    _prog,
                                )
                                msg = f"Done — {t_upd} tickets updated, {s_upd} snapshots linked, {errs} errors."
                                cb_status.set_text(msg)
                                cb_progress.set_value(1.0)
                                ui.notify(msg, type="positive" if errs == 0 else "warning")
                            except Exception as exc:
                                cb_status.set_text(f"Migration error: {exc}")
                                ui.notify(str(exc), type="negative")
                            finally:
                                btn_migrate_links.set_enabled(True)

                        btn_migrate_links = ui.button(
                            "Migrate Ticket ↔ Snapshot Links",
                            on_click=_do_migrate_links,
                            icon="link",
                        ).props("color=teal outline").classes("mt-1")

                        # ── Load FROM Couchbase (skip re-scrape for re-embedding) ─────────
                        ui.separator().classes("mt-4")
                        ui.label("Load tickets from Couchbase").classes("text-sm font-semibold text-gray-600 mt-1")
                        ui.label(
                            "Reload previously stored tickets into the session — useful for re-embedding "
                            "without re-scraping the site."
                        ).classes("text-xs text-gray-400")

                        with ui.row().classes("w-full gap-3 mt-2 items-start"):
                            with ui.column().classes("gap-1 flex-1"):
                                ui.label("Filter by organization (optional)").classes("text-xs text-gray-500")
                                cb_load_filter = (
                                    ui.input(placeholder="Type name fragment, select, or leave blank for all")
                                    .props("outlined clearable")
                                    .classes("w-full")
                                )
                            cb_load_org_results = ui.select([], label="Matching organizations").classes("w-72")
                            cb_load_org_results.set_visibility(False)
                            btn_load_from_cb = ui.button("Load from Couchbase", icon="download_for_offline").props("outline color=primary")

                        async def _do_cb_load_org_lookup():
                            q = (cb_load_filter.value or "").strip()
                            if len(q) < 2:
                                cb_load_org_results.set_visibility(False)
                                return
                            if not _CB_AVAILABLE:
                                return
                            try:
                                orgs = await run.io_bound(
                                    search_orgs_from_cb,
                                    cb_url_input.value.strip(), cb_bucket_input.value.strip(),
                                    cb_user_input.value.strip(), cb_pass_input.value,
                                    cb_tls_toggle.value,
                                    cb_scope_input.value.strip() or "_default",
                                    cb_collection_input.value.strip() or "tickets",
                                    q,
                                )
                                cb_load_org_results.options = orgs
                                cb_load_org_results.update()
                                cb_load_org_results.set_visibility(bool(orgs))
                            except Exception:
                                pass

                        def _apply_cb_load_org_select():
                            if cb_load_org_results.value:
                                cb_load_filter.set_value(cb_load_org_results.value)
                                cb_load_org_results.set_visibility(False)

                        cb_load_filter.on_value_change(lambda _: asyncio.ensure_future(_do_cb_load_org_lookup()))
                        cb_load_org_results.on_value_change(lambda _: _apply_cb_load_org_select())

                        cb_load_status   = ui.label("").classes("text-sm text-gray-500 mt-1")
                        cb_load_progress = _PctBar(value=0).classes("mt-1")
                        cb_load_progress.set_visibility(False)

                        async def _do_load_from_cb():
                            btn_load_from_cb.set_enabled(False)
                            cb_load_progress.set_visibility(True)
                            cb_load_progress.set_value(0)
                            cb_load_status.set_text("Connecting …")
                            loop = asyncio.get_event_loop()

                            def _prog(msg: str, pct: float):
                                asyncio.run_coroutine_threadsafe(
                                    _upd_load(msg, pct), loop
                                )

                            async def _upd_load(msg: str, pct: float):
                                cb_load_status.set_text(msg)
                                cb_load_progress.set_value(pct)

                            try:
                                tickets = await run.io_bound(
                                    load_tickets_from_cb,
                                    cb_url_input.value.strip(),
                                    cb_bucket_input.value.strip(),
                                    cb_user_input.value.strip(),
                                    cb_pass_input.value,
                                    cb_tls_toggle.value,
                                    cb_scope_input.value.strip() or "_default",
                                    cb_collection_input.value.strip() or "tickets",
                                    cb_load_filter.value,
                                    _prog,
                                )
                                state["results"] = tickets
                                _results_empty.set_visibility(False)
                                _results_card.set_visibility(True)
                                cb_cust = (cb_load_filter.value or "").strip() or "All Customers"
                                _old_cust2 = state.get("customer_name", "")
                                if _CB_AVAILABLE and state.get("chat_history") and _old_cust2 != cb_cust:
                                    await run.io_bound(
                                        save_customer_chat_history, _old_cust2, list(state["chat_history"]),
                                        cb_url_input.value.strip(), cb_bucket_input.value.strip(),
                                        cb_user_input.value.strip(), cb_pass_input.value, cb_tls_toggle.value,
                                    )
                                state["customer_name"] = cb_cust
                                _SERVER_STATE["results"] = tickets
                                _SERVER_STATE["customer_name"] = cb_cust
                                _set_customer_banner(cb_cust)
                                if _CB_AVAILABLE and _old_cust2 != cb_cust:
                                    _loaded_hist2 = await run.io_bound(
                                        load_customer_chat_history, cb_cust,
                                        cb_url_input.value.strip(), cb_bucket_input.value.strip(),
                                        cb_user_input.value.strip(), cb_pass_input.value, cb_tls_toggle.value,
                                    )
                                    state["chat_history"] = _loaded_hist2
                                    state["chat_session_turns"] = []
                                    state["chat_session_id"] = str(uuid.uuid4())
                                    state["prior_session_block"] = ""
                                    _render_chat()

                                # Auto-extract scores if already present on the docs
                                auto_scores = {
                                    str(t["ticket_id"]): t["score"]
                                    for t in tickets
                                    if t.get("score") and t.get("ticket_id")
                                }
                                if auto_scores:
                                    state["scores"] = auto_scores
                                    _SERVER_STATE["scores"] = auto_scores

                                _refresh_table(tickets)
                                btn_embed.set_enabled(_CB_AVAILABLE)
                                btn_dl_json.set_enabled(True)
                                btn_dl_csv.set_enabled(True)
                                btn_dl_xls.set_enabled(True)
                                btn_score.set_enabled(True)
                                btn_load_scores.set_enabled(True)
                                btn_rescore_all.set_enabled(True)
                                btn_render_charts.set_enabled(True)
                                scored_note = f", {len(auto_scores)} already scored" if auto_scores else ""
                                msg = f"Loaded {len(tickets)} tickets from Couchbase{scored_note}."
                                cb_load_status.set_text(msg)
                                cb_load_progress.set_value(1.0)
                                ui.notify(msg, type="positive")
                            except Exception as exc:
                                cb_load_status.set_text(f"Error: {exc}")
                                ui.notify(str(exc), type="negative")
                            finally:
                                btn_load_from_cb.set_enabled(True)

                        btn_load_from_cb.on("click", _do_load_from_cb)

                        # Enable load button when results are available (mirrored alongside download buttons)
                        # Done by patching enable logic at the scrape-done site — see below.

                    with ui.tab_panel(cfg_embed):
                        with ui.row().classes("items-center justify-between w-full"):
                            ui.label("Data Operations").classes("text-base font-semibold")
                            ui.label("Configure embedding provider in the AI Models tab").classes("text-xs text-gray-400")

                        def _get_embed_config() -> tuple[str, str, str, str, int, int | None]:
                            """Returns (provider, model, api_key, base_url, dims, num_ctx).
                            num_ctx is only used by Ollama; all other providers return None."""
                            provider = ai_emb_provider.value or "Ollama"
                            if provider == "LMStudio":
                                return (
                                    "lmstudio",
                                    (emb_lms_model_input.value or "").strip(),
                                    "",
                                    emb_lms_url_input.value.strip() or "http://localhost:1234",
                                    int(emb_lms_dims_input.value or 768),
                                    None,
                                )
                            elif provider == "Gemini":
                                return (
                                    "gemini",
                                    (emb_gemini_model_input.value or "text-embedding-004").strip(),
                                    emb_gemini_key_input.value,
                                    "",
                                    int(emb_gemini_dims_input.value or 768),
                                    None,
                                )
                            elif provider == "MLX":
                                return (
                                    "mlx",
                                    (emb_mlx_model_input.value or "mixedbread-ai/mxbai-embed-large-v1").strip(),
                                    "",
                                    "",
                                    int(emb_mlx_dims_input.value or 1024),
                                    None,
                                )
                            elif provider == "OpenAI":
                                return (
                                    "openai",
                                    (emb_openai_model_input.value or "text-embedding-3-small").strip(),
                                    emb_openai_key_input.value,
                                    "",
                                    int(emb_openai_dims_input.value or 1536),
                                    None,
                                )
                            else:  # Ollama
                                return (
                                    "ollama",
                                    (emb_ollama_model_input.value or "nomic-embed-text").strip(),
                                    "",
                                    emb_ollama_url_input.value.strip() or "http://localhost:11434",
                                    int(emb_dims_input.value or 768),
                                    int(emb_num_ctx_input.value or 8) * 1024,
                                )

                        emb_status   = ui.label("").classes("text-sm text-gray-500 mt-1")
                        emb_progress = _PctBar(value=0).classes("mt-1")
                        emb_progress.set_visibility(False)

                        async def _do_embed():
                            if not state["results"]:
                                ui.notify("Scrape tickets first.", type="warning")
                                return
                            _cancel.clear()
                            btn_embed.set_enabled(False)
                            btn_create_idx.set_enabled(False)
                            btn_stop_embed.set_enabled(True)
                            emb_progress.set_visibility(True)
                            emb_progress.set_value(0)
                            emb_status.set_text("Starting …")
                            loop = asyncio.get_event_loop()

                            def _prog(msg: str, pct: float):
                                asyncio.run_coroutine_threadsafe(
                                    _upd_emb(msg, pct), loop
                                )

                            async def _upd_emb(msg: str, pct: float):
                                emb_status.set_text(msg)
                                emb_progress.set_value(pct)

                            try:
                                emb_provider, emb_model, emb_api_key, emb_base_url, emb_dims, emb_num_ctx = _get_embed_config()
                                done, errs = await run.io_bound(
                                    embed_all_tickets,
                                    state["results"],
                                    cb_url_input.value.strip(),
                                    cb_bucket_input.value.strip(),
                                    cb_user_input.value.strip(),
                                    cb_pass_input.value,
                                    cb_tls_toggle.value,
                                    cb_scope_input.value.strip() or "_default",
                                    cb_collection_input.value.strip() or "tickets",
                                    emb_provider,
                                    emb_model,
                                    emb_api_key,
                                    emb_base_url,
                                    emb_dims,
                                    _prog,
                                    _cancel,
                                    emb_num_ctx,
                                    int(embed_parallel_input.value or 1),
                                )
                                msg = f"Done — {done} embedded, {errs} errors."
                                emb_status.set_text(msg)
                                emb_progress.set_value(1.0)
                                ui.notify(msg, type="positive" if errs == 0 else "warning")
                                btn_create_idx.set_enabled(True)
                            except Exception as exc:
                                emb_status.set_text(f"Error: {exc}")
                                ui.notify(str(exc), type="negative")
                            finally:
                                btn_embed.set_enabled(True)
                                btn_stop_embed.set_enabled(False)

                        async def _do_create_index():
                            btn_create_idx.set_enabled(False)
                            emb_status.set_text("Creating vector index …")
                            try:
                                await run.io_bound(
                                    create_vector_index,
                                    cb_url_input.value.strip(),
                                    cb_bucket_input.value.strip(),
                                    cb_user_input.value.strip(),
                                    cb_pass_input.value,
                                    cb_tls_toggle.value,
                                    cb_scope_input.value.strip() or "_default",
                                    cb_collection_input.value.strip() or "tickets",
                                    _get_embed_config()[4],
                                )
                                emb_status.set_text("Vector index created — ready for chat.")
                                ui.notify("Vector index created!", type="positive")
                            except Exception as exc:
                                emb_status.set_text(f"Index error: {exc}")
                                ui.notify(str(exc), type="negative")
                            finally:
                                btn_create_idx.set_enabled(True)

                        async def _do_embed_snaps_from_cb():
                            btn_embed_snaps.set_enabled(False)
                            emb_status.set_text("Embedding snapshots from Couchbase …")
                            emb_progress.set_visibility(True)
                            emb_progress.set_value(0)
                            loop = asyncio.get_event_loop()

                            async def _upd_snaps(msg: str, pct: float):
                                emb_status.set_text(msg)
                                emb_progress.set_value(pct)

                            def _prog(msg: str, pct: float):
                                asyncio.run_coroutine_threadsafe(_upd_snaps(msg, pct), loop)

                            try:
                                ep, em, ek, eu, ed, _ = _get_embed_config()
                                done, errs = await run.io_bound(
                                    embed_snapshots_from_cb,
                                    cb_url_input.value.strip(),
                                    cb_bucket_input.value.strip(),
                                    cb_user_input.value.strip(),
                                    cb_pass_input.value,
                                    cb_tls_toggle.value,
                                    cb_scope_input.value.strip() or "_default",
                                    ch_snap_coll.value.strip() or "snapshots",
                                    ep, em, ek, eu, int(ed or 1024),
                                    _prog,
                                    int(embed_parallel_input.value or 1),
                                )
                                msg = f"Snapshot embed complete — {done} embedded, {errs} errors."
                                emb_status.set_text(msg)
                                emb_progress.set_value(1.0)
                                ui.notify(msg, type="positive" if errs == 0 else "warning")
                            except Exception as exc:
                                emb_status.set_text(f"Snapshot embed error: {exc}")
                                ui.notify(str(exc), type="negative")
                            finally:
                                btn_embed_snaps.set_enabled(True)

                        async def _do_summarize():
                            btn_summarize.set_enabled(False)
                            emb_status.set_text("Summarizing tickets …")
                            emb_progress.set_visibility(True)
                            emb_progress.set_value(0)
                            loop = asyncio.get_event_loop()

                            async def _upd_sum(msg: str, pct: float):
                                emb_status.set_text(msg)
                                emb_progress.set_value(pct)

                            def _prog(msg: str, pct: float):
                                asyncio.run_coroutine_threadsafe(_upd_sum(msg, pct), loop)

                            try:
                                llm_p, llm_m, llm_k, llm_u = _get_llm_config()
                                done, errs = await run.io_bound(
                                    summarize_tickets_from_cb,
                                    cb_url_input.value.strip(),
                                    cb_bucket_input.value.strip(),
                                    cb_user_input.value.strip(),
                                    cb_pass_input.value,
                                    cb_tls_toggle.value,
                                    cb_scope_input.value.strip() or "_default",
                                    cb_collection_input.value.strip() or "supportal",
                                    cb_summary_coll.value.strip() or "summary",
                                    llm_p, llm_m, llm_k, llm_u,
                                    _prog,
                                    main_cust_input.value.strip(),
                                    summarize_force_cb.value,  # force — overwrite existing if checked
                                    1,      # max_workers — LLM calls are sequential by default
                                )
                                msg = f"Summarization complete — {done} written, {errs} errors."
                                emb_status.set_text(msg)
                                emb_progress.set_value(1.0)
                                ui.notify(msg, type="positive" if errs == 0 else "warning")
                            except Exception as exc:
                                emb_status.set_text(f"Summarize error: {exc}")
                                ui.notify(str(exc), type="negative")
                            finally:
                                btn_summarize.set_enabled(True)

                        async def _do_backfill():
                            btn_backfill.set_enabled(False)
                            emb_status.set_text("Backfilling analytics fields …")
                            emb_progress.set_visibility(True)
                            emb_progress.set_value(0)
                            loop = asyncio.get_event_loop()

                            async def _upd_backfill(msg: str, pct: float):
                                emb_status.set_text(msg)
                                emb_progress.set_value(pct)

                            def _prog(msg: str, pct: float):
                                asyncio.run_coroutine_threadsafe(
                                    _upd_backfill(msg, pct), loop
                                )

                            try:
                                updated, errs = await run.io_bound(
                                    backfill_analytics_fields,
                                    cb_url_input.value.strip(),
                                    cb_bucket_input.value.strip(),
                                    cb_user_input.value.strip(),
                                    cb_pass_input.value,
                                    cb_tls_toggle.value,
                                    cb_scope_input.value.strip() or "_default",
                                    cb_collection_input.value.strip() or "tickets",
                                    _prog,
                                )
                                msg = f"Backfill complete — {updated} updated, {errs} errors."
                                emb_status.set_text(msg)
                                emb_progress.set_value(1.0)
                                ui.notify(msg, type="positive" if errs == 0 else "warning")
                            except Exception as exc:
                                emb_status.set_text(f"Backfill error: {exc}")
                                ui.notify(str(exc), type="negative")
                            finally:
                                btn_backfill.set_enabled(True)

                        async def _do_backfill_cbse():
                            btn_backfill_cbse.set_enabled(False)
                            emb_status.set_text("Backfilling missing CBSE/JIRA fields …")
                            emb_progress.set_visibility(True)
                            emb_progress.set_value(0)
                            loop = asyncio.get_event_loop()

                            async def _upd_cbse(msg: str, pct: float):
                                emb_status.set_text(msg)
                                emb_progress.set_value(pct)

                            def _prog_cbse(msg: str, pct: float):
                                asyncio.run_coroutine_threadsafe(_upd_cbse(msg, pct), loop)

                            try:
                                updated, errs = await run.io_bound(
                                    backfill_missing_cbse_fields,
                                    cb_url_input.value.strip(),
                                    cb_bucket_input.value.strip(),
                                    cb_user_input.value.strip(),
                                    cb_pass_input.value,
                                    cb_tls_toggle.value,
                                    cb_scope_input.value.strip() or "_default",
                                    cb_collection_input.value.strip() or "tickets",
                                    _prog_cbse,
                                )
                                msg = f"CBSE backfill complete — {updated} updated, {errs} errors."
                                emb_status.set_text(msg)
                                emb_progress.set_value(1.0)
                                ui.notify(msg, type="positive" if errs == 0 else "warning")
                            except Exception as exc:
                                emb_status.set_text(f"CBSE backfill error: {exc}")
                                ui.notify(str(exc), type="negative")
                            finally:
                                btn_backfill_cbse.set_enabled(True)

                        async def _do_backfill_last_reply():
                            btn_backfill_last_reply.set_enabled(False)
                            emb_status.set_text("Backfilling last_comment_at from stored comments …")
                            emb_progress.set_visibility(True)
                            emb_progress.set_value(0)
                            loop = asyncio.get_event_loop()

                            async def _upd_lr(msg: str, pct: float):
                                emb_status.set_text(msg)
                                emb_progress.set_value(pct)

                            def _prog_lr(msg: str, pct: float):
                                asyncio.run_coroutine_threadsafe(_upd_lr(msg, pct), loop)

                            try:
                                updated, errs = await run.io_bound(
                                    backfill_last_comment_at,
                                    cb_url_input.value.strip(),
                                    cb_bucket_input.value.strip(),
                                    cb_user_input.value.strip(),
                                    cb_pass_input.value,
                                    cb_tls_toggle.value,
                                    cb_scope_input.value.strip() or "_default",
                                    cb_collection_input.value.strip() or "tickets",
                                    _prog_lr,
                                )
                                msg = f"Last Reply backfill complete — {updated} tickets updated, {errs} errors."
                                emb_status.set_text(msg)
                                emb_progress.set_value(1.0)
                                ui.notify(msg, type="positive" if errs == 0 else "warning")
                            except Exception as exc:
                                emb_status.set_text(f"Last Reply backfill error: {exc}")
                                ui.notify(str(exc), type="negative")
                            finally:
                                btn_backfill_last_reply.set_enabled(True)

                        with ui.row().classes("items-center gap-4 mt-2"):
                            summarize_force_cb = ui.checkbox("Force re-summarize (overwrite existing)", value=False)

                        with ui.row().classes("gap-3 mt-2 flex-wrap"):
                            btn_embed       = ui.button("Embed Tickets",             on_click=_do_embed,                  icon="model_training").props("outline color=primary")
                            btn_embed_snaps = ui.button("Embed All Snapshots",       on_click=_do_embed_snaps_from_cb,    icon="hub").props("outline color=indigo")
                            btn_summarize   = ui.button("Summarize Tickets",         on_click=_do_summarize,              icon="summarize").props("outline color=teal")
                            btn_create_idx  = ui.button("Create Vector Index",       on_click=_do_create_index,           icon="manage_search").props("outline color=secondary")
                            btn_backfill           = ui.button("Backfill Analytics Fields", on_click=_do_backfill,              icon="auto_fix_high").props("outline color=brown")
                            btn_backfill_cbse      = ui.button("Backfill Missing CBSEs",    on_click=_do_backfill_cbse,         icon="bug_report").props("outline color=deep-orange")
                            btn_backfill_last_reply = ui.button("Backfill Last Reply Dates", on_click=_do_backfill_last_reply,  icon="chat_bubble_outline").props("outline color=purple")
                            btn_stop_embed    = ui.button("Stop", icon="stop_circle", on_click=lambda: (_cancel.set(), btn_stop_embed.set_enabled(False))).props("outline color=red")
                            btn_stop_embed.set_enabled(False)
                        btn_embed.set_enabled(_CB_AVAILABLE)
                        btn_embed_snaps.set_enabled(_CB_AVAILABLE)
                        btn_summarize.set_enabled(_CB_AVAILABLE)
                        btn_create_idx.set_enabled(_CB_AVAILABLE)

                    with ui.tab_panel(cfg_chat_mem):
                        ui.label("Chat Memory & Cache").classes("text-base font-semibold")
                        ui.label(
                            "Cache embeddings and vector-search results in Couchbase to avoid repeating "
                            "expensive operations for the same question. Permanent memory summaries preserve "
                            "Q&A knowledge even after short-lived cache entries expire."
                        ).classes("text-xs text-gray-400 mt-1")

                        ui.separator().classes("my-3")

                        # Cache collection override
                        with ui.row().classes("items-center gap-3 w-full mt-1"):
                            cache_collection_input = ui.input(
                                "Cache collection",
                                placeholder="(leave blank to use global collection above)",
                            ).classes("flex-1").tooltip(
                                "Dedicated Couchbase collection for chat_cache::* and chat_memory::* docs. "
                                "Leave blank to store them alongside your ticket documents."
                            )

                        ui.separator().classes("my-3")
                        ui.label("TTL settings").classes("text-xs font-semibold text-gray-600")
                        ui.label(
                            "Set to 0 to store permanently (recommended default). "
                            "A positive value causes the entry to expire automatically."
                        ).classes("text-xs text-gray-400")

                        with ui.row().classes("items-end gap-6 mt-2 flex-wrap"):
                            with ui.column().classes("gap-1"):
                                ui.label("Embed cache TTL").classes("text-xs text-gray-500")
                                embed_cache_ttl = ui.number(
                                    "days  (0 = forever)", value=0, min=0, max=365, format="%.0f"
                                ).classes("w-44").tooltip(
                                    "How many days to keep a cached embedding vector. "
                                    "0 = never expire."
                                )
                            with ui.column().classes("gap-1"):
                                ui.label("Search cache TTL").classes("text-xs text-gray-500")
                                search_cache_ttl = ui.number(
                                    "hours  (0 = forever)", value=0, min=0, max=720, format="%.0f"
                                ).classes("w-44").tooltip(
                                    "How many hours to keep cached vector-search results. "
                                    "0 = never expire. Lower values refresh results as tickets change."
                                )

                        ui.separator().classes("my-3")
                        ui.label("Permanent memory").classes("text-xs font-semibold text-gray-600")
                        ui.label(
                            "When a cache entry has a TTL (will eventually expire), "
                            "a permanent summary of the question and answer is stored under "
                            "chat_cache::memory::* with no expiry, so the knowledge is never lost."
                        ).classes("text-xs text-gray-400")

                        with ui.row().classes("items-center gap-3 mt-2"):
                            ui.label("Store memory summary when TTL is set").classes("text-sm")
                            store_memory_toggle = ui.switch(value=True).tooltip(
                                "Write a permanent chat_cache::memory:: doc (question + answer summary + "
                                "ticket IDs) whenever a cache entry is created with a TTL > 0."
                            )

                        ui.separator().classes("my-3")
                        cache_mgmt_status = ui.label("").classes("text-sm text-gray-500")

                        def _cache_cb_args_cfg():
                            raw = cache_collection_input.value.strip()
                            col = raw if raw else (cb_collection_input.value.strip() or "tickets")
                            return (
                                cb_url_input.value.strip(),
                                cb_bucket_input.value.strip(),
                                cb_user_input.value.strip(),
                                cb_pass_input.value,
                                cb_tls_toggle.value,
                                cb_scope_input.value.strip() or "_default",
                                col,
                            )

                        async def _do_clear_cache():
                            btn_clear_cache_cfg.set_enabled(False)
                            cache_mgmt_status.set_text("Clearing embed + search cache…")
                            n = await run.io_bound(chat_cache_clear, *_cache_cb_args_cfg())
                            cache_mgmt_status.set_text(
                                f"Cache cleared — {n} entr{'y' if n == 1 else 'ies'} removed."
                            )
                            btn_clear_cache_cfg.set_enabled(True)

                        async def _do_clear_memory():
                            btn_clear_memory_cfg.set_enabled(False)
                            cache_mgmt_status.set_text("Clearing permanent memory summaries…")
                            n = await run.io_bound(chat_memory_clear, *_cache_cb_args_cfg())
                            cache_mgmt_status.set_text(
                                f"Memory cleared — {n} entr{'y' if n == 1 else 'ies'} removed."
                            )
                            btn_clear_memory_cfg.set_enabled(True)

                        with ui.row().classes("gap-3 mt-1"):
                            btn_clear_cache_cfg = ui.button(
                                "Clear Cache", icon="delete_sweep", on_click=_do_clear_cache
                            ).props("outline color=orange")
                            btn_clear_memory_cfg = ui.button(
                                "Clear Memory", icon="psychology_alt", on_click=_do_clear_memory
                            ).props("outline color=deep-purple")

                    with ui.tab_panel(cfg_analytics):
                        ui.label("Org Name Consolidation").classes("text-base font-semibold")
                        ui.label(
                            "Automatically merge tickets from the same customer even when their "
                            "organization names differ in casing, punctuation, or legal suffixes "
                            "(e.g. \"Acme Inc\", \"ACME CORP\", \"acme corp\"). "
                            "Raise the threshold to only merge near-identical names; lower it to "
                            "be more aggressive. Tip: if two companies share a similar short name "
                            "(e.g. \"Acme\" vs \"Acme Technologies\"), raise the threshold to 95-100%."
                        ).classes("text-xs text-gray-400 mt-1")

                        ui.separator().classes("my-3")

                        with ui.row().classes("items-center gap-6 flex-wrap mt-1"):
                            org_consolidation_toggle = ui.switch(
                                "Consolidate similar org names", value=True
                            ).tooltip(
                                "When enabled, variant spellings of the same customer are grouped "
                                "under a single canonical name in all charts and exports."
                            )
                            with ui.column().classes("gap-1"):
                                org_threshold_label = ui.label(
                                    "Similarity threshold: 90%"
                                ).classes("text-xs text-gray-500")
                                org_threshold_slider = ui.slider(
                                    min=80, max=100, value=90, step=1
                                ).classes("w-52").tooltip(
                                    "How similar two normalized names must be to be merged. "
                                    "90% (default) merges near-identical names. "
                                    "80% is more aggressive; 98% is conservative."
                                )

                        def _org_threshold_changed(e):
                            org_threshold_label.set_text(
                                f"Similarity threshold: {int(e.value)}%"
                            )

                        org_threshold_slider.on("update:model-value", _org_threshold_changed)

                        org_consolidation_status = ui.label("").classes(
                            "text-sm text-gray-500 mt-2"
                        )
                        org_preview_area = ui.column().classes("w-full mt-2 gap-0")

                        async def _preview_org_merges():
                            results = state.get("results") or []
                            if not results:
                                org_consolidation_status.set_text(
                                    "No tickets loaded — load data first to preview merges."
                                )
                                return
                            org_consolidation_status.set_text("Computing…")
                            org_preview_area.clear()
                            enabled   = org_consolidation_toggle.value
                            threshold = org_threshold_slider.value / 100.0
                            org_map   = await run.io_bound(
                                build_org_name_map, results, enabled, threshold
                            )
                            # Build groups: canonical → sorted list of raw names
                            groups: dict[str, list] = {}
                            for raw, canon in org_map.items():
                                groups.setdefault(canon, [])
                                if raw not in groups[canon]:
                                    groups[canon].append(raw)
                            merged_groups = {
                                c: sorted(raws) for c, raws in groups.items() if len(raws) > 1
                            }
                            if not merged_groups:
                                org_consolidation_status.set_text(
                                    "No merges found at this threshold — all org names are "
                                    "already distinct."
                                )
                                return
                            total_variants = sum(len(v) for v in merged_groups.values())
                            org_consolidation_status.set_text(
                                f"{len(merged_groups)} group(s) — {total_variants} variant names "
                                f"consolidated."
                            )
                            with org_preview_area:
                                rows = [
                                    {"Canonical name": canon, "Merged variants": ", ".join(raws)}
                                    for canon, raws in sorted(merged_groups.items())
                                ]
                                ui.table(
                                    columns=[
                                        {"name": "canon",    "label": "Canonical name",  "field": "Canonical name",  "align": "left"},
                                        {"name": "variants", "label": "Merged variants", "field": "Merged variants", "align": "left"},
                                    ],
                                    rows=rows,
                                ).classes("w-full text-xs").props("dense flat bordered")

                        def _save_org_settings():
                            profiles = _load_settings_file()
                            profiles["__org_consolidation__"] = {
                                "enabled":   org_consolidation_toggle.value,
                                "threshold": org_threshold_slider.value,
                            }
                            _save_settings_file(profiles)
                            org_consolidation_status.set_text("Settings saved.")

                        with ui.row().classes("gap-3 mt-3"):
                            ui.button(
                                "Preview Merges", icon="preview", on_click=_preview_org_merges
                            ).props("outline color=blue")
                            ui.button(
                                "Save Settings", icon="save", on_click=_save_org_settings
                            ).props("outline color=green")

                        # Auto-load saved org consolidation settings on page open
                        _org_cfg_init = _load_settings_file().get("__org_consolidation__", {})
                        if _org_cfg_init:
                            org_consolidation_toggle.set_value(
                                bool(_org_cfg_init.get("enabled", True))
                            )
                            _thr_init = int(_org_cfg_init.get("threshold", 90))
                            org_threshold_slider.set_value(_thr_init)
                            org_threshold_label.set_text(f"Similarity threshold: {_thr_init}%")

                    with ui.tab_panel(cfg_ai):
                        with ui.row().classes("items-center justify-between w-full"):
                            ui.label("AI Provider Configuration").classes("text-base font-semibold")
                            ui.label("Configure embedding and LLM models for each provider. Use Fetch Models to populate available options.").classes("text-xs text-gray-400")

                        ai_status = ui.label("").classes("text-sm text-gray-500 mt-1")

                        # Active provider selectors
                        with ui.row().classes("gap-6 mt-3 items-center flex-wrap"):
                            with ui.column().classes("gap-1"):
                                ui.label("Active Embedding Provider").classes("text-xs font-semibold text-blue-600")
                                ai_emb_provider = ui.select(
                                    ["Ollama", "LMStudio", "OpenAI", "Gemini", "MLX"],
                                    value="Ollama",
                                ).classes("w-44")
                            with ui.column().classes("gap-1"):
                                ui.label("Active LLM Provider").classes("text-xs font-semibold text-purple-600")
                                ai_llm_provider = ui.select(
                                    ["Claude", "Ollama", "LMStudio", "OpenAI", "Gemini"],
                                    value="Claude",
                                ).classes("w-44")

                        ui.separator().classes("my-3")

                        # Provider tabs
                        with ui.tabs().classes("w-full") as ai_prov_tabs:
                            ai_tab_ollama   = ui.tab("Ollama")
                            ai_tab_lmstudio = ui.tab("LMStudio")
                            ai_tab_openai   = ui.tab("OpenAI")
                            ai_tab_gemini   = ui.tab("Gemini")
                            ai_tab_claude   = ui.tab("Claude")
                            ai_tab_mlx      = ui.tab("MLX")

                        with ui.tab_panels(ai_prov_tabs, value=ai_tab_ollama).classes("w-full border-t"):

                            # ── Ollama ────────────────────────────────────────────────────
                            with ui.tab_panel(ai_tab_ollama):
                                with ui.row().classes("items-end gap-3 w-full flex-wrap"):
                                    emb_ollama_url_input = ui.input("Base URL", value="http://localhost:11434").classes("flex-1")
                                    async def _fetch_ollama():
                                        url = emb_ollama_url_input.value.strip() or "http://localhost:11434"
                                        ai_status.set_text("Fetching Ollama models …")
                                        try:
                                            models = await run.io_bound(fetch_ollama_models, url)
                                            emb_ollama_model_input.options = models
                                            emb_ollama_model_input.update()
                                            ollama_chat_model_input.options = models
                                            ollama_chat_model_input.update()
                                            ai_status.set_text(f"Ollama: {len(models)} model(s) found.")
                                        except Exception as exc:
                                            ai_status.set_text(f"Ollama fetch error: {exc}")
                                    ui.button("Fetch Models", icon="refresh", on_click=_fetch_ollama).props("outline color=teal").classes("h-10")

                                ui.label("Embedding").classes("text-xs font-semibold text-blue-600 mt-3")
                                with ui.grid(columns=3).classes("w-full gap-3"):
                                    emb_ollama_model_input = ui.select(
                                        ["nomic-embed-text"], label="Embedding Model",
                                        value="nomic-embed-text",
                                    ).props('use-input hide-selected fill-input input-debounce=0 new-value-mode=add').classes("w-full")
                                    emb_dims_input = ui.number("Vector Dims", value=768, min=64, max=8192).classes("w-full")
                                    emb_num_ctx_input = ui.number(
                                        "num_ctx (K)", value=8, min=2, max=512,
                                    ).classes("w-full").props(
                                        'hint="Auto-set when model is selected"'
                                    ).tooltip(
                                        "Input context window for the embedding model (thousands of tokens). "
                                        "Increase if tickets are long — text beyond this limit is silently truncated before embedding."
                                    )

                                ui.label("Chat / LLM").classes("text-xs font-semibold text-purple-600 mt-3")
                                ollama_chat_model_input = ui.select(
                                    ["llama3.2"], label="LLM Model",
                                    value="llama3.2",
                                ).props('use-input hide-selected fill-input input-debounce=0 new-value-mode=add').classes("w-full")
                                # Keep alias for profile compat
                                ollama_chat_url_input = emb_ollama_url_input

                                async def _on_ollama_chat_model_change(e):
                                    model = (e.value or "").strip()
                                    if not model:
                                        return
                                    url = emb_ollama_url_input.value.strip() or "http://localhost:11434"
                                    info = await run.io_bound(fetch_ollama_model_info, url, model)
                                    ctx = info.get("num_ctx")
                                    if ctx:
                                        ctx_k = max(8, round(ctx / 1024))
                                        score_ctx_input.set_value(ctx_k)
                                        score_ctx_input.props(
                                            f'hint="Model default: {ctx:,} tokens ({ctx_k}K)"'
                                        )
                                    else:
                                        score_ctx_input.props('hint="Could not detect model default"')
                                    # Auto-enable "disable thinking" if model supports it
                                    thinking = info.get("thinking", False)
                                    score_no_think_toggle.set_value(thinking)
                                    caps = info.get("caps", [])
                                    cap_str = ", ".join(caps) if caps else "unknown"
                                    score_no_think_toggle.tooltip(
                                        f"Capabilities: {cap_str}. "
                                        "Uses Ollama native API with think=false — suppresses "
                                        "Qwen3/QwQ reasoning traces for faster JSON output."
                                    )

                                ollama_chat_model_input.on_value_change(_on_ollama_chat_model_change)

                                async def _on_ollama_emb_model_change(e):
                                    model = (e.value or "").strip()
                                    if not model:
                                        return
                                    url = emb_ollama_url_input.value.strip() or "http://localhost:11434"
                                    info = await run.io_bound(fetch_ollama_model_info, url, model)
                                    ctx = info.get("num_ctx")
                                    if ctx:
                                        ctx_k = max(2, round(ctx / 1024))
                                        emb_num_ctx_input.set_value(ctx_k)
                                        emb_num_ctx_input.props(
                                            f'hint="Model default: {ctx:,} tokens ({ctx_k}K)"'
                                        )
                                    else:
                                        emb_num_ctx_input.props('hint="Could not detect model default"')

                                emb_ollama_model_input.on_value_change(_on_ollama_emb_model_change)

                            # ── LMStudio ──────────────────────────────────────────────────
                            with ui.tab_panel(ai_tab_lmstudio):
                                with ui.row().classes("items-end gap-3 w-full flex-wrap"):
                                    emb_lms_url_input = ui.input("Base URL", value="http://localhost:1234").classes("flex-1")
                                    async def _fetch_lmstudio():
                                        url = emb_lms_url_input.value.strip() or "http://localhost:1234"
                                        ai_status.set_text("Fetching LMStudio models …")
                                        try:
                                            models = await run.io_bound(fetch_openai_compat_models, url, "")
                                            emb_lms_model_input.options = models
                                            emb_lms_model_input.update()
                                            lms_model_input.options = models
                                            lms_model_input.update()
                                            ai_status.set_text(f"LMStudio: {len(models)} model(s) found.")
                                        except Exception as exc:
                                            ai_status.set_text(f"LMStudio fetch error: {exc}")
                                    ui.button("Fetch Models", icon="refresh", on_click=_fetch_lmstudio).props("outline color=teal").classes("h-10")

                                ui.label("Embedding").classes("text-xs font-semibold text-blue-600 mt-3")
                                with ui.grid(columns=2).classes("w-full gap-3"):
                                    emb_lms_model_input = ui.select(
                                        ["text-embedding-nomic-embed-text-v1.5"], label="Embedding Model",
                                        value="text-embedding-nomic-embed-text-v1.5",
                                    ).props('use-input hide-selected fill-input input-debounce=0 new-value-mode=add').classes("w-full")
                                    emb_lms_dims_input = ui.number("Vector Dims", value=768, min=64, max=8192).classes("w-full")

                                ui.label("Chat / LLM").classes("text-xs font-semibold text-purple-600 mt-3")
                                lms_model_input = ui.select(
                                    ["local-model"], label="LLM Model",
                                    value="local-model",
                                ).props('use-input hide-selected fill-input input-debounce=0 new-value-mode=add').classes("w-full")
                                lms_url_input = emb_lms_url_input

                            # ── OpenAI ────────────────────────────────────────────────────
                            with ui.tab_panel(ai_tab_openai):
                                with ui.row().classes("items-end gap-3 w-full flex-wrap"):
                                    emb_openai_key_input = ui.input("API Key").props("type=password").classes("flex-1")
                                    async def _fetch_openai():
                                        key = emb_openai_key_input.value
                                        ai_status.set_text("Fetching OpenAI models …")
                                        try:
                                            models = await run.io_bound(fetch_openai_compat_models, "https://api.openai.com", key)
                                            emb_openai_model_input.options = models
                                            emb_openai_model_input.update()
                                            openai_llm_model_input.options = models
                                            openai_llm_model_input.update()
                                            ai_status.set_text(f"OpenAI: {len(models)} model(s) found.")
                                        except Exception as exc:
                                            ai_status.set_text(f"OpenAI fetch error: {exc}")
                                    ui.button("Fetch Models", icon="refresh", on_click=_fetch_openai).props("outline color=teal").classes("h-10")

                                ui.label("Embedding").classes("text-xs font-semibold text-blue-600 mt-3")
                                with ui.grid(columns=2).classes("w-full gap-3"):
                                    emb_openai_model_input = ui.select(
                                        ["text-embedding-3-small", "text-embedding-3-large", "text-embedding-ada-002"],
                                        label="Embedding Model",
                                        value="text-embedding-3-small",
                                    ).props('use-input hide-selected fill-input input-debounce=0 new-value-mode=add').classes("w-full")
                                    emb_openai_dims_input = ui.number("Vector Dims", value=1536, min=64, max=3072).classes("w-full")

                                ui.label("Chat / LLM").classes("text-xs font-semibold text-purple-600 mt-3")
                                openai_llm_model_input = ui.select(
                                    ["gpt-4o", "gpt-4o-mini", "gpt-4-turbo", "gpt-3.5-turbo"],
                                    label="LLM Model",
                                    value="gpt-4o",
                                ).props('use-input hide-selected fill-input input-debounce=0 new-value-mode=add').classes("w-full")

                            # ── Gemini ────────────────────────────────────────────────────
                            with ui.tab_panel(ai_tab_gemini):
                                emb_gemini_key_input = ui.input("API Key").props("type=password").classes("w-full")
                                gemini_key_input = emb_gemini_key_input

                                ui.label("Embedding").classes("text-xs font-semibold text-blue-600 mt-3")
                                with ui.grid(columns=2).classes("w-full gap-3"):
                                    emb_gemini_model_input = ui.select(
                                        ["text-embedding-004", "text-embedding-preview-0409"],
                                        label="Embedding Model",
                                        value="text-embedding-004",
                                    ).props('use-input hide-selected fill-input input-debounce=0 new-value-mode=add').classes("w-full")
                                    emb_gemini_dims_input = ui.number("Vector Dims", value=768, min=64, max=3072).classes("w-full")

                                ui.label("Chat / LLM").classes("text-xs font-semibold text-purple-600 mt-3")
                                gemini_model_input = ui.select(
                                    ["gemini-2.0-flash", "gemini-2.5-pro-preview-03-25", "gemini-1.5-pro"],
                                    label="LLM Model",
                                    value="gemini-2.0-flash",
                                ).props('use-input hide-selected fill-input input-debounce=0 new-value-mode=add').classes("w-full")

                            # ── Claude ────────────────────────────────────────────────────
                            with ui.tab_panel(ai_tab_claude):
                                ui.label("Anthropic Claude — LLM only (no embedding support)").classes("text-xs text-gray-500 mb-2")
                                with ui.grid(columns=2).classes("w-full gap-3"):
                                    claude_key_input = ui.input("API Key").props("type=password").classes("w-full")
                                    claude_model_input = ui.select(
                                        ["claude-opus-4-6", "claude-sonnet-4-6", "claude-haiku-4-5-20251001"],
                                        label="Model",
                                        value="claude-sonnet-4-6",
                                    ).props('use-input hide-selected fill-input input-debounce=0 new-value-mode=add').classes("w-full")

                            # ── MLX ───────────────────────────────────────────────────────
                            with ui.tab_panel(ai_tab_mlx):
                                ui.label(
                                    "Runs locally via mlx-embeddings — no server needed. "
                                    "Embedding only. Use any mlx-community model from HuggingFace."
                                ).classes("text-xs text-gray-500 mb-2")
                                with ui.grid(columns=2).classes("w-full gap-3"):
                                    emb_mlx_model_input = ui.input(
                                        "HuggingFace model ID",
                                        value="mixedbread-ai/mxbai-embed-large-v1",
                                    ).classes("w-full")
                                    emb_mlx_dims_input = ui.number("Vector Dims", value=1024, min=64, max=8192).classes("w-full")

                        # Keep legacy tab aliases for profile compatibility
                        emb_tabs = ai_prov_tabs
                        emb_tab_ollama   = ai_tab_ollama
                        emb_tab_lmstudio = ai_tab_lmstudio
                        emb_tab_gemini   = ai_tab_gemini
                        emb_tab_mlx      = ai_tab_mlx
                        emb_tab_openai   = ai_tab_openai
                        llm_tabs         = ai_prov_tabs
                        tab_claude       = ai_tab_claude
                        tab_gemini       = ai_tab_gemini
                        tab_ollama       = ai_tab_ollama
                        tab_lmstudio     = ai_tab_lmstudio

                    def _warn_if_small_model(model: str) -> None:
                        """Notify if the model name suggests a very small / 8B-or-under variant."""
                        _low = (model or "").lower()
                        _small_hints = ("3b", "7b", "8b", "1b", "mini", "nano", "tiny", "phi-2", "phi2")
                        if any(h in _low for h in _small_hints):
                            ui.notify(
                                f"⚠ Small model detected ({model}). Scoring quality may be reduced.",
                                type="warning",
                                timeout=6000,
                            )

                    def _get_llm_config():
                        """Return (provider, model, api_key, base_url) from current UI inputs."""
                        provider = (ai_llm_provider.value or "Claude").lower()
                        if provider == "claude":
                            return (
                                provider,
                                claude_model_input.value or "claude-sonnet-4-6",
                                claude_key_input.value or "",
                                "",
                            )
                        elif provider == "gemini":
                            return (
                                provider,
                                gemini_model_input.value or "gemini-2.0-flash",
                                gemini_key_input.value or "",
                                "",
                            )
                        elif provider == "openai":
                            return (
                                provider,
                                openai_llm_model_input.value or "gpt-4o",
                                emb_openai_key_input.value or "",
                                "",
                            )
                        elif provider in ("lmstudio",):
                            return (
                                provider,
                                lms_model_input.value or "local-model",
                                "",
                                emb_lms_url_input.value or "http://localhost:1234",
                            )
                        else:  # ollama
                            return (
                                "ollama",
                                ollama_chat_model_input.value or "llama3.2",
                                "",
                                emb_ollama_url_input.value or "http://localhost:11434",
                            )

                    # ── Preflight tab ─────────────────────────────────────────────────────
                    with ui.tab_panel(cfg_preflight):
                        ui.label("Preflight Checks").classes("text-base font-semibold mb-1")
                        ui.label(
                            "Verify that all configured endpoints and models are reachable before scraping or scoring."
                        ).classes("text-xs text-gray-500 mb-3")

                        # Each check: (key, label)
                        _PF_CHECKS = [
                            ("supportal",     "Supportal (customer page reachable)"),
                            ("analytics_api", "Analytics API (query endpoint)"),
                            ("emb_model",     "Embedding model"),
                            ("llm_model",     "LLM model (test generation)"),
                            ("cb_sdk",        "Couchbase SDK connection"),
                        ]

                        pf_rows: dict = {}   # key → {"icon": ui.icon, "label": ui.label, "ts": ui.label}
                        with ui.column().classes("gap-2 w-full"):
                            for _pf_key, _pf_label in _PF_CHECKS:
                                with ui.row().classes("items-center gap-3"):
                                    _ico = ui.icon("radio_button_unchecked").classes("text-gray-300 text-xl")
                                    _lbl = ui.label(_pf_label).classes("text-sm")
                                    _ts  = ui.label("").classes("text-xs text-gray-400 ml-auto")
                                    pf_rows[_pf_key] = {"icon": _ico, "label": _lbl, "ts": _ts}

                        pf_summary = ui.label("").classes("text-sm mt-3 font-semibold")

                        with ui.row().classes("items-center gap-3 mt-1"):
                            pf_last_run = ui.label("Never run").classes("text-xs text-gray-400")
                            pf_auto_refresh = ui.switch("Auto-refresh (60s)").classes("text-xs")
                            pf_auto_refresh.tooltip("Re-run all checks every 60 seconds")

                        def _pf_set(key: str, ok: bool | None, detail: str = ""):
                            row = pf_rows.get(key)
                            if not row:
                                return
                            if ok is None:
                                row["icon"].set_name("hourglass_empty")
                                row["icon"].classes(replace="text-yellow-500 text-xl")
                                suffix = " — checking…"
                                row["ts"].set_text("")
                            elif ok:
                                row["icon"].set_name("check_circle")
                                row["icon"].classes(replace="text-green-500 text-xl")
                                suffix = f" — OK{(' (' + detail + ')') if detail else ''}"
                                row["ts"].set_text(datetime.datetime.now().strftime("%H:%M:%S"))
                            else:
                                row["icon"].set_name("cancel")
                                row["icon"].classes(replace="text-red-500 text-xl")
                                suffix = f" — FAIL{(': ' + detail) if detail else ''}"
                                row["ts"].set_text(datetime.datetime.now().strftime("%H:%M:%S"))
                            base = dict(_PF_CHECKS).get(key, key)
                            row["label"].set_text(base + suffix)

                        def _pf_reset():
                            for _k, _lbl in _PF_CHECKS:
                                row = pf_rows[_k]
                                row["icon"].set_name("radio_button_unchecked")
                                row["icon"].classes(replace="text-gray-300 text-xl")
                                row["label"].set_text(_lbl)
                            pf_summary.set_text("")

                        async def _run_preflight():
                            import urllib.request as _ur
                            _pf_reset()
                            btn_preflight.props("loading disabled")
                            pf_summary.set_text("Running checks…")
                            loop = asyncio.get_event_loop()

                            def _upd(key, ok, detail=""):
                                asyncio.run_coroutine_threadsafe(
                                    run.io_bound(lambda: None), loop  # dummy flush
                                )
                                _pf_set(key, ok, detail)

                            results: dict[str, bool] = {}

                            # ── 1. Supportal ────────────────────────────────────────────
                            _pf_set("supportal", None)
                            def _chk_supportal():
                                cookie = cookie_input.value or None
                                try:
                                    sess = requests.Session()
                                    h = {"User-Agent": UA}
                                    if cookie:
                                        h["Cookie"] = cookie
                                    r = sess.get(BASE_URL, headers=h, timeout=10, verify=False, allow_redirects=True)
                                    if r.status_code < 400:
                                        return True, f"HTTP {r.status_code}"
                                    return False, f"HTTP {r.status_code}"
                                except Exception as e:
                                    return False, str(e)
                            ok, detail = await run.io_bound(_chk_supportal)
                            _pf_set("supportal", ok, detail)
                            results["supportal"] = ok

                            # ── 2. Analytics API ─────────────────────────────────────────
                            _pf_set("analytics_api", None)
                            def _chk_analytics():
                                try:
                                    rows = query_supportal_analytics(
                                        "select count(1) from cluster;",
                                        cookie_input.value or None,
                                    )
                                    return True, f"{len(rows)} row(s)"
                                except Exception as e:
                                    return False, str(e)
                            ok, detail = await run.io_bound(_chk_analytics)
                            _pf_set("analytics_api", ok, detail)
                            results["analytics_api"] = ok

                            # ── 3. Embedding model ────────────────────────────────────────
                            _pf_set("emb_model", None)
                            def _chk_embed():
                                try:
                                    ep, em, ek, eu, ed, enctx = _get_embed_config()
                                    vec = embed_text("preflight test", ep, em, ek, eu, dims=ed, num_ctx=enctx)
                                    if vec and len(vec) > 0:
                                        return True, f"{ep}/{em}: {len(vec)}-dim"
                                    return False, "empty vector returned"
                                except Exception as e:
                                    return False, str(e)
                            ok, detail = await run.io_bound(_chk_embed)
                            _pf_set("emb_model", ok, detail)
                            results["emb_model"] = ok

                            # ── 4. LLM model ──────────────────────────────────────────────
                            _pf_set("llm_model", None)
                            def _chk_llm():
                                try:
                                    provider = (ai_llm_provider.value or "Claude").lower()
                                    if provider == "claude":
                                        model   = claude_model_input.value or "claude-sonnet-4-6"
                                        api_key = claude_key_input.value or ""
                                        base_url = ""
                                    elif provider == "gemini":
                                        model   = gemini_model_input.value or "gemini-2.0-flash"
                                        api_key = gemini_key_input.value or ""
                                        base_url = ""
                                    elif provider == "openai":
                                        model   = openai_llm_model_input.value or "gpt-4o"
                                        api_key = emb_openai_key_input.value or ""
                                        base_url = ""
                                    elif provider in ("LMStudio", "lmstudio"):
                                        provider = "lmstudio"
                                        model   = lms_model_input.value or "local-model"
                                        api_key = ""
                                        base_url = emb_lms_url_input.value or "http://localhost:1234"
                                        # Ensure the LLM model is loaded before testing
                                        _lms_base = base_url.rstrip("/v1").rstrip("/")
                                        _loaded = lmstudio_ensure_model_loaded(_lms_base, model, timeout_s=120, model_type="llm")
                                        if _loaded:
                                            model = _loaded
                                        else:
                                            return False, "no LLM/VLM model loaded in LMStudio — load one manually or enable autoload"
                                    else:  # ollama
                                        provider = "ollama"
                                        model   = ollama_chat_model_input.value or "llama3.2"
                                        api_key = ""
                                        base_url = emb_ollama_url_input.value or "http://localhost:11434"
                                    msgs = [
                                        {"role": "system", "content": "You are a test assistant."},
                                        {"role": "user",   "content": "Reply with just the word: OK"},
                                    ]
                                    # Use 128 tokens — 16 is too small for thinking-mode models
                                    # (Qwen3/QwQ reasoning traces exhaust the budget before output)
                                    result = call_llm(msgs, provider, model, api_key, base_url, max_tokens=128)
                                    if result and result.strip():
                                        short = result.strip()[:40].replace("\n", " ")
                                        return True, f"{provider}/{model}: \"{short}\""
                                    return False, "empty response — model may be in thinking mode; check LMStudio logs"
                                except Exception as e:
                                    return False, str(e)
                            ok, detail = await run.io_bound(_chk_llm)
                            _pf_set("llm_model", ok, detail)
                            results["llm_model"] = ok

                            # ── 5. Couchbase SDK ──────────────────────────────────────────
                            _pf_set("cb_sdk", None)
                            def _chk_cb():
                                if not _CB_AVAILABLE:
                                    return False, "SDK not installed"
                                try:
                                    import couchbase.cluster as _cbc
                                    from couchbase.auth import PasswordAuthenticator
                                    from couchbase.options import ClusterOptions
                                    host = cb_url_input.value.strip()
                                    user = cb_user_input.value.strip()
                                    pwd  = cb_pass_input.value
                                    if not host:
                                        return False, "no host configured"
                                    tls = cb_tls_toggle.value
                                    scheme = "couchbases" if tls else "couchbase"
                                    conn_str = host if "://" in host else f"{scheme}://{host}"
                                    cluster = _cbc.Cluster(
                                        conn_str,
                                        ClusterOptions(PasswordAuthenticator(user, pwd)),
                                    )
                                    cluster.wait_until_ready(datetime.timedelta(seconds=5))
                                    cluster.close()
                                    return True, host
                                except Exception as e:
                                    return False, str(e)
                            ok, detail = await run.io_bound(_chk_cb)
                            _pf_set("cb_sdk", ok, detail)
                            results["cb_sdk"] = ok

                            passed = sum(1 for v in results.values() if v)
                            total_checks = len(results)
                            now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                            pf_last_run.set_text(f"Last run: {now_str}")
                            if passed == total_checks:
                                pf_summary.set_text(f"All {total_checks} checks passed ✓")
                                pf_summary.classes(replace="text-sm mt-3 font-semibold text-green-600")
                            else:
                                failed = total_checks - passed
                                pf_summary.set_text(f"{passed}/{total_checks} passed — {failed} check(s) failed")
                                pf_summary.classes(replace="text-sm mt-3 font-semibold text-red-600")
                            btn_preflight.props(remove="loading disabled")

                        _pf_timer: list = []

                        def _on_pf_auto_toggle():
                            for t in _pf_timer:
                                t.cancel()
                            _pf_timer.clear()
                            if pf_auto_refresh.value:
                                t = ui.timer(60.0, lambda: asyncio.ensure_future(_run_preflight()))
                                _pf_timer.append(t)

                        pf_auto_refresh.on_value_change(lambda _: _on_pf_auto_toggle())

                        btn_preflight = ui.button(
                            "Run Preflight Checks", icon="checklist",
                            on_click=lambda: asyncio.ensure_future(_run_preflight()),
                        ).props("color=indigo").classes("mt-4")

                        with ui.row().classes("items-center gap-3 mt-3"):
                            alias_status_lbl = ui.label("").classes("text-xs text-gray-500")
                            ui.button(
                                "Refresh App Aliases", icon="sync",
                                on_click=lambda: asyncio.ensure_future(
                                    _refresh_cluster_map(alias_status_lbl)
                                ),
                            ).props("color=teal outline size=sm").tooltip(
                                "Query CB snapshots + tickets to build dynamic cluster→app name mappings. "
                                "Improves search labelling for any customer, not just AmEx."
                            )

            _render_chat = lambda: None  # chat moved to Corax iframe
            with ui.tab_panel(tab_chat):
                _corax_port = int(os.environ.get("CORAX_PORT", 8766))
                ui.html(
                    f'<iframe src="http://localhost:{_corax_port}"'
                    ' style="width:100%;height:calc(100vh - 130px);'
                    'border:none;border-radius:8px;display:block;">'
                    '</iframe>'
                ).classes("w-full")

            with ui.tab_panel(tab_scoring):
                with ui.column().classes("w-full gap-0"):
                    # ── Banner + sub-tab navigation ────────────────────────────────────
                    with ui.row().classes("w-full items-center gap-3 px-4 py-2 bg-indigo-50 border-b border-indigo-200"):
                        ui.icon("business").classes("text-indigo-400")
                        scoring_banner = ui.label("No customer loaded").classes("text-sm font-semibold text-indigo-700 flex-1")

                    def _set_customer_banner(name: str) -> None:
                        """Update both tab banners and the header chip whenever the active customer changes."""
                        text = f"Viewing: {name}"
                        results_banner.set_text(text)
                        scoring_banner.set_text(text)
                        if name and name.lower() not in ("", "all customers"):
                            _hdr_cust.set_text(f"▸ {name}")
                            _hdr_cust.classes(remove="hidden")
                        else:
                            _hdr_cust.classes(add="hidden")

                    # Populate banners from any already-loaded state (e.g. page refresh)
                    if state.get("customer_name"):
                        _set_customer_banner(state["customer_name"])

                    with ui.tabs().classes("w-full bg-white border-b border-gray-200") as scoring_sub_tabs:
                        sub_score      = ui.tab("Score Tickets",      icon="psychology")
                        sub_analytics  = ui.tab("Analytics",          icon="bar_chart")
                        sub_profile    = ui.tab("Customer Profile",   icon="person_search")
                        sub_comparison = ui.tab("Comparison",         icon="radar")
                        sub_drill      = ui.tab("Cluster Drill-Down", icon="timeline")
                        sub_cluster    = ui.tab("Cluster Health",     icon="monitor_heart")

                with ui.tab_panels(scoring_sub_tabs, value=sub_analytics).classes("w-full"):

                    # ── Score Tickets ────────────────────────────────────────────────
                    with ui.tab_panel(sub_score):
                        with ui.row().classes("items-center justify-between w-full"):
                            ui.label("Score Tickets (Sentiment & Complexity)").classes("text-base font-semibold")
                            ui.label("Uses the LLM provider configured above").classes("text-xs text-gray-400")

                        ui.label(
                            "Scores each ticket for stars (1-5), temperature (cold/warm/hot), "
                            "resolution quality, timeliness, communication clarity, and complexity "
                            "using few-shot prompting."
                        ).classes("text-xs text-gray-500 mt-1")

                        # ── Action buttons (top) ───────────────────────────────────────
                        with ui.row().classes("gap-3 mt-3 flex-wrap items-center"):
                            def _click_score():
                                _c = ui.context.client
                                asyncio.ensure_future(_do_score(_c))
                            btn_score = ui.button(
                                "Score Tickets", icon="psychology",
                                on_click=_click_score,
                            ).props("color=deep-purple")
                            def _click_load_scores():
                                _c = ui.context.client
                                asyncio.ensure_future(_do_load_scores(_c))
                            btn_load_scores = ui.button(
                                "Load Scores from CB", icon="download",
                                on_click=_click_load_scores,
                            ).props("outline color=deep-purple")
                            def _click_rescore_all():
                                _c = ui.context.client
                                asyncio.ensure_future(_do_rescore_all(_c))
                            btn_rescore_all = ui.button(
                                "Bulk Rescore All Customers", icon="refresh",
                                on_click=_click_rescore_all,
                            ).props("outline color=orange")
                            btn_stop_score = ui.button(
                                "Stop", icon="stop_circle",
                                on_click=lambda: (_cancel.set(), btn_stop_score.set_enabled(False)),
                            ).props("outline color=red")
                            async def _do_recover_clusters(client=None):
                                _c2 = client or ui.context.client
                                btn_recover_clusters.set_enabled(False)
                                score_progress.set_value(0)
                                score_status.set_text("Restoring cluster fields …")
                                _loop2 = asyncio.get_event_loop()
                                async def _upd_rc(msg: str, pct: float):
                                    score_status.set_text(msg)
                                    score_progress.set_value(pct)
                                def _prog_rc(msg: str, pct: float):
                                    asyncio.run_coroutine_threadsafe(_upd_rc(msg, pct), _loop2)
                                try:
                                    recovered, errs = await run.io_bound(
                                        recover_score_cluster_fields_cb,
                                        cb_url_input.value.strip(),
                                        cb_bucket_input.value.strip(),
                                        cb_user_input.value.strip(),
                                        cb_pass_input.value,
                                        cb_tls_toggle.value,
                                        cb_scope_input.value.strip() or "_default",
                                        cb_collection_input.value.strip() or "tickets",
                                        _prog_rc,
                                    )
                                    with _c2:
                                        ui.notify(f"Cluster fields restored: {recovered} tickets, {errs} errors.", type="positive" if not errs else "warning")
                                except Exception as exc:
                                    with _c2:
                                        ui.notify(f"Recovery failed: {exc}", type="negative")
                                finally:
                                    btn_recover_clusters.set_enabled(True)
                            btn_recover_clusters = ui.button(
                                "Restore Cluster Fields", icon="healing",
                                on_click=lambda: asyncio.ensure_future(_do_recover_clusters(ui.context.client)),
                            ).props("outline color=teal").tooltip("Restore score.cluster_names/cluster_ids wiped by a broken rescore run")

                            async def _do_enrich_app_labels(client=None):
                                _c3 = client or ui.context.client
                                btn_enrich_app_labels.set_enabled(False)
                                score_progress.set_value(0)
                                score_status.set_text("Enriching app labels via Analytics API + LLM…")
                                _loop3 = asyncio.get_event_loop()
                                async def _upd_ea(msg: str, pct: float):
                                    score_status.set_text(msg)
                                    score_progress.set_value(pct)
                                def _prog_ea(msg: str, pct: float):
                                    asyncio.run_coroutine_threadsafe(_upd_ea(msg, pct), _loop3)
                                _cust = state.get("customer_name", "")
                                _cookie_ea = (cookie_input.value or "").strip() or os.environ.get("SUPPORTAL_COOKIE", "")
                                llm_prov, llm_mod, llm_key, llm_base = _get_llm_config()
                                try:
                                    enriched_ea, errs_ea = await run.io_bound(
                                        enrich_ticket_apps_via_analytics,
                                        _cust,
                                        _cookie_ea or None,
                                        cb_url_input.value.strip(),
                                        cb_bucket_input.value.strip(),
                                        cb_user_input.value.strip(),
                                        cb_pass_input.value,
                                        cb_tls_toggle.value,
                                        cb_scope_input.value.strip() or "_default",
                                        cb_collection_input.value.strip() or "tickets",
                                        llm_prov, llm_mod, llm_key, llm_base,
                                        _prog_ea,
                                    )
                                    with _c3:
                                        ui.notify(
                                            f"App label enrichment: {enriched_ea} labels written"
                                            + (f", {errs_ea} errors" if errs_ea else " — done"),
                                            type="positive" if not errs_ea else "warning",
                                        )
                                except Exception as exc:
                                    with _c3:
                                        ui.notify(f"Enrichment failed: {exc}", type="negative")
                                finally:
                                    btn_enrich_app_labels.set_enabled(True)
                                    score_progress.set_value(0)

                            btn_enrich_app_labels = ui.button(
                                "Enrich App Labels", icon="label",
                                on_click=lambda: asyncio.ensure_future(_do_enrich_app_labels(ui.context.client)),
                            ).props("outline color=indigo").tooltip(
                                "For tickets missing [Application: X] labels: query the Analytics API "
                                "for linked snapshot cluster names, then use LLM to extract app names "
                                "from ticket subjects (e.g. 'Enterprise Wallet', 'DQF', 'Griffin'). "
                                "Requires a valid session cookie and LLM configured."
                            )

                        btn_score.set_enabled(False)
                        btn_load_scores.set_enabled(False)
                        btn_rescore_all.set_enabled(_CB_AVAILABLE)
                        btn_recover_clusters.set_enabled(_CB_AVAILABLE)
                        btn_enrich_app_labels.set_enabled(_CB_AVAILABLE)
                        btn_stop_score.set_enabled(False)

                        ui.separator().classes("my-2")

                        with ui.row().classes("items-center gap-4 mt-2 flex-wrap"):
                            score_ctx_input = ui.number(
                                "Ollama num_ctx (K)", value=131, min=8, max=512,
                            ).classes("w-44").props(
                                'hint="Auto-set when model is selected"'
                            ).tooltip(
                                "Input context window for Ollama/LMStudio (thousands of tokens). "
                                "Auto-populated from model defaults — edit freely. "
                                "Ignored for Claude/Gemini."
                            )
                            score_autosave_toggle = ui.checkbox("Auto-save scores to Couchbase", value=True)

                        score_status   = ui.label("").classes("text-sm text-gray-500 mt-1")
                        score_progress = _PctBar(value=0).classes("mt-1")
                        score_progress.set_visibility(False)
                        score_error_log = ui.log(max_lines=200).classes("w-full h-32 text-xs font-mono border rounded mt-1")
                        score_error_log.set_visibility(False)

                        async def _do_score(client=None):
                            if not state["results"]:
                                ui.notify("No tickets loaded.", type="warning")
                                return
                            _cancel.clear()
                            btn_score.set_enabled(False)
                            btn_load_scores.set_enabled(False)
                            btn_stop_score.set_enabled(True)
                            score_progress.set_visibility(True)
                            score_progress.set_value(0)
                            score_error_log.set_visibility(False)
                            score_status.set_text("Starting …")
                            loop = asyncio.get_event_loop()
                            _score_ts = lambda: time.strftime("%H:%M:%S")

                            def _prog(msg: str, pct: float):
                                _OP_STATUS["op"] = "score"
                                _OP_STATUS["status"] = msg
                                _OP_STATUS["progress"] = pct
                                _OP_STATUS["done"] = (pct >= 1.0)
                                asyncio.run_coroutine_threadsafe(_upd_score(msg, pct), loop)

                            async def _upd_score(msg: str, pct: float):
                                import time as _time
                                score_status.set_text(msg)
                                score_progress.set_value(pct)
                                _msg_lo = msg.lower()
                                if ("error" in _msg_lo or "fail" in _msg_lo or "empty" in _msg_lo
                                        or "missing" in _msg_lo or "warn" in _msg_lo):
                                    score_error_log.set_visibility(True)
                                    score_error_log.push(f"{_time.strftime('%H:%M:%S')}  {msg}")

                            provider, model, api_key, base_url = _get_llm_config()
                            _warn_if_small_model(model)
                            if provider == "lmstudio":
                                _lms_base = (base_url or "http://localhost:1234").rstrip("/v1").rstrip("/")
                                score_status.set_text("Ensuring LMStudio model is loaded…")
                                _lms_llm = await run.io_bound(lmstudio_ensure_model_loaded, _lms_base, model, 120, "llm")
                                if _lms_llm:
                                    if _lms_llm != model:
                                        score_status.set_text(f"LMStudio: using '{_lms_llm}' for scoring")
                                    model = _lms_llm
                                else:
                                    ui.notify("No LLM/VLM model loaded in LMStudio — load one in the LMStudio UI, then retry.", type="warning", timeout=8000)
                                    btn_score.set_enabled(True)
                                    btn_load_scores.set_enabled(True)
                                    btn_stop_score.set_enabled(False)
                                    return
                            try:
                                scores = await run.io_bound(
                                    score_all_tickets,
                                    state["results"],
                                    provider,
                                    model,
                                    api_key,
                                    base_url,
                                    int(score_batch_input.value or 20),
                                    _prog,
                                    _cancel,
                                    int(score_ctx_input.value or 131) * 1024,
                                    score_no_think_toggle.value,
                                    int(score_parallel_input.value or 1),
                                )
                                state["scores"] = scores
                                _SERVER_STATE["scores"] = scores
                                _OP_STATUS["op"] = None
                                _OP_STATUS["done"] = True
                                msg = f"Scored {len(scores)}/{len(state['results'])} tickets."
                                score_status.set_text(msg)
                                score_progress.set_value(1.0)
                                _safe_notify(client, msg, type="positive")
                                btn_render_charts.set_enabled(True)

                                if score_autosave_toggle.value and scores:
                                    score_status.set_text("Saving scores to Couchbase …")
                                    saved, errs = await run.io_bound(
                                        persist_scores_to_cb,
                                        scores,
                                        cb_url_input.value.strip(),
                                        cb_bucket_input.value.strip(),
                                        cb_user_input.value.strip(),
                                        cb_pass_input.value,
                                        cb_tls_toggle.value,
                                        cb_scope_input.value.strip() or "_default",
                                        cb_collection_input.value.strip() or "tickets",
                                        _prog,
                                        state["results"],
                                    )
                                    score_status.set_text(
                                        f"{msg} — {saved} scores saved to Couchbase"
                                        + (f", {errs} errors." if errs else ".")
                                    )
                                    if state.get("customer_name"):
                                        await run.io_bound(
                                            upsert_inventory_doc,
                                            state["customer_name"],
                                            {"score": {"at": datetime.datetime.now(datetime.timezone.utc).isoformat().replace("+00:00", "Z"), "done": len(scores), "total": len(state["results"]), "errors": errs}},
                                            cb_url_input.value.strip(),
                                            cb_bucket_input.value.strip(),
                                            cb_user_input.value.strip(),
                                            cb_pass_input.value,
                                            cb_tls_toggle.value,
                                            cb_scope_input.value.strip() or "_default",
                                            cb_collection_input.value.strip() or "tickets",
                                        )
                            except Exception as exc:
                                score_status.set_text(f"Error: {exc}")
                                _safe_notify(client, str(exc), type="negative")
                            finally:
                                btn_score.set_enabled(True)
                                btn_load_scores.set_enabled(True)
                                btn_rescore_all.set_enabled(True)
                                btn_stop_score.set_enabled(False)

                        async def _do_load_scores(client=None):
                            if not state["results"]:
                                ui.notify("No tickets loaded.", type="warning")
                                return
                            btn_load_scores.set_enabled(False)
                            score_progress.set_visibility(True)
                            score_progress.set_value(0)
                            score_status.set_text("Loading scores from Couchbase …")
                            loop = asyncio.get_event_loop()

                            def _prog2(msg: str, pct: float):
                                asyncio.run_coroutine_threadsafe(_upd_score(msg, pct), loop)

                            async def _upd_score(msg: str, pct: float):
                                score_status.set_text(msg)
                                score_progress.set_value(pct)

                            try:
                                ticket_ids = [str(t.get("ticket_id")) for t in state["results"]]
                                loaded = await run.io_bound(
                                    load_scores_from_cb,
                                    ticket_ids,
                                    cb_url_input.value.strip(),
                                    cb_bucket_input.value.strip(),
                                    cb_user_input.value.strip(),
                                    cb_pass_input.value,
                                    cb_tls_toggle.value,
                                    cb_scope_input.value.strip() or "_default",
                                    cb_collection_input.value.strip() or "tickets",
                                    _prog2,
                                )
                                state["scores"] = loaded
                                _SERVER_STATE["scores"] = loaded
                                msg = f"Loaded {len(loaded)} scores from Couchbase."
                                score_status.set_text(msg)
                                score_progress.set_value(1.0)
                                _safe_notify(client, msg, type="positive")
                                if loaded:
                                    btn_render_charts.set_enabled(True)
                            except Exception as exc:
                                score_status.set_text(f"Error: {exc}")
                                _safe_notify(client, str(exc), type="negative")
                            finally:
                                btn_load_scores.set_enabled(True)
                                btn_rescore_all.set_enabled(True)

                        async def _do_rescore_all(client=None):
                            _cancel.clear()
                            btn_rescore_all.set_enabled(False)
                            btn_score.set_enabled(False)
                            btn_load_scores.set_enabled(False)
                            btn_stop_score.set_enabled(True)
                            score_progress.set_visibility(True)
                            score_progress.set_value(0)
                            score_status.set_text("Starting bulk rescore …")
                            loop = asyncio.get_event_loop()

                            async def _upd_rescore(msg: str, pct: float):
                                score_status.set_text(msg)
                                score_progress.set_value(pct)

                            def _prog_rescore(msg: str, pct: float):
                                asyncio.run_coroutine_threadsafe(_upd_rescore(msg, pct), loop)

                            provider, model, api_key, base_url = _get_llm_config()
                            _warn_if_small_model(model)
                            try:
                                scored, errs, err_log = await run.io_bound(
                                    rescore_all_customers_cb,
                                    cb_url_input.value.strip(),
                                    cb_bucket_input.value.strip(),
                                    cb_user_input.value.strip(),
                                    cb_pass_input.value,
                                    cb_tls_toggle.value,
                                    cb_scope_input.value.strip() or "_default",
                                    cb_collection_input.value.strip() or "tickets",
                                    provider,
                                    model,
                                    api_key,
                                    base_url,
                                    int(score_batch_input.value or 20),
                                    _prog_rescore,
                                    _cancel,
                                    int(score_parallel_input.value or 1),
                                )
                                msg = f"Bulk rescore complete — {scored} scored, {errs} errors."
                                score_status.set_text(msg)
                                score_progress.set_value(1.0)
                                _safe_notify(client, msg, type="positive" if errs == 0 else "warning")
                                if err_log:
                                    try:
                                        from nicegui.client import Client as _NC
                                        if client and client.id in _NC.instances:
                                            with client:
                                                with ui.dialog() as err_dialog, ui.card().classes("w-full max-w-2xl"):
                                                    ui.label(f"Rescore Errors ({len(err_log)})").classes("text-base font-semibold text-red-600")
                                                    ui.separator()
                                                    with ui.scroll_area().classes("w-full h-64"):
                                                        for e in err_log:
                                                            ui.label(e).classes("text-xs font-mono text-red-700 break-all")
                                                    ui.button("Close", on_click=err_dialog.close).classes("mt-2")
                                                err_dialog.open()
                                    except Exception:
                                        pass
                            except Exception as exc:
                                score_status.set_text(f"Bulk rescore error: {exc}")
                                _safe_notify(client, str(exc), type="negative")
                            finally:
                                btn_rescore_all.set_enabled(True)
                                btn_score.set_enabled(True)
                                btn_load_scores.set_enabled(True)
                                btn_stop_score.set_enabled(False)

                    # ── Analytics ────────────────────────────────────────────────────
                    with ui.tab_panel(sub_analytics):

                        # Customer filter for main charts
                        with ui.row().classes("items-end gap-3 mt-2 w-full flex-wrap"):
                            with ui.column().classes("gap-1 flex-1"):
                                ui.label("Filter by Customer (optional)").classes("text-xs text-gray-500")
                                main_cust_input = ui.input(
                                    placeholder="Type name fragment, select, or leave blank for all"
                                ).props("outlined clearable").classes("w-full")
                            main_cust_results = ui.select([], label="Matching customers").classes("w-64")
                            main_cust_results.set_visibility(False)

                        async def _do_main_cust_lookup():
                            q = (main_cust_input.value or "").strip()
                            if len(q) < 2:
                                main_cust_results.set_visibility(False)
                                return
                            if not _CB_AVAILABLE:
                                return
                            try:
                                orgs = await run.io_bound(
                                    search_orgs_from_cb,
                                    cb_url_input.value.strip(), cb_bucket_input.value.strip(),
                                    cb_user_input.value.strip(), cb_pass_input.value,
                                    cb_tls_toggle.value,
                                    cb_scope_input.value.strip() or "_default",
                                    cb_collection_input.value.strip() or "tickets",
                                    q,
                                )
                                main_cust_results.options = orgs
                                main_cust_results.update()
                                main_cust_results.set_visibility(bool(orgs))
                            except Exception:
                                pass

                        def _apply_main_cust_select():
                            if main_cust_results.value:
                                main_cust_input.set_value(main_cust_results.value)
                                main_cust_results.set_visibility(False)

                        main_cust_input.on_value_change(lambda _: asyncio.ensure_future(_do_main_cust_lookup()))
                        main_cust_results.on_value_change(lambda _: _apply_main_cust_select())

                        # ── Action buttons (top) ───────────────────────────────────────
                        with ui.row().classes("gap-3 mt-3 flex-wrap items-center"):
                            btn_render_charts = ui.button(
                                "Generate Charts", icon="bar_chart",
                                on_click=lambda: asyncio.ensure_future(_render_charts(ui.context.client)),
                            ).props("color=teal")
                            btn_render_charts.set_enabled(_CB_AVAILABLE)
                            btn_export_pdf = ui.button(
                                "Export PDF", icon="picture_as_pdf",
                                on_click=lambda: asyncio.ensure_future(_export_charts_pdf(ui.context.client)),
                            ).props("outline color=red-7")
                            btn_export_pdf.set_enabled(False)
                            ui.button(
                                "Diag Charts", icon="bug_report",
                                on_click=lambda: asyncio.ensure_future(_diag_charts()),
                            ).props("outline color=grey").tooltip(
                                "Run JS diagnostics to check Highcharts availability"
                            )

                        # ── Date range filter ──────────────────────────────────────────
                        with ui.row().classes("gap-4 mt-2 flex-wrap items-end"):
                            with ui.input("From", placeholder="YYYY-MM-DD").classes("w-36") as chart_date_from:
                                with ui.menu().props("no-parent-event") as _date_from_menu:
                                    with ui.date(mask="YYYY-MM-DD").bind_value(chart_date_from):
                                        with ui.row().classes("justify-end"):
                                            ui.button("Close", on_click=_date_from_menu.close).props("flat")
                                with chart_date_from.add_slot("append"):
                                    ui.icon("event").on("click", _date_from_menu.open).classes("cursor-pointer")
                            with ui.input("To", placeholder="YYYY-MM-DD").classes("w-36") as chart_date_to:
                                with ui.menu().props("no-parent-event") as _date_to_menu:
                                    with ui.date(mask="YYYY-MM-DD").bind_value(chart_date_to):
                                        with ui.row().classes("justify-end"):
                                            ui.button("Close", on_click=_date_to_menu.close).props("flat")
                                with chart_date_to.add_slot("append"):
                                    ui.icon("event").on("click", _date_to_menu.open).classes("cursor-pointer")
                            ui.button(
                                "Clear dates", icon="clear",
                                on_click=lambda: (chart_date_from.set_value(""), chart_date_to.set_value("")),
                            ).props("flat color=grey-7 size=sm")

                        chart_status = ui.label("").classes("text-sm text-gray-500 mt-1")
                        charts_area  = ui.column().classes("w-full gap-4 mt-3")

                        # ── Chart drill-down dialog ────────────────────────────────────────
                        # Shared modal — chart clicks open this; rows click to ticket detail
                        _drill_dlg = ui.dialog().props("maximized=false").classes("q-pa-none")
                        with _drill_dlg, ui.card().classes("w-full").style("min-width:700px;max-width:1000px;max-height:85vh;overflow-y:auto;padding:16px"):
                            with ui.row().classes("w-full items-center justify-between mb-2"):
                                _drill_label = ui.label("").classes("text-sm font-semibold text-blue-800")
                                ui.button(icon="close", on_click=lambda: _drill_dlg.close()).props("flat round dense color=grey-6 size=sm")
                            _drill_rows_area = ui.column().classes("w-full")

                        # Ticket detail dialog
                        _ticket_dlg = ui.dialog().props("maximized=false").classes("q-pa-none")
                        with _ticket_dlg, ui.card().classes("w-full").style("min-width:600px;max-width:900px;max-height:85vh;overflow-y:auto"):
                            _ticket_dlg_body = ui.column().classes("w-full gap-3 p-4")

                        def _make_chart(container, cfg: dict, height: int = 380):
                            with container:
                                ui.echart(cfg).classes("w-full").style(f"height:{height}px")

                        async def _render_charts(client=None):
                            cust_filter = (main_cust_input.value or "").strip()
                            if not state["results"]:
                                if not _CB_AVAILABLE:
                                    ui.notify("No tickets loaded and Couchbase not available.", type="warning")
                                    return
                                chart_status.set_text("Loading tickets from Couchbase …")
                                btn_render_charts.set_enabled(False)
                                loop_rc = asyncio.get_event_loop()

                                def _rc_prog(msg: str, pct: float):
                                    asyncio.run_coroutine_threadsafe(_rc_upd(msg, pct), loop_rc)

                                async def _rc_upd(msg: str, pct: float):
                                    chart_status.set_text(msg)

                                try:
                                    loaded = await run.io_bound(
                                        load_tickets_from_cb,
                                        cb_url_input.value.strip(), cb_bucket_input.value.strip(),
                                        cb_user_input.value.strip(), cb_pass_input.value,
                                        cb_tls_toggle.value,
                                        cb_scope_input.value.strip() or "_default",
                                        cb_collection_input.value.strip() or "tickets",
                                        "",
                                        _rc_prog,
                                    )
                                    state["results"] = loaded
                                    _SERVER_STATE["results"] = loaded
                                    auto_sc = {
                                        str(t["ticket_id"]): t["score"]
                                        for t in loaded
                                        if t.get("score") and t.get("ticket_id")
                                    }
                                    if auto_sc:
                                        state["scores"].update(auto_sc)
                                    chart_status.set_text(
                                        f"Loaded {len(loaded)} tickets from Couchbase — building charts …"
                                    )
                                except Exception as exc:
                                    chart_status.set_text(f"CB load error: {exc}")
                                    btn_render_charts.set_enabled(True)
                                    return

                            btn_render_charts.set_enabled(False)
                            chart_status.set_text("Building charts …")
                            charts_area.clear()

                            # Compute responsive chart heights & font sizes from browser window size
                            _win_h, _win_w = 768, 1280
                            if client:
                                try:
                                    _dims  = await client.run_javascript("[window.innerWidth, window.innerHeight]")
                                    _win_w = int((_dims or [1280, 768])[0] or 1280)
                                    _win_h = int((_dims or [1280, 768])[1] or 768)
                                except Exception:
                                    pass
                            ch      = max(360, int(_win_h * 0.40))   # standard full/half charts
                            ch_sm   = max(300, int(_win_h * 0.32))   # smaller/pie side-by-side charts
                            ch_bbl  = max(440, int(_win_h * 0.48))   # bubble scatter chart

                            # Font sizes — scale with window height, capped to avoid extremes
                            _fs      = max(11, min(15, int(_win_h * 0.013)))   # normal labels
                            _fs_sm   = max(10, _fs - 1)                        # dense/rotated labels
                            _fs_str  = f"{_fs}px"
                            _fs_sm_str = f"{_fs_sm}px"
                            _title_sz = f"{max(13, min(18, _fs + 2))}px"
                            _sub_sz   = f"{max(10, _fs - 1)}px"

                            # Build org-name consolidation map for this ticket set
                            _oc = _load_settings_file().get("__org_consolidation__", {})
                            _org_map = build_org_name_map(
                                state["results"],
                                enabled=bool(_oc.get("enabled", True)),
                                threshold=int(_oc.get("threshold", 90)) / 100.0,
                            )
                            state["_org_map"] = _org_map

                            # Filter tickets for display if a customer is selected.
                            # Match against the *canonical* org name so "Acme Inc",
                            # "ACME CORP" etc. all satisfy a filter typed as "Acme".
                            if cust_filter:
                                _cf_lower = cust_filter.lower()
                                display_tickets = [
                                    t for t in state["results"]
                                    if _cf_lower in _apply_org_map(
                                        (t.get("organization") or ""), _org_map
                                    ).lower()
                                ]
                                display_scores  = {
                                    tid: sc for tid, sc in state["scores"].items()
                                    if any(
                                        str(t.get("ticket_id")) == tid
                                        for t in display_tickets
                                    )
                                }
                                # Use the canonical name of the first matched ticket as label
                                _canon = _apply_org_map(cust_filter, _org_map)
                                state["_main_chart_label"] = _canon
                                scoring_banner.set_text(f"Viewing: {_canon}")
                            else:
                                display_tickets = state["results"]
                                display_scores  = state["scores"]
                                state["_main_chart_label"] = "All Customers"
                                scoring_banner.set_text(
                                    f"Viewing: All Customers ({len(display_tickets)} tickets)"
                                )

                            # ── Date range filter ─────────────────────────────────────────
                            _df = (chart_date_from.value or "").strip()[:10]
                            _dt = (chart_date_to.value   or "").strip()[:10]
                            if _df or _dt:
                                _pre_filter = len(display_tickets)
                                def _in_range(t):
                                    raw = (t.get("created") or t.get("created_at") or "")[:10]
                                    if not raw:
                                        return True
                                    if _df and raw < _df:
                                        return False
                                    if _dt and raw > _dt:
                                        return False
                                    return True
                                display_tickets = [t for t in display_tickets if _in_range(t)]
                                display_scores  = {
                                    tid: sc for tid, sc in display_scores.items()
                                    if any(str(t.get("ticket_id")) == tid for t in display_tickets)
                                }
                                _date_label = f"{_df or '…'} → {_dt or '…'}"
                                chart_status.set_text(
                                    f"Date filter: {_date_label} — {len(display_tickets)} of {_pre_filter} tickets"
                                )

                            # ── Drill-down helpers (close over display_tickets) ───────────
                            def _open_ticket_detail(ticket: dict):
                                _ticket_dlg_body.clear()
                                with _ticket_dlg_body:
                                    tf = _parse_ticket_fields(ticket)
                                    with ui.row().classes("w-full items-start justify-between gap-2"):
                                        with ui.column().classes("flex-1 gap-0"):
                                            ui.label(f"#{ticket.get('ticket_id')} · {ticket.get('organization','')}").classes("text-xs text-gray-400")
                                            ui.label(ticket.get("subject") or "").classes("text-base font-semibold")
                                        ui.button(icon="close", on_click=_ticket_dlg.close).props("flat round dense color=grey-6")
                                    with ui.row().classes("gap-4 flex-wrap text-xs text-gray-500"):
                                        for _lbl, _val in [
                                            ("Priority",  (ticket.get("priority") or "—").upper()),
                                            ("Status",    (ticket.get("status")   or "—").capitalize()),
                                            ("Created",   (ticket.get("created")  or "")[:10]),
                                            ("Version",   extract_ticket_version(ticket)),
                                            ("Component", tf.get("Component") or "—"),
                                        ]:
                                            with ui.column().classes("gap-0"):
                                                ui.label(_lbl).classes("text-xs text-gray-400")
                                                ui.label(_val).classes("text-xs font-medium")
                                    sc = state["scores"].get(str(ticket.get("ticket_id", ""))) or {}
                                    if sc:
                                        with ui.row().classes("gap-4 flex-wrap text-xs"):
                                            for _lbl, _key in [("Stars","stars"),("Temp","temperature"),("Complexity","complexity")]:
                                                if sc.get(_key):
                                                    ui.badge(f"{_lbl}: {sc[_key]}").props("color=blue-grey-6")
                                    desc = (ticket.get("description") or "").strip()
                                    if desc:
                                        ui.separator()
                                        ui.label("Description").classes("text-xs font-semibold text-gray-500")
                                        ui.label(desc[:3000]).classes("text-xs text-gray-700 whitespace-pre-wrap")
                                    summ = (ticket.get("summary_text") or "").strip()
                                    if summ:
                                        ui.separator()
                                        ui.label("AI Summary").classes("text-xs font-semibold text-gray-500")
                                        ui.label(summ).classes("text-xs text-gray-700 whitespace-pre-wrap")
                                    # ── Cluster topology from snapshot ────────────────────
                                    _dtopo = ticket.get("snapshot_topology") or {}
                                    if isinstance(_dtopo, str):
                                        try:
                                            _dtopo = json.loads(_dtopo)
                                        except Exception:
                                            _dtopo = {}
                                    if isinstance(_dtopo, dict) and _dtopo and (
                                        _dtopo.get("total_nodes") or _dtopo.get("cb_version") or _dtopo.get("cpus_per_node")
                                    ):
                                        ui.separator()
                                        ui.label("Cluster Topology (snapshot)").classes("text-xs font-semibold text-gray-500")
                                        _topo_chips = []
                                        if _dtopo.get("cb_version"):
                                            _topo_chips.append(("CB Version", _dtopo["cb_version"]))
                                        if _dtopo.get("total_nodes"):
                                            _topo_chips.append(("Nodes", str(_dtopo["total_nodes"])))
                                        if _dtopo.get("cpus_per_node"):
                                            _topo_chips.append(("CPUs/node", str(_dtopo["cpus_per_node"])))
                                        if _dtopo.get("ram_used_per_node_mib") and _dtopo.get("ram_per_node_mib"):
                                            _topo_chips.append(("RAM used/node", f"{_dtopo['ram_used_per_node_mib']}/{_dtopo['ram_per_node_mib']} MiB"))
                                        elif _dtopo.get("ram_per_node_mib"):
                                            _topo_chips.append(("RAM/node", f"{_dtopo['ram_per_node_mib']} MiB"))
                                        if _dtopo.get("os_name"):
                                            _topo_chips.append(("OS", _dtopo["os_name"]))
                                        if _dtopo.get("n2n_encryption") is not None:
                                            _topo_chips.append(("N2N Enc", _dtopo["n2n_encryption"]))
                                        if _dtopo.get("data_quota_mib"):
                                            _topo_chips.append(("Data Quota", f"{_dtopo['data_quota_mib']} MiB"))
                                        if _dtopo.get("auto_failover_seconds") is not None:
                                            _topo_chips.append(("Auto-failover", f"{_dtopo['auto_failover_seconds']}s"))
                                        _svc_parts = []
                                        for _sn, _sk in [("KV", "data_nodes"), ("Index", "index_nodes"),
                                                         ("Query", "query_nodes"), ("FTS", "fts_nodes"),
                                                         ("Eventing", "eventing_nodes"), ("Analytics", "analytics_nodes")]:
                                            if _dtopo.get(_sk):
                                                _svc_parts.append(f"{_sn}×{_dtopo[_sk]}")
                                        if _svc_parts:
                                            _topo_chips.append(("Services", ", ".join(_svc_parts)))
                                        if _dtopo.get("global_index_count") is not None:
                                            _topo_chips.append(("GSI Indexes", str(_dtopo["global_index_count"])))
                                        if _dtopo.get("fts_index_count") is not None:
                                            _topo_chips.append(("FTS Indexes", str(_dtopo["fts_index_count"])))
                                        if _dtopo.get("eventing_function_count") is not None:
                                            _topo_chips.append(("Eventing Fns", str(_dtopo["eventing_function_count"])))
                                        _dbc = _dtopo.get("bad_count") or len(_dtopo.get("bad_items") or [])
                                        _dwc = _dtopo.get("warn_count") or len(_dtopo.get("warn_items") or [])
                                        if _dbc or _dwc:
                                            _topo_chips.append(("Health", f"{_dbc} bad / {_dwc} warn"))
                                        with ui.row().classes("gap-2 flex-wrap mt-1"):
                                            for _lbl, _val in _topo_chips:
                                                with ui.column().classes("gap-0"):
                                                    ui.label(_lbl).classes("text-xs text-gray-400")
                                                    ui.label(_val).classes("text-xs font-medium")
                                        _dbi = _dtopo.get("bad_items") or []
                                        if _dbi:
                                            ui.label(f"Bad checks: {', '.join(_dbi[:12])}").classes("text-xs text-red-600 mt-1")
                                _ticket_dlg.open()

                            def _show_drill(title: str, tickets: list[dict]):
                                _drill_label.set_text(title)
                                _drill_rows_area.clear()
                                if not tickets:
                                    with _drill_rows_area:
                                        ui.label("No matching tickets.").classes("text-sm text-gray-400 p-2")
                                    _drill_dlg.open()
                                    return
                                _sorted = sorted(tickets, key=lambda x: x.get("created") or "", reverse=True)[:200]
                                _cols = [
                                    {"name": "date",     "label": "Date",     "field": "date",     "align": "left", "sortable": True},
                                    {"name": "id",       "label": "#",        "field": "id",       "align": "left"},
                                    {"name": "priority", "label": "Pri",      "field": "priority", "align": "center", "sortable": True},
                                    {"name": "status",   "label": "Status",   "field": "status",   "align": "left"},
                                    {"name": "org",      "label": "Customer", "field": "org",      "align": "left"},
                                    {"name": "subject",  "label": "Subject",  "field": "subject",  "align": "left"},
                                ]
                                _rows = [
                                    {
                                        "date":     (t.get("created") or "")[:10],
                                        "id":       str(t.get("ticket_id", "")),
                                        "priority": (t.get("priority") or "—").upper(),
                                        "status":   (t.get("status")   or "—").capitalize(),
                                        "org":      (t.get("organization") or "")[:35],
                                        "subject":  (t.get("subject")   or "")[:90],
                                        "_tid":     str(t.get("ticket_id", "")),
                                    }
                                    for t in _sorted
                                ]
                                _tid_map = {str(t.get("ticket_id","")): t for t in _sorted}
                                with _drill_rows_area:
                                    _dtbl = ui.table(columns=_cols, rows=_rows, row_key="id").classes("w-full text-xs").props("dense flat")
                                    _dtbl.on("rowClick", lambda e: _open_ticket_detail(
                                        _tid_map.get((e.args[1] if isinstance(e.args, list) else e.args.get("row", {})).get("_tid", ""), {})
                                    ))
                                _drill_dlg.open()

                            # Run in a thread so the NiceGUI event loop isn't blocked
                            # while processing large ticket sets (e.g. 1 000+ tickets).
                            chart_status.set_text("Crunching analytics data …")
                            data = await run.io_bound(build_analytics_data, display_tickets, display_scores)

                            with charts_area:
                                # ── Row 1: Stacked volume by origin over time ─────────────
                                if data["month_keys"]:
                                    _mchart = ui.echart({
                                        "title":    {"text": "Ticket Volume Over Time by Origin", "subtext": "Drag to zoom · Click bar to drill down"},
                                        "tooltip":  {"trigger": "axis", "axisPointer": {"type": "shadow"}},
                                        "legend":   {"bottom": 0},
                                        "dataZoom": [{"type": "inside"}, {"type": "slider", "bottom": 30}],
                                        "grid":     {"bottom": 80},
                                        "xAxis":    {"type": "category", "data": data["month_keys"], "axisLabel": {"rotate": 45, "fontSize": _fs_sm}},
                                        "yAxis":    {"type": "value", "name": "Tickets"},
                                        "color":    ["#1E88E5", "#FB8C00", "#6D4C41"],
                                        "series": [
                                            {"name": "Customer-Initiated",  "type": "bar", "stack": "total", "data": data["month_customer"]},
                                            {"name": "Agent-Initiated",     "type": "bar", "stack": "total", "data": data["month_agent"]},
                                            {"name": "Proactive/Automated", "type": "bar", "stack": "total", "data": data["month_proactive"]},
                                        ],
                                    }).classes("w-full").style(f"height:{ch}px")
                                    def _on_month_click(e, _dts=display_tickets):
                                        mo = e.name or ""
                                        if not mo:
                                            return
                                        _f = [t for t in _dts if _parse_created([t]) and _parse_created([t])[0] == mo]
                                        _show_drill(f"{len(_f)} tickets — {mo}", _f)
                                    _mchart.on_point_click(_on_month_click)
                                else:
                                    ui.label("No parseable dates for frequency chart.").classes("text-sm text-gray-400")

                                # ── Row 1b: Tickets per year ──────────────────────────────
                                if data["year_keys"]:
                                    with ui.card().classes("w-full"):
                                        _ychart = ui.echart({
                                            "title":   {"text": "Tickets per Year", "subtext": "Click bar to drill down"},
                                            "tooltip": {"trigger": "axis"},
                                            "xAxis":   {"type": "category", "data": data["year_keys"], "name": "Year"},
                                            "yAxis":   {"type": "value", "name": "Tickets", "minInterval": 1},
                                            "color":   ["#039BE5"],
                                            "series":  [{"name": "Tickets", "type": "bar", "data": data["year_values"], "label": {"show": True, "position": "top"}}],
                                        }).classes("w-full").style(f"height:{ch_sm}px")
                                        def _on_year_click(e, _dts=display_tickets):
                                            yr = e.name or ""
                                            if not yr:
                                                return
                                            _f = [t for t in _dts if (t.get("created") or "")[:4] == yr]
                                            _show_drill(f"{len(_f)} tickets — {yr}", _f)
                                        _ychart.on_point_click(_on_year_click)

                                # ── Row 2: Priority + Status side by side ─────────────────
                                with ui.row().classes("w-full gap-4"):
                                    with ui.card().classes("flex-1"):
                                        _pchart = ui.echart({
                                            "title":   {"text": "Priority Distribution", "subtext": "Click slice to drill down"},
                                            "tooltip": {"trigger": "item", "formatter": "{b}: {c} ({d}%)"},
                                            "color":   ["#43A047","#FB8C00","#E53935","#8E24AA"],
                                            "series":  [{"name": "Tickets", "type": "pie", "radius": "62%", "label": {"fontSize": _fs_sm}, "data": [{"name": l, "value": v} for l, v in zip(data["priority_labels"], data["priority_values"])]}],
                                        }).classes("w-full").style(f"height:{ch_sm}px")
                                        def _on_priority_click(e, _dts=display_tickets):
                                            pri = e.name or ""
                                            if not pri:
                                                return
                                            _f = [t for t in _dts if (t.get("priority") or "unknown").capitalize() == pri]
                                            _show_drill(f"{len(_f)} tickets — Priority: {pri}", _f)
                                        _pchart.on_point_click(_on_priority_click)
                                    with ui.card().classes("flex-1"):
                                        _schart = ui.echart({
                                            "title":   {"text": "Status Breakdown", "subtext": "Click slice to drill down"},
                                            "tooltip": {"trigger": "item", "formatter": "{b}: {c} ({d}%)"},
                                            "series":  [{"name": "Tickets", "type": "pie", "radius": ["45%", "68%"], "label": {"fontSize": _fs_sm}, "data": [{"name": l, "value": v} for l, v in zip(data["status_labels"], data["status_values"])]}],
                                        }).classes("w-full").style(f"height:{ch_sm}px")
                                        def _on_status_click(e, _dts=display_tickets):
                                            st = e.name or ""
                                            if not st:
                                                return
                                            _f = [t for t in _dts if (t.get("status") or "unknown").capitalize() == st]
                                            _show_drill(f"{len(_f)} tickets — Status: {st}", _f)
                                        _schart.on_point_click(_on_status_click)

                                # ── Row 3: Comment distribution + Escalation rate ─────────
                                with ui.row().classes("w-full gap-4"):
                                    with ui.card().classes("flex-1"):
                                        _comm_labels = data["comment_labels"]
                                        _cchart = ui.echart({
                                            "title":   {"text": "Comment Count Distribution", "subtext": "Click bar to drill down"},
                                            "tooltip": {"trigger": "axis"},
                                            "xAxis":   {"type": "category", "data": _comm_labels},
                                            "yAxis":   {"type": "value", "name": "Tickets"},
                                            "color":   ["#00ACC1"],
                                            "series":  [{"name": "Tickets", "type": "bar", "data": data["comment_values"]}],
                                        }).classes("w-full").style(f"height:{ch_sm}px")
                                        _comm_bucket_map = {
                                            "1": lambda c: c <= 1,
                                            "2-5": lambda c: 2 <= c <= 5,
                                            "6-10": lambda c: 6 <= c <= 10,
                                            "11-20": lambda c: 11 <= c <= 20,
                                            "21+": lambda c: c >= 21,
                                        }
                                        def _on_comm_click(e, _dts=display_tickets, _bm=_comm_bucket_map):
                                            lbl = e.name or ""
                                            fn = _bm.get(lbl)
                                            if not fn:
                                                return
                                            _f = [t for t in _dts if fn(int(t.get("comment_count") or 0))]
                                            _show_drill(f"{len(_f)} tickets — Comments: {lbl}", _f)
                                        _cchart.on_point_click(_on_comm_click)
                                    with ui.card().classes("flex-1"):
                                        _eschart = ui.echart({
                                            "title":   {"text": "Escalation Rate", "subtext": "Click slice to drill down"},
                                            "tooltip": {"trigger": "item", "formatter": "{b}: {c} ({d}%)"},
                                            "color":   ["#E53935","#43A047"],
                                            "series":  [{"name": "Tickets", "type": "pie", "radius": ["45%", "68%"], "label": {"fontSize": _fs_sm}, "data": [{"name": l, "value": v} for l, v in zip(data["esc_labels"], data["esc_values"])]}],
                                        }).classes("w-full").style(f"height:{ch_sm}px")
                                        def _on_esc_click(e, _dts=display_tickets):
                                            lbl = e.name or ""
                                            if not lbl:
                                                return
                                            want_esc = lbl.lower().startswith("escalat")
                                            _f = [t for t in _dts if bool(t.get("escalations")) == want_esc]
                                            _show_drill(f"{len(_f)} tickets — {lbl}", _f)
                                        _eschart.on_point_click(_on_esc_click)

                                # ── Row 4: Ticket origin ──────────────────────────────────
                                with ui.card().classes("w-full"):
                                    _origchart = ui.echart({
                                        "title":   {"text": "Ticket Origin", "subtext": "How the ticket was opened · Click slice to drill down"},
                                        "tooltip": {"trigger": "item", "formatter": "{b}: {c} ({d}%)"},
                                        "color":   ["#1E88E5", "#FB8C00", "#6D4C41"],
                                        "series":  [{"name": "Tickets", "type": "pie", "radius": ["45%", "68%"], "label": {"formatter": "{b}: {c} ({d}%)", "fontSize": _fs_sm}, "data": [{"name": l, "value": v} for l, v in zip(data["origin_labels"], data["origin_values"])]}],
                                    }).classes("w-full").style(f"height:{ch_sm}px")
                                    def _on_origin_click(e, _dts=display_tickets):
                                        orig = e.name or ""
                                        if not orig:
                                            return
                                        _f = [t for t in _dts if classify_ticket_origin(t) == orig]
                                        _show_drill(f"{len(_f)} tickets — Origin: {orig}", _f)
                                    _origchart.on_point_click(_on_origin_click)

                                # Proactive diagnostic breakdown
                                if data["origin_values"][2] > 0:  # Proactive/Automated count
                                    with ui.expansion(
                                        f"Proactive ticket breakdown ({data['origin_values'][2]} tickets) — click to inspect",
                                        icon="search",
                                    ).classes("w-full text-sm text-orange-700 border border-orange-200 rounded"):
                                        with ui.column().classes("gap-1 p-2 w-full"):
                                            with ui.row().classes("gap-6 text-xs text-gray-500 mb-1"):
                                                ui.label(f"By subject prefix 'Proactive ticket:': {data['proactive_by_subject']}")
                                                ui.label(f"By oldest comment phrase: {data['proactive_by_comment']}")
                                            if data["proactive_tickets_sample"]:
                                                cols = [
                                                    {"name": "signal",  "label": "Signal",   "field": "signal",  "align": "left"},
                                                    {"name": "id",      "label": "Ticket",   "field": "id",      "align": "left"},
                                                    {"name": "org",     "label": "Customer", "field": "org",     "align": "left"},
                                                    {"name": "subject", "label": "Subject",  "field": "subject", "align": "left"},
                                                ]
                                                ui.table(
                                                    columns=cols,
                                                    rows=data["proactive_tickets_sample"],
                                                    row_key="id",
                                                ).classes("w-full text-xs").props("dense flat")

                                # ── Row 5: Version + Feature area ─────────────────────────
                                with ui.row().classes("w-full gap-4"):
                                    with ui.card().classes("flex-1"):
                                        if data["version_breakdown"]:
                                            _VER_COLORS = {
                                                "version": "#0277BD",
                                                "eol":     "#FF8F00",
                                                "admin":   "#9E9E9E",
                                                "blank":   "#78909C",
                                            }
                                            _vf_state = {"eol": True, "admin": True, "blank": True}

                                            def _ver_chart_opts(_vf=_vf_state, _bd=data["version_breakdown"]):
                                                shown = [
                                                    (lbl, cnt, cat) for lbl, cnt, cat in _bd
                                                    if cat == "version" or _vf.get(cat, True)
                                                ]
                                                return {
                                                    "title":    {"text": "Tickets by Couchbase Version", "subtext": "Drag to zoom"},
                                                    "tooltip":  {"trigger": "axis", "formatter": "{b}: {c} tickets"},
                                                    "dataZoom": [{"type": "inside"}, {"type": "slider", "bottom": 5}],
                                                    "grid":     {"bottom": 60},
                                                    "xAxis":    {"type": "category", "data": [l for l,_,_ in shown], "axisLabel": {"rotate": 45, "fontSize": _fs_sm}},
                                                    "yAxis":    {"type": "value", "name": "Tickets"},
                                                    "series":   [{"name": "Tickets", "type": "bar", "data": [
                                                        {"value": cnt, "itemStyle": {"color": _VER_COLORS.get(cat, "#0277BD")}}
                                                        for _, cnt, cat in shown
                                                    ]}],
                                                }

                                            _vchart = ui.echart(_ver_chart_opts()).classes("w-full").style(f"height:{ch}px")

                                            def _refresh_vchart():
                                                _vchart.options.clear()
                                                _vchart.options.update(_ver_chart_opts())
                                                _vchart.update()

                                            def _on_ver_click(e, _dts=display_tickets, _bd=data["version_breakdown"]):
                                                lbl = e.name or ""
                                                if not lbl:
                                                    return
                                                _cat = next((cat for l2, _, cat in _bd if l2 == lbl), "version")
                                                if _cat == "version":
                                                    _f = [t for t in _dts if extract_ticket_version(t) == lbl]
                                                elif _cat == "eol":
                                                    _f = [t for t in _dts if "end of life" in (_parse_ticket_fields(t).get("Couchbase_Server") or "").lower()]
                                                elif _cat == "admin":
                                                    _f = [t for t in _dts if "Couchbase_Server" not in _parse_ticket_fields(t)]
                                                else:  # blank
                                                    _f = [t for t in _dts if "Couchbase_Server" in _parse_ticket_fields(t) and not (_parse_ticket_fields(t).get("Couchbase_Server") or "").strip()]
                                                _show_drill(f"{len(_f)} tickets — {lbl}", _f)
                                            _vchart.on_point_click(_on_ver_click)

                                            # Colour legend + filter checkboxes
                                            with ui.row().classes("gap-4 items-center mt-1 flex-wrap"):
                                                for _cat_lbl, _cat_key, _cat_clr, _cat_cnt in [
                                                    ("Known version",       "version", "#0277BD", None),
                                                    ("EOL",                 "eol",     "#FF8F00", data["version_eol_count"]),
                                                    ("Admin/No-product",    "admin",   "#9E9E9E", data["version_admin_count"]),
                                                    ("Version unspecified", "blank",   "#78909C", data["version_blank_count"]),
                                                ]:
                                                    if _cat_cnt is None or _cat_cnt > 0:
                                                        with ui.row().classes("gap-1 items-center"):
                                                            ui.element("div").style(
                                                                f"width:10px;height:10px;border-radius:2px;background:{_cat_clr}"
                                                            )
                                                            if _cat_key == "version":
                                                                ui.label(_cat_lbl).classes("text-xs text-gray-600")
                                                            else:
                                                                def _make_toggle(_k=_cat_key):
                                                                    def _toggle(e):
                                                                        v = e.args
                                                                        _vf_state[_k] = v[0] if isinstance(v, (list, tuple)) else v
                                                                        _refresh_vchart()
                                                                    return _toggle
                                                                ui.checkbox(
                                                                    f"{_cat_lbl} ({_cat_cnt})",
                                                                    value=True,
                                                                ).classes("text-xs").on("update:model-value", _make_toggle(_cat_key))
                                        else:
                                            ui.label("No version data found in ticket fields.").classes("text-sm text-gray-400 p-4")
                                    with ui.card().classes("flex-1"):
                                        if data["feature_labels"]:
                                            _fchart = ui.echart({
                                                "title":    {"text": "Tickets by Feature Area", "subtext": "Drag to zoom · Click bar to drill down"},
                                                "tooltip":  {"trigger": "axis", "axisPointer": {"type": "shadow"}},
                                                "dataZoom": [{"type": "inside", "yAxisIndex": 0}, {"type": "slider", "yAxisIndex": 0, "right": 10}],
                                                "grid":     {"left": 170, "right": 60},
                                                "xAxis":    {"type": "value", "name": "Tickets"},
                                                "yAxis":    {"type": "category", "data": data["feature_labels"], "axisLabel": {"overflow": "truncate", "width": 140, "fontSize": _fs_sm}},
                                                "color":    ["#00838F"],
                                                "series":   [{"name": "Tickets", "type": "bar", "data": data["feature_values"]}],
                                            }).classes("w-full").style(f"height:{ch}px")
                                            def _on_feature_click(e, _dts=display_tickets):
                                                feat = e.name or ""
                                                if not feat:
                                                    return
                                                _f = [t for t in _dts if classify_ticket_feature(t) == feat]
                                                _show_drill(f"{len(_f)} tickets — Feature: {feat}", _f)
                                            _fchart.on_point_click(_on_feature_click)
                                        else:
                                            ui.label("No component/feature data found.").classes("text-sm text-gray-400 p-4")

                                # ── Scored metrics (only if scores available) ─────────────
                                if display_scores:
                                    ui.label("— Scored Metrics —").classes("text-sm font-semibold text-gray-500 text-center w-full")

                                    # Row 4: Stars + Temperature
                                    with ui.row().classes("w-full gap-4"):
                                        with ui.card().classes("flex-1"):
                                            _stchart = ui.echart({
                                                "title":   {"text": "Experience Stars Distribution", "subtext": "Click bar to drill down"},
                                                "tooltip": {"trigger": "axis"},
                                                "xAxis":   {"type": "category", "data": ["★1","★2","★3","★4","★5"]},
                                                "yAxis":   {"type": "value", "name": "Tickets"},
                                                "color":   ["#FDD835"],
                                                "series":  [{"name": "Tickets", "type": "bar", "data": data["stars_values"]}],
                                            }).classes("w-full").style(f"height:{ch_sm}px")
                                            def _on_stars_click(e, _dts=display_tickets, _sc=display_scores):
                                                lbl = e.name or ""
                                                n = lbl.replace("★","").strip()
                                                if not n.isdigit():
                                                    return
                                                _f = [t for t in _dts if str(_sc.get(str(t.get("ticket_id","")), {}).get("stars") or "") == n]
                                                _show_drill(f"{len(_f)} tickets — {lbl} stars", _f)
                                            _stchart.on_point_click(_on_stars_click)
                                        with ui.card().classes("flex-1"):
                                            _tempchart = ui.echart({
                                                "title":   {"text": "Temperature Distribution", "subtext": "Click slice to drill down"},
                                                "tooltip": {"trigger": "item", "formatter": "{b}: {c} ({d}%)"},
                                                "color":   ["#42A5F5","#FFA726","#EF5350"],
                                                "series":  [{"name": "Tickets", "type": "pie", "radius": ["45%", "68%"], "label": {"fontSize": _fs_sm}, "data": [{"name": l, "value": v} for l, v in zip(data["temp_labels"], data["temp_values"])]}],
                                            }).classes("w-full").style(f"height:{ch_sm}px")
                                            def _on_temp_click(e, _dts=display_tickets, _sc=display_scores):
                                                lbl = (e.name or "").lower()
                                                if not lbl:
                                                    return
                                                _f = [t for t in _dts if ((_sc.get(str(t.get("ticket_id","")), {}) or {}).get("temperature") or "").lower() == lbl]
                                                _show_drill(f"{len(_f)} tickets — Temperature: {e.name}", _f)
                                            _tempchart.on_point_click(_on_temp_click)

                                    # Row 5: Complexity + Dimension averages
                                    with ui.row().classes("w-full gap-4"):
                                        with ui.card().classes("flex-1"):
                                            _compchart = ui.echart({
                                                "title":   {"text": "Complexity Score Distribution", "subtext": "Click bar to drill down"},
                                                "tooltip": {"trigger": "axis"},
                                                "xAxis":   {"type": "category", "data": ["1","2","3","4","5"]},
                                                "yAxis":   {"type": "value", "name": "Tickets"},
                                                "color":   ["#8E24AA"],
                                                "series":  [{"name": "Tickets", "type": "bar", "data": data["complexity_values"]}],
                                            }).classes("w-full").style(f"height:{ch_sm}px")
                                            def _on_complexity_click(e, _dts=display_tickets, _sc=display_scores):
                                                lbl = e.name or ""
                                                if not lbl.isdigit():
                                                    return
                                                _f = [t for t in _dts if str((_sc.get(str(t.get("ticket_id","")), {}) or {}).get("complexity") or "") == lbl]
                                                _show_drill(f"{len(_f)} tickets — Complexity: {lbl}", _f)
                                            _compchart.on_point_click(_on_complexity_click)
                                        with ui.card().classes("flex-1"):
                                            ui.echart({
                                                "title":   {"text": "Avg Dimension Scores (1-5)"},
                                                "tooltip": {"trigger": "axis", "axisPointer": {"type": "shadow"}},
                                                "grid":    {"left": 180},
                                                "xAxis":   {"type": "value", "name": "Avg Score", "max": 5},
                                                "yAxis":   {"type": "category", "data": data["dim_categories"], "axisLabel": {"overflow": "truncate", "width": 160, "fontSize": _fs_sm}},
                                                "color":   ["#26A69A"],
                                                "series":  [{"name": "Avg Score", "type": "bar", "data": data["dim_avg"]}],
                                            }).classes("w-full").style(f"height:{ch_sm}px")

                                    # Row 6: Customer portfolio scatter
                                    cust_data = build_customer_analytics(display_tickets, display_scores)
                                    if cust_data:
                                        bubble_pts = [
                                            {
                                                "x": v["ticket_count"],
                                                "y": v["avg_stars"],
                                                "z": max(v["avg_complexity"] * 10, 3),
                                                "name": org,
                                            }
                                            for org, v in cust_data.items()
                                            if v["scored_count"] > 0
                                        ]
                                        if bubble_pts:
                                            with ui.card().classes("w-full"):
                                                _bbchart = ui.echart({
                                                    "title":     {"text": "Customer Portfolio: Volume vs Satisfaction", "subtext": "Bubble size = avg complexity · Click bubble to drill down"},
                                                    "tooltip":   {"trigger": "item", "formatter": "{b}<br/>Tickets: {c[0]}<br/>Avg Stars: {c[1]}<br/>Avg Complexity: {c[2]}"},
                                                    "xAxis":     {"type": "value", "name": "Ticket Count"},
                                                    "yAxis":     {"type": "value", "name": "Avg Stars (1-5)", "min": 0, "max": 5},
                                                    "visualMap": {"show": False, "dimension": 2, "min": 3, "max": 50, "inRange": {"symbolSize": [8, 50]}},
                                                    "color":     ["rgba(30,136,229,0.65)"],
                                                    "series":    [{"name": "Customers", "type": "scatter", "data": [{"name": pt["name"], "value": [pt["x"], pt["y"], pt["z"]]} for pt in bubble_pts]}],
                                                }).classes("w-full").style(f"height:{ch_bbl}px")
                                                def _on_bubble_click(e, _dts=display_tickets, _om=_org_map):
                                                    org = e.name or ""
                                                    if not org:
                                                        return
                                                    _f = [t for t in _dts if _apply_org_map(t.get("organization",""), _om) == org]
                                                    _show_drill(f"{len(_f)} tickets — {org}", _f)
                                                _bbchart.on_point_click(_on_bubble_click)

                                # ── Cluster & Snapshot metrics ──────────────────────────
                                ui.label("— Cluster & Snapshot Metrics —").classes("text-sm font-semibold text-gray-500 text-center w-full mt-2")

                                _snap_bucket_fn = {
                                    "0":    lambda t: int(t.get("snapshot_count") or 0) == 0,
                                    "1":    lambda t: int(t.get("snapshot_count") or 0) == 1,
                                    "2-5":  lambda t: 2 <= int(t.get("snapshot_count") or 0) <= 5,
                                    "6-10": lambda t: 6 <= int(t.get("snapshot_count") or 0) <= 10,
                                    "11+":  lambda t: int(t.get("snapshot_count") or 0) >= 11,
                                }

                                with ui.row().classes("w-full gap-4"):
                                    # Snapshot count distribution
                                    with ui.card().classes("flex-1"):
                                        _snchart = ui.echart({
                                            "title":   {"text": "Snapshots per Ticket", "subtext": f"{data['tickets_with_snapshots']} tickets have ≥1 snapshot · Click bar to drill down"},
                                            "tooltip": {"trigger": "axis"},
                                            "xAxis":   {"type": "category", "data": data["snap_bucket_labels"], "name": "Snapshot Count"},
                                            "yAxis":   {"type": "value", "name": "Tickets"},
                                            "color":   ["#00ACC1"],
                                            "series":  [{"name": "Tickets", "type": "bar", "data": data["snap_bucket_values"]}],
                                        }).classes("w-full").style(f"height:{ch_sm}px")
                                        def _on_snap_click(e, _dts=display_tickets, _bm=_snap_bucket_fn):
                                            lbl = e.name or ""
                                            fn = _bm.get(lbl)
                                            if not fn:
                                                return
                                            try:
                                                _f = [t for t in _dts if fn(t)]
                                            except Exception:
                                                _f = []
                                            _show_drill(f"{len(_f)} tickets — Snapshots: {lbl}", _f)
                                        _snchart.on_point_click(_on_snap_click)

                                    # Cluster names (if any detected)
                                    if data["cluster_name_labels"]:
                                        with ui.card().classes("flex-1"):
                                            _cnchart = ui.echart({
                                                "title":    {"text": "Top Cluster Names by Ticket Count", "subtext": "Drag to zoom · Click bar to drill down"},
                                                "tooltip":  {"trigger": "axis", "axisPointer": {"type": "shadow"}},
                                                "dataZoom": [{"type": "inside", "yAxisIndex": 0}, {"type": "slider", "yAxisIndex": 0, "right": 10}],
                                                "grid":     {"left": 190, "right": 60},
                                                "xAxis":    {"type": "value", "name": "Tickets"},
                                                "yAxis":    {"type": "category", "data": data["cluster_name_labels"], "axisLabel": {"overflow": "truncate", "width": 170, "fontSize": _fs_sm}},
                                                "color":    ["#5E35B1"],
                                                "series":   [{"name": "Tickets", "type": "bar", "data": data["cluster_name_values"]}],
                                            }).classes("w-full").style(f"height:{ch_sm}px")
                                            def _on_cname_click(e, _dts=display_tickets):
                                                cn = (e.name or "").lower()
                                                if not cn:
                                                    return
                                                _f = [t for t in _dts if _topo_str((t.get("snapshot_topology") or {}).get("cluster_name")).lower() == cn]
                                                _show_drill(f"{len(_f)} tickets — Cluster: {e.name}", _f)
                                            _cnchart.on_point_click(_on_cname_click)
                                    elif data["cluster_id_labels"]:
                                        with ui.card().classes("flex-1"):
                                            _cidchart = ui.echart({
                                                "title":    {"text": "Top Cluster IDs by Ticket Count", "subtext": "Drag to zoom · Click bar to drill down"},
                                                "tooltip":  {"trigger": "axis", "axisPointer": {"type": "shadow"}},
                                                "dataZoom": [{"type": "inside", "yAxisIndex": 0}, {"type": "slider", "yAxisIndex": 0, "right": 10}],
                                                "grid":     {"left": 110, "right": 60},
                                                "xAxis":    {"type": "value", "name": "Tickets"},
                                                "yAxis":    {"type": "category", "data": data["cluster_id_labels"], "axisLabel": {"fontFamily": "monospace", "overflow": "truncate", "width": 90, "fontSize": _fs_sm}},
                                                "color":    ["#5E35B1"],
                                                "series":   [{"name": "Tickets", "type": "bar", "data": [{"value": v, "name": data["cluster_id_full"][i]} for i, v in enumerate(data["cluster_id_values"])]}],
                                            }).classes("w-full").style(f"height:{ch_sm}px")
                                            def _on_cid_click(e, _dts=display_tickets):
                                                # e.name is the full UUID embedded in the data item
                                                cid = (e.name or "").lower()
                                                if not cid:
                                                    return
                                                _f = [t for t in _dts if cid in _topo_str((t.get("snapshot_topology") or {}).get("cluster_uuid")).lower() or cid in _topo_str((t.get("snapshot_topology") or {}).get("capella_cluster_id")).lower()]
                                                _show_drill(f"{len(_f)} tickets — Cluster ID: {e.name[:12]}…", _f)
                                            _cidchart.on_point_click(_on_cid_click)
                                    else:
                                        with ui.card().classes("flex-1"):
                                            ui.label("No cluster names or IDs detected in ticket data.").classes("text-sm text-gray-400 p-4")

                                # Cluster IDs (separate row, only if both names and IDs exist)
                                if data["cluster_name_labels"] and data["cluster_id_labels"]:
                                    with ui.card().classes("w-full"):
                                        _cid2chart = ui.echart({
                                            "title":    {"text": "Top Cluster IDs by Ticket Count", "subtext": "Drag to zoom · Click bar to drill down"},
                                            "tooltip":  {"trigger": "axis", "axisPointer": {"type": "shadow"}},
                                            "dataZoom": [{"type": "inside", "yAxisIndex": 0}, {"type": "slider", "yAxisIndex": 0, "right": 10}],
                                            "grid":     {"left": 110, "right": 60},
                                            "xAxis":    {"type": "value", "name": "Tickets"},
                                            "yAxis":    {"type": "category", "data": data["cluster_id_labels"], "axisLabel": {"fontFamily": "monospace", "overflow": "truncate", "width": 90, "fontSize": _fs_sm}},
                                            "color":    ["#3949AB"],
                                            "series":   [{"name": "Tickets", "type": "bar", "data": [{"value": v, "name": data["cluster_id_full"][i]} for i, v in enumerate(data["cluster_id_values"])]}],
                                        }).classes("w-full").style(f"height:{ch}px")
                                        def _on_cid2_click(e, _dts=display_tickets):
                                            cid = (e.name or "").lower()
                                            if not cid:
                                                return
                                            _f = [t for t in _dts if cid in _topo_str((t.get("snapshot_topology") or {}).get("cluster_uuid")).lower() or cid in _topo_str((t.get("snapshot_topology") or {}).get("capella_cluster_id")).lower()]
                                            _show_drill(f"{len(_f)} tickets — Cluster ID: {e.name[:12]}…", _f)
                                        _cid2chart.on_point_click(_on_cid2_click)

                                # Unique clusters vs tickets per version
                                if data["clusters_by_version_labels"]:
                                    _cv_h = max(ch, len(data["clusters_by_version_labels"]) * 36 + 80)
                                    with ui.row().classes("w-full gap-4 items-center mt-2"):
                                        with ui.card().classes("px-6 py-3 text-center"):
                                            ui.label(str(data["unique_cluster_total"])).classes("text-3xl font-bold text-teal-600")
                                            ui.label("Unique Clusters Seen (all tickets)").classes("text-xs text-gray-500")
                                    with ui.card().classes("w-full"):
                                        _cvchart = ui.echart({
                                            "title":   {"text": "Unique Clusters vs Tickets — by Version", "subtext": f"{data['unique_cluster_total']} unique cluster UUIDs · drag to zoom · click to drill down"},
                                            "tooltip": {"trigger": "axis", "axisPointer": {"type": "shadow"}},
                                            "legend":  {"bottom": 0},
                                            "grid":    {"left": 100, "bottom": 50},
                                            "xAxis":   {"type": "value", "name": "Count", "minInterval": 1},
                                            "yAxis":   {"type": "category", "data": data["clusters_by_version_labels"], "axisLabel": {"fontSize": _fs_sm}},
                                            "color":   ["#00897B", "#1E88E5"],
                                            "series":  [
                                                {"name": "Unique Clusters", "type": "bar", "data": data["clusters_by_version_values"]},
                                                {"name": "Tickets",         "type": "bar", "data": data["tickets_by_version_for_clusters"]},
                                            ],
                                        }).classes("w-full").style(f"height:{_cv_h}px")
                                        def _on_cv_click(e, _dts=display_tickets):
                                            ver = e.name or ""
                                            if not ver:
                                                return
                                            _f = [t for t in _dts if extract_ticket_version(t) == ver]
                                            _show_drill(f"{len(_f)} tickets — Version: {ver}", _f)
                                        _cvchart.on_point_click(_on_cv_click)

                                # ── CBSE Document Analytics ──────────────────────────────────────
                                if data["cbse_total"] > 0:
                                    ui.label("— CBSE Document Analytics —").classes(
                                        "text-sm font-semibold text-gray-500 text-center w-full mt-4"
                                    )
                                    # Stat row: lifetime total + avg per ticket
                                    with ui.row().classes("w-full gap-4 justify-center mb-2"):
                                        with ui.card().classes("px-6 py-3 text-center"):
                                            ui.label(str(data["cbse_total"])).classes("text-3xl font-bold text-purple-600")
                                            ui.label("CBSEs Generated (Lifetime)").classes("text-xs text-gray-500")
                                        with ui.card().classes("px-6 py-3 text-center"):
                                            ui.label(str(data["cbse_avg_per_ticket"])).classes("text-3xl font-bold text-indigo-500")
                                            ui.label("CBSEs per Ticket (avg)").classes("text-xs text-gray-500")

                                    # Charts: per-year bar + per-month line
                                    _cbse_re2 = re.compile(r"cbse[-_]?\d+", re.IGNORECASE)
                                    with ui.row().classes("w-full gap-4"):
                                        if data["cbse_year_labels"]:
                                            with ui.card().classes("flex-1"):
                                                _cbseyrchart = ui.echart({
                                                    "title":   {"text": "CBSEs Generated per Year", "subtext": "Click bar to drill down"},
                                                    "tooltip": {"trigger": "axis"},
                                                    "xAxis":   {"type": "category", "data": data["cbse_year_labels"], "name": "Year"},
                                                    "yAxis":   {"type": "value", "name": "CBSE Count", "minInterval": 1},
                                                    "color":   ["#7B1FA2"],
                                                    "series":  [{"name": "CBSEs", "type": "bar", "data": data["cbse_year_values"], "label": {"show": True, "position": "top"}}],
                                                }).classes("w-full").style(f"height:{ch_sm}px")
                                                def _on_cbse_yr_click(e, _dts=display_tickets, _rx=_cbse_re2):
                                                    yr = e.name or ""
                                                    if not yr:
                                                        return
                                                    _f = [t for t in _dts if (t.get("created") or "")[:4] == yr and _rx.search(_topo_str(_parse_ticket_fields(t).get("CBSE")))]
                                                    _show_drill(f"{len(_f)} tickets with CBSEs — {yr}", _f)
                                                _cbseyrchart.on_point_click(_on_cbse_yr_click)

                                        if data["cbse_month_keys"]:
                                            with ui.card().classes("flex-1"):
                                                _cbsemochart = ui.echart({
                                                    "title":    {"text": "CBSEs Generated per Month", "subtext": "Drag to zoom · Click to drill down"},
                                                    "tooltip":  {"trigger": "axis"},
                                                    "dataZoom": [{"type": "inside"}, {"type": "slider", "bottom": 5}],
                                                    "grid":     {"bottom": 60},
                                                    "xAxis":    {"type": "category", "data": data["cbse_month_keys"], "name": "Month", "axisLabel": {"rotate": 45, "fontSize": _fs_sm}},
                                                    "yAxis":    {"type": "value", "name": "CBSE Count", "minInterval": 1},
                                                    "color":    ["#5C6BC0"],
                                                    "series":   [{"name": "CBSEs", "type": "line", "smooth": True, "data": data["cbse_month_values"]}],
                                                }).classes("w-full").style(f"height:{ch_sm}px")
                                                def _on_cbse_mo_click(e, _dts=display_tickets, _rx=_cbse_re2):
                                                    mo = e.name or ""
                                                    if not mo:
                                                        return
                                                    _f = [t for t in _dts if _parse_created([t]) and _parse_created([t])[0] == mo and _rx.search(_topo_str(_parse_ticket_fields(t).get("CBSE")))]
                                                    _show_drill(f"{len(_f)} tickets with CBSEs — {mo}", _f)
                                                _cbsemochart.on_point_click(_on_cbse_mo_click)

                                # ── Enriched topology charts (only when ≥1 ticket enriched) ──────
                                if data["enriched_ticket_count"] > 0:
                                    ui.label("— Enriched Snapshot Topology Metrics —").classes(
                                        "text-sm font-semibold text-gray-500 text-center w-full mt-4"
                                    )
                                    ui.label(
                                        f"Based on {data['enriched_ticket_count']} tickets enriched from #nutshell-alternative tab"
                                    ).classes("text-xs text-gray-400 text-center w-full mb-2")

                                    # Row 1: Node count distribution + Bucket count distribution
                                    with ui.row().classes("w-full gap-4"):
                                        with ui.card().classes("flex-1"):
                                            ui.echart({
                                                "title":   {"text": "Node Count Distribution", "subtext": "Cluster size across enriched tickets"},
                                                "tooltip": {"trigger": "axis"},
                                                "xAxis":   {"type": "category", "data": data["node_dist_labels"], "name": "Nodes"},
                                                "yAxis":   {"type": "value", "name": "Tickets"},
                                                "color":   ["#1E88E5"],
                                                "series":  [{"name": "Tickets", "type": "bar", "data": data["node_dist_values"]}],
                                            }).classes("w-full").style(f"height:{ch_sm}px")

                                        if data["bucket_dist_labels"]:
                                            with ui.card().classes("flex-1"):
                                                ui.echart({
                                                    "title":   {"text": "Bucket Count per Cluster", "subtext": "Number of buckets configured"},
                                                    "tooltip": {"trigger": "axis"},
                                                    "xAxis":   {"type": "category", "data": data["bucket_dist_labels"], "name": "Buckets"},
                                                    "yAxis":   {"type": "value", "name": "Tickets"},
                                                    "color":   ["#43A047"],
                                                    "series":  [{"name": "Tickets", "type": "bar", "data": data["bucket_dist_values"]}],
                                                }).classes("w-full").style(f"height:{ch_sm}px")

                                    # Row 2: RAM per node tier + Auto-failover distribution
                                    with ui.row().classes("w-full gap-4"):
                                        with ui.card().classes("flex-1"):
                                            ui.echart({
                                                "title":   {"text": "RAM per Node", "subtext": "Memory tier across enriched clusters"},
                                                "tooltip": {"trigger": "axis"},
                                                "xAxis":   {"type": "category", "data": data["ram_labels"]},
                                                "yAxis":   {"type": "value", "name": "Tickets"},
                                                "color":   ["#FB8C00"],
                                                "series":  [{"name": "Tickets", "type": "bar", "data": data["ram_values"]}],
                                            }).classes("w-full").style(f"height:{ch_sm}px")

                                        with ui.card().classes("flex-1"):
                                            ui.echart({
                                                "title":   {"text": "Auto-Failover Setting", "subtext": "Configured threshold across enriched clusters"},
                                                "tooltip": {"trigger": "axis"},
                                                "xAxis":   {"type": "category", "data": data["af_labels"]},
                                                "yAxis":   {"type": "value", "name": "Tickets"},
                                                "color":   ["#E53935"],
                                                "series":  [{"name": "Tickets", "type": "bar", "data": data["af_values"]}],
                                            }).classes("w-full").style(f"height:{ch_sm}px")

                                    # Row 3: LDAP status + CB version from snapshot
                                    with ui.row().classes("w-full gap-4"):
                                        _ldap_total = sum(data["ldap_values"])
                                        if _ldap_total > 0:
                                            with ui.card().classes("flex-1"):
                                                ui.echart({
                                                    "title":   {"text": "LDAP Status", "subtext": "Across enriched clusters"},
                                                    "tooltip": {"trigger": "item", "formatter": "{b}: {c} ({d}%)"},
                                                    "color":   ["#43A047", "#E53935", "#9E9E9E"],
                                                    "series":  [{"name": "Tickets", "type": "pie", "radius": "62%", "data": [{"name": l, "value": v} for l, v in zip(data["ldap_labels"], data["ldap_values"]) if v > 0]}],
                                                }).classes("w-full").style(f"height:{ch_sm}px")

                                        if data["topo_version_labels"]:
                                            with ui.card().classes("flex-1"):
                                                ui.echart({
                                                    "title":   {"text": "CB Version Distribution", "subtext": "Ticket fields primary, snapshot fallback"},
                                                    "tooltip": {"trigger": "axis", "axisPointer": {"type": "shadow"}},
                                                    "grid":    {"left": 100},
                                                    "xAxis":   {"type": "value", "name": "Tickets"},
                                                    "yAxis":   {"type": "category", "data": data["topo_version_labels"], "axisLabel": {"fontSize": _fs_sm}},
                                                    "color":   ["#8E24AA"],
                                                    "series":  [{"name": "Tickets", "type": "bar", "data": data["topo_version_values"]}],
                                                }).classes("w-full").style(f"height:{ch_sm}px")

                                    # Row 4: Orchestrator hotspot (full width if data present)
                                    if data["orchestrator_labels"]:
                                        with ui.card().classes("w-full"):
                                            ui.echart({
                                                "title":    {"text": "Orchestrator Node Hotspot", "subtext": "Top 10 nodes most often acting as orchestrator · drag to zoom"},
                                                "tooltip":  {"trigger": "axis", "axisPointer": {"type": "shadow"}},
                                                "dataZoom": [{"type": "inside", "yAxisIndex": 0}, {"type": "slider", "yAxisIndex": 0, "right": 10}],
                                                "grid":     {"left": 190, "right": 60},
                                                "xAxis":    {"type": "value", "name": "Tickets"},
                                                "yAxis":    {"type": "category", "data": data["orchestrator_labels"], "axisLabel": {"fontFamily": "monospace", "overflow": "truncate", "width": 170, "fontSize": _fs_sm}},
                                                "color":    ["#00ACC1"],
                                                "series":   [{"name": "Tickets", "type": "bar", "data": data["orchestrator_values"]}],
                                            }).classes("w-full").style(f"height:{ch}px")

                            cust_label = state.get("_main_chart_label", "All Customers")
                            chart_status.set_text(
                                f"{len(display_tickets)} tickets · "
                                f"{len(display_scores)} scored · "
                                f"Customer: {cust_label}"
                            )
                            # Record how many charts belong to the main section
                            js_count = "(function(){var EC=window.echarts;if(!EC)return 0;var n=0;document.querySelectorAll('.nicegui-echart').forEach(function(el){if(EC.getInstanceByDom(el))n++;});return n;})()"
                            if client:
                                try:
                                    state["_main_chart_count"] = await client.run_javascript(js_count)
                                except Exception:
                                    state["_main_chart_count"] = 0
                            else:
                                state["_main_chart_count"] = 0
                            btn_render_charts.set_enabled(True)
                            btn_export_pdf.set_enabled(True)

                        # ── Chart descriptions & section breaks for PDF export ────────
                        _CHART_DESC: dict[str, str] = {
                            "Ticket Volume Over Time by Origin": (
                                "Monthly ticket volume split by who opened the ticket. Spikes in "
                                "Customer-Initiated activity may indicate product friction or an "
                                "incident. Agent-Initiated and Proactive tickets reflect internal "
                                "monitoring and support-driven outreach."
                            ),
                            "Tickets per Year": (
                                "Annual ticket count trend. Sustained growth may reflect an expanding "
                                "deployment footprint; a sudden spike in a single year often points to "
                                "a major incident or a product-wide issue."
                            ),
                            "Priority Distribution": (
                                "Proportion of tickets by severity. A high share of Urgent or High "
                                "priority cases can indicate systemic instability or SLA risk."
                            ),
                            "Status Breakdown": (
                                "Current state of all tickets — open, pending, solved, or closed. "
                                "Elevated pending counts may signal approval bottlenecks or awaiting "
                                "customer responses."
                            ),
                            "Comment Count Distribution": (
                                "Number of comments per ticket is a proxy for resolution complexity. "
                                "Tickets requiring many back-and-forth exchanges often involve unclear "
                                "initial diagnosis or hard-to-reproduce problems."
                            ),
                            "Escalation Rate": (
                                "Share of tickets that were formally escalated. Escalations typically "
                                "involve engineering engagement and correlate with customer dissatisfaction "
                                "when they occur frequently."
                            ),
                            "Ticket Origin": (
                                "How tickets entered the system: directly by the customer, opened by a "
                                "support agent, or generated by automated monitoring. A healthy account "
                                "trend shows proactive issue detection before customer impact."
                            ),
                            "Tickets by Couchbase Version": (
                                "Issue frequency by database version. Disproportionate activity on older "
                                "versions may signal end-of-life risk or delayed upgrade adoption."
                            ),
                            "Tickets by Feature Area": (
                                "Which product areas or services generate the most support load. Consistent "
                                "concentration in one area warrants targeted review of documentation, "
                                "configuration defaults, or known bugs."
                            ),
                            "Experience Stars Distribution": (
                                "AI-scored overall support experience per ticket on a 1–5 star scale. "
                                "Scores reflect resolution quality, response timeliness, and communication "
                                "clarity — not direct customer survey responses."
                            ),
                            "Temperature Distribution": (
                                "Ticket urgency classification: Cold (routine, low risk), Warm (moderate "
                                "complexity or friction), Hot (high urgency, escalation risk, or customer "
                                "impact). Hot tickets warrant leadership awareness."
                            ),
                            "Complexity Score Distribution": (
                                "Technical complexity of each issue on a 1–5 scale. Score 1 = simple "
                                "configuration question; Score 5 = multi-system or architectural failure "
                                "requiring deep engineering investigation."
                            ),
                            "Avg Dimension Scores (1-5)": (
                                "Average AI quality scores across three dimensions: how completely the "
                                "issue was resolved (Resolution Quality), how promptly the team responded "
                                "(Response Timeliness), and how clearly the team communicated throughout "
                                "(Communication Clarity)."
                            ),
                            "Customer Portfolio: Volume vs Satisfaction": (
                                "Each bubble represents one customer: x-axis = ticket volume, y-axis = "
                                "average satisfaction stars, bubble size = average complexity. Customers "
                                "in the lower-right quadrant (high volume, low satisfaction) are the "
                                "highest-priority engagement opportunities."
                            ),
                            "Snapshots per Ticket": (
                                "Number of cluster diagnostic snapshots attached per ticket. Multiple "
                                "snapshots on a single ticket indicate iterative debugging or a long-running "
                                "investigation."
                            ),
                            "Top Cluster Names by Ticket Count": (
                                "Named clusters ranked by how many tickets they have generated. Clusters "
                                "appearing repeatedly at the top may have recurring unresolved issues or "
                                "configuration patterns that need attention."
                            ),
                            "Top Cluster IDs by Ticket Count": (
                                "Cluster identifiers (Capella UUIDs) ranked by ticket volume, shown where "
                                "human-readable names are not yet available. Hover each bar for the full UUID."
                            ),
                            "Unique Clusters vs Tickets — by Version": (
                                "Compares the number of distinct clusters to the total ticket count for "
                                "each Couchbase version. A low cluster-to-ticket ratio means the same "
                                "clusters are generating repeated issues — a signal of persistent problems."
                            ),
                            "CBSEs Generated per Year": (
                                "Couchbase Support Escalation (CBSE) documents created annually. CBSEs "
                                "represent formally tracked engineering escalations and are a leading "
                                "indicator of the most severe or complex issues."
                            ),
                            "CBSEs Generated per Month": (
                                "Monthly CBSE volume trend. Sudden spikes often correlate with product "
                                "releases, migrations, or major incidents."
                            ),
                            "Node Count Distribution": (
                                "Cluster sizes across enriched tickets. Enterprise deployments typically "
                                "run larger clusters; single-node clusters in production represent a "
                                "high-availability risk."
                            ),
                            "Bucket Count per Cluster": (
                                "Number of data buckets per cluster. Very high bucket counts can indicate "
                                "multi-tenant architectures or complex application designs that increase "
                                "operational overhead."
                            ),
                            "RAM per Node": (
                                "Memory tier distribution across enriched clusters. Nodes with less than "
                                "16 GiB RAM are often associated with performance-related tickets."
                            ),
                            "Auto-Failover Setting": (
                                "Automatic failover threshold configured across enriched clusters. Disabled "
                                "or very long thresholds increase downtime risk during node failures."
                            ),
                            "LDAP Status": (
                                "External authentication (LDAP/SAML) adoption across enriched clusters. "
                                "Clusters without external auth rely on local Couchbase credentials, which "
                                "may conflict with enterprise security policies."
                            ),
                            "CB Version Distribution": (
                                "Couchbase version spread across all tickets, sourced from ticket fields "
                                "and supplemented with snapshot data. A wide spread indicates fragmented "
                                "upgrade posture across the customer environment."
                            ),
                            "Orchestrator Node Hotspot": (
                                "Nodes most frequently acting as cluster orchestrator across enriched "
                                "snapshots. Persistent orchestrator concentration on a single node can "
                                "mask underlying rebalance or failover issues."
                            ),
                        }

                        # Charts that trigger a section title page *before* them
                        _SECTION_BEFORE: dict[str, tuple[str, str]] = {
                            "Priority Distribution": (
                                "Ticket Characteristics",
                                "Analysis of ticket severity, status, escalation patterns, and "
                                "communication depth. These metrics reflect the operational profile "
                                "of support activity and highlight areas of process friction.",
                            ),
                            "Tickets by Couchbase Version": (
                                "Product & Version Insights",
                                "Breakdown by Couchbase version and product feature area. Identifies "
                                "which parts of the platform and which release branches generate the "
                                "most support demand.",
                            ),
                            "Experience Stars Distribution": (
                                "Quality & Satisfaction Metrics",
                                "AI-scored quality analysis across all tickets. Scores cover the overall "
                                "support experience, issue complexity, urgency classification, and "
                                "individual quality dimensions — resolution, timeliness, and communication.",
                            ),
                            "Snapshots per Ticket": (
                                "Cluster & Infrastructure",
                                "Cluster-level analysis derived from live diagnostic snapshot data. "
                                "Covers cluster identification, node topology, and version distribution "
                                "at the time each ticket was opened.",
                            ),
                            "CBSEs Generated per Year": (
                                "CBSE Escalation Documents",
                                "Metrics on formally tracked engineering escalations (CBSEs). "
                                "These represent the most severe or complex issues requiring direct "
                                "engineering team involvement.",
                            ),
                            "Node Count Distribution": (
                                "Enriched Snapshot Topology",
                                "Detailed infrastructure metrics extracted from live cluster snapshots "
                                "via the Supportal nutshell API. Provides architectural context — "
                                "cluster size, memory, failover configuration — at the time of each ticket.",
                            ),
                        }

                        async def _export_charts_pdf(client=None):
                            # ── Use asyncio.Event to avoid dialog-close resetting parent UI ──
                            _done      = asyncio.Event()
                            _opts: dict = {}
                            _cl = client or ui.context.client

                            # Create dialog in client layout slot, then build content separately
                            _cl.layout.__enter__()
                            try:
                                export_dlg = ui.dialog()
                            finally:
                                _cl.layout.__exit__(None, None, None)

                            with export_dlg, ui.card().classes("w-96"):
                                ui.label("Print / Export Charts").classes("text-base font-semibold")
                                ui.separator()
                                incl_main  = ui.checkbox("Main analytics charts", value=True)
                                incl_comp  = ui.checkbox("Customer comparison charts",
                                                          value=bool(state.get("_comparison_orgs")))
                                incl_title = ui.checkbox("Include cover & section pages", value=True)
                                ui.separator()
                                orient_select = ui.select(
                                    ["Landscape", "Portrait"], value="Landscape",
                                    label="Page orientation",
                                ).props("outlined dense")
                                size_select = ui.select(
                                    ['Letter (8.5×11")', "A4 (210×297mm)", 'Tabloid (11×17")'],
                                    value='Letter (8.5×11")', label="Page size",
                                ).props("outlined dense")
                                ui.label(
                                    "Opens a print-ready page in a new tab — one chart per page, "
                                    "vector quality. Use ⌘P / Ctrl+P → Save as PDF."
                                ).classes("text-xs text-gray-400 mt-2")
                                ui.separator()
                                with ui.row().classes("gap-3 justify-end w-full mt-2"):
                                    def _do_cancel():
                                        export_dlg.close()
                                        _done.set()
                                    async def _do_confirm():
                                        _opts["confirmed"]   = True
                                        _opts["main"]        = incl_main.value
                                        _opts["comp"]        = incl_comp.value
                                        _opts["title_pages"] = incl_title.value
                                        _opts["landscape"]   = orient_select.value == "Landscape"
                                        _opts["size"]        = size_select.value
                                        export_dlg.close()
                                        _done.set()
                                    ui.button("Cancel", on_click=_do_cancel).props("flat")
                                    ui.button("Open Print Preview",
                                              on_click=_do_confirm, icon="print").props("color=primary")

                            export_dlg.open()
                            await _done.wait()

                            if not _opts.get("confirmed"):
                                return

                            btn_export_pdf.set_enabled(False)
                            chart_status.set_text("Collecting chart data …")
                            try:
                                # ── Collect PNG data URLs from every live ECharts instance ──
                                collect_js = """
                                (function() {
                                    var out = [], EC = window.echarts;
                                    if (!EC) return out;
                                    document.querySelectorAll('.nicegui-echart').forEach(function(el) {
                                        var inst = EC.getInstanceByDom(el);
                                        if (!inst) return;
                                        var opts  = inst.getOption();
                                        var tArr  = opts.title || [];
                                        var title = (tArr[0] && tArr[0].text)    || '';
                                        var sub   = (tArr[0] && tArr[0].subtext) || '';
                                        try {
                                            var img = inst.getDataURL({type: 'png', pixelRatio: 2, backgroundColor: '#fff'});
                                            out.push({img: img, title: title, subtitle: sub});
                                        } catch(e) {}
                                    });
                                    return out;
                                })()
                                """
                                charts = await _cl.run_javascript(collect_js, timeout=20.0)
                                if not charts:
                                    with _cl:
                                        ui.notify("No charts found — generate charts first.",
                                                  type="warning")
                                    btn_export_pdf.set_enabled(True)
                                    return

                                # ── Split main vs comparison sections ─────────────────────────
                                _stored   = state.get("_main_chart_count", len(charts))
                                comp_orgs = state.get("_comparison_orgs", [])
                                if comp_orgs and 0 < _stored < len(charts):
                                    main_charts = charts[:_stored]
                                    comp_charts = charts[_stored:]
                                else:
                                    main_charts = charts
                                    comp_charts = []

                                selected = (main_charts if _opts["main"] else []) + \
                                           (comp_charts if _opts["comp"] else [])
                                if not selected:
                                    with _cl:
                                        ui.notify("No charts selected.", type="warning")
                                    btn_export_pdf.set_enabled(True)
                                    return

                                import datetime as _dt, html as _html, uuid as _uuid

                                main_label = state.get("_main_chart_label") or "All Customers"
                                # If the label is generic but all loaded tickets belong to one
                                # org (e.g. the scrape was scoped to a single customer), use
                                # that org name instead of "All Customers".
                                if main_label == "All Customers":
                                    _results = state.get("results") or []
                                    _oc2  = _load_settings_file().get("__org_consolidation__", {})
                                    _omap = state.get("_org_map") or build_org_name_map(
                                        _results,
                                        enabled=bool(_oc2.get("enabled", True)),
                                        threshold=int(_oc2.get("threshold", 90)) / 100.0,
                                    )
                                    _orgs    = {
                                        _apply_org_map((t.get("organization") or "").strip(), _omap)
                                        for t in _results
                                        if (t.get("organization") or "").strip()
                                    }
                                    if len(_orgs) == 1:
                                        main_label = next(iter(_orgs))
                                report_date = _dt.datetime.now().strftime("%B %d, %Y  %H:%M")
                                ticket_cnt  = len(state.get("results") or [])

                                _landscape = _opts["landscape"]
                                _css_size  = {
                                    'Letter (8.5×11")': "letter",
                                    "A4 (210×297mm)":   "A4",
                                    'Tabloid (11×17")': "11in 17in",
                                }.get(_opts["size"], "letter")
                                _page_rule = (
                                    f"@page {{ size: {_css_size} "
                                    f"{'landscape' if _landscape else 'portrait'}; margin: 0.5in; }}"
                                )

                                def _section_page(title: str, desc: str) -> str:
                                    return (
                                        f'<div class="page section-page">'
                                        f'<div class="section-inner">'
                                        f'<div class="section-rule"></div>'
                                        f'<h2>{_html.escape(title)}</h2>'
                                        f'<p>{_html.escape(desc)}</p>'
                                        f'</div></div>'
                                    )

                                def _chart_page(ch: dict) -> str:
                                    title = (ch.get("title") or "").strip()
                                    desc  = _CHART_DESC.get(title, "")
                                    img   = ch.get("img", "")
                                    cap   = (
                                        f'<p class="caption">{_html.escape(desc)}</p>'
                                        if desc else ""
                                    )
                                    return (
                                        f'<div class="page chart-page">'
                                        f'<div class="chart-wrap"><img src="{img}" alt="{_html.escape(title)}"></div>'
                                        f'{cap}</div>'
                                    )

                                pages: list[str] = []

                                # Cover page
                                if _opts["title_pages"]:
                                    pages.append(
                                        f'<div class="page cover-page">'
                                        f'<div class="cover-logo">COUCHBASE</div>'
                                        f'<h1>{_html.escape(main_label)}</h1>'
                                        f'<p class="cover-sub">Support Analytics Report</p>'
                                        f'<div class="cover-meta">'
                                        f'<span>{ticket_cnt:,} tickets analysed</span>'
                                        f'<span class="sep">·</span>'
                                        f'<span>{_html.escape(report_date)}</span>'
                                        f'</div></div>'
                                    )
                                    # Volume & Trends section opener (first section always)
                                    pages.append(_section_page(
                                        "Ticket Volume & Trends",
                                        "This section examines how support ticket volume has changed "
                                        "over time, broken down by year, month, and ticket origin. "
                                        "Trends here reflect overall account health and engagement "
                                        "patterns with the Couchbase support team.",
                                    ))

                                # Chart pages — insert section breaks where appropriate
                                for ch in selected:
                                    title = (ch.get("title") or "").strip()
                                    if _opts["title_pages"] and title in _SECTION_BEFORE:
                                        sec_title, sec_desc = _SECTION_BEFORE[title]
                                        pages.append(_section_page(sec_title, sec_desc))
                                    pages.append(_chart_page(ch))

                                css = f"""
  {_page_rule}
  *,*::before,*::after{{box-sizing:border-box;margin:0;padding:0}}
  body{{background:#fff;font-family:'Helvetica Neue',Arial,sans-serif;margin:0;color:#1a1a2e}}
  .page{{display:flex;flex-direction:column;align-items:center;justify-content:center;
         width:100%;min-height:100vh;page-break-after:always;break-after:page;
         padding:0.3in 0.4in}}
  .page:last-child{{page-break-after:avoid;break-after:avoid}}

  /* Cover */
  .cover-page{{background:linear-gradient(145deg,#1a237e 0%,#283593 60%,#1565C0 100%);
               color:#fff;gap:0.6rem;text-align:center}}
  .cover-logo{{font-size:0.75rem;letter-spacing:0.25em;text-transform:uppercase;
               opacity:0.7;margin-bottom:0.5rem}}
  .cover-page h1{{font-size:2.6rem;font-weight:800;line-height:1.15}}
  .cover-sub{{font-size:1.1rem;opacity:0.8;margin-top:0.25rem}}
  .cover-meta{{display:flex;gap:0.75rem;align-items:center;margin-top:1.5rem;
               font-size:0.8rem;opacity:0.65}}
  .sep{{opacity:0.4}}

  /* Section break */
  .section-page{{background:#f5f7ff;justify-content:center}}
  .section-inner{{max-width:6in;text-align:left}}
  .section-rule{{width:2.5rem;height:4px;background:#1a237e;border-radius:2px;margin-bottom:1rem}}
  .section-page h2{{font-size:1.9rem;font-weight:700;color:#1a237e;margin-bottom:0.75rem}}
  .section-page p{{font-size:0.95rem;color:#444;line-height:1.6}}

  /* Chart */
  .chart-page{{justify-content:flex-start;padding-top:0.4in}}
  .chart-wrap{{width:100%;flex:1;display:flex;align-items:center;justify-content:center}}
  .chart-wrap img{{width:100%;height:auto;max-height:80vh;object-fit:contain}}
  .caption{{width:100%;margin-top:0.3rem;font-size:0.72rem;color:#555;
            line-height:1.5;border-top:1px solid #e0e0e0;padding-top:0.25rem;
            max-width:9in}}

  /* Print toolbar (hidden when printing) */
  #print-bar{{position:fixed;top:0;left:0;right:0;z-index:9999;
              display:flex;align-items:center;gap:1rem;padding:10px 20px;
              background:#1a237e;color:#fff;box-shadow:0 2px 8px rgba(0,0,0,.35)}}
  #print-bar button{{padding:8px 20px;background:#4CAF50;color:#fff;border:none;
                     border-radius:6px;font-size:14px;font-weight:600;cursor:pointer}}
  #print-bar button:hover{{background:#388E3C}}
  #print-bar .hint{{font-size:12px;color:#90CAF9}}
  @media print{{
    #print-bar{{display:none!important}}
    .page{{min-height:unset;height:100vh}}
  }}
"""
                                html_doc = (
                                    f'<!DOCTYPE html><html><head><meta charset="utf-8">'
                                    f'<title>{_html.escape(main_label)} — Support Analytics</title>'
                                    f'<style>{css}</style></head><body>'
                                    f'<div id="print-bar">'
                                    f'<button onclick="window.print()">🖨️ Print / Save as PDF</button>'
                                    f'<span class="hint">Choose "Save as PDF" in the print dialog &nbsp;·&nbsp; '
                                    f'{len(selected)} chart{"s" if len(selected)!=1 else ""}'
                                    f'</span></div>'
                                    f'<div style="padding-top:52px">{"".join(pages)}</div>'
                                    f'</body></html>'
                                )

                                _pid   = _uuid.uuid4().hex[:12]
                                _route = f"/print/{_pid}"
                                from nicegui import app as _ng_app
                                from starlette.responses import HTMLResponse as _HTMLResp
                                async def _serve(_req):
                                    return _HTMLResp(content=html_doc)
                                _ng_app.add_route(_route, _serve)
                                await _cl.run_javascript(f"window.open('{_route}', '_blank')")
                                chart_status.set_text(
                                    f"Print preview opened — {len(selected)} chart(s). "
                                    "Use ⌘P / Ctrl+P in the new tab to save as PDF."
                                )
                            except Exception as exc:
                                chart_status.set_text(f"Export error: {exc}")
                                with _cl:
                                    ui.notify(str(exc), type="negative")
                            finally:
                                btn_export_pdf.set_enabled(True)

                        async def _diag_charts():
                            diag_js = """
                            (function() {
                                var EC = window.echarts;
                                var containers = document.querySelectorAll('.nicegui-echart');
                                var live = 0, dataUrlLen = 0, dataUrlErr = null;
                                containers.forEach(function(el) {
                                    var inst = EC && EC.getInstanceByDom(el);
                                    if (!inst) return;
                                    live++;
                                    if (dataUrlLen === 0) {
                                        try {
                                            var d = inst.getDataURL({type:'png', pixelRatio:1, backgroundColor:'#fff'});
                                            dataUrlLen = d ? d.length : 0;
                                        } catch(e) { dataUrlErr = e.toString(); }
                                    }
                                });
                                return {
                                    ecContainers: containers.length,
                                    liveCharts:   live,
                                    has_ECharts:  typeof EC !== 'undefined',
                                    dataUrlLen:   dataUrlLen,
                                    dataUrlErr:   dataUrlErr,
                                };
                            })()
                            """
                            try:
                                result = await ui.run_javascript(diag_js, timeout=10.0)
                                msg = (
                                    f"EC containers: {result.get('ecContainers')} | "
                                    f"live instances: {result.get('liveCharts')} | "
                                    f"echarts global: {result.get('has_ECharts')} | "
                                    f"dataUrl len: {result.get('dataUrlLen')}"
                                    + (f" | err: {result.get('dataUrlErr')}" if result.get('dataUrlErr') else "")
                                )
                                chart_status.set_text(msg)
                                ui.notify(msg, timeout=15000)
                            except Exception as exc:
                                ui.notify(f"Diag error: {exc}", type="negative")

                    # ── Customer Comparison ──────────────────────────────────────────
                    with ui.tab_panel(sub_comparison):
                        ui.label(
                            "Search for customers by name, then select 2–6 to compare across all scored dimensions."
                        ).classes("text-xs text-gray-500 mt-1")

                        # ── Action buttons (top) ───────────────────────────────────────
                        with ui.row().classes("gap-3 mt-3 flex-wrap items-center"):
                            btn_radar = ui.button(
                                "Compare", icon="radar",
                                on_click=lambda: asyncio.ensure_future(_do_radar()),
                            ).props("color=indigo")
                            btn_refresh_orgs = ui.button(
                                "Refresh Customer List", icon="refresh",
                                on_click=lambda: asyncio.ensure_future(_refresh_cust_select()),
                            ).props("outline color=grey")

                        ui.separator().classes("my-2")

                        # ── Customer Lookup ───────────────────────────────────────────────
                        with ui.expansion("Customer Lookup", icon="search").classes("w-full mb-2"):
                            ui.label(
                                "Type a name fragment to search Couchbase. Results appear below — "
                                "select one then click Add to Comparison."
                            ).classes("text-xs text-gray-500 mb-2")
                            lookup_input   = ui.input("Search customer name", placeholder="e.g. Acme").props("outlined clearable").classes("w-full")
                            lookup_results = ui.select([], label="Matching customers").classes("w-full mt-2")
                            lookup_results.set_visibility(False)
                            lookup_status  = ui.label("").classes("text-xs text-gray-400 mt-1")

                            async def _do_customer_lookup():
                                q = (lookup_input.value or "").strip()
                                if len(q) < 2:
                                    lookup_results.set_visibility(False)
                                    return
                                if not _CB_AVAILABLE:
                                    lookup_status.set_text("Couchbase not available.")
                                    return
                                lookup_status.set_text("Searching …")
                                try:
                                    orgs = await run.io_bound(
                                        search_orgs_from_cb,
                                        cb_url_input.value.strip(),
                                        cb_bucket_input.value.strip(),
                                        cb_user_input.value.strip(),
                                        cb_pass_input.value,
                                        cb_tls_toggle.value,
                                        cb_scope_input.value.strip() or "_default",
                                        cb_collection_input.value.strip() or "tickets",
                                        q,
                                    )
                                    lookup_results.options = orgs
                                    lookup_results.update()
                                    lookup_results.set_visibility(bool(orgs))
                                    lookup_status.set_text(f"{len(orgs)} match(es)." if orgs else "No matches.")
                                except Exception as exc:
                                    lookup_status.set_text(f"Error: {exc}")

                            lookup_input.on_value_change(lambda _: asyncio.ensure_future(_do_customer_lookup()))

                            def _add_to_comparison():
                                val = lookup_results.value
                                if not val:
                                    ui.notify("Select a customer from the results first.", type="warning")
                                    return
                                existing = list(cust_select.options or [])
                                if val not in existing:
                                    existing.append(val)
                                    cust_select.options = sorted(existing)
                                    cust_select.update()
                                selected = list(cust_select.value or [])
                                if val not in selected:
                                    selected.append(val)
                                    cust_select.set_value(selected)
                                ui.notify(f"Added '{val}' to comparison.", type="positive")

                            ui.button("Add to Comparison", icon="add", on_click=_add_to_comparison).props("outline color=indigo").classes("mt-2")

                        ui.separator().classes("my-2")

                        cust_select = ui.select(
                            [], label="Select customers to compare",
                            multiple=True,
                        ).classes("w-full")
                        radar_status = ui.label("").classes("text-sm text-gray-500 mt-1")
                        radar_area   = ui.column().classes("w-full mt-3")

                        async def _refresh_cust_select():
                            if state["results"]:
                                orgs = sorted({
                                    (t.get("organization") or "Unknown").strip()
                                    for t in state["results"]
                                    if (t.get("organization") or "").strip()
                                })
                                cust_select.options = orgs
                                cust_select.update()
                                # Keep profile selector in sync
                                profile_org_select.options = orgs
                                profile_org_select.update()
                            elif _CB_AVAILABLE:
                                radar_status.set_text("Querying Couchbase for customer list …")
                                try:
                                    orgs = await run.io_bound(
                                        list_orgs_from_cb,
                                        cb_url_input.value.strip(),
                                        cb_bucket_input.value.strip(),
                                        cb_user_input.value.strip(),
                                        cb_pass_input.value,
                                        cb_tls_toggle.value,
                                        cb_scope_input.value.strip() or "_default",
                                        cb_collection_input.value.strip() or "tickets",
                                    )
                                    cust_select.options = orgs
                                    cust_select.update()
                                    radar_status.set_text(f"Loaded {len(orgs)} customers from Couchbase.")
                                except Exception as exc:
                                    radar_status.set_text(f"CB query error: {exc}")
                            else:
                                radar_status.set_text("No tickets loaded and Couchbase not available.")

                        async def _do_radar():
                            selected = cust_select.value or []
                            if len(selected) < 2:
                                radar_status.set_text("Select at least 2 customers to compare.")
                                return

                            btn_radar.set_enabled(False)
                            radar_area.clear()

                            # Load any missing customers directly from CB
                            available_orgs = {
                                (t.get("organization") or "").strip()
                                for t in state["results"]
                            }
                            missing = [o for o in selected if o not in available_orgs]
                            if missing and _CB_AVAILABLE:
                                radar_status.set_text(f"Loading tickets for {len(missing)} customer(s) from Couchbase …")
                                loop_rd = asyncio.get_event_loop()

                                def _rd_prog(msg: str, pct: float):
                                    asyncio.run_coroutine_threadsafe(
                                        _rd_upd(msg), loop_rd
                                    )

                                async def _rd_upd(msg: str):
                                    radar_status.set_text(msg)

                                try:
                                    new_tickets = await run.io_bound(
                                        load_tickets_for_orgs_from_cb,
                                        missing,
                                        cb_url_input.value.strip(),
                                        cb_bucket_input.value.strip(),
                                        cb_user_input.value.strip(),
                                        cb_pass_input.value,
                                        cb_tls_toggle.value,
                                        cb_scope_input.value.strip() or "_default",
                                        cb_collection_input.value.strip() or "tickets",
                                        _rd_prog,
                                    )
                                    # Deduplicate by ticket_id before merging
                                    existing_ids = {str(t.get("ticket_id")) for t in state["results"]}
                                    for t in new_tickets:
                                        if str(t.get("ticket_id")) not in existing_ids:
                                            state["results"].append(t)
                                            existing_ids.add(str(t.get("ticket_id")))
                                    # Pull scores embedded in docs
                                    for t in new_tickets:
                                        if t.get("score") and t.get("ticket_id"):
                                            state["scores"].setdefault(str(t["ticket_id"]), t["score"])
                                except Exception as exc:
                                    radar_status.set_text(f"CB load error: {exc}")
                                    btn_radar.set_enabled(True)
                                    return
                            elif missing:
                                ui.notify(f"{len(missing)} customer(s) have no loaded tickets and Couchbase is unavailable.", type="warning")

                            cust_data = build_customer_analytics(state["results"], state["scores"])
                            dims      = ["Avg Stars", "Complexity", "Resolution Quality", "Response Timeliness", "Comm. Clarity"]
                            dim_keys  = ["avg_stars", "avg_complexity", "avg_resolution_quality",
                                         "avg_response_timeliness", "avg_communication_clarity"]

                            ec_series = []
                            for org in selected:
                                v = cust_data.get(org)
                                if not v or v["scored_count"] == 0:
                                    radar_status.set_text(f"No scored tickets for: {org}")
                                    continue
                                ec_series.append({
                                    "name":  org,
                                    "value": [round(v[k], 2) for k in dim_keys],
                                })

                            if not ec_series:
                                radar_status.set_text("No scored data for selected customers.")
                                btn_radar.set_enabled(True)
                                return

                            with radar_area:
                                ui.echart({
                                    "title":   {"text": "Customer Dimension Comparison"},
                                    "tooltip": {"trigger": "item"},
                                    "legend":  {"bottom": 0, "data": [s["name"] for s in ec_series]},
                                    "radar":   {
                                        "indicator": [{"name": d, "max": 5} for d in dims],
                                        "shape": "polygon",
                                        "splitNumber": 5,
                                    },
                                    "series":  [{"type": "radar", "data": ec_series}],
                                }).classes("w-full").style("height:420px")

                                # Summary table below radar
                                rows = []
                                for org in selected:
                                    v = cust_data.get(org, {})
                                    rows.append({
                                        "org":         org,
                                        "tickets":     v.get("ticket_count", 0),
                                        "scored":      v.get("scored_count", 0),
                                        "stars":       v.get("avg_stars", 0),
                                        "complexity":  v.get("avg_complexity", 0),
                                        "hot_pct":     f"{v.get('hot_pct', 0)}%",
                                        "escalations": v.get("escalations", 0),
                                    })
                                ui.table(
                                    columns=[
                                        {"name": "org",        "label": "Customer",      "field": "org",        "align": "left"},
                                        {"name": "tickets",    "label": "Tickets",       "field": "tickets"},
                                        {"name": "scored",     "label": "Scored",        "field": "scored"},
                                        {"name": "stars",      "label": "Avg Stars",     "field": "stars"},
                                        {"name": "complexity", "label": "Avg Complexity","field": "complexity"},
                                        {"name": "hot_pct",    "label": "Hot %",         "field": "hot_pct"},
                                        {"name": "escalations","label": "Escalations",   "field": "escalations"},
                                    ],
                                    rows=rows,
                                    row_key="org",
                                ).props("flat dense").classes("w-full mt-3")

                                # ── Ticket count per customer ─────────────────────────────
                                with ui.card().classes("w-full mt-4"):
                                    ui.echart({
                                        "title":   {"text": "Ticket Count by Customer"},
                                        "tooltip": {"trigger": "axis"},
                                        "xAxis":   {"type": "category", "data": selected, "axisLabel": {"rotate": 30, "fontSize": 11, "overflow": "truncate", "width": 100}},
                                        "yAxis":   {"type": "value", "name": "Tickets"},
                                        "color":   ["#1E88E5"],
                                        "series":  [{"name": "Total Tickets", "type": "bar", "data": [cust_data.get(o, {}).get("ticket_count", 0) for o in selected]}],
                                    }).classes("w-full").style("height:280px")

                                # ── Priority breakdown per customer ───────────────────────
                                from collections import Counter as _Counter
                                pri_counts: dict[str, dict[str, int]] = {o: {} for o in selected}
                                for t in state["results"]:
                                    org = (t.get("organization") or "Unknown").strip()
                                    if org in pri_counts:
                                        p = (t.get("priority") or "Unknown").capitalize()
                                        pri_counts[org][p] = pri_counts[org].get(p, 0) + 1
                                _CMP_PRI_COLOR_MAP = {
                                    "urgent": "#D32F2F", "p1": "#D32F2F",
                                    "high":   "#F57C00", "p2": "#F57C00",
                                    "normal": "#1976D2", "p3": "#1976D2", "medium": "#1976D2",
                                    "low":    "#388E3C", "p4": "#388E3C",
                                    "unknown": "#9E9E9E",
                                }
                                all_priorities = sorted({p for d in pri_counts.values() for p in d})
                                with ui.card().classes("w-full mt-4"):
                                    ui.echart({
                                        "title":   {"text": "Priority Breakdown by Customer"},
                                        "tooltip": {"trigger": "axis", "axisPointer": {"type": "shadow"}},
                                        "legend":  {"bottom": 0},
                                        "grid":    {"bottom": 50},
                                        "xAxis":   {"type": "category", "data": selected, "axisLabel": {"rotate": 30, "fontSize": 11, "overflow": "truncate", "width": 100}},
                                        "yAxis":   {"type": "value", "name": "Tickets"},
                                        "series":  [
                                            {"name": p, "type": "bar", "stack": "total",
                                             "itemStyle": {"color": _CMP_PRI_COLOR_MAP.get(p.lower(), "#9E9E9E")},
                                             "data": [pri_counts[o].get(p, 0) for o in selected]}
                                            for p in all_priorities
                                        ],
                                    }).classes("w-full").style("height:300px")

                                # ── Stars distribution per customer ───────────────────────
                                if state["scores"]:
                                    star_by_org: dict[str, list[int]] = {o: [0] * 5 for o in selected}
                                    for t in state["results"]:
                                        org = (t.get("organization") or "Unknown").strip()
                                        if org not in star_by_org:
                                            continue
                                        tid = str(t.get("ticket_id", ""))
                                        sc  = state["scores"].get(tid, {})
                                        if sc.get("stars"):
                                            idx = min(max(int(float(sc["stars"])), 1), 5) - 1
                                            star_by_org[org][idx] += 1
                                    with ui.card().classes("w-full mt-4"):
                                        ui.echart({
                                            "title":   {"text": "Experience Stars Distribution by Customer"},
                                            "tooltip": {"trigger": "axis", "axisPointer": {"type": "shadow"}},
                                            "legend":  {"bottom": 0},
                                            "grid":    {"bottom": 50},
                                            "xAxis":   {"type": "category", "data": ["★1","★2","★3","★4","★5"]},
                                            "yAxis":   {"type": "value", "name": "Tickets"},
                                            "series":  [
                                                {"name": org, "type": "bar", "data": star_by_org[org]}
                                                for org in selected
                                            ],
                                        }).classes("w-full").style("height:320px")

                            radar_status.set_text(f"Comparing {len(ec_series)} customers.")
                            state["_comparison_orgs"] = selected
                            btn_radar.set_enabled(True)

                    # ── Customer Profile ─────────────────────────────────────────────
                    with ui.tab_panel(sub_profile):
                        ui.label(
                            "Single-customer deep-dive: volume, priority trends, feature mix, "
                            "version usage, cluster count, resolution time, and CARR status."
                        ).classes("text-xs text-gray-500 mt-1")

                        # ── Action bar (top) ───────────────────────────────────────────
                        with ui.row().classes("w-full gap-3 mt-3 items-end flex-wrap"):
                            profile_org_select = ui.select(
                                [], label="Select customer", with_input=True,
                            ).classes("flex-1 min-w-72").props("outlined dense clearable")
                            btn_profile = ui.button(
                                "View Profile", icon="person_search",
                                on_click=lambda: asyncio.ensure_future(_do_profile(ui.context.client)),
                            ).props("color=deep-purple")
                            ui.button(
                                "Refresh List", icon="refresh",
                                on_click=lambda: asyncio.ensure_future(_refresh_profile_orgs()),
                            ).props("outline dense color=grey")
                        profile_status = ui.label("").classes("text-xs text-gray-400 mt-1")
                        profile_area   = ui.column().classes("w-full gap-4 mt-3")

                        # Populate org select from loaded tickets
                        async def _refresh_profile_orgs():
                            tickets_src = state.get("results") or []
                            orgs = sorted({
                                (t.get("organization") or "").strip()
                                for t in tickets_src
                                if (t.get("organization") or "").strip()
                            })
                            profile_org_select.options = orgs
                            profile_org_select.update()

                        async def _do_profile(client=None):
                            org = (profile_org_select.value or "").strip()
                            if not org:
                                _safe_notify(client, "Select a customer first.", type="warning")
                                return
                            btn_profile.set_enabled(False)
                            profile_status.set_text(f"Building profile for {org}…")
                            profile_area.clear()

                            tickets_src = state.get("results") or []
                            scores_src  = state.get("scores") or {}
                            prof = await run.io_bound(
                                build_customer_profile, tickets_src, scores_src, org
                            )
                            if not prof:
                                profile_status.set_text(f"No tickets found for: {org}")
                                btn_profile.set_enabled(True)
                                return

                            # Case-insensitive priority → color lookup
                            _PRIORITY_COLOR_MAP = {
                                "urgent": "#D32F2F", "p1": "#D32F2F",
                                "high":   "#F57C00", "p2": "#F57C00",
                                "normal": "#1976D2", "p3": "#1976D2", "medium": "#1976D2",
                                "low":    "#388E3C", "p4": "#388E3C",
                                "unknown": "#9E9E9E",
                            }
                            def _pri_color(p: str) -> str:
                                return _PRIORITY_COLOR_MAP.get(p.lower(), "#9E9E9E")

                            with profile_area:
                                # ── Header stat strip ─────────────────────────────────────
                                with ui.row().classes("w-full gap-3 flex-wrap items-stretch"):
                                    # CARR badge
                                    with ui.card().classes("px-5 py-3 text-center" + (
                                        " bg-amber-50 border border-amber-300" if prof["is_carr"] else ""
                                    )):
                                        if prof["is_carr"]:
                                            ui.label("CARR Account").classes("text-xs font-bold text-amber-700 uppercase tracking-wide")
                                            ui.label(
                                                f"Since {prof['carr_first_year']}" if prof["carr_first_year"] else "Tagged"
                                            ).classes("text-lg font-semibold text-amber-600")
                                        else:
                                            ui.label("CARR Status").classes("text-xs text-gray-400 uppercase tracking-wide")
                                            ui.label("Not flagged").classes("text-lg text-gray-500")

                                    with ui.card().classes("px-5 py-3 text-center"):
                                        ui.label("Total Tickets").classes("text-xs text-gray-400 uppercase tracking-wide")
                                        ui.label(str(prof["ticket_count"])).classes("text-2xl font-bold text-blue-700")

                                    with ui.card().classes("px-5 py-3 text-center"):
                                        ui.label("Tenure").classes("text-xs text-gray-400 uppercase tracking-wide")
                                        first = prof["first_ticket"] or "—"
                                        last  = prof["last_ticket"]  or "—"
                                        ui.label(first).classes("text-sm font-semibold text-gray-700")
                                        ui.label("→ " + last).classes("text-sm text-gray-500")

                                    with ui.card().classes("px-5 py-3 text-center"):
                                        ui.label("Unique Clusters").classes("text-xs text-gray-400 uppercase tracking-wide")
                                        ui.label(str(prof["unique_clusters"])).classes("text-2xl font-bold text-teal-600")

                                    avg_res = prof["avg_resolution_days"]
                                    with ui.card().classes("px-5 py-3 text-center"):
                                        ui.label("Avg Resolution").classes("text-xs text-gray-400 uppercase tracking-wide")
                                        if avg_res is not None:
                                            weeks = round(avg_res / 7, 1)
                                            ui.label(f"{avg_res}d").classes("text-2xl font-bold text-indigo-600")
                                            ui.label(f"≈ {weeks} wk").classes("text-xs text-gray-400")
                                        else:
                                            ui.label("—").classes("text-2xl text-gray-400")

                                    with ui.card().classes("px-5 py-3 text-center"):
                                        ui.label("Escalations").classes("text-xs text-gray-400 uppercase tracking-wide")
                                        ui.label(str(prof["escalation_count"])).classes("text-2xl font-bold text-red-600")

                                    if prof["proactive_count"]:
                                        with ui.card().classes("px-5 py-3 text-center bg-blue-50"):
                                            ui.label("Proactive Tickets").classes("text-xs text-blue-500 uppercase tracking-wide")
                                            ui.label(str(prof["proactive_count"])).classes("text-2xl font-bold text-blue-600")

                                # ── Row 1: Volume ─────────────────────────────────────────
                                with ui.row().classes("w-full gap-4"):
                                    if prof["year_keys"]:
                                        with ui.card().classes("flex-1"):
                                            ui.echart({
                                                "title":   {"text": "Tickets per Year"},
                                                "tooltip": {"trigger": "axis"},
                                                "xAxis":   {"type": "category", "data": prof["year_keys"]},
                                                "yAxis":   {"type": "value", "name": "Tickets", "minInterval": 1},
                                                "color":   ["#3949AB"],
                                                "series":  [{"name": "Tickets", "type": "bar", "data": prof["year_values"], "label": {"show": True, "position": "top"}}],
                                            }).classes("w-full").style("height:240px")

                                    if prof["month_keys"]:
                                        with ui.card().classes("flex-1"):
                                            ui.echart({
                                                "title":    {"text": "Tickets per Month", "subtext": "Drag to zoom"},
                                                "tooltip":  {"trigger": "axis"},
                                                "dataZoom": [{"type": "inside"}, {"type": "slider", "bottom": 5}],
                                                "grid":     {"bottom": 50},
                                                "xAxis":    {"type": "category", "data": prof["month_keys"], "axisLabel": {"rotate": 45, "fontSize": 10}},
                                                "yAxis":    {"type": "value", "name": "Tickets", "minInterval": 1},
                                                "color":    ["#1E88E5"],
                                                "series":   [{"name": "Tickets", "type": "bar", "data": prof["month_values"]}],
                                            }).classes("w-full").style("height:240px")

                                # ── Row 2: Priority per month (stacked) ───────────────────
                                if prof["month_keys"] and prof["all_priorities"]:
                                    with ui.card().classes("w-full"):
                                        ui.echart({
                                            "title":    {"text": "Ticket Priority per Month", "subtext": "Drag to zoom"},
                                            "tooltip":  {"trigger": "axis", "axisPointer": {"type": "shadow"}},
                                            "legend":   {"bottom": 0},
                                            "dataZoom": [{"type": "inside"}, {"type": "slider", "bottom": 30}],
                                            "grid":     {"bottom": 80},
                                            "xAxis":    {"type": "category", "data": prof["pri_month_keys"], "axisLabel": {"rotate": 45, "fontSize": 10}},
                                            "yAxis":    {"type": "value", "name": "Tickets"},
                                            "series":   [
                                                {"name": p, "type": "bar", "stack": "total",
                                                 "itemStyle": {"color": _pri_color(p)},
                                                 "data": prof["pri_by_month"][p]}
                                                for p in prof["all_priorities"]
                                            ],
                                        }).classes("w-full").style("height:300px")

                                # ── Row 3: Composition ────────────────────────────────────
                                with ui.row().classes("w-full gap-4"):
                                    if prof["feature_labels"]:
                                        with ui.card().classes("flex-1"):
                                            ui.echart({
                                                "title":   {"text": "Feature Area Breakdown"},
                                                "tooltip": {"trigger": "item", "formatter": "{b}: {c} ({d}%)"},
                                                "series":  [{"name": "Tickets", "type": "pie", "radius": ["38%", "65%"], "label": {"formatter": "{b}: {d}%", "fontSize": 10}, "data": [{"name": l, "value": v} for l, v in zip(prof["feature_labels"], prof["feature_values"])]}],
                                            }).classes("w-full").style("height:280px")

                                    if prof["version_breakdown"]:
                                        with ui.card().classes("flex-1"):
                                            _PROF_VER_COLORS = {
                                                "version": "#0277BD",
                                                "eol":     "#FF8F00",
                                                "admin":   "#9E9E9E",
                                                "blank":   "#78909C",
                                            }
                                            _pvf_state = {"eol": True, "admin": True, "blank": True}

                                            def _prof_ver_opts(_vf=_pvf_state, _bd=prof["version_breakdown"]):
                                                shown = [
                                                    (lbl, cnt, cat) for lbl, cnt, cat in _bd
                                                    if cat == "version" or _vf.get(cat, True)
                                                ]
                                                _h = max(280, len(shown) * 28 + 60)
                                                return {
                                                    "title":   {"text": "CB Version Distribution", "subtext": "From ticket fields + snapshots"},
                                                    "tooltip": {"trigger": "axis", "axisPointer": {"type": "shadow"}, "formatter": "{b}: {c} tickets"},
                                                    "grid":    {"left": 160},
                                                    "xAxis":   {"type": "value", "name": "Tickets", "minInterval": 1},
                                                    "yAxis":   {"type": "category", "data": [l for l,_,_ in shown], "axisLabel": {"fontSize": 10}},
                                                    "series":  [{"name": "Tickets", "type": "bar", "data": [
                                                        {"value": cnt, "itemStyle": {"color": _PROF_VER_COLORS.get(cat, "#0277BD")}}
                                                        for _, cnt, cat in shown
                                                    ]}],
                                                }

                                            _pvchart = ui.echart(_prof_ver_opts()).classes("w-full").style("height:280px")

                                            def _refresh_pvchart():
                                                _pvchart.options.clear()
                                                _pvchart.options.update(_prof_ver_opts())
                                                _pvchart.update()

                                            with ui.row().classes("gap-4 items-center mt-1 flex-wrap"):
                                                for _pcat_lbl, _pcat_key, _pcat_clr, _pcat_cnt in [
                                                    ("Known version",       "version", "#0277BD", None),
                                                    ("EOL",                 "eol",     "#FF8F00", prof["version_eol_count"]),
                                                    ("Admin/No-product",    "admin",   "#9E9E9E", prof["version_admin_count"]),
                                                    ("Version unspecified", "blank",   "#78909C", prof["version_blank_count"]),
                                                ]:
                                                    if _pcat_cnt is None or _pcat_cnt > 0:
                                                        with ui.row().classes("gap-1 items-center"):
                                                            ui.element("div").style(
                                                                f"width:10px;height:10px;border-radius:2px;background:{_pcat_clr}"
                                                            )
                                                            if _pcat_key == "version":
                                                                ui.label(_pcat_lbl).classes("text-xs text-gray-600")
                                                            else:
                                                                def _make_ptoggle(_k=_pcat_key):
                                                                    def _ptoggle(e):
                                                                        v = e.args
                                                                        _pvf_state[_k] = v[0] if isinstance(v, (list, tuple)) else v
                                                                        _refresh_pvchart()
                                                                    return _ptoggle
                                                                ui.checkbox(
                                                                    f"{_pcat_lbl} ({_pcat_cnt})",
                                                                    value=True,
                                                                ).classes("text-xs").on("update:model-value", _make_ptoggle(_pcat_key))

                                # ── Row 4: Satisfaction trend ─────────────────────────────
                                if prof["stars_trend_keys"]:
                                    with ui.card().classes("w-full"):
                                        ui.echart({
                                            "title":    {"text": "Avg Satisfaction Stars per Month", "subtext": "AI-scored · drag to zoom"},
                                            "tooltip":  {"trigger": "axis"},
                                            "dataZoom": [{"type": "inside"}, {"type": "slider", "bottom": 5}],
                                            "grid":     {"bottom": 50},
                                            "xAxis":    {"type": "category", "data": prof["stars_trend_keys"], "axisLabel": {"rotate": 45, "fontSize": 10}},
                                            "yAxis":    {"type": "value", "name": "Avg Stars", "min": 1, "max": 5, "minInterval": 1},
                                            "color":    ["#F9A825"],
                                            "series":   [{"name": "Avg Stars", "type": "line", "smooth": True, "symbol": "circle", "data": prof["stars_trend_values"]}],
                                        }).classes("w-full").style("height:240px")

                                # ── Cluster list (if any) ─────────────────────────────────
                                if prof["cluster_list"]:
                                    with ui.expansion("Known Clusters", icon="device_hub").classes("w-full text-sm"):
                                        with ui.row().classes("flex-wrap gap-2 p-2"):
                                            for cn in prof["cluster_list"]:
                                                ui.chip(cn, color="teal").props("outline dense")

                            profile_status.set_text(
                                f"Profile: {prof['ticket_count']} tickets · "
                                f"{prof['unique_clusters']} clusters · "
                                + ("CARR account · " if prof["is_carr"] else "")
                                + (f"avg {prof['avg_resolution_days']}d resolution" if prof["avg_resolution_days"] else "resolution time unavailable")
                            )
                            btn_profile.set_enabled(True)

                    # ── Cluster Drill-Down ───────────────────────────────────────────
                    with ui.tab_panel(sub_drill):
                        ui.label(
                            "Select a cluster enriched from snapshot topology to see how its "
                            "metrics changed across tickets over time. Run Enrich step first."
                        ).classes("text-xs text-gray-500 mt-1")

                        with ui.row().classes("w-full gap-3 items-end flex-wrap"):
                            drill_org_input = ui.input(
                                label="Filter by customer (optional)",
                                placeholder="e.g. Acme Corp",
                            ).classes("flex-1 min-w-48").props("outlined dense clearable")

                            cluster_drill_select = ui.select(
                                [], label="Select cluster", with_input=True,
                            ).classes("flex-1 min-w-64").props("outlined dense clearable")
                            cluster_drill_select.tooltip(
                                "Populated from snapshot_topology cluster_name / cluster_uuid. "
                                "Use customer filter to narrow by org name."
                            )

                            async def _populate_cluster_drill_select():
                                """Build the cluster dropdown filtered by org if specified."""
                                tickets_src = state.get("results") or []
                                if not tickets_src and _CB_AVAILABLE:
                                    try:
                                        tickets_src = await run.io_bound(
                                            load_tickets_from_cb,
                                            cb_url_input.value.strip(),
                                            cb_bucket_input.value.strip(),
                                            cb_user_input.value.strip(),
                                            cb_pass_input.value,
                                            cb_tls_toggle.value,
                                            cb_scope_input.value.strip() or "_default",
                                            cb_collection_input.value.strip() or "tickets",
                                            None,
                                        )
                                    except Exception:
                                        tickets_src = []

                                # Apply org filter if specified
                                org_filter = (drill_org_input.value or "").strip().lower()
                                if org_filter:
                                    tickets_src = [
                                        t for t in tickets_src
                                        if org_filter in (t.get("organization") or "").lower()
                                    ]

                                seen: set[str] = set()
                                options: list[str] = []
                                by_name: list[str] = []
                                by_uuid: list[str] = []
                                _std_uuid_re = re.compile(
                                    r"^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$",
                                    re.IGNORECASE,
                                )
                                for t in tickets_src:
                                    topo = t.get("snapshot_topology") or {}
                                    if not isinstance(topo, dict):
                                        topo = {}
                                    name       = _topo_str(topo.get("cluster_name"))
                                    capella_id = _topo_str(topo.get("capella_cluster_id"))
                                    # Also pull Cluster_ID from ticket fields — it's the
                                    # Capella UUID users see in Zendesk (matches capella_cluster_id)
                                    tf    = _parse_ticket_fields(t)
                                    tf_id = (tf.get("Cluster_ID") or "").strip()
                                    if name and name not in seen:
                                        seen.add(name)
                                        options.append(name)
                                        by_name.append(name)
                                    # Only add standard UUID-format IDs (Capella style, with dashes)
                                    # Skip bare 32-hex CB internal UUIDs — users don't work with those
                                    for uuid_val in [capella_id, tf_id]:
                                        if (uuid_val
                                                and uuid_val not in seen
                                                and _std_uuid_re.match(uuid_val)):
                                            seen.add(uuid_val)
                                            options.append(uuid_val)
                                            by_uuid.append(uuid_val)
                                options.sort()
                                cluster_drill_select.options = options
                                cluster_drill_select.update()
                                suffix = f' for "{org_filter}"' if org_filter else ""
                                if options:
                                    drill_status.set_text(
                                        f"{len(options)} cluster(s) found{suffix}: "
                                        f"{len(by_name)} by name, {len(by_uuid)} by UUID only."
                                    )
                                else:
                                    drill_status.set_text(
                                        f"No enriched clusters found{suffix} — run Enrich step first."
                                    )

                            ui.button(
                                "Refresh Clusters", icon="refresh",
                                on_click=_populate_cluster_drill_select,
                            ).props("outline dense color=grey")
                            btn_cluster_report = ui.button(
                                "Generate Cluster Report", icon="timeline",
                                on_click=lambda: asyncio.ensure_future(_generate_cluster_report()),
                            ).props("color=deep-purple")

                        drill_status = ui.label("").classes("text-xs text-gray-400 mt-1")
                        drill_chart_area = ui.column().classes("w-full gap-4 mt-2")

                        async def _generate_cluster_report():
                            key = (cluster_drill_select.value or "").strip()
                            if not key:
                                ui.notify("Select a cluster first.", type="warning")
                                return

                            drill_status.set_text("Building report …")
                            btn_cluster_report.set_enabled(False)
                            drill_chart_area.clear()

                            tickets_src = state.get("results") or []
                            if not tickets_src and _CB_AVAILABLE:
                                try:
                                    drill_status.set_text("Loading tickets from Couchbase …")
                                    tickets_src = await run.io_bound(
                                        load_tickets_from_cb,
                                        cb_url_input.value.strip(),
                                        cb_bucket_input.value.strip(),
                                        cb_user_input.value.strip(),
                                        cb_pass_input.value,
                                        cb_tls_toggle.value,
                                        cb_scope_input.value.strip() or "_default",
                                        cb_collection_input.value.strip() or "tickets",
                                        None,
                                    )
                                except Exception as exc:
                                    drill_status.set_text(f"CB load error: {exc}")
                                    btn_cluster_report.set_enabled(True)
                                    return

                            points = await run.io_bound(build_cluster_timeline, tickets_src, key)

                            if not points:
                                drill_status.set_text(f'No enriched tickets found for "{key}".')
                                btn_cluster_report.set_enabled(True)
                                return

                            chart_specs = await run.io_bound(_cluster_timeline_charts, points)

                            with drill_chart_area:
                                # ── Summary header ────────────────────────────────────────
                                with ui.card().classes("w-full bg-blue-50"):
                                    with ui.row().classes("gap-6 flex-wrap items-start"):
                                        with ui.column().classes("gap-1"):
                                            ui.label("Cluster").classes("text-xs text-gray-500")
                                            ui.label(key).classes("text-sm font-mono font-semibold")
                                        with ui.column().classes("gap-1"):
                                            ui.label("Tickets").classes("text-xs text-gray-500")
                                            ui.label(str(len(points))).classes("text-sm font-semibold")
                                        with ui.column().classes("gap-1"):
                                            ui.label("Date Range").classes("text-xs text-gray-500")
                                            ui.label(
                                                f"{points[0]['date_label']} → {points[-1]['date_label']}"
                                            ).classes("text-sm font-semibold")
                                        # Latest snapshot values
                                        last = points[-1]
                                        for label, val in [
                                            ("Nodes",        last["node_count"]),
                                            ("Buckets",      last["bucket_count"]),
                                            ("CB Version",   last["cb_version"]),
                                            ("Auto-failover", f"{last['auto_failover_sec']}s" if last["auto_failover_sec"] is not None else None),
                                            ("LDAP",         "Enabled" if last["ldap_enabled"] is True else ("Disabled" if last["ldap_enabled"] is False else None)),
                                        ]:
                                            if val is not None:
                                                with ui.column().classes("gap-1"):
                                                    ui.label(label).classes("text-xs text-gray-500")
                                                    ui.label(str(val)).classes("text-sm font-semibold")

                                # ── Ticket detail opener (reuses shared scoring-tab dialog) ──
                                _cr_tid_map = {str(t.get("ticket_id","")): t for t in tickets_src}

                                def _open_cr_ticket(ticket: dict):
                                    _ticket_dlg_body.clear()
                                    with _ticket_dlg_body:
                                        tf2 = _parse_ticket_fields(ticket)
                                        with ui.row().classes("w-full items-start justify-between gap-2"):
                                            with ui.column().classes("flex-1 gap-0"):
                                                ui.label(f"#{ticket.get('ticket_id')} · {ticket.get('organization','')}").classes("text-xs text-gray-400")
                                                ui.label(ticket.get("subject") or "").classes("text-base font-semibold")
                                            ui.button(icon="close", on_click=_ticket_dlg.close).props("flat round dense color=grey-6")
                                        with ui.row().classes("gap-4 flex-wrap text-xs text-gray-500"):
                                            for _l2, _v2 in [
                                                ("Priority",  (ticket.get("priority") or "—").upper()),
                                                ("Status",    (ticket.get("status")   or "—").capitalize()),
                                                ("Created",   (ticket.get("created")  or "")[:10]),
                                                ("Version",   extract_ticket_version(ticket)),
                                                ("Component", tf2.get("Component") or "—"),
                                            ]:
                                                with ui.column().classes("gap-0"):
                                                    ui.label(_l2).classes("text-xs text-gray-400")
                                                    ui.label(_v2).classes("text-xs font-medium")
                                        _dtopo2 = ticket.get("snapshot_topology") or {}
                                        if isinstance(_dtopo2, dict) and _dtopo2 and (
                                            _dtopo2.get("total_nodes") or _dtopo2.get("cb_version")
                                        ):
                                            ui.separator()
                                            ui.label("Cluster Topology (snapshot)").classes("text-xs font-semibold text-gray-500")
                                            _chips2 = []
                                            for _l2, _k2 in [("CB Version","cb_version"),("Nodes","total_nodes"),("Buckets","bucket_count"),("GSI","global_index_count"),("FTS Idx","fts_index_count"),("Eventing Fns","eventing_function_count"),("N2N Enc","n2n_encryption"),("Auto-failover","auto_failover_seconds")]:
                                                if _dtopo2.get(_k2) is not None:
                                                    _chips2.append((_l2, str(_dtopo2[_k2]) + ("s" if _k2 == "auto_failover_seconds" else "")))
                                            if _dtopo2.get("bad_count") or _dtopo2.get("warn_count"):
                                                _chips2.append(("Health", f"{_dtopo2.get('bad_count',0)} bad / {_dtopo2.get('warn_count',0)} warn"))
                                            with ui.row().classes("gap-2 flex-wrap mt-1"):
                                                for _l2, _v2 in _chips2:
                                                    with ui.column().classes("gap-0"):
                                                        ui.label(_l2).classes("text-xs text-gray-400")
                                                        ui.label(_v2).classes("text-xs font-medium")
                                        desc2 = (ticket.get("description") or "").strip()
                                        if desc2:
                                            ui.separator()
                                            ui.label(desc2[:2000]).classes("text-xs text-gray-700 whitespace-pre-wrap")
                                    _ticket_dlg.open()

                                # ── Time-series charts ────────────────────────────────────
                                for spec in chart_specs:
                                    _ch = spec.pop("_height", 280)
                                    _is_scatter = any(
                                        s.get("type") == "scatter"
                                        for s in (spec.get("series") or [])
                                    )
                                    with ui.card().classes("w-full"):
                                        _ec = ui.echart(spec).classes("w-full").style(f"height:{_ch}px")
                                    if _is_scatter:
                                        # Scatter: ticket_id is embedded at data[3]
                                        def _on_scatter_click(e, _tm=_cr_tid_map):
                                            d = e.data
                                            if isinstance(d, list) and len(d) > 3:
                                                t2 = _tm.get(str(d[3]))
                                                if t2:
                                                    _open_cr_ticket(t2)
                                        _ec.on_point_click(_on_scatter_click)
                                    else:
                                        # Line/bar: data_index corresponds to points[idx]
                                        def _on_tl_click(e, _pts=points, _tm=_cr_tid_map):
                                            idx = e.data_index
                                            if idx is not None and 0 <= idx < len(_pts):
                                                t2 = _tm.get(str(_pts[idx]["ticket_id"]))
                                                if t2:
                                                    _open_cr_ticket(t2)
                                        _ec.on_point_click(_on_tl_click)

                                # ── CB Version / SDK / Orchestrator change log ────────────
                                version_changes = [
                                    (p["date_label"], p["ticket_id"], p["cb_version"], p.get("sdk_version") or "")
                                    for p in points if p["cb_version"]
                                ]
                                orch_changes = [
                                    (p["date_label"], p["ticket_id"], p["orchestrator"])
                                    for p in points if p["orchestrator"]
                                ]

                                if version_changes or orch_changes:
                                    with ui.row().classes("w-full gap-4"):
                                        if version_changes:
                                            with ui.card().classes("flex-1"):
                                                ui.label("CB Version & SDK Change Log").classes("text-sm font-semibold mb-2")
                                                cols = [
                                                    {"name": "date",    "label": "Date",        "field": "date",    "align": "left"},
                                                    {"name": "tid",     "label": "Ticket",      "field": "tid",     "align": "left"},
                                                    {"name": "version", "label": "CB Version",  "field": "version", "align": "left"},
                                                    {"name": "sdk",     "label": "SDK Version", "field": "sdk",     "align": "left"},
                                                ]
                                                rows = [
                                                    {"date": d, "tid": str(tid), "version": v, "sdk": sdk}
                                                    for d, tid, v, sdk in version_changes
                                                ]
                                                ui.table(columns=cols, rows=rows, row_key="tid").classes("w-full text-xs")

                                        if orch_changes:
                                            with ui.card().classes("flex-1"):
                                                ui.label("Orchestrator History").classes("text-sm font-semibold mb-2")
                                                cols = [
                                                    {"name": "date",  "label": "Date",        "field": "date",  "align": "left"},
                                                    {"name": "tid",   "label": "Ticket",      "field": "tid",   "align": "left"},
                                                    {"name": "orch",  "label": "Orchestrator","field": "orch",  "align": "left"},
                                                ]
                                                rows = [
                                                    {"date": d, "tid": str(tid), "orch": o}
                                                    for d, tid, o in orch_changes
                                                ]
                                                ui.table(columns=cols, rows=rows, row_key="tid").classes("w-full text-xs")

                                # ── Ticket list ───────────────────────────────────────────
                                with ui.card().classes("w-full"):
                                    ui.label("Matching Tickets (chronological)").classes("text-sm font-semibold mb-2")
                                    cols = [
                                        {"name": "date",    "label": "Date",        "field": "date",    "align": "left"},
                                        {"name": "tid",     "label": "Ticket",      "field": "tid",     "align": "left"},
                                        {"name": "subject", "label": "Subject",     "field": "subject", "align": "left"},
                                        {"name": "nodes",   "label": "Nodes",       "field": "nodes",   "align": "right"},
                                        {"name": "bkts",    "label": "Buckets",     "field": "bkts",    "align": "right"},
                                        {"name": "ver",     "label": "CB Version",  "field": "ver",     "align": "left"},
                                        {"name": "sdk",     "label": "SDK Version", "field": "sdk",     "align": "left"},
                                        {"name": "bad",     "label": "BAD",         "field": "bad",     "align": "right"},
                                        {"name": "warn",    "label": "WARN",        "field": "warn",    "align": "right"},
                                    ]
                                    rows = [
                                        {
                                            "date":    p["date_label"],
                                            "tid":     str(p["ticket_id"]),
                                            "subject": p["subject"],
                                            "nodes":   p["node_count"] if p["node_count"] is not None else "",
                                            "bkts":    p["bucket_count"] if p["bucket_count"] is not None else "",
                                            "ver":     p["cb_version"] or "",
                                            "sdk":     p.get("sdk_version") or "",
                                            "bad":     p["bad_count"],
                                            "warn":    p["warn_count"],
                                        }
                                        for p in points
                                    ]
                                    ui.table(columns=cols, rows=rows, row_key="tid").classes("w-full text-xs")

                            drill_status.set_text(
                                f'Report generated for "{key}" — {len(points)} ticket(s), '
                                f"{len(chart_specs)} chart(s)."
                            )
                            btn_cluster_report.set_enabled(True)

                    with ui.tab_panel(sub_cluster):

                        with ui.column().classes("w-full gap-0"):
                            # ── Sub-tab bar ──────────────────────────────────────────
                            with ui.tabs().classes("w-full bg-white border-b border-gray-200") as ch_tabs:
                                ch_overview = ui.tab("Overview",      icon="dashboard")
                                ch_timeline = ui.tab("Timeline",      icon="timeline")
                                ch_issues   = ui.tab("Issue Tracker", icon="warning_amber")
                                ch_drift    = ui.tab("Config Drift",  icon="compare")
                                ch_llm      = ui.tab("LLM Report",    icon="psychology")
                        with ui.tab_panels(ch_tabs, value=ch_overview).classes("w-full"):

                            # ── Overview ─────────────────────────────────────────────
                            with ui.tab_panel(ch_overview):
                                with ui.column().classes("w-full gap-4 p-4"):

                                    # Controls row
                                    # Stat cards row
                                    with ui.row().classes("w-full gap-4 flex-wrap"):
                                        with ui.card().classes("px-6 py-3 text-center flex-1 min-w-32"):
                                            ch_stat_snaps = ui.label("—").classes("text-3xl font-bold text-teal-600")
                                            ui.label("Snapshots").classes("text-xs text-gray-500")
                                        with ui.card().classes("px-6 py-3 text-center flex-1 min-w-32"):
                                            ch_stat_clusters = ui.label("—").classes("text-3xl font-bold text-blue-600")
                                            ui.label("Clusters").classes("text-xs text-gray-500")
                                        with ui.card().classes("px-6 py-3 text-center flex-1 min-w-32"):
                                            ch_stat_active = ui.label("—").classes("text-3xl font-bold text-green-600")
                                            ui.label("Active (≤90d)").classes("text-xs text-gray-500")
                                        with ui.card().classes("px-6 py-3 text-center flex-1 min-w-32"):
                                            ch_stat_retired = ui.label("—").classes("text-3xl font-bold text-yellow-500")
                                            ui.label("Stale (>90d)").classes("text-xs text-gray-500")
                                        with ui.card().classes("px-6 py-3 text-center flex-1 min-w-32"):
                                            ch_stat_deprecated = ui.label("—").classes("text-3xl font-bold text-gray-400")
                                            ui.label("Deprecated").classes("text-xs text-gray-500")
                                        with ui.card().classes("px-6 py-3 text-center flex-1 min-w-32"):
                                            ch_stat_bad = ui.label("—").classes("text-3xl font-bold text-red-600")
                                            ui.label("Total Bad Issues").classes("text-xs text-gray-500")
                                        with ui.card().classes("px-6 py-3 text-center flex-1 min-w-32"):
                                            ch_stat_warn = ui.label("—").classes("text-3xl font-bold text-orange-500")
                                            ui.label("Total Warn Issues").classes("text-xs text-gray-500")

                                    # Cluster table
                                    with ui.card().classes("w-full"):
                                        ui.label("Cluster Summary").classes("text-sm font-semibold text-gray-600 mb-2")
                                        ch_cluster_table = ui.table(
                                            columns=[
                                                {"name": "cluster_name",  "label": "Cluster Name",      "field": "cluster_name",  "align": "left"},
                                                {"name": "cluster_id",    "label": "Cluster ID (short)", "field": "cluster_id",    "align": "left"},
                                                {"name": "snapshots",     "label": "Snapshots",          "field": "snapshots",     "align": "right"},
                                                {"name": "first_seen",    "label": "First Seen",         "field": "first_seen",    "align": "left"},
                                                {"name": "last_seen",     "label": "Last Seen",          "field": "last_seen",     "align": "left"},
                                                {"name": "nodes",         "label": "Nodes (min/max)",    "field": "nodes",         "align": "center"},
                                                {"name": "versions",      "label": "Versions",           "field": "versions",      "align": "left"},
                                                {"name": "avg_bad",       "label": "Avg Bad",            "field": "avg_bad",       "align": "right"},
                                                {"name": "avg_warn",      "label": "Avg Warn",           "field": "avg_warn",      "align": "right"},
                                                {"name": "status",        "label": "Status",             "field": "status",        "align": "center"},
                                            ],
                                            rows=[],
                                            row_key="cluster_id",
                                        ).classes("w-full text-xs").props("dense flat")
                                        ch_cluster_table.add_slot("body-cell-status", """
                                            <q-td :props="props">
                                              <q-badge
                                                :color="props.row.status === 'Active' ? 'green'
                                                      : props.row.status === 'Deprecated' ? 'grey'
                                                      : 'orange'">
                                                {{ props.row.status }}
                                              </q-badge>
                                            </q-td>
                                        """)

                            # ── Timeline ──────────────────────────────────────────────
                            with ui.tab_panel(ch_timeline):
                                with ui.column().classes("w-full gap-4 p-4"):
                                    with ui.card().classes("w-full"):
                                        with ui.row().classes("items-center gap-4"):
                                            ch_tl_cluster_select = ui.select(
                                                [], label="Select cluster",
                                            ).classes("flex-1 min-w-64")
                                            btn_ch_render_tl = ui.button(
                                                "Render Timeline", icon="timeline",
                                                on_click=lambda: asyncio.ensure_future(
                                                    _ch_render_timeline(ui.context.client)
                                                ),
                                            ).props("color=teal")
                                            btn_ch_render_tl.set_enabled(False)
                                        ch_tl_status = ui.label("Load snapshots in Overview first.").classes("text-sm text-gray-400")

                                    ch_tl_area = ui.column().classes("w-full gap-4")

                            # ── Issue Tracker ──────────────────────────────────────────
                            with ui.tab_panel(ch_issues):
                                with ui.column().classes("w-full gap-4 p-4"):
                                    with ui.card().classes("w-full"):
                                        with ui.row().classes("items-center gap-4"):
                                            ch_iss_cluster_select = ui.select(
                                                [], label="Cluster (blank = all)",
                                            ).classes("flex-1 min-w-64")
                                            btn_ch_render_iss = ui.button(
                                                "Render Issue Charts", icon="warning_amber",
                                                on_click=lambda: asyncio.ensure_future(
                                                    _ch_render_issues(ui.context.client)
                                                ),
                                            ).props("color=orange")
                                            btn_ch_render_iss.set_enabled(False)
                                        ch_iss_status = ui.label("Load snapshots in Overview first.").classes("text-sm text-gray-400")
                                    ch_iss_area = ui.column().classes("w-full gap-4")

                            # ── Config Drift ───────────────────────────────────────────
                            with ui.tab_panel(ch_drift):
                                with ui.column().classes("w-full gap-4 p-4"):
                                    with ui.card().classes("w-full"):
                                        ui.label("Config Drift").classes("text-base font-semibold")
                                        with ui.row().classes("items-end gap-4 flex-wrap mt-2"):
                                            ch_drift_cluster = ui.select(
                                                [], label="Cluster",
                                            ).classes("flex-1 min-w-48")
                                            ch_drift_snap_a = ui.select(
                                                [], label="Snapshot A (earlier)",
                                            ).classes("flex-1 min-w-48")
                                            ch_drift_snap_b = ui.select(
                                                [], label="Snapshot B (later)",
                                            ).classes("flex-1 min-w-48")
                                            btn_ch_diff = ui.button(
                                                "Compare", icon="compare",
                                                on_click=lambda: _ch_show_diff(),
                                            ).props("color=indigo")
                                            btn_ch_diff.set_enabled(False)
                                    ch_drift_out = ui.column().classes("w-full gap-2 p-2")

                            # ── LLM Report ─────────────────────────────────────────────
                            with ui.tab_panel(ch_llm):
                                with ui.column().classes("w-full gap-4 p-4"):
                                    with ui.card().classes("w-full"):
                                        ui.label("LLM Cluster Health Report").classes("text-base font-semibold")
                                        ui.label(
                                            "Generates a structured health narrative per cluster using the "
                                            "configured AI model. Includes upgrade detection, lifecycle "
                                            "assessment, and correlation with ticket history."
                                        ).classes("text-xs text-gray-500 mt-1")
                                        with ui.row().classes("items-end gap-4 flex-wrap mt-3"):
                                            ch_llm_cluster = ui.select(
                                                [], label="Cluster (blank = all)",
                                            ).classes("flex-1 min-w-48")
                                            btn_ch_llm = ui.button(
                                                "Generate Report", icon="psychology",
                                                on_click=lambda: asyncio.ensure_future(_ch_llm_report()),
                                            ).props("color=deep-purple")
                                            btn_ch_llm.set_enabled(False)
                                    ch_llm_status = ui.label("").classes("text-sm text-gray-500 px-4")
                                    ch_llm_breakdown = ui.column().classes("w-full gap-4 px-4")
                                    ch_llm_out = ui.markdown("").classes(
                                        "w-full px-4 prose prose-sm max-w-none "
                                        "[&_h2]:text-base [&_h2]:font-bold [&_h2]:mt-6 [&_h2]:mb-1 "
                                        "[&_h3]:text-sm [&_h3]:font-semibold [&_h3]:mt-3 [&_h3]:mb-1 "
                                        "[&_ul]:list-disc [&_ul]:pl-5 [&_ol]:list-decimal [&_ol]:pl-5 "
                                        "[&_hr]:my-4 [&_strong]:font-semibold"
                                    )

                        # ── Cluster Health async handlers ─────────────────────────────

                        def _ch_update_stats():
                            hd = ch_snap_state.get("health_data") or {}
                            ci = ch_snap_state.get("cluster_index") or {}
                            ch_stat_snaps.set_text(str(hd.get("total_snapshots", 0)))
                            ch_stat_clusters.set_text(str(hd.get("total_clusters", 0)))
                            ch_stat_active.set_text(str(hd.get("active_clusters", 0)))
                            ch_stat_retired.set_text(str(hd.get("stale_clusters", 0)))
                            ch_stat_deprecated.set_text(str(hd.get("deprecated_clusters", 0)))
                            total_bad  = sum(c.get("total_bad",  0) for c in ci.values())
                            total_warn = sum(c.get("total_warn", 0) for c in ci.values())
                            ch_stat_bad.set_text(str(total_bad))
                            ch_stat_warn.set_text(str(total_warn))

                        def _ch_update_cluster_table():
                            ci = ch_snap_state.get("cluster_index") or {}
                            rows = []
                            for cid, c in sorted(
                                ci.items(),
                                key=lambda kv: kv[1].get("last_seen") or "",
                                reverse=True,
                            ):
                                nmin = c.get("node_count_min")
                                nmax = c.get("node_count_max")
                                nodes_str = (
                                    f"{nmin}/{nmax}" if nmin is not None and nmax is not None
                                    else "—"
                                )
                                rows.append({
                                    "cluster_id":   cid[:16] + "…" if len(cid) > 16 else cid,
                                    "cluster_name": c.get("cluster_name") or "—",
                                    "snapshots":    c.get("snapshot_count", 0),
                                    "first_seen":   (c.get("first_seen") or "")[:10],
                                    "last_seen":    (c.get("last_seen") or "")[:10],
                                    "nodes":        nodes_str,
                                    "versions":     ", ".join(c.get("version_history") or [])[:30],
                                    "avg_bad":      c.get("avg_bad", 0),
                                    "avg_warn":     c.get("avg_warn", 0),
                                    "status":       "Active" if c.get("is_active") else ("Deprecated" if c.get("is_deprecated") else "Stale"),
                                })
                            ch_cluster_table.rows = rows
                            ch_cluster_table.update()

                        def _ch_populate_selectors():
                            ci = ch_snap_state.get("cluster_index") or {}
                            options = {
                                cid: f"{c.get('cluster_name') or cid[:16]} ({c.get('snapshot_count',0)} snaps)"
                                for cid, c in ci.items()
                            }
                            opts_with_blank = {"": "— all —", **options}
                            ch_tl_cluster_select.set_options(options)
                            ch_iss_cluster_select.set_options(opts_with_blank)
                            ch_drift_cluster.set_options(options)
                            ch_llm_cluster.set_options(opts_with_blank)
                            btn_ch_render_tl.set_enabled(bool(options))
                            btn_ch_render_iss.set_enabled(True)
                            btn_ch_llm.set_enabled(True)
                            btn_ch_diff.set_enabled(bool(options))

                        def _ch_refresh_drift_snaps(cid: str):
                            """Populate Snapshot A/B selectors when a cluster is chosen in Config Drift."""
                            by_cluster = ch_snap_state.get("health_data", {}).get("by_cluster", {})
                            snaps = by_cluster.get(cid, [])
                            opts = {s["snap_id"]: f"{s['date'][:10] if s['date'] else '?'} — {s['snap_id'][:12]}"
                                    for s in snaps}
                            ch_drift_snap_a.set_options(opts)
                            ch_drift_snap_b.set_options(opts)

                        ch_drift_cluster.on("update:model-value", lambda _: _ch_refresh_drift_snaps(ch_drift_cluster.value))

                        def _ch_clear_cache():
                            ch_snap_state["snapshots"] = []
                            ch_snap_state["health_data"] = {}
                            ch_snap_state["cluster_index"] = {}
                            ch_snap_state["last_customer"] = ""
                            ch_status.set_text("Cache cleared — ready to re-scrape.")
                            btn_ch_save_cb.set_enabled(False)
                            _ch_update_stats()
                            _ch_update_cluster_table()
                            ui.notify("Snapshot cache cleared.", type="info")

                        async def _ch_scrape(client=None):
                            customer = (ch_cust_input.value or main_cust_input.value or "").strip()
                            if not customer:
                                ui.notify("Enter a customer name or URL.", type="warning")
                                return
                            btn_ch_scrape.set_enabled(False)
                            ch_progress.set_visibility(True)
                            ch_progress.set_value(0)
                            loop = asyncio.get_event_loop()

                            def _prog(msg: str, pct: float):
                                asyncio.run_coroutine_threadsafe(
                                    _ch_prog_upd(msg, pct), loop
                                )

                            async def _ch_prog_upd(msg: str, pct: float):
                                ch_status.set_text(msg)
                                ch_progress.set_value(pct)

                            new_snaps: list = []
                            # Clear accumulated state when scraping a different customer
                            _last_cust = ch_snap_state.get("last_customer", "")
                            if _last_cust and _last_cust.lower() != customer.lower():
                                ch_snap_state["snapshots"] = []
                                ch_snap_state["health_data"] = {}
                                ch_snap_state["cluster_index"] = {}
                            ch_snap_state["last_customer"] = customer
                            # Build skip set: complete snaps in CB + in-memory complete snaps
                            snap_signals: dict = {}
                            max_snaps = int(ch_max_snapshots.value or 0)
                            if _CB_AVAILABLE and cb_url_input.value.strip():
                                ch_status.set_text("Checking Couchbase for already-complete snapshots…")
                                snap_signals = await run.io_bound(
                                    fetch_snapshot_signals_from_cb,
                                    cb_url_input.value.strip(),
                                    cb_bucket_input.value.strip() or "supportal",
                                    cb_user_input.value.strip(),
                                    cb_pass_input.value,
                                    cb_tls_toggle.value,
                                    cb_scope_input.value.strip() or "_default",
                                    ch_snap_coll.value.strip() or "snapshots",
                                )
                            # Also skip in-memory complete snaps not yet in CB
                            for s in ch_snap_state["snapshots"]:
                                sid = s.get("snap_id", "")
                                if sid and (s.get("nutshell_html") or s.get("cb_ver") or s.get("structured_data")):
                                    if sid not in snap_signals:
                                        snap_signals[sid] = {"complete": True}
                            skip = {sid for sid, sig in snap_signals.items() if sig.get("complete")}
                            skip.discard("")

                            try:
                                max_p = int(ch_max_pages.value or 0)
                                workers = int(ch_workers.value or 4)
                                new_snaps = await run.io_bound(
                                    scrape_snapshots_for_customer,
                                    customer,
                                    cookie_input.value or None,
                                    max_p,
                                    workers,
                                    _prog,
                                    skip,
                                    max_snaps,
                                )
                                # Merge: scraped results replace stubs with same snap_id.
                                scraped_ids = {s.get("snap_id", "") for s in new_snaps}
                                kept = [s for s in ch_snap_state["snapshots"] if s.get("snap_id", "") not in scraped_ids]
                                ch_snap_state["snapshots"] = kept + new_snaps
                            except Exception as exc:
                                import traceback as _tb
                                _tb.print_exc()
                                ch_status.set_text(f"Scrape error: {exc}")
                                try:
                                    ui.notify(str(exc), type="negative")
                                except Exception:
                                    pass
                            # Always rebuild health data and enable save if we have anything
                            try:
                                tickets = state.get("results") or []
                                ch_snap_state["health_data"] = build_cluster_health_data(
                                    ch_snap_state["snapshots"], tickets
                                )
                                ch_snap_state["cluster_index"] = ch_snap_state["health_data"]["cluster_index"]
                                _ch_update_stats()
                                _ch_update_cluster_table()
                                _ch_populate_selectors()
                            except Exception as _hd_exc:
                                import traceback as _tb2
                                _tb2.print_exc()
                                print(f"[CH] Health data build error: {_hd_exc}")
                            # Enable save whenever there are snapshots in memory
                            all_snaps = ch_snap_state.get("snapshots") or []
                            btn_ch_save_cb.set_enabled(bool(all_snaps) and _CB_AVAILABLE)
                            btn_ch_embed_cb.set_enabled(bool(all_snaps) and _CB_AVAILABLE)
                            if new_snaps:
                                ch_status.set_text(
                                    f"Scraped {len(new_snaps)} new snapshots "
                                    f"({len(all_snaps)} total)."
                                )
                                try:
                                    ui.notify(f"Scraped {len(new_snaps)} new snapshots.", type="positive")
                                except Exception:
                                    pass
                            elif all_snaps and not ch_status.text.startswith("Scrape error"):
                                ch_status.set_text(f"{len(all_snaps)} snapshots in memory (no new).")

                            if new_snaps and ch_auto_save_cb.value and _CB_AVAILABLE and cb_url_input.value.strip():
                                ch_status.set_text(f"Auto-saving {len(new_snaps)} snapshots to Couchbase…")
                                try:
                                    _saved, _errs = await run.io_bound(
                                        load_snapshots_to_couchbase,
                                        new_snaps,
                                        cb_url_input.value.strip(),
                                        cb_bucket_input.value.strip(),
                                        cb_user_input.value.strip(),
                                        cb_pass_input.value,
                                        cb_tls_toggle.value,
                                        cb_scope_input.value.strip() or "_default",
                                        ch_snap_coll.value.strip() or "snapshots",
                                        lambda msg, pct: None,
                                    )
                                    ch_status.set_text(
                                        f"Saved {_saved} snapshots ({_errs} errors). {len(all_snaps)} total in memory."
                                    )
                                except Exception as _ae:
                                    ch_status.set_text(f"Auto-save error: {_ae}")

                            btn_ch_scrape.set_enabled(True)
                            ch_progress.set_visibility(False)

                        async def _ch_fetch_analytics(client=None):
                            """Fetch snapshot listing + ticket IDs via analytics API (fast, no Playwright)."""
                            customer = (ch_cust_input.value or main_cust_input.value or "").strip()
                            if not customer:
                                _safe_notify(client, "Enter a customer name or URL.", type="warning")
                                return
                            btn_ch_fetch_analytics.set_enabled(False)
                            ch_progress.set_visibility(True)
                            ch_progress.set_value(0)
                            loop = asyncio.get_event_loop()

                            def _prog(msg: str, pct: float):
                                asyncio.run_coroutine_threadsafe(
                                    _ch_analytics_upd(msg, pct), loop
                                )

                            async def _ch_analytics_upd(msg: str, pct: float):
                                ch_status.set_text(msg)
                                ch_progress.set_value(pct)

                            _last_cust = ch_snap_state.get("last_customer", "")
                            if _last_cust and _last_cust.lower() != customer.lower():
                                ch_snap_state["snapshots"] = []
                                ch_snap_state["health_data"] = {}
                                ch_snap_state["cluster_index"] = {}
                            ch_snap_state["last_customer"] = customer

                            # Strip URL prefix if user pasted a full Supportal URL
                            cust_name = customer
                            if customer.startswith("http"):
                                cust_name = urllib.parse.unquote(
                                    customer.rstrip("/").rsplit("/", 1)[-1]
                                ).replace("-", " ").title()

                            new_snaps: list = []
                            try:
                                limit = int(ch_analytics_limit.value or 200)
                                cookie = (cookie_input.value or "").strip() or _get_profile_cookie()

                                new_snaps = await run.io_bound(
                                    fetch_snapshots_via_analytics,
                                    cust_name, cookie, limit, _prog,
                                )
                                # Merge with existing — skip snap_ids already present
                                existing_ids = {s.get("snap_id", "") for s in ch_snap_state["snapshots"]}
                                added = [s for s in new_snaps if s.get("snap_id", "") not in existing_ids]
                                ch_snap_state["snapshots"] = ch_snap_state["snapshots"] + added
                                new_snaps = added
                            except Exception as exc:
                                import traceback as _tb
                                _tb.print_exc()
                                ch_status.set_text(f"Analytics fetch error: {exc}")
                                try:
                                    if client:
                                        _safe_notify(client, str(exc), type="negative")
                                    else:
                                        ui.notify(str(exc), type="negative")
                                except Exception:
                                    pass
                                btn_ch_fetch_analytics.set_enabled(True)
                                ch_progress.set_visibility(False)
                                return

                            try:
                                tickets = state.get("results") or []
                                ch_snap_state["health_data"] = build_cluster_health_data(
                                    ch_snap_state["snapshots"], tickets
                                )
                                ch_snap_state["cluster_index"] = ch_snap_state["health_data"]["cluster_index"]
                                _ch_update_stats()
                                _ch_update_cluster_table()
                                _ch_populate_selectors()
                            except Exception as _hd_exc:
                                print(f"[CH] Health data build error: {_hd_exc}")

                            all_snaps = ch_snap_state.get("snapshots") or []
                            btn_ch_save_cb.set_enabled(bool(all_snaps) and _CB_AVAILABLE)
                            btn_ch_embed_cb.set_enabled(bool(all_snaps) and _CB_AVAILABLE)
                            msg = (
                                f"Analytics: {len(new_snaps)} new snapshots fetched "
                                f"({len(all_snaps)} total). Topology not loaded — use Scrape for full detail."
                            )
                            ch_status.set_text(msg)
                            btn_ch_fetch_analytics.set_enabled(True)
                            ch_progress.set_visibility(False)

                        async def _ch_scrape_stubs(client=None):
                            """Scrape full topology for analytics stubs, skipping CB-complete snaps."""
                            all_stubs = [
                                s for s in ch_snap_state.get("snapshots", [])
                                if not (s.get("nutshell_html") or s.get("cb_ver") or s.get("structured_data"))
                                and s.get("snap_id")
                            ]
                            if not all_stubs:
                                _safe_notify(client, "No analytics stubs to scrape — fetch via Analytics API first, or all snaps already have topology.", type="info")
                                return
                            btn_ch_scrape_stubs.set_enabled(False)
                            ch_progress.set_visibility(True)
                            ch_progress.set_value(0)
                            loop = asyncio.get_event_loop()

                            def _prog(msg: str, pct: float):
                                asyncio.run_coroutine_threadsafe(
                                    _upd(msg, pct), loop
                                )

                            async def _upd(msg: str, pct: float):
                                ch_status.set_text(msg)
                                ch_progress.set_value(pct)

                            # Check CB for already-complete snapshots and skip them
                            stubs = all_stubs
                            max_snaps = int(ch_max_snapshots.value or 0)
                            if _CB_AVAILABLE and cb_url_input.value.strip():
                                ch_status.set_text("Checking Couchbase for already-complete snapshots…")
                                snap_signals = await run.io_bound(
                                    fetch_snapshot_signals_from_cb,
                                    cb_url_input.value.strip(),
                                    cb_bucket_input.value.strip() or "supportal",
                                    cb_user_input.value.strip(),
                                    cb_pass_input.value,
                                    cb_tls_toggle.value,
                                    cb_scope_input.value.strip() or "_default",
                                    ch_snap_coll.value.strip() or "snapshots",
                                )
                                stubs, n_new, n_incomplete, n_skipped = _filter_incomplete_snapshots(
                                    all_stubs, snap_signals, max_snaps
                                )
                                ch_status.set_text(
                                    f"Completeness check: {n_new} new, {n_incomplete} incomplete, "
                                    f"{n_skipped} complete (skipped)"
                                    + (f", capped at {max_snaps}" if max_snaps > 0 else "") + "."
                                )
                                if not stubs:
                                    _safe_notify(client, f"All {n_skipped} snapshots already complete in Couchbase.", type="info")
                                    btn_ch_scrape_stubs.set_enabled(True)
                                    ch_progress.set_visibility(False)
                                    return
                            elif max_snaps > 0:
                                stubs = stubs[:max_snaps]

                            try:
                                workers = int(ch_workers.value or 4)
                                scraped = await run.io_bound(
                                    scrape_snapshots_from_stubs,
                                    stubs,
                                    cookie_input.value or None,
                                    workers,
                                    _prog,
                                )
                                # Replace stubs with fully-scraped docs
                                scraped_ids = {s.get("snap_id", "") for s in scraped}
                                kept = [s for s in ch_snap_state["snapshots"] if s.get("snap_id", "") not in scraped_ids]
                                ch_snap_state["snapshots"] = kept + scraped
                            except Exception as exc:
                                import traceback as _tb
                                _tb.print_exc()
                                ch_status.set_text(f"Scrape error: {exc}")
                                _safe_notify(client, str(exc), type="negative")
                                btn_ch_scrape_stubs.set_enabled(True)
                                ch_progress.set_visibility(False)
                                return

                            try:
                                tickets = state.get("results") or []
                                ch_snap_state["health_data"] = build_cluster_health_data(
                                    ch_snap_state["snapshots"], tickets
                                )
                                ch_snap_state["cluster_index"] = ch_snap_state["health_data"]["cluster_index"]
                                _ch_update_stats()
                                _ch_update_cluster_table()
                                _ch_populate_selectors()
                            except Exception as _hd_exc:
                                print(f"[CH] Health data build error: {_hd_exc}")

                            all_snaps = ch_snap_state.get("snapshots") or []
                            btn_ch_save_cb.set_enabled(bool(all_snaps) and _CB_AVAILABLE)
                            btn_ch_embed_cb.set_enabled(bool(all_snaps) and _CB_AVAILABLE)
                            ch_status.set_text(
                                f"Scraped topology for {len(scraped)} snapshots ({len(all_snaps)} total)."
                            )
                            _safe_notify(client, f"Scraped {len(scraped)} snapshots.", type="positive")

                            if scraped and ch_auto_save_cb.value and _CB_AVAILABLE and cb_url_input.value.strip():
                                ch_status.set_text(f"Auto-saving {len(scraped)} snapshots to Couchbase…")
                                try:
                                    _saved, _errs = await run.io_bound(
                                        load_snapshots_to_couchbase,
                                        scraped,
                                        cb_url_input.value.strip(),
                                        cb_bucket_input.value.strip(),
                                        cb_user_input.value.strip(),
                                        cb_pass_input.value,
                                        cb_tls_toggle.value,
                                        cb_scope_input.value.strip() or "_default",
                                        ch_snap_coll.value.strip() or "snapshots",
                                        lambda msg, pct: None,
                                    )
                                    ch_status.set_text(
                                        f"Saved {_saved} snapshots ({_errs} errors). {len(all_snaps)} total in memory."
                                    )
                                except Exception as _ae:
                                    ch_status.set_text(f"Auto-save error: {_ae}")

                            btn_ch_scrape_stubs.set_enabled(True)
                            ch_progress.set_visibility(False)

                        async def _ch_load_cb():
                            if not _CB_AVAILABLE:
                                ui.notify("Couchbase SDK not available.", type="warning")
                                return
                            ch_status.set_text("Loading snapshots from Couchbase…")
                            btn_ch_load_cb.set_enabled(False)
                            try:
                                customer_f = (ch_cust_input.value or main_cust_input.value or "").strip()

                                def _prog(msg, pct):
                                    pass

                                snaps = await run.io_bound(
                                    load_snapshots_from_couchbase,
                                    cb_url_input.value.strip(),
                                    cb_bucket_input.value.strip(),
                                    cb_user_input.value.strip(),
                                    cb_pass_input.value,
                                    cb_tls_toggle.value,
                                    cb_scope_input.value.strip() or "_default",
                                    ch_snap_coll.value.strip() or "snapshots",
                                    customer_f,
                                    _prog,
                                )
                                ch_snap_state["snapshots"] = snaps
                                tickets = state.get("results") or []
                                ch_snap_state["health_data"] = build_cluster_health_data(snaps, tickets)
                                ch_snap_state["cluster_index"] = ch_snap_state["health_data"]["cluster_index"]
                                _ch_update_stats()
                                _ch_update_cluster_table()
                                _ch_populate_selectors()
                                ch_status.set_text(f"Loaded {len(snaps)} snapshots from Couchbase.")
                                try:
                                    ui.notify(f"Loaded {len(snaps)} snapshots.", type="positive")
                                except Exception:
                                    pass
                            except Exception as exc:
                                ch_status.set_text(f"CB load error: {exc}")
                                try:
                                    ui.notify(str(exc), type="negative")
                                except Exception:
                                    pass
                            finally:
                                btn_ch_load_cb.set_enabled(True)

                        async def _ch_save_cb():
                            if not _CB_AVAILABLE:
                                ui.notify("Couchbase SDK not available.", type="warning")
                                return
                            snaps = ch_snap_state.get("snapshots") or []
                            if not snaps:
                                ui.notify("No snapshots to save.", type="warning")
                                return
                            ch_status.set_text(f"Saving {len(snaps)} snapshots to Couchbase…")
                            btn_ch_save_cb.set_enabled(False)
                            try:
                                def _prog(msg, pct):
                                    pass

                                upserted, errors = await run.io_bound(
                                    load_snapshots_to_couchbase,
                                    snaps,
                                    cb_url_input.value.strip(),
                                    cb_bucket_input.value.strip(),
                                    cb_user_input.value.strip(),
                                    cb_pass_input.value,
                                    cb_tls_toggle.value,
                                    cb_scope_input.value.strip() or "_default",
                                    ch_snap_coll.value.strip() or "snapshots",
                                    _prog,
                                )
                                ch_status.set_text(f"Saved {upserted} snapshots ({errors} errors).")
                                btn_ch_embed_cb.set_enabled(bool(snaps) and _CB_AVAILABLE)
                                try:
                                    ui.notify(f"Saved {upserted} snapshots.", type="positive")
                                except Exception:
                                    pass
                            except Exception as exc:
                                ch_status.set_text(f"Save error: {exc}")
                                try:
                                    ui.notify(str(exc), type="negative")
                                except Exception:
                                    pass
                            finally:
                                btn_ch_save_cb.set_enabled(True)

                        async def _ch_embed_cb():
                            if not _CB_AVAILABLE:
                                ui.notify("Couchbase SDK not available.", type="warning")
                                return
                            snaps = ch_snap_state.get("snapshots") or []
                            if not snaps:
                                ui.notify("No snapshots in memory — scrape or load first.", type="warning")
                                return
                            ep, em, ek, eu, ed, _enctx = _get_embed_config()
                            if not ep or not em:
                                ui.notify("Configure embedding model first (Configuration tab).", type="warning")
                                return
                            ch_status.set_text(f"Embedding {len(snaps)} snapshots …")
                            btn_ch_embed_cb.set_enabled(False)
                            ch_progress.set_visibility(True)
                            ch_progress.set_value(0)
                            loop = asyncio.get_event_loop()

                            def _prog(msg: str, pct: float):
                                asyncio.run_coroutine_threadsafe(
                                    _ch_emb_upd(msg, pct), loop
                                )

                            async def _ch_emb_upd(msg: str, pct: float):
                                ch_status.set_text(msg)
                                ch_progress.set_value(pct)

                            try:
                                done, errs = await run.io_bound(
                                    embed_all_snapshots,
                                    snaps,
                                    cb_url_input.value.strip(),
                                    cb_bucket_input.value.strip(),
                                    cb_user_input.value.strip(),
                                    cb_pass_input.value,
                                    cb_tls_toggle.value,
                                    cb_scope_input.value.strip() or "_default",
                                    ch_snap_coll.value.strip() or "snapshots",
                                    ep, em, ek, eu, int(ed or 1024),
                                    _prog,
                                )
                                ch_status.set_text(f"Embedded {done} snapshots ({errs} errors).")
                                try:
                                    ui.notify(f"Embedded {done} snapshots.", type="positive")
                                except Exception:
                                    pass
                            except Exception as exc:
                                ch_status.set_text(f"Embed error: {exc}")
                                try:
                                    ui.notify(str(exc), type="negative")
                                except Exception:
                                    pass
                            finally:
                                btn_ch_embed_cb.set_enabled(True)
                                ch_progress.set_visibility(False)

                        async def _ch_render_timeline(client=None):
                            cid = ch_tl_cluster_select.value or ""
                            if not cid:
                                ui.notify("Select a cluster first.", type="warning")
                                return
                            by_cluster = ch_snap_state.get("health_data", {}).get("by_cluster", {})
                            series_data = by_cluster.get(cid, [])
                            if not series_data:
                                ch_tl_status.set_text("No timeline data for this cluster.")
                                return

                            ch_tl_status.set_text(f"Rendering {len(series_data)} snapshots…")
                            ch_tl_area.clear()

                            ci = ch_snap_state.get("cluster_index", {}).get(cid, {})
                            cluster_label = ci.get("cluster_name") or cid[:16]
                            dates    = [s["date"][:10] if s.get("date") else str(i) for i, s in enumerate(series_data)]
                            bad_vals  = [s["bad_count"]  for s in series_data]
                            warn_vals = [s["warn_count"] for s in series_data]
                            node_vals = [s["node_count"] for s in series_data]

                            _win_h = 768
                            if client:
                                try:
                                    _win_h = int(await client.run_javascript("window.innerHeight") or 768)
                                except Exception:
                                    pass
                            ch_h = max(300, int(_win_h * 0.36))

                            with ch_tl_area:
                                # Bad + Warn over time
                                ui.echart({
                                    "title":    {"text": f"Issue Count Over Time — {cluster_label}", "subtext": "bad=red, warn=orange · drag to zoom"},
                                    "tooltip":  {"trigger": "axis"},
                                    "legend":   {"bottom": 0},
                                    "dataZoom": [{"type": "inside"}, {"type": "slider", "bottom": 30}],
                                    "grid":     {"bottom": 70},
                                    "xAxis":    {"type": "category", "data": dates, "axisLabel": {"rotate": 45}},
                                    "yAxis":    {"type": "value", "name": "Issue Count", "minInterval": 1, "min": 0},
                                    "color":    ["#E53935", "#FB8C00"],
                                    "series":   [
                                        {"name": "Bad",  "type": "line", "smooth": True, "data": bad_vals,  "itemStyle": {"color": "#E53935"}},
                                        {"name": "Warn", "type": "line", "smooth": True, "data": warn_vals, "itemStyle": {"color": "#FB8C00"}},
                                    ],
                                }).classes("w-full").style(f"height:{ch_h}px")

                                # Node count over time
                                if any(n > 0 for n in node_vals):
                                    _nc_h = max(240, int(_win_h * 0.25))
                                    ui.echart({
                                        "title":    {"text": f"Node Count Over Time — {cluster_label}", "subtext": f"min={ci.get('node_count_min')} · max={ci.get('node_count_max')}"},
                                        "tooltip":  {"trigger": "axis"},
                                        "dataZoom": [{"type": "inside"}, {"type": "slider", "bottom": 30}],
                                        "grid":     {"bottom": 70},
                                        "xAxis":    {"type": "category", "data": dates, "axisLabel": {"rotate": 45}},
                                        "yAxis":    {"type": "value", "name": "Nodes", "minInterval": 1, "min": 0},
                                        "color":    ["#1E88E5"],
                                        "series":   [{"name": "Nodes", "type": "line", "smooth": True, "data": node_vals}],
                                    }).classes("w-full").style(f"height:{_nc_h}px")

                                # Version markers (show where version changed)
                                version_changes = []
                                last_ver = None
                                for i, s in enumerate(series_data):
                                    ver = s.get("cb_version") or ""
                                    if ver and ver != last_ver:
                                        version_changes.append({"date": dates[i], "version": ver, "idx": i})
                                        last_ver = ver
                                if len(version_changes) > 1:
                                    with ui.card().classes("w-full"):
                                        ui.label("Version History").classes("text-sm font-semibold text-gray-600 mb-1")
                                        with ui.row().classes("gap-4 flex-wrap"):
                                            for vc in version_changes:
                                                ui.label(f"{vc['date']}: {vc['version']}").classes(
                                                    "text-xs px-2 py-1 bg-blue-50 border border-blue-200 rounded"
                                                )

                            ch_tl_status.set_text(
                                f"Timeline rendered: {len(series_data)} snapshots for {cluster_label}."
                            )

                        async def _ch_render_issues(client=None):
                            cid_filter = (ch_iss_cluster_select.value or "").strip()
                            hd = ch_snap_state.get("health_data") or {}
                            by_cluster = hd.get("by_cluster") or {}
                            ci = ch_snap_state.get("cluster_index") or {}

                            if cid_filter:
                                clusters_to_show = {cid_filter: by_cluster.get(cid_filter, [])}
                            else:
                                clusters_to_show = by_cluster

                            if not clusters_to_show:
                                ui.notify("No snapshot data. Load snapshots in Overview first.", type="warning")
                                return

                            ch_iss_area.clear()
                            ch_iss_status.set_text("Rendering issue charts…")

                            _win_h = 768
                            if client:
                                try:
                                    _win_h = int(await client.run_javascript("window.innerHeight") or 768)
                                except Exception:
                                    pass
                            ch_h = max(300, int(_win_h * 0.36))

                            with ch_iss_area:
                                # Heatmap table: months × clusters
                                heatmap = hd.get("heatmap") or {}
                                months  = sorted(heatmap.keys())
                                cids    = sorted(ci.keys()) if not cid_filter else [cid_filter]

                                if months and cids:
                                    with ui.card().classes("w-full"):
                                        ui.label("Issue Heatmap (Bad + Warn per month × cluster)").classes(
                                            "text-sm font-semibold text-gray-600 mb-2"
                                        )
                                        # Render as a simple table using NiceGUI ui.table
                                        hm_cols = [{"name": "month", "label": "Month", "field": "month", "align": "left"}]
                                        for cid2 in cids[:20]:  # cap at 20 clusters for readability
                                            cname = ci.get(cid2, {}).get("cluster_name") or cid2[:12]
                                            hm_cols.append({"name": cid2, "label": cname, "field": cid2, "align": "right"})
                                        hm_rows = []
                                        for month in months:
                                            row = {"month": month}
                                            for cid2 in cids[:20]:
                                                row[cid2] = heatmap.get(month, {}).get(cid2, 0)
                                            hm_rows.append(row)
                                        ui.table(columns=hm_cols, rows=hm_rows, row_key="month").classes(
                                            "w-full text-xs"
                                        ).props("dense flat")

                                # Per-cluster bad+warn aggregate bar chart (top 20)
                                sorted_ci = sorted(
                                    ci.items(),
                                    key=lambda kv: kv[1].get("total_bad", 0) + kv[1].get("total_warn", 0),
                                    reverse=True,
                                )
                                top20 = sorted_ci[:20] if not cid_filter else [(cid_filter, ci.get(cid_filter, {}))]
                                if top20:
                                    labels = [c.get("cluster_name") or cid2[:12] for cid2, c in top20]
                                    bad_data  = [c.get("total_bad",  0) for _, c in top20]
                                    warn_data = [c.get("total_warn", 0) for _, c in top20]
                                    _iss_h = max(ch_h, len(top20) * 28 + 80)
                                    ui.echart({
                                        "title":   {"text": "Total Bad + Warn Issues by Cluster"},
                                        "tooltip": {"trigger": "axis", "axisPointer": {"type": "shadow"}},
                                        "legend":  {"bottom": 0},
                                        "grid":    {"left": 140, "right": 20, "bottom": 40},
                                        "xAxis":   {"type": "value", "name": "Issue Count"},
                                        "yAxis":   {"type": "category", "data": labels, "axisLabel": {"overflow": "truncate", "width": 120}},
                                        "color":   ["#E53935", "#FB8C00"],
                                        "series":  [
                                            {"name": "Bad",  "type": "bar", "data": bad_data},
                                            {"name": "Warn", "type": "bar", "data": warn_data},
                                        ],
                                    }).classes("w-full").style(f"height:{_iss_h}px")

                            ch_iss_status.set_text(
                                f"Issue charts rendered for "
                                f"{'cluster ' + ci.get(cid_filter, {}).get('cluster_name', cid_filter[:12]) if cid_filter else 'all clusters'}."
                            )

                        def _ch_show_diff():
                            cid  = ch_drift_cluster.value or ""
                            sid_a = ch_drift_snap_a.value or ""
                            sid_b = ch_drift_snap_b.value or ""
                            if not (cid and sid_a and sid_b):
                                ui.notify("Select cluster and both snapshots.", type="warning")
                                return
                            if sid_a == sid_b:
                                ui.notify("Select two different snapshots.", type="warning")
                                return

                            by_cluster = ch_snap_state.get("health_data", {}).get("by_cluster", {})
                            snaps = {s["snap_id"]: s for s in by_cluster.get(cid, [])}

                            # Also look in the flat list
                            for s in ch_snap_state.get("snapshots", []):
                                snaps.setdefault(s["snap_id"], s)

                            def _topo(sid):
                                s = snaps.get(sid, {})
                                t = s.get("topology") or {}
                                return {
                                    "cb_version":          t.get("cb_version") or s.get("cb_version") or "—",
                                    "total_nodes":         t.get("total_nodes") or s.get("node_count") or "—",
                                    "bucket_count":        t.get("bucket_count") or s.get("bucket_count") or "—",
                                    "bucket_names":        ", ".join(t.get("bucket_names") or s.get("bucket_names") or []) or "—",
                                    "ram_per_node_mib":    t.get("ram_per_node_mib") or s.get("ram_per_node_mib") or "—",
                                    "auto_failover_seconds": t.get("auto_failover_seconds") if t.get("auto_failover_seconds") is not None else (s.get("auto_failover_seconds") if s.get("auto_failover_seconds") is not None else "—"),
                                    "bad_count":           t.get("bad_count", 0) or s.get("bad_count", 0),
                                    "warn_count":          t.get("warn_count", 0) or s.get("warn_count", 0),
                                    "server_groups":       ", ".join(t.get("server_groups") or s.get("server_groups") or []) or "—",
                                }

                            topo_a = _topo(sid_a)
                            topo_b = _topo(sid_b)
                            fields = list(topo_a.keys())
                            date_a = (snaps.get(sid_a) or {}).get("date", "?")[:10]
                            date_b = (snaps.get(sid_b) or {}).get("date", "?")[:10]

                            ch_drift_out.clear()
                            with ch_drift_out:
                                with ui.card().classes("w-full"):
                                    ui.label(
                                        f"Comparing {sid_a[:14]}… ({date_a})  vs  {sid_b[:14]}… ({date_b})"
                                    ).classes("text-sm font-semibold text-gray-600 mb-2")
                                    diff_cols = [
                                        {"name": "field",   "label": "Field",      "field": "field",   "align": "left"},
                                        {"name": "snap_a",  "label": f"A  {date_a}", "field": "snap_a", "align": "left"},
                                        {"name": "snap_b",  "label": f"B  {date_b}", "field": "snap_b", "align": "left"},
                                        {"name": "changed", "label": "Changed",    "field": "changed", "align": "center"},
                                    ]
                                    diff_rows = [
                                        {
                                            "field":   f,
                                            "snap_a":  str(topo_a[f]),
                                            "snap_b":  str(topo_b[f]),
                                            "changed": "✓" if str(topo_a[f]) != str(topo_b[f]) else "",
                                        }
                                        for f in fields
                                    ]
                                    ui.table(columns=diff_cols, rows=diff_rows, row_key="field").classes(
                                        "w-full text-xs"
                                    ).props("dense flat")

                        async def _ch_llm_report():
                            cid_filter = (ch_llm_cluster.value or "").strip()
                            ci = ch_snap_state.get("cluster_index") or {}
                            if not ci:
                                ui.notify("Load snapshots first.", type="warning")
                                return

                            btn_ch_llm.set_enabled(False)
                            ch_llm_status.set_text("Generating report…")
                            ch_llm_out.set_content("")
                            ch_llm_breakdown.clear()

                            # Build structured summary prompt
                            if cid_filter and cid_filter in ci:
                                clusters_to_report = {cid_filter: ci[cid_filter]}
                            else:
                                # Limit to top 10 by snapshot count to fit in context
                                clusters_to_report = dict(
                                    sorted(ci.items(), key=lambda kv: kv[1].get("snapshot_count", 0), reverse=True)[:10]
                                )

                            # ── Pre-report breakdown UI ───────────────────────────────
                            by_cluster = (ch_snap_state.get("health_data") or {}).get("by_cluster", {})
                            with ch_llm_breakdown:
                                for cid2, c in clusters_to_report.items():
                                    cname = c.get("cluster_name") or cid2[:16]
                                    with ui.card().classes("w-full"):
                                        ui.label(f"{cname}  ·  {cid2[:16]}").classes("text-sm font-semibold mb-2")
                                        # Bad/Warn timeline chart
                                        series = by_cluster.get(cid2) or []
                                        if series:
                                            dates     = [s["date"][:10] if s["date"] else "?" for s in series]
                                            bad_vals  = [s.get("bad_count",  0) for s in series]
                                            warn_vals = [s.get("warn_count", 0) for s in series]
                                            ui.echart({
                                                "tooltip": {"trigger": "axis"},
                                                "legend": {"data": ["Bad", "Warn"], "top": 0},
                                                "xAxis": {"type": "category", "data": dates,
                                                          "axisLabel": {"rotate": 45, "fontSize": 9}},
                                                "yAxis": {"type": "value"},
                                                "series": [
                                                    {"name": "Bad",  "type": "bar", "stack": "issues",
                                                     "data": bad_vals,  "color": "#EF5350"},
                                                    {"name": "Warn", "type": "bar", "stack": "issues",
                                                     "data": warn_vals, "color": "#FFA726"},
                                                ],
                                            }).classes("w-full h-40")
                                        # Recurring checkers tables
                                        bad_items  = c.get("top_bad_items")  or []
                                        warn_items = c.get("top_warn_items") or []
                                        if bad_items or warn_items:
                                            with ui.row().classes("w-full gap-4 mt-2 flex-wrap"):
                                                if bad_items:
                                                    with ui.column().classes("flex-1 min-w-48"):
                                                        ui.label("Recurring BAD checkers").classes("text-xs font-semibold text-red-600 mb-1")
                                                        for name in bad_items:
                                                            ui.label(f"• {name}").classes("text-xs text-red-700")
                                                if warn_items:
                                                    with ui.column().classes("flex-1 min-w-48"):
                                                        ui.label("Recurring WARN checkers").classes("text-xs font-semibold text-orange-500 mb-1")
                                                        for name in warn_items:
                                                            ui.label(f"• {name}").classes("text-xs text-orange-600")

                            lines = ["# Cluster Health Summary\n"]
                            for cid2, c in clusters_to_report.items():
                                lines.append(f"## Cluster: {c.get('cluster_name') or cid2[:16]}")
                                lines.append(f"- Cluster ID: `{cid2}`")
                                lines.append(f"- Organization: {c.get('organization') or '—'}")
                                lines.append(f"- Snapshots: {c.get('snapshot_count', 0)}")
                                lines.append(f"- First seen: {c.get('first_seen', '—')[:10]}")
                                lines.append(f"- Last seen: {c.get('last_seen', '—')[:10]}")
                                lines.append(f"- Status: {'Active' if c.get('is_active') else ('Deprecated' if c.get('is_deprecated') else 'Stale')}")
                                lines.append(f"- Nodes (min/max/last): {c.get('node_count_min')}/{c.get('node_count_max')}/{c.get('node_count_last')}")
                                lines.append(f"- Version history: {', '.join(c.get('version_history') or [])}")
                                lines.append(f"- Bucket names seen: {', '.join(c.get('bucket_names_seen') or [])}")
                                lines.append(f"- Total bad issues: {c.get('total_bad', 0)}  |  avg per snapshot: {c.get('avg_bad', 0)}")
                                lines.append(f"- Total warn issues: {c.get('total_warn', 0)}  |  avg per snapshot: {c.get('avg_warn', 0)}")
                                _bi = c.get("top_bad_items") or []
                                _wi = c.get("top_warn_items") or []
                                if _bi:
                                    lines.append(f"- Recurring BAD checkers: {', '.join(_bi)}")
                                if _wi:
                                    lines.append(f"- Recurring WARN checkers: {', '.join(_wi)}")
                                lines.append("")

                            system_prompt = (
                                "You are a Couchbase support engineer analysing cluster health data. "
                                "Given structured snapshot summaries, produce a concise health report. "
                                "STRICT FORMATTING RULES:\n"
                                "- Use `## Cluster Name` (h2) as the top-level heading for each cluster, followed by the cluster ID in backticks on the same line.\n"
                                "- Use `### Section` (h3) for subsections: Overall Health, Upgrade History, Activity Status, Issue Trend, Key Observations, Recommended Actions.\n"
                                "- Use bullet lists (`- item`) for observations and actions.\n"
                                "- Use numbered lists (`1. item`) only for ordered steps.\n"
                                "- Separate clusters with a horizontal rule `---`.\n"
                                "- Use **bold** for labels (e.g. **Status:**, **Nodes:**).\n"
                                "- Do NOT use LaTeX or math notation ($...$). Write plain numbers (e.g. 'approximately 40', '~40').\n"
                                "- Do NOT use HTML tags.\n"
                                "- Keep each cluster report under 250 words."
                            )
                            user_msg = "\n".join(lines)

                            # Use the configured LLM (same provider/model as the Chat tab)
                            try:
                                provider, model, api_key, base_url = _get_llm_config()
                                messages = [
                                    {"role": "system", "content": system_prompt},
                                    {"role": "user",   "content": user_msg},
                                ]
                                ch_llm_status.set_text(f"Asking {provider} ({model})…")
                                response = await run.io_bound(
                                    call_llm, messages, provider, model, api_key, base_url, 4096
                                )
                                ch_llm_out.set_content(response)
                                ch_llm_status.set_text(f"Report generated via {provider} ({model}).")
                            except Exception as exc:
                                ch_llm_status.set_text(f"LLM error: {exc}")
                                try:
                                    ui.notify(str(exc), type="negative")
                                except Exception:
                                    pass
                            finally:
                                btn_ch_llm.set_enabled(True)

            # ══════════════════════════════════════════════════════════════════
            # Customers tab — global directory from tickets + snapshots
            # ══════════════════════════════════════════════════════════════════
            with ui.tab_panel(tab_custs):
                with ui.column().classes("w-full gap-4 p-4"):
                    with ui.card().classes("w-full"):
                        with ui.row().classes("w-full items-center gap-3 flex-wrap"):
                            ui.label("Customer Directory").classes("text-base font-semibold flex-1")
                            btn_dir_refresh = ui.button(
                                "Load from Couchbase", icon="refresh",
                            ).props("color=teal")
                        dir_status = ui.label(
                            "Click 'Load from Couchbase' to aggregate stats from tickets + snapshots."
                        ).classes("text-sm text-gray-500 mt-1")

                    dir_table = ui.table(
                        columns=[
                            {"name": "health_score",    "label": "Health",             "field": "health_score",    "align": "center", "sortable": True},
                            {"name": "organization",    "label": "Customer",           "field": "organization",    "align": "left",   "sortable": True},
                            {"name": "active_clusters", "label": "Active (≤90d)",      "field": "active_clusters", "align": "center", "sortable": True},
                            {"name": "stale_clusters",  "label": "Stale (>90d)",       "field": "stale_clusters",  "align": "center", "sortable": True},
                            {"name": "total_clusters",  "label": "Total Clusters",     "field": "total_clusters",  "align": "center", "sortable": True},
                            {"name": "total_snapshots", "label": "Snapshots",          "field": "total_snapshots", "align": "center", "sortable": True},
                            {"name": "total_tickets",   "label": "Tickets",            "field": "total_tickets",   "align": "center", "sortable": True},
                            {"name": "open_p1",         "label": "Open P1",            "field": "open_p1",         "align": "center", "sortable": True},
                            {"name": "last_scraped_at", "label": "Last Scraped",       "field": "last_scraped_at", "align": "left",   "sortable": True},
                            {"name": "customer_url",    "label": "Supportal URL",      "field": "customer_url",    "align": "left"},
                        ],
                        rows=[],
                        row_key="organization",
                    ).classes("w-full").props("flat bordered dense")

                    dir_table.add_slot("body-row", """
                        <q-tr :props="props" class="cursor-pointer hover:bg-blue-50"
                              @click="$emit('rowclick', props.row)">
                          <q-td key="health_score" :props="props" class="text-center">
                            <q-badge v-if="props.row.health_score != null"
                              :color="props.row.health_score >= 70 ? 'green' : props.row.health_score >= 40 ? 'orange' : 'red'"
                              class="text-white font-bold px-2">
                              {{ props.row.health_score }}
                            </q-badge>
                            <span v-else class="text-gray-300 text-xs">—</span>
                          </q-td>
                          <q-td key="organization" :props="props">
                            <span class="font-medium">{{ props.row.organization }}</span>
                          </q-td>
                          <q-td key="active_clusters" :props="props" class="text-center">
                            <q-badge :color="props.row.active_clusters > 0 ? 'green' : 'grey'">
                              {{ props.row.active_clusters }}
                            </q-badge>
                          </q-td>
                          <q-td key="stale_clusters" :props="props" class="text-center">
                            <q-badge v-if="props.row.stale_clusters > 0" color="orange">
                              {{ props.row.stale_clusters }}
                            </q-badge>
                            <span v-else class="text-gray-400">0</span>
                          </q-td>
                          <q-td key="total_clusters"  :props="props" class="text-center">{{ props.row.total_clusters }}</q-td>
                          <q-td key="total_snapshots" :props="props" class="text-center">{{ props.row.total_snapshots }}</q-td>
                          <q-td key="total_tickets"   :props="props" class="text-center">{{ props.row.total_tickets }}</q-td>
                          <q-td key="open_p1" :props="props" class="text-center">
                            <q-badge v-if="props.row.open_p1 > 0" color="red" class="font-bold">
                              {{ props.row.open_p1 }}
                            </q-badge>
                            <span v-else class="text-gray-400">0</span>
                          </q-td>
                          <q-td key="last_scraped_at" :props="props">
                            {{ props.row.last_scraped_at ? (typeof props.row.last_scraped_at === 'number' ? new Date(props.row.last_scraped_at * 1000) : new Date(props.row.last_scraped_at)).toISOString().substring(0,16).replace('T',' ') : '—' }}
                          </q-td>
                          <q-td key="customer_url" :props="props">
                            <a v-if="props.row.customer_url" :href="props.row.customer_url" target="_blank"
                               class="text-blue-600 text-xs" @click.stop>{{ props.row.customer_url }}</a>
                            <span v-else class="text-gray-400 text-xs">—</span>
                          </q-td>
                        </q-tr>
                    """)

                    async def _dir_load():
                        if not _CB_AVAILABLE:
                            ui.notify("Couchbase SDK not available.", type="warning")
                            return
                        btn_dir_refresh.set_enabled(False)
                        dir_status.set_text("Querying Couchbase…")
                        try:
                            rows = await run.io_bound(
                                query_customer_directory_from_cb,
                                cb_url_input.value.strip(),
                                cb_bucket_input.value.strip(),
                                cb_user_input.value.strip(),
                                cb_pass_input.value,
                                cb_tls_toggle.value,
                                cb_scope_input.value.strip() or "_default",
                                ch_snap_coll.value.strip() or "snapshots",
                                cb_collection_input.value.strip() or "tickets",
                            )
                            dir_table.rows = rows
                            dir_table.update()
                            dir_status.set_text(f"{len(rows)} customer(s) found.")
                            # Update Customers tab badge
                            tab_custs._props["label"] = f"Customers ({len(rows)})"
                            tab_custs.update()
                        except Exception as exc:
                            dir_status.set_text(f"Error: {exc}")
                            ui.notify(str(exc), type="negative")
                        finally:
                            btn_dir_refresh.set_enabled(True)

                    def _dir_pick(e):
                        row = e.args
                        if not isinstance(row, dict):
                            return
                        url = row.get("customer_url") or row.get("organization", "")
                        ch_cust_input.set_value(url)
                        main_tabs.set_value(tab_scoring)
                        scoring_sub_tabs.set_value(sub_cluster)

                    btn_dir_refresh.on_click(lambda: asyncio.ensure_future(_dir_load()))
                    dir_table.on("rowclick", _dir_pick)

            # ══════════════════════════════════════════════════════════════════
            # Assets tab — persistent artifacts (charts, reports, CSV, JSON…)
            # ══════════════════════════════════════════════════════════════════
            with ui.tab_panel(tab_assets):
                with ui.column().classes("w-full gap-4 p-4"):
                    with ui.card().classes("w-full"):
                        # ── Header ────────────────────────────────────────────────────
                        with ui.row().classes("w-full items-center gap-3 flex-wrap"):
                            ui.icon("folder", color="amber").classes("text-2xl")
                            ui.label("Assets").classes("text-base font-semibold flex-1")
                            _assets_status = ui.label("").classes("text-xs text-gray-400")
                            _btn_assets_refresh = ui.button(
                                "Refresh", icon="refresh"
                            ).props("outline size=sm color=primary")

                        ui.label(
                            "Charts and reports generated during chat are auto-saved here. "
                            "Ask the agent to 'save this as an asset' for any text content."
                        ).classes("text-xs text-gray-400 mt-1 mb-2")

                        # ── Filter row ────────────────────────────────────────────────
                        with ui.row().classes("w-full gap-3 mt-1 flex-wrap items-end"):
                            _af_org = ui.input(
                                "Filter by org", placeholder="all"
                            ).props("dense outlined clearable").classes("w-48")
                            _af_type = ui.select(
                                ["all", "chart", "report", "table", "csv", "json", "js", "html"],
                                value="all", label="Type",
                            ).props("dense outlined").classes("w-36")
                            _af_search = ui.input(
                                "Search title", placeholder="keyword"
                            ).props("dense outlined clearable").classes("w-48")

                        # ── Asset list ────────────────────────────────────────────────
                        _assets_area = ui.column().classes("w-full gap-2 mt-3")

                        def _render_asset_card(row: dict) -> None:
                            _aid       = row.get("id", "")
                            _atype     = row.get("asset_type", "report")
                            _atitle    = row.get("title") or row.get("filename") or "Untitled"
                            _aorg      = row.get("org") or ""
                            _ats       = row.get("created_at") or 0
                            _afname    = row.get("filename") or f"{_atitle}.{_atype}"
                            _amime     = row.get("mime_type") or _ASSET_MIME.get(_atype, "text/plain")
                            _aicon     = _ASSET_ICONS.get(_atype, "description")
                            _athumb    = row.get("thumbnail") or ""
                            import datetime as _dtt
                            _ts_str = (
                                _dtt.datetime.fromtimestamp(_ats).strftime("%Y-%m-%d %H:%M")
                                if _ats else "—"
                            )

                            def _cb_args_assets():
                                return (
                                    cb_url_input.value.strip(),
                                    cb_bucket_input.value.strip(),
                                    cb_user_input.value.strip(),
                                    cb_pass_input.value,
                                    cb_tls_toggle.value,
                                    cb_scope_input.value.strip() or "_default",
                                )

                            with ui.card().classes("w-full overflow-hidden"):
                                # ── Thumbnail strip ────────────────────────────────────
                                if _athumb:
                                    if _atype in ("chart", "echart"):
                                        try:
                                            _topt = json.loads(_athumb)
                                            ui.echart(_topt).classes("w-full pointer-events-none").style(
                                                "height:110px;border-bottom:1px solid #e5e7eb"
                                            )
                                        except Exception:
                                            pass
                                    else:
                                        # Text snippet preview
                                        _snip = _athumb[:220].replace("\n", " ").strip()
                                        with ui.element("div").classes(
                                            "w-full px-3 py-2 bg-gray-50 border-b border-gray-200"
                                        ).style("height:56px;overflow:hidden"):
                                            ui.label(_snip).classes(
                                                "font-mono text-xs text-gray-500 break-all"
                                            ).style(
                                                "line-height:1.4;"
                                                "display:-webkit-box;"
                                                "-webkit-line-clamp:3;"
                                                "-webkit-box-orient:vertical;"
                                                "overflow:hidden"
                                            ).tooltip(_snip)

                                # ── Metadata + actions row ─────────────────────────────
                                with ui.row().classes("w-full items-center gap-2 px-3 py-2"):
                                    ui.icon(_aicon, color="blue-grey").classes("text-xl shrink-0")
                                    with ui.column().classes("flex-1 gap-0 min-w-0"):
                                        ui.label(_atitle).classes(
                                            "text-sm font-semibold leading-tight truncate"
                                        )
                                        with ui.row().classes("gap-2 items-center flex-wrap"):
                                            if _aorg:
                                                ui.badge(_aorg[:28]).props(
                                                    "color=teal outline"
                                                ).classes("text-xs")
                                            ui.badge(_atype).props(
                                                "color=blue-grey"
                                            ).classes("text-xs")
                                            ui.label(_ts_str).classes("text-xs text-gray-400")

                                    # ── Per-asset actions ──────────────────────────
                                    with ui.row().classes("gap-1 shrink-0"):
                                        async def _preview(aid=_aid, atype=_atype, atitle=_atitle):
                                            doc = await run.io_bound(
                                                _get_asset_content_from_cb, *_cb_args_assets(), aid
                                            )
                                            content = doc.get("content", "")
                                            with ui.dialog() as _dlg, ui.card().classes(
                                                "w-full max-w-4xl max-h-screen overflow-auto"
                                            ):
                                                with ui.row().classes("w-full items-center sticky top-0 bg-white z-10 pb-2"):
                                                    ui.label(atitle).classes("text-base font-semibold flex-1")
                                                    ui.button(icon="close", on_click=_dlg.close).props("flat round dense")
                                                if atype in ("chart", "echart"):
                                                    try:
                                                        _opt = json.loads(content)
                                                        _opt_c = {k: v for k, v in _opt.items() if not k.startswith("_")}
                                                        ui.echart(_opt_c).classes("w-full").style("height:380px")
                                                    except Exception:
                                                        ui.code(content, language="json").classes("w-full")
                                                elif atype == "report":
                                                    ui.markdown(content).classes("prose prose-sm max-w-none")
                                                elif atype in ("csv", "table"):
                                                    import csv as _cv2, io as _io2
                                                    import html as _hm2
                                                    _rdr = list(_cv2.reader(_io2.StringIO(content)))
                                                    if _rdr:
                                                        _th2 = "".join(
                                                            f'<th class="border border-gray-300 px-2 py-1 bg-gray-100 text-xs font-semibold">{_hm2.escape(str(c))}</th>'
                                                            for c in _rdr[0]
                                                        )
                                                        _tb2 = "".join(
                                                            "<tr>" + "".join(
                                                                f'<td class="border border-gray-300 px-2 py-1 text-xs">{_hm2.escape(str(c))}</td>'
                                                                for c in r
                                                            ) + "</tr>"
                                                            for r in _rdr[1:]
                                                        )
                                                        ui.html(
                                                            f'<div class="overflow-x-auto"><table class="border-collapse">'
                                                            f'<thead><tr>{_th2}</tr></thead><tbody>{_tb2}</tbody></table></div>'
                                                        )
                                                else:
                                                    _lang = {"json": "json", "js": "javascript", "javascript": "javascript", "html": "html"}.get(atype, "text")
                                                    ui.code(content, language=_lang).classes("w-full")
                                            _dlg.open()

                                        ui.button(icon="visibility", on_click=_preview).props(
                                            "flat round dense color=primary"
                                        ).tooltip("Preview")

                                        async def _download(aid=_aid, afname=_afname, amime=_amime):
                                            doc = await run.io_bound(
                                                _get_asset_content_from_cb, *_cb_args_assets(), aid
                                            )
                                            ui.download(
                                                doc.get("content", "").encode(), afname, amime
                                            )

                                        ui.button(icon="download", on_click=_download).props(
                                            "flat round dense color=green"
                                        ).tooltip("Download")

                                        async def _print_asset(aid=_aid, atitle=_atitle, atype=_atype):
                                            doc = await run.io_bound(
                                                _get_asset_content_from_cb, *_cb_args_assets(), aid
                                            )
                                            content = doc.get("content", "")
                                            _esc_content = json.dumps(content)
                                            _esc_title   = json.dumps(atitle)
                                            await ui.run_javascript(f"""
(function() {{
  var w = window.open('', '_blank');
  var c = {_esc_content};
  var t = {_esc_title};
  w.document.write('<html><head><title>' + t + '</title>');
  w.document.write('<style>body{{font-family:system-ui,sans-serif;padding:2rem;max-width:960px;margin:auto;line-height:1.6}}');
  w.document.write('pre{{background:#f5f5f5;padding:1rem;border-radius:4px;overflow-x:auto;font-size:12px}}');
  w.document.write('table{{border-collapse:collapse;width:100%}}td,th{{border:1px solid #ccc;padding:4px 8px;font-size:12px}}');
  w.document.write('h1,h2,h3{{margin-top:1.5rem}}</style></head><body>');
  w.document.write('<h2>' + t + '</h2><pre>' + c.replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;') + '</pre>');
  w.document.write('</body></html>');
  w.document.close();
  setTimeout(function(){{w.print();}}, 400);
}})();
""")

                                        ui.button(icon="print", on_click=_print_asset).props(
                                            "flat round dense color=blue-grey"
                                        ).tooltip("Print / Save as PDF")

                                        async def _delete(aid=_aid):
                                            ok = await run.io_bound(
                                                _delete_asset_from_cb, *_cb_args_assets(), aid
                                            )
                                            if ok:
                                                ui.notify("Asset deleted.", type="positive")
                                                await _load_assets()
                                            else:
                                                ui.notify("Delete failed.", type="warning")

                                        ui.button(icon="delete_outline", on_click=_delete).props(
                                            "flat round dense color=red"
                                        ).tooltip("Delete")

                        async def _load_assets():
                            _org_f  = (_af_org.value or "").strip()
                            _type_f = _af_type.value if _af_type.value != "all" else ""
                            _srch   = (_af_search.value or "").strip().lower()
                            _assets_status.set_text("Loading…")
                            _assets_area.clear()
                            if not (_CB_AVAILABLE and cb_url_input.value.strip()):
                                _assets_status.set_text("Couchbase not configured.")
                                return
                            try:
                                rows = await run.io_bound(
                                    _list_assets_from_cb,
                                    cb_url_input.value.strip(),
                                    cb_bucket_input.value.strip(),
                                    cb_user_input.value.strip(),
                                    cb_pass_input.value,
                                    cb_tls_toggle.value,
                                    cb_scope_input.value.strip() or "_default",
                                    _org_f, _type_f,
                                )
                                if _srch:
                                    rows = [r for r in rows if _srch in (r.get("title") or "").lower()]
                                _assets_status.set_text(f"{len(rows)} asset(s)")
                                with _assets_area:
                                    if not rows:
                                        ui.label(
                                            "No assets yet. Charts and reports from chat are saved here automatically."
                                        ).classes("text-sm text-gray-400 mt-4 text-center w-full")
                                    else:
                                        for _row in rows:
                                            _render_asset_card(_row)
                            except Exception as _exc:
                                _assets_status.set_text(f"Error: {_classify_agent_error(_exc)}")

                        _btn_assets_refresh.on_click(lambda: asyncio.ensure_future(_load_assets()))
                        _af_org.on("change", lambda: asyncio.ensure_future(_load_assets()))
                        _af_type.on("update:model-value", lambda: asyncio.ensure_future(_load_assets()))
                        _af_search.on("change", lambda: asyncio.ensure_future(_load_assets()))

    # ── Settings profile logic ───────────────────────────────────────────────
    def _tab_name(tab_val) -> str:
        """Extract the string label from a NiceGUI Tab object (or pass through if already a str)."""
        if isinstance(tab_val, str):
            return tab_val
        if hasattr(tab_val, '_props'):
            return tab_val._props.get('name') or tab_val._props.get('label') or str(tab_val)
        return str(tab_val)

    def _collect_profile() -> dict:
        """Gather every configurable field into a dict."""
        return {
            # Auth
            "cookie":        cookie_input.value,
            # Couchbase
            "cb_url":          cb_url_input.value,
            "cb_bucket":       cb_bucket_input.value,
            "cb_user":         cb_user_input.value,
            "cb_pass":         cb_pass_input.value,
            "cb_tls":          cb_tls_toggle.value,
            "cb_scope":        cb_scope_input.value,
            "cb_collection":   cb_collection_input.value,
            "ch_snap_coll":    ch_snap_coll.value,
            "cb_summary_coll": cb_summary_coll.value,
            # Embedding
            "emb_provider":       ai_emb_provider.value,
            "emb_ollama_url":     emb_ollama_url_input.value,
            "emb_ollama_model":   emb_ollama_model_input.value or "",
            "emb_ollama_dims":    emb_dims_input.value,
            "emb_ollama_num_ctx": emb_num_ctx_input.value,
            "emb_lms_url":        emb_lms_url_input.value,
            "emb_lms_model":      emb_lms_model_input.value or "",
            "emb_lms_dims":       emb_lms_dims_input.value,
            "emb_gemini_key":     emb_gemini_key_input.value,
            "emb_gemini_model":   emb_gemini_model_input.value or "",
            "emb_gemini_dims":    emb_gemini_dims_input.value,
            "emb_mlx_model":      emb_mlx_model_input.value,
            "emb_mlx_dims":       emb_mlx_dims_input.value,
            "emb_openai_key":     emb_openai_key_input.value,
            "emb_openai_model":   emb_openai_model_input.value or "",
            "emb_openai_dims":    emb_openai_dims_input.value,
            "embed_parallel":     embed_parallel_input.value,
            "enrich_workers":     enrich_workers_input.value,
            # LLM (chat/scoring)
            "llm_provider":       ai_llm_provider.value,
            "claude_key":         claude_key_input.value,
            "claude_model":       claude_model_input.value or "",
            "gemini_llm_key":     gemini_key_input.value,
            "gemini_llm_model":   gemini_model_input.value or "",
            "ollama_chat_model":  ollama_chat_model_input.value or "",
            "lms_model":          lms_model_input.value or "",
            "openai_llm_model":   openai_llm_model_input.value or "",
            # Scoring
            "score_batch":        score_batch_input.value,
            "score_parallel":     score_parallel_input.value,
            "score_ctx":          score_ctx_input.value,
            "score_no_think":     score_no_think_toggle.value,
            "score_autosave":     score_autosave_toggle.value,
            # Pipeline
            "pipeline_save":      pipeline_save_toggle.value,
            "pipeline_embed":     pipeline_embed_toggle.value,
            "pipeline_score":     pipeline_score_toggle.value,
            "pipeline_enrich":    pipeline_enrich_toggle.value,
            "pipeline_validate":   pipeline_validate_toggle.value,
            "pipeline_reconcile":  pipeline_reconcile_toggle.value,
            "snap_auto_save_cb":   ch_auto_save_cb.value,
            # CH scrape settings
            "ch_max_pages":       ch_max_pages.value,
            "ch_workers":         ch_workers.value,
            "ch_max_snapshots":   ch_max_snapshots.value,
            "ch_analytics_limit": ch_analytics_limit.value,
            # Chat settings (retired — Corax handles chat; keys kept for profile compat)
            # Chat cache / memory
            "cache_collection":   cache_collection_input.value,
            "embed_cache_ttl":    embed_cache_ttl.value,
            "search_cache_ttl":   search_cache_ttl.value,
            "store_memory":       store_memory_toggle.value,
        }

    def _apply_profile(p: dict) -> None:
        """Populate all UI fields from a saved profile dict."""
        def _set(widget, key, default=None):
            val = p.get(key, default)
            if val is not None:
                widget.set_value(val)

        _set(cookie_input,        "cookie")
        _set(cb_url_input,        "cb_url")
        _set(cb_bucket_input,     "cb_bucket")
        _set(cb_user_input,       "cb_user")
        _set(cb_pass_input,       "cb_pass")
        _set(cb_tls_toggle,       "cb_tls")
        _set(cb_scope_input,      "cb_scope")
        _set(cb_collection_input, "cb_collection")
        _set(ch_snap_coll,        "ch_snap_coll")
        _set(cb_summary_coll,     "cb_summary_coll")

        _set(emb_ollama_url_input,   "emb_ollama_url")
        if p.get("emb_ollama_model"):
            emb_ollama_model_input.options = [p["emb_ollama_model"]]
            emb_ollama_model_input.set_value(p["emb_ollama_model"])
        _set(emb_dims_input,         "emb_ollama_dims")
        _set(emb_num_ctx_input,      "emb_ollama_num_ctx")
        _set(emb_lms_url_input,      "emb_lms_url")
        if p.get("emb_lms_model"):
            emb_lms_model_input.options = [p["emb_lms_model"]]
            emb_lms_model_input.set_value(p["emb_lms_model"])
        _set(emb_lms_dims_input,     "emb_lms_dims")
        _set(emb_gemini_key_input,   "emb_gemini_key")
        if p.get("emb_gemini_model"):
            emb_gemini_model_input.set_value(p["emb_gemini_model"])
        _set(emb_gemini_dims_input,  "emb_gemini_dims")
        _set(emb_mlx_model_input,    "emb_mlx_model")
        _set(emb_mlx_dims_input,     "emb_mlx_dims")
        _set(emb_openai_key_input,   "emb_openai_key")
        if p.get("emb_openai_model"):
            emb_openai_model_input.set_value(p["emb_openai_model"])
        _set(emb_openai_dims_input,  "emb_openai_dims")
        _set(embed_parallel_input,   "embed_parallel")
        _set(enrich_workers_input,   "enrich_workers")
        if p.get("emb_provider"):
            ai_emb_provider.set_value(p["emb_provider"])

        _set(claude_key_input,   "claude_key")
        if p.get("claude_model"):
            claude_model_input.set_value(p["claude_model"])
        _set(gemini_key_input,   "gemini_llm_key")
        if p.get("gemini_llm_model"):
            gemini_model_input.set_value(p["gemini_llm_model"])
        if p.get("ollama_chat_model"):
            ollama_chat_model_input.options = [p["ollama_chat_model"]]
            ollama_chat_model_input.set_value(p["ollama_chat_model"])
        if p.get("lms_model"):
            lms_model_input.options = [p["lms_model"]]
            lms_model_input.set_value(p["lms_model"])
        if p.get("openai_llm_model"):
            openai_llm_model_input.set_value(p["openai_llm_model"])
        if p.get("llm_provider"):
            ai_llm_provider.set_value(p["llm_provider"])

        _set(score_batch_input,       "score_batch")
        _set(score_parallel_input,    "score_parallel")
        _set(score_ctx_input,         "score_ctx")
        # If no saved value, auto-enable based on the currently saved model name
        if "score_no_think" in p:
            _set(score_no_think_toggle, "score_no_think")
        else:
            _saved_llm = p.get("ollama_chat_model") or p.get("lms_model") or ""
            if _model_has_thinking_by_name(_saved_llm):
                score_no_think_toggle.set_value(True)
        _set(score_autosave_toggle,   "score_autosave")
        if _CB_AVAILABLE:
            score_autosave_toggle.set_value(True)
        _set(pipeline_save_toggle,     "pipeline_save")
        _set(pipeline_embed_toggle,   "pipeline_embed")
        _set(pipeline_score_toggle,   "pipeline_score")
        _set(pipeline_enrich_toggle,  "pipeline_enrich")
        _set(pipeline_validate_toggle, "pipeline_validate")
        _set(pipeline_reconcile_toggle,"pipeline_reconcile")
        _set(ch_auto_save_cb,          "snap_auto_save_cb")
        _set(ch_max_pages,            "ch_max_pages")
        _set(ch_workers,              "ch_workers")
        _set(ch_max_snapshots,        "ch_max_snapshots")
        _set(ch_analytics_limit,      "ch_analytics_limit")
        # chat_mode / top_k / batch_size / compact_context / deep_reason retired
        _set(cache_collection_input,  "cache_collection")
        _set(embed_cache_ttl,        "embed_cache_ttl")
        _set(search_cache_ttl,       "search_cache_ttl")
        _set(store_memory_toggle,    "store_memory")

    def _cb_args():
        """Return CB connection args from the current UI inputs, or None if unconfigured."""
        if not (_CB_AVAILABLE and cb_url_input.value.strip()):
            return None
        return (
            cb_url_input.value.strip(), cb_bucket_input.value.strip(),
            cb_user_input.value.strip(), cb_pass_input.value,
            cb_tls_toggle.value,
            cb_scope_input.value.strip() or "_default",
            cb_collection_input.value.strip() or "tickets",
        )

    async def _refresh_cluster_map(status_label=None) -> str:
        """Query CB snapshots + tickets to populate dynamic cluster↔app alias maps."""
        args = _cb_args()
        if not args:
            msg = "CB not configured — aliases unchanged"
            if status_label:
                status_label.set_text(msg)
            return msg
        cb_url, bucket, cb_user, cb_pass, cb_tls, scope, _col = args
        snap_col = ch_snap_coll.value.strip() or "snapshots"
        if status_label:
            status_label.set_text("Querying CB for cluster↔app mappings…")
        n_snaps, n_tickets = await run.io_bound(
            _load_cluster_app_map,
            cb_url, bucket, cb_user, cb_pass, cb_tls, scope,
            snap_col, _col,
        )
        total = len(_cluster_app_dynamic)
        msg = f"Aliases refreshed — {total} cluster→app mappings ({n_snaps} from snaps, {n_tickets} from tickets)"
        if status_label:
            status_label.set_text(msg)
        return msg

    async def _save_profile():
        profiles   = _load_settings_file()
        name       = (profile_name_input.value or profile_select.value or "default").strip()
        profiles[name]     = _collect_profile()
        profiles["__last__"] = name
        _save_settings_file(profiles)
        names = sorted(k for k in profiles if k != "__last__")
        profile_select.options = names
        profile_select.set_value(name)
        profile_name_input.set_value("")
        args = _cb_args()
        if args:
            profile_status.set_text(f"Saving \"{name}\" to Couchbase…")
            try:
                await run.io_bound(_cb_save_settings, *args, profiles)
                profile_status.set_text(f"Saved \"{name}\" locally + Couchbase")
            except Exception as exc:
                profile_status.set_text(f"Saved \"{name}\" locally (CB error: {exc})")
        else:
            profile_status.set_text(f"Saved \"{name}\"")

    async def _load_profile():
        name = profile_select.value or "default"
        args = _cb_args()
        if args:
            profile_status.set_text(f"Loading \"{name}\" from Couchbase…")
            try:
                cb_profiles = await run.io_bound(_cb_load_settings, *args)
                if name in cb_profiles:
                    _apply_profile(cb_profiles[name])
                    # Sync CB snapshot back to local (merge, don't overwrite special keys)
                    local = _load_settings_file()
                    local.update({k: v for k, v in cb_profiles.items() if not k.startswith("__")})
                    local["__last__"] = name
                    _save_settings_file(local)
                    profile_status.set_text(f"Loaded \"{name}\" from Couchbase")
                    asyncio.ensure_future(_refresh_cluster_map())
                    return
                profile_status.set_text(f"\"{name}\" not in Couchbase — checking local…")
            except Exception as exc:
                profile_status.set_text(f"CB error: {exc} — falling back to local…")
        profiles = _load_settings_file()
        p = profiles.get(name)
        if not p:
            profile_status.set_text(f"No profile \"{name}\" found.")
            return
        _apply_profile(p)
        profiles["__last__"] = name
        _save_settings_file(profiles)
        profile_status.set_text(f"Loaded \"{name}\"")
        asyncio.ensure_future(_refresh_cluster_map())

    async def _delete_profile():
        profiles = _load_settings_file()
        name     = profile_select.value or ""
        if name and name in profiles:
            del profiles[name]
            if profiles.get("__last__") == name:
                profiles.pop("__last__", None)
            _save_settings_file(profiles)
            args = _cb_args()
            if args:
                try:
                    await run.io_bound(_cb_save_settings, *args, profiles)
                except Exception:
                    pass
            names = sorted(k for k in profiles if k != "__last__")
            profile_select.options = names or ["default"]
            profile_select.set_value(names[0] if names else "default")
            profile_status.set_text(f"Deleted \"{name}\"")
        else:
            profile_status.set_text(f"Profile \"{name}\" not found.")

    # Auto-load the last-used profile on page open, then refresh cluster→app aliases
    _last = _load_settings_file().get("__last__")
    if _last:
        _saved = _load_settings_file().get(_last, {})
        if _saved:
            _apply_profile(_saved)
            profile_status.set_text(f"Auto-loaded \"{_last}\"")
            # Fire alias refresh in background so it doesn't block page paint
            asyncio.ensure_future(_refresh_cluster_map())












# If any of these appear at the start or as a dominant pattern the rewriter fires.




def _job_fail(job: dict, stage: str, err, ticket_id: str | None = None) -> None:
    """Record a pipeline failure on the job: bump the error counter AND capture
    detail into job["error_log"] (capped) so failures become durable knowledge
    instead of an opaque count. Never raises."""
    try:
        from supportal.cb_helpers import classify_error
        job["errors"] = job.get("errors", 0) + 1
        log = job.setdefault("error_log", [])
        if len(log) < 200:
            entry = {"stage": stage, "at": time.time(), **classify_error(err)}
            if ticket_id:
                entry["ticket_id"] = str(ticket_id)
            log.append(entry)
    except Exception:
        pass


def _persist_failure_log(
    job: dict,
    cb_url: str, bucket: str, username: str, password: str,
    use_tls: bool, scope: str,
) -> None:
    """If the job had any errors, write a PERMANENT failure-knowledge doc to the
    `markers` collection (unlike scrape_job:: docs, which expire after 48h).
    Key: failurelog::<job_id>. Never raises."""
    if not job.get("errors") and not job.get("error_log"):
        return
    if not _CB_AVAILABLE or not cb_url:
        return
    try:
        from couchbase.cluster import Cluster as _Cl
        from couchbase.options import ClusterOptions as _CO
        from couchbase.auth import PasswordAuthenticator as _PA
        _c = _Cl(_cb_conn_str(cb_url, use_tls), _CO(_PA(username, password)))
        _c.wait_until_ready(timedelta(seconds=5))
        try:
            cm = _c.bucket(bucket).collections()
            existing = {s.name: {cc.name for cc in s.collections} for s in cm.get_all_scopes()}
            if "markers" not in existing.get(scope, set()):
                from couchbase.management.collections import CollectionSpec
                cm.create_collection(CollectionSpec("markers", scope_name=scope))
        except Exception:
            pass
        doc = {
            "type":         "failure_log",
            "job_id":       job.get("job_id"),
            "organization": job.get("org"),
            "mode":         job.get("mode"),
            "finished_at":  job.get("finished_at"),
            "total":        job.get("total"),
            "errors":       job.get("errors", 0),
            "error_log":    (job.get("error_log") or [])[:200],
            "last_message": job.get("last_message", ""),
        }
        _c.bucket(bucket).scope(scope).collection("markers").upsert(
            f"failurelog::{job.get('job_id')}", doc
        )
        _c.close()
    except Exception:
        pass


_JOBRUN_DEFAULT_TTL_S = 45 * 60   # reap deadline when total is unknown
_JOBRUN_PER_TICKET_S  = 6         # scrape+embed+score allowance per ticket
_JOBRUN_MARGIN_S      = 15 * 60


def _jobrun_deadline(started_at: float, total: int | None) -> float:
    """Computed conclude-by time: past this, a still-'started' jobrun is
    presumed lost (process died / silently dropped) and gets reaped."""
    if total:
        return started_at + max(_JOBRUN_DEFAULT_TTL_S,
                                total * _JOBRUN_PER_TICKET_S + _JOBRUN_MARGIN_S)
    return started_at + _JOBRUN_DEFAULT_TTL_S


def _persist_job_run(
    job: dict,
    cb_url: str, bucket: str, username: str, password: str,
    use_tls: bool, scope: str, collection: str = "",
    conclude: bool = False,
) -> None:
    """PERMANENT jobrun::<job_id> lifecycle record in `markers`.

    Written once at job start (status 'started' + computed conclude-by
    deadline) and again at conclusion. A jobrun still 'started' past its
    deadline is evidence of a silently dropped job — the start record is the
    only witness, since scrape_job:: docs expire and in-memory state dies
    with the process. On conclusion of an org-scoped job, also re-verifies
    freshness so drift → rescrape → VERIFIED fresh, not assumed. Never raises.
    """
    if not _CB_AVAILABLE or not cb_url:
        return
    try:
        from couchbase.cluster import Cluster as _Cl
        from couchbase.options import ClusterOptions as _CO
        from couchbase.auth import PasswordAuthenticator as _PA
        _c = _Cl(_cb_conn_str(cb_url, use_tls), _CO(_PA(username, password)))
        _c.wait_until_ready(timedelta(seconds=5))
        try:
            from supportal.cb_helpers import _ensure_collection
            _ensure_collection(_c, bucket, scope, "markers")
        except Exception:
            pass
        col = _c.bucket(bucket).scope(scope).collection("markers")
        key = f"jobrun::{job.get('job_id')}"
        doc = {
            "type":         "job_run",
            "job_id":       job.get("job_id"),
            "organization": job.get("org"),
            "mode":         job.get("mode"),
            "status":       job.get("status") if conclude else "started",
            "started_at":   job.get("started_at"),
            "expected_deadline": _jobrun_deadline(
                job.get("started_at") or time.time(), job.get("total")),
        }
        if conclude:
            doc.update({
                "finished_at":  job.get("finished_at"),
                "total":        job.get("total"),
                "processed":    job.get("processed", 0),
                "saved":        job.get("saved", 0),
                "embedded":     job.get("embedded", 0),
                "scored":       job.get("scored", 0),
                "errors":       job.get("errors", 0),
                "last_message": job.get("last_message", ""),
            })
        col.upsert(key, doc)

        # Post-job freshness verification — only on a real conclusion of an
        # org-scoped job that saved something (skip fatal-at-startup noise).
        if conclude and job.get("org") and job.get("status") == "done" and collection:
            try:
                from supportal.cb_helpers import compute_and_mark_freshness
                fresh = compute_and_mark_freshness(
                    job["org"], cb_url, bucket, username, password,
                    use_tls, scope, collection,
                    verified_by=f"jobrun::{job.get('job_id')}",
                )
                doc["freshness_after"] = {
                    "status": fresh.get("status"),
                    "missing_count": fresh.get("missing_count"),
                    "checked_at": fresh.get("checked_at"),
                }
                col.upsert(key, doc)
            except Exception as exc:
                doc["freshness_after"] = {"status": "unverified", "error": str(exc)[:160]}
                col.upsert(key, doc)
            # Contacts enrichment — refresh the org-scoped contacts:: marker
            # from live zdorg so AE/TSE/renewal flags stay current alongside
            # the tickets. Org-scoped only; never denormalized per-ticket.
            try:
                from supportal.cb_helpers import refresh_org_contacts
                refresh_org_contacts(
                    job["org"], cb_url, bucket, username, password,
                    use_tls, scope,
                    refreshed_by=f"jobrun::{job.get('job_id')}",
                )
            except Exception:
                pass  # best-effort — never fail a completed job on enrichment
        _c.close()
    except Exception:
        pass


def _upsert_job_doc(job: dict, col) -> None:
    """Write job state using an already-open CB collection handle. Never raises."""
    if col is None:
        return
    try:
        from couchbase.options import UpsertOptions as _UO
        col.upsert(
            f"scrape_job::{job['job_id']}",
            {**job, "type": "scrape_job"},
            _UO(expiry=timedelta(hours=48)),
        )
    except Exception:
        pass


def _persist_job_state(
    job: dict,
    cb_url: str, bucket: str, username: str, password: str,
    use_tls: bool, scope: str, collection: str,
) -> None:
    """Write job state to CB by opening a short-lived connection. Use at phase transitions."""
    if not _CB_AVAILABLE or not cb_url:
        return
    try:
        from couchbase.cluster import Cluster as _Cl
        from couchbase.options import ClusterOptions as _CO
        from couchbase.auth import PasswordAuthenticator as _PA
        from couchbase.options import UpsertOptions as _UO
        _conn = _cb_conn_str(cb_url, use_tls)
        _c    = _Cl(_conn, _CO(_PA(username, password)))
        _c.wait_until_ready(timedelta(seconds=5))
        _c.bucket(bucket).scope(scope).collection(collection).upsert(
            f"scrape_job::{job['job_id']}",
            {**job, "type": "scrape_job"},
            _UO(expiry=timedelta(hours=48)),
        )
        _c.close()
    except Exception:
        pass


def _make_scrape_job(org: str, mode: str) -> dict:
    """Create a new scrape job record, register it, and return it."""
    import secrets
    job_id = secrets.token_hex(3)  # 6-char hex, e.g. "a3f9c1"
    job: dict = {
        "job_id":       job_id,
        "org":          org,
        "mode":         mode,          # "scrape" | "rescrape"
        "phase":        "queued",      # queued → scraping → saving → embedding → scoring → done
        "status":       "running",     # running | done | error | cancelled | interrupted
        "total":        None,          # total tickets to process (set once known)
        "processed":    0,             # tickets processed so far
        "saved":        0,
        "embedded":     0,
        "scored":       0,
        "errors":       0,
        "last_message": "Queued…",
        "started_at":   time.time(),
        "finished_at":  None,
        "heartbeat_at": time.time(),   # updated every ticket; stale = process died
    }
    _SCRAPE_JOBS[job_id] = job
    _JOB_CANCEL_EVENTS[job_id] = threading.Event()
    # Trim to MAX
    if len(_SCRAPE_JOBS) > _MAX_SCRAPE_JOBS:
        oldest = next(iter(_SCRAPE_JOBS))
        del _SCRAPE_JOBS[oldest]
        _JOB_CANCEL_EVENTS.pop(oldest, None)
    return job


def _run_scrape_job_bg(
    job: dict,
    org: str,
    cookie: str,
    max_tickets: int,
    cb_params: dict,
    emb_params: dict,
    score_params: dict,
) -> None:
    """Background worker for scrape_customer_tickets. Runs full scrape→save→embed→score pipeline."""
    cb_url    = cb_params["url"]
    bucket    = cb_params["bucket"]
    username  = cb_params["username"]
    password  = cb_params["password"]
    use_tls   = cb_params["use_tls"]
    scope     = cb_params["scope"]
    collection = cb_params["collection"]

    def _set_op(msg: str, pct: float, done: bool = False):
        job["last_message"] = msg
        _OP_STATUS["op"]       = "scrape"
        _OP_STATUS["status"]   = f"[{job['job_id']}] {msg}"
        _OP_STATUS["progress"] = pct
        _OP_STATUS["done"]     = done

    try:
        _persist_job_run(job, cb_url, bucket, username, password, use_tls, scope)
        # ── Phase 1: scrape ─────────────────────────────────────────────────
        job["phase"] = "scraping"
        _set_op(f"Scraping tickets for '{org}'…", 0.0)

        def _scrape_prog(msg: str, pct: float):
            job["last_message"] = msg
            _OP_STATUS["status"]   = f"[{job['job_id']}] {msg}"
            _OP_STATUS["progress"] = pct * 0.4   # scrape = 0–40%
            m = re.search(r"(\d+)\s*/\s*(\d+)", msg)
            if m:
                job["processed"] = int(m.group(1))
                job["total"]     = int(m.group(2))

        scraped = scrape_with_cookie(org, cookie, progress_cb=_scrape_prog, max_tickets=max_tickets)
        job["total"]     = len(scraped)
        job["processed"] = len(scraped)
        _set_op(f"Scraped {len(scraped)} tickets — saving to Couchbase…", 0.40)

        # ── Phase 2: save ───────────────────────────────────────────────────
        job["phase"] = "saving"
        _saved = 0
        now_epoch = int(time.time())
        try:
            from couchbase.cluster import Cluster as _Cl
            from couchbase.options import ClusterOptions as _CO
            from couchbase.auth import PasswordAuthenticator as _PA
            _conn = _cb_conn_str(cb_url, use_tls)
            _cluster = _Cl(_conn, _CO(_PA(username, password)))
            _cluster.wait_until_ready(timedelta(seconds=15))
            _col = _cluster.bucket(bucket).scope(scope).collection(collection)
            _upsert_job_doc(job, _col)   # persist "saving" phase start
            for t in scraped:
                tid = t.get("ticket_id")
                if not tid:
                    continue
                try:
                    _col.upsert(f"ticket::{tid}", {**t, "type": "ticket", "last_scraped_at": now_epoch})
                    _saved += 1
                except Exception as _texc:
                    _job_fail(job, "save", _texc, ticket_id=tid)
            _upsert_job_doc(job, _col)   # persist end-of-save state
            _cluster.close()
        except Exception as exc:
            _job_fail(job, "save", f"CB save failed: {exc}")
            job["last_message"] = f"CB save failed: {exc}"
        job["saved"] = _saved
        _set_op(f"Saved {_saved} tickets — embedding…", 0.55)

        # ── Phase 3: embed ──────────────────────────────────────────────────
        emb_p = emb_params.get("provider", "").lower().strip()
        emb_m = emb_params.get("model", "")
        emb_k = emb_params.get("api_key", "")
        emb_u = emb_params.get("base_url", "")
        emb_d = int(emb_params.get("dims") or 0)
        emb_workers = int(emb_params.get("max_workers") or emb_params.get("embed_parallel") or 1)
        if emb_p and emb_m and emb_d and _saved > 0:
            job["phase"] = "embedding"
            if emb_p == "lmstudio":
                _lms_base = (emb_u or "http://localhost:1234").rstrip("/v1").rstrip("/")
                _lms_emb_id = lmstudio_ensure_model_loaded(_lms_base, emb_m, timeout_s=45, model_type="embeddings")
                if _lms_emb_id:
                    emb_m = _lms_emb_id
            try:
                _saved_data = [{**t, "type": "ticket", "last_scraped_at": now_epoch} for t in scraped]

                def _emb_prog(msg: str, pct: float):
                    job["last_message"] = msg
                    _OP_STATUS["status"]   = f"[{job['job_id']}] {msg}"
                    _OP_STATUS["progress"] = 0.55 + pct * 0.30  # embed = 55–85%

                _emb_errs: list = job.setdefault("error_log", [])
                _done_emb, _errs_emb = embed_all_tickets(
                    _saved_data, cb_url, bucket, username, password,
                    use_tls, scope, collection,
                    emb_p, emb_m, emb_k, emb_u, emb_d,
                    _emb_prog,
                    max_workers=emb_workers,
                    error_sink=_emb_errs,
                )
                job["embedded"] = _done_emb
                job["errors"]  += _errs_emb
            except Exception as exc:
                job["last_message"] = f"Embedding failed: {exc}"
                _job_fail(job, "embed", exc)
        _set_op(f"Embedded {job['embedded']} tickets — scoring…", 0.85)

        # ── Phase 4: score ──────────────────────────────────────────────────
        s_prov = score_params.get("provider", "")
        s_mod  = score_params.get("model", "")
        s_key  = score_params.get("api_key", "")
        s_url  = score_params.get("base_url", "")
        if s_prov and s_mod and _saved > 0:
            job["phase"] = "scoring"
            try:
                _score_data = [{**t, "type": "ticket"} for t in scraped[:_saved]]
                _scores = score_tickets_batch(
                    _score_data[:10],
                    s_prov, s_mod, s_key, s_url,
                    cb_url, bucket, username, password, use_tls, scope, collection,
                    save_to_cb=True,
                )
                job["scored"] = len(_scores)
            except Exception as exc:
                job["last_message"] = f"Scoring failed: {exc}"
                _job_fail(job, "score", exc)

        # ── Done ────────────────────────────────────────────────────────────
        job["status"]       = "done"
        job["phase"]        = None
        job["finished_at"]  = time.time()
        summary = (
            f"Done — {job['saved']} scraped, {job['embedded']} embedded, "
            f"{job['scored']} scored"
            + (f", {job['errors']} errors" if job["errors"] else "")
        )
        job["last_message"] = summary
        _set_op(summary, 1.0, done=True)
        _persist_job_state(job, cb_url, bucket, username, password, use_tls, scope, collection)
        _persist_failure_log(job, cb_url, bucket, username, password, use_tls, scope)
        _persist_job_run(job, cb_url, bucket, username, password, use_tls, scope, collection, conclude=True)

    except Exception as exc:
        job["status"]      = "error"
        job["phase"]       = None
        job["finished_at"] = time.time()
        job["last_message"] = f"Fatal error: {exc}"
        _job_fail(job, "fatal", exc)
        _OP_STATUS.update({"op": None, "status": str(exc), "progress": 0.0, "done": True})
        _persist_job_state(job, cb_url, bucket, username, password, use_tls, scope, collection)
        _persist_failure_log(job, cb_url, bucket, username, password, use_tls, scope)
        _persist_job_run(job, cb_url, bucket, username, password, use_tls, scope, collection, conclude=True)


def _run_rescrape_job_bg(
    job: dict,
    org: str,
    to_scrape: list[dict],
    cookie: str,
    cb_params: dict,
    emb_params: dict | None = None,
) -> None:
    """Background worker for rescrape_customer_tickets. Refreshes individual ticket docs."""
    from couchbase.options import GetOptions, UpsertOptions  # type: ignore

    cb_url     = cb_params["url"]
    bucket     = cb_params["bucket"]
    username   = cb_params["username"]
    password   = cb_params["password"]
    use_tls    = cb_params["use_tls"]
    scope      = cb_params["scope"]
    collection = cb_params["collection"]
    _cb_op_timeout = timedelta(seconds=5)

    total = len(to_scrape)
    job["total"] = total

    def _set_op(msg: str, pct: float, done: bool = False):
        job["last_message"] = msg
        _OP_STATUS["op"]       = "scrape"
        _OP_STATUS["status"]   = f"[{job['job_id']}] {msg}"
        _OP_STATUS["progress"] = pct
        _OP_STATUS["done"]     = done

    try:
        _persist_job_run(job, cb_url, bucket, username, password, use_tls, scope)
        job["phase"] = "scraping"
        _set_op(f"Rescraping 0/{total} tickets for '{org}'…", 0.0)

        from couchbase.cluster import Cluster as _Cl
        from couchbase.options import ClusterOptions as _CO
        from couchbase.auth import PasswordAuthenticator as _PA
        conn_str  = _cb_conn_str(cb_url, use_tls)
        _bcluster = _Cl(conn_str, _CO(_PA(username, password)))
        _bcluster.wait_until_ready(timedelta(seconds=10))
        _bcol = _bcluster.bucket(bucket).scope(scope).collection(collection)
        _upsert_job_doc(job, _bcol)   # persist initial "scraping" state

        # Create session once — reused across all ticket fetches
        _sess = _make_api_session(cookie)

        ok = skipped = errors = 0
        job["enriched"] = 0
        refreshed_tickets: list[dict] = []
        _cancel_ev = _JOB_CANCEL_EVENTS.get(job["job_id"])
        for i, t in enumerate(to_scrape, 1):
            # Check for cancellation request before each ticket
            if _cancel_ev and _cancel_ev.is_set():
                job["status"]       = "cancelled"
                job["phase"]        = None
                job["finished_at"]  = time.time()
                job["last_message"] = (
                    f"Cancelled at ticket {i}/{total}. "
                    f"{ok} saved so far. To resume, rescrape with stale_hours=1 — "
                    f"the {ok} already-refreshed tickets will be skipped automatically."
                )
                _upsert_job_doc(job, _bcol)
                _persist_job_run(job, cb_url, bucket, username, password, use_tls, scope, collection, conclude=True)
                return

            tid = str(t.get("ticket_id") or "").strip()
            job["heartbeat_at"] = time.time()   # alive signal for zombie detection
            if not tid:
                skipped += 1
                job["processed"] = i
                continue
            try:
                fresh = fetch_ticket_api(tid, _sess)
                if not fresh or not fresh.get("ticket_id"):
                    skipped += 1
                    job["processed"] = i
                    continue
                fresh["last_scraped_at"] = int(time.time())
                fresh["type"]            = "ticket"
                fresh["cb_version"]      = extract_ticket_version(fresh)
                fresh["feature_area"]    = classify_ticket_feature(fresh)
                fresh["ticket_origin"]   = classify_ticket_origin(fresh)
                doc_key = f"ticket::{tid}"
                try:
                    existing = _bcol.get(
                        doc_key, GetOptions(timeout=_cb_op_timeout)
                    ).content_as[dict]
                    merged = {**existing}
                    for k, v in fresh.items():
                        if v not in (None, "", [], {}):
                            merged[k] = v
                        elif k not in merged:
                            merged[k] = v
                    _bcol.upsert(doc_key, merged, UpsertOptions(timeout=_cb_op_timeout))
                    refreshed_tickets.append(merged)
                except Exception:
                    _bcol.upsert(doc_key, fresh, UpsertOptions(timeout=_cb_op_timeout))
                    refreshed_tickets.append(fresh)
                ok += 1
                job["saved"] = ok
            except Exception as exc:
                errors += 1
                job["errors"] = errors
                _log = job.setdefault("error_log", [])
                if len(_log) < 200:
                    from supportal.cb_helpers import classify_error as _clf
                    _log.append({"stage": "scrape", "ticket_id": str(tid),
                                 "at": time.time(), **_clf(exc)})
            job["processed"] = i
            pct = i / total * 0.80   # scrape phase = 0–80%
            if i % 10 == 0 or i == total:
                _set_op(f"Rescraping {i}/{total} for '{org}'…", pct)
            if i % 20 == 0 or i == total:
                _upsert_job_doc(job, _bcol)   # persist progress every 20 tickets
            time.sleep(0.35)

        try:
            _bcluster.close()
        except Exception:
            pass

        # ── Enrich with snapshot topology ─────────────────────────────────
        # Fetch snapshot topology for any refreshed ticket that has snap IDs,
        # then patch the topology fields back into the already-saved CB docs.
        if refreshed_tickets:
            job["phase"] = "enriching"
            _set_op(f"Enriching {len(refreshed_tickets)} tickets with snapshot topology…", 0.81)
            try:
                _enrich_cancel = threading.Event()  # never set — runs to completion

                # Reconnect to CB for topology write-back
                from couchbase.cluster import Cluster as _ECl
                from couchbase.options import ClusterOptions as _ECO
                from couchbase.auth import PasswordAuthenticator as _EPA
                from couchbase.options import UpsertOptions as _EUO
                _econn  = _cb_conn_str(cb_url, use_tls)
                _eclust = _ECl(_econn, _ECO(_EPA(username, password)))
                _eclust.wait_until_ready(timedelta(seconds=10))
                _ecol   = _eclust.bucket(bucket).scope(scope).collection(collection)
                _snapcol = None
                try:
                    _snapcol = _eclust.bucket(bucket).scope(scope).collection("snapshots")
                except Exception:
                    pass

                def _snap_upsert(snap_doc: dict):
                    if _snapcol is None:
                        return
                    _sid = snap_doc.get("snap_id") or ""
                    if _sid:
                        try:
                            _snapcol.upsert(f"snapshot::{_sid}", snap_doc)
                        except Exception:
                            pass

                def _enrich_prog(msg: str, pct: float):
                    job["last_message"] = msg
                    _OP_STATUS["status"] = f"[{job['job_id']}] {msg}"

                _enriched_n, _enrich_errs = enrich_tickets_with_snapshots(
                    refreshed_tickets, cookie, _enrich_prog, _enrich_cancel,
                    max_workers=4, snap_upsert_fn=_snap_upsert,
                )
                job["enriched"] = _enriched_n

                # Write topology fields back to the ticket docs in CB
                _topo_op_to = timedelta(seconds=5)
                for _et in refreshed_tickets:
                    if not _et.get("snapshot_topology"):
                        continue
                    _etid = str(_et.get("ticket_id") or "").strip()
                    if not _etid:
                        continue
                    try:
                        _ex2 = _ecol.get(f"ticket::{_etid}", GetOptions(timeout=_topo_op_to)).content_as[dict]
                        _ex2["snapshot_topology"] = _et["snapshot_topology"]
                        _ex2["snapshot_summary"]  = _et.get("snapshot_summary") or {}
                        _ex2["snap_ids"]          = _et.get("snap_ids") or []
                        _ecol.upsert(f"ticket::{_etid}", _ex2, UpsertOptions(timeout=_topo_op_to))
                    except Exception:
                        pass

                try:
                    _eclust.close()
                except Exception:
                    pass
            except Exception as exc:
                job["last_message"] = f"Enrichment failed: {exc}"
                _job_fail(job, "enrich", exc)

        # ── Embed refreshed tickets ────────────────────────────────────────
        emb_p = (emb_params or {}).get("provider", "").lower().strip()
        emb_m = (emb_params or {}).get("model", "")
        emb_k = (emb_params or {}).get("api_key", "")
        emb_u = (emb_params or {}).get("base_url", "")
        emb_d = int((emb_params or {}).get("dims") or 0)
        emb_workers = int((emb_params or {}).get("max_workers") or (emb_params or {}).get("embed_parallel") or 1)
        if emb_p and emb_m and emb_d and refreshed_tickets:
            job["phase"] = "embedding"
            _set_op(f"Embedding {len(refreshed_tickets)} refreshed tickets…", 0.82)
            if emb_p == "lmstudio":
                _lms_base = (emb_u or "http://localhost:1234").rstrip("/v1").rstrip("/")
                _lms_emb_id = lmstudio_ensure_model_loaded(_lms_base, emb_m, timeout_s=45, model_type="embeddings")
                if _lms_emb_id:
                    emb_m = _lms_emb_id
            try:
                def _emb_prog(msg: str, pct: float):
                    job["last_message"] = msg
                    _OP_STATUS["status"]   = f"[{job['job_id']}] {msg}"
                    _OP_STATUS["progress"] = 0.82 + pct * 0.15

                _done_emb, _errs_emb = embed_all_tickets(
                    refreshed_tickets, cb_url, bucket, username, password,
                    use_tls, scope, collection,
                    emb_p, emb_m, emb_k, emb_u, emb_d,
                    _emb_prog,
                    max_workers=emb_workers,
                    error_sink=job.setdefault("error_log", []),
                )
                job["embedded"] = _done_emb
                job["errors"]  += _errs_emb
            except Exception as exc:
                job["last_message"] = f"Embedding failed: {exc}"
                _job_fail(job, "embed", exc)

        # ── Score refreshed tickets ───────────────────────────────────────
        s_prov    = (emb_params or {}).get("score_provider", "").lower().strip()
        s_mod     = (emb_params or {}).get("score_model", "")
        s_key     = (emb_params or {}).get("score_api_key", "")
        s_url     = (emb_params or {}).get("score_base_url", "")
        s_num_ctx = int((emb_params or {}).get("score_ctx") or 0) or None
        s_no_think = bool((emb_params or {}).get("score_no_think", False))
        if s_prov and s_mod and refreshed_tickets:
            job["phase"] = "scoring"
            _score_batch_sz = 20
            _total_scored = 0
            _scored_at = int(time.time())
            # Scoring runs after the scrape-phase cluster (_bcluster) has already
            # been closed above — open a fresh connection rather than reusing it,
            # the same way the enrich phase opens its own _eclust connection.
            _scol = None
            try:
                from couchbase.cluster import Cluster as _SCl
                from couchbase.options import ClusterOptions as _SCO
                from couchbase.auth import PasswordAuthenticator as _SPA
                _sconn   = _cb_conn_str(cb_url, use_tls)
                _sclust  = _SCl(_sconn, _SCO(_SPA(username, password)))
                _sclust.wait_until_ready(timedelta(seconds=10))
                _scol = _sclust.bucket(bucket).scope(scope).collection(collection)
            except Exception as exc:
                job["last_message"] = f"Scoring save-back connection failed: {exc}"
                _job_fail(job, "score", exc)
            for _si in range(0, len(refreshed_tickets), _score_batch_sz):
                _chunk = refreshed_tickets[_si:_si + _score_batch_sz]
                _set_op(
                    f"Scoring tickets {_si + 1}–{min(_si + _score_batch_sz, len(refreshed_tickets))}"
                    f" of {len(refreshed_tickets)}…",
                    0.97 + (_si / max(len(refreshed_tickets), 1)) * 0.02,
                )
                try:
                    _score_results = score_tickets_batch(
                        _chunk,
                        s_prov, s_mod, s_key, s_url,
                        num_ctx=s_num_ctx,
                        no_think=s_no_think,
                    )
                    # Persist scores back into ticket docs in CB
                    if _scol and _score_results:
                        for _sc in _score_results:
                            _stid = str(_sc.get("ticket_id") or "").strip()
                            if not _stid:
                                continue
                            try:
                                _sdoc_key = f"ticket::{_stid}"
                                _sdoc = _scol.get(_sdoc_key).content_as[dict]
                                _sdoc["score"] = {**(_sdoc.get("score") or {}), **_sc, "scored_at": _scored_at}
                                _scol.upsert(_sdoc_key, _sdoc)
                                _total_scored += 1
                            except Exception as _wexc:
                                _job_fail(job, "score_save", _wexc, ticket_id=_stid)
                                job["last_message"] = f"Failed to save score for ticket {_stid}: {_wexc}"
                    elif _score_results:
                        # No working CB connection — do not fake success.
                        job["errors"] += len(_score_results)
                        _log = job.setdefault("error_log", [])
                        if len(_log) < 200:
                            _log.append({"stage": "score_save", "error_type": "str",
                                         "error_code": "CONN_REFUSED",
                                         "abridged": f"{len(_score_results)} scores lost — no CB connection",
                                         "at": time.time()})
                        job["last_message"] = (
                            f"Scored {len(_score_results)} tickets but had no CB connection "
                            f"to save them — see earlier error."
                        )
                    job["scored"] = _total_scored
                except Exception as exc:
                    job["last_message"] = f"Scoring batch {_si // _score_batch_sz + 1} failed: {exc}"
                    _job_fail(job, "score", exc)
            try:
                if _scol is not None:
                    _sclust.close()
            except Exception:
                pass

        job["status"]      = "done"
        job["phase"]       = None
        job["finished_at"] = time.time()
        _new_n = job.get("new_count", 0)
        _new_label = f"{_new_n} new + " if _new_n else ""
        summary = (
            f"Done — {_new_label}{ok}/{total} tickets updated, "
            f"{job.get('enriched', 0)} enriched with topology, "
            f"{job['embedded']} embedded, {job['scored']} scored"
            + (f", {skipped} skipped" if skipped else "")
            + (f", {job['errors']} errors" if job["errors"] else "")
        )
        job["last_message"] = summary
        _set_op(summary, 1.0, done=True)
        _persist_job_state(job, cb_url, bucket, username, password, use_tls, scope, collection)
        _persist_failure_log(job, cb_url, bucket, username, password, use_tls, scope)
        _persist_job_run(job, cb_url, bucket, username, password, use_tls, scope, collection, conclude=True)

    except Exception as exc:
        job["status"]      = "error"
        job["phase"]       = None
        job["finished_at"] = time.time()
        job["last_message"] = f"Fatal error: {exc}"
        _job_fail(job, "fatal", exc)
        _OP_STATUS.update({"op": None, "status": str(exc), "progress": 0.0, "done": True})
        _persist_job_state(job, cb_url, bucket, username, password, use_tls, scope, collection)
        _persist_failure_log(job, cb_url, bucket, username, password, use_tls, scope)
        _persist_job_run(job, cb_url, bucket, username, password, use_tls, scope, collection, conclude=True)


_PROFILE_SCOPE      = "chat"
_PROFILE_COLLECTION = "profiles"


def _ensure_profiles_collection(cluster, bucket_name: str) -> None:
    try:
        from couchbase.management.collections import CollectionSpec  # type: ignore
        cm = cluster.bucket(bucket_name).collections()
        existing = {s.name: {c.name for c in s.collections} for s in cm.get_all_scopes()}
        if _PROFILE_SCOPE not in existing:
            cm.create_scope(_PROFILE_SCOPE)
            existing[_PROFILE_SCOPE] = set()
        if _PROFILE_COLLECTION not in existing[_PROFILE_SCOPE]:
            cm.create_collection(CollectionSpec(_PROFILE_COLLECTION, scope_name=_PROFILE_SCOPE))
        cluster.query(
            f"CREATE PRIMARY INDEX IF NOT EXISTS "
            f"ON `{bucket_name}`.`{_PROFILE_SCOPE}`.`{_PROFILE_COLLECTION}`"
        ).execute()
    except Exception:
        pass


def _record_customer_access(
    org: str,
    cb_url: str, bucket: str, username: str, password: str,
    use_tls: bool, profile_user: str,
) -> None:
    """Increment access_count and update last_accessed_at for org in the user profile."""
    if not _CB_AVAILABLE or not cb_url or not org:
        return
    try:
        from couchbase.cluster import Cluster as _Cl  # type: ignore
        from couchbase.options import ClusterOptions as _CO  # type: ignore
        from couchbase.auth import PasswordAuthenticator as _PA               # type: ignore
        conn = _cb_conn_str(cb_url, use_tls)
        cl = _Cl(conn, _CO(_PA(username, password)))
        cl.wait_until_ready(timedelta(seconds=10))
        _ensure_profiles_collection(cl, bucket)
        col = cl.bucket(bucket).scope(_PROFILE_SCOPE).collection(_PROFILE_COLLECTION)
        key = f"profile::{profile_user}"
        now = int(time.time())
        try:
            doc = col.get(key).content_as[dict]
        except Exception:
            doc = {"username": profile_user, "top_customers": [], "alert_thresholds": {
                "new_p1": True, "score_drop_pts": 10, "stale_hours": 12,
            }, "last_validated_at": 0, "updated_at": 0}
        customers = doc.get("top_customers") or []
        entry = next((c for c in customers if (c.get("name") or "").lower() == org.lower()), None)
        if entry:
            entry["access_count"] = (entry.get("access_count") or 0) + 1
            entry["last_accessed_at"] = now
        else:
            customers.append({
                "name": org, "access_count": 1,
                "last_accessed_at": now, "validated_at": 0, "is_valid": True,
            })
        def _score(c: dict) -> float:
            days = max(0, (now - (c.get("last_accessed_at") or 0)) / 86400)
            return (c.get("access_count") or 1) / (1.0 + days)
        customers.sort(key=_score, reverse=True)
        doc["top_customers"] = customers[:20]
        doc["updated_at"] = now
        col.upsert(key, doc)
        cl.close()
    except Exception:
        pass


_SETTINGS_SCOPE      = "chat"
_SETTINGS_COLLECTION = "settings"
_SETTINGS_KEY        = "strabo::profiles"


def _cb_save_settings(
    cb_url: str,
    bucket: str,
    username: str,
    password: str,
    use_tls: bool,
    scope: str,
    collection: str,
    profiles: dict,
) -> None:
    """Persist the profiles dict to a fixed CB doc (best-effort, does not raise)."""
    if not _CB_AVAILABLE or not cb_url:
        return
    try:
        from couchbase.cluster import Cluster as _Cl
        from couchbase.options import ClusterOptions as _CO
        from couchbase.auth import PasswordAuthenticator as _PA
        conn = _cb_conn_str(cb_url, use_tls)
        cl = _Cl(conn, _CO(_PA(username, password)))
        cl.wait_until_ready(timedelta(seconds=15))
        col = cl.bucket(bucket).scope(scope).collection(collection)
        col.upsert(_SETTINGS_KEY, profiles)
        cl.close()
    except Exception:
        pass


def _cb_load_settings(
    cb_url: str,
    bucket: str,
    username: str,
    password: str,
    use_tls: bool,
    scope: str,
    collection: str,
) -> dict:
    """Load the profiles dict from CB. Returns {} on any error."""
    if not _CB_AVAILABLE or not cb_url:
        return {}
    try:
        from couchbase.cluster import Cluster as _Cl
        from couchbase.options import ClusterOptions as _CO
        from couchbase.auth import PasswordAuthenticator as _PA
        conn = _cb_conn_str(cb_url, use_tls)
        cl = _Cl(conn, _CO(_PA(username, password)))
        cl.wait_until_ready(timedelta(seconds=15))
        col = cl.bucket(bucket).scope(scope).collection(collection)
        result = col.get(_SETTINGS_KEY).content_as[dict]
        cl.close()
        return result
    except Exception:
        return {}


def _load_customer_profile(
    cb_url: str,
    bucket: str,
    username: str,
    password: str,
    use_tls: bool,
    profile_user: str,
) -> dict:
    """Read the profile doc for profile_user from CB. Returns empty profile on error."""
    _default = {
        "username": profile_user,
        "top_customers": [],
        "alert_thresholds": {"new_p1": True, "score_drop_pts": 10, "stale_hours": 12},
    }
    if not _CB_AVAILABLE or not cb_url:
        return _default
    try:
        from couchbase.cluster import Cluster as _Cl
        from couchbase.options import ClusterOptions as _CO
        from couchbase.auth import PasswordAuthenticator as _PA
        conn = _cb_conn_str(cb_url, use_tls)
        cl = _Cl(conn, _CO(_PA(username, password)))
        cl.wait_until_ready(timedelta(seconds=10))
        _ensure_profiles_collection(cl, bucket)
        col = cl.bucket(bucket).scope(_PROFILE_SCOPE).collection(_PROFILE_COLLECTION)
        doc = col.get(f"profile::{profile_user}").content_as[dict]
        cl.close()
        return doc
    except Exception:
        return _default


def _get_briefing_data(
    top_customers: list[dict],
    cb_url: str,
    bucket: str,
    username: str,
    password: str,
    use_tls: bool,
    scope: str,
    collection: str,
    stale_hours: float = 12.0,
) -> list[dict]:
    """Return one health-summary row per customer in top_customers."""
    rows = []
    for c in top_customers:
        org = (c.get("name") or "").strip()
        if not org:
            continue
        try:
            h = _compute_health_score(org, cb_url, bucket, username, password,
                                      use_tls, scope, collection)
            hours_stale = h.get("hours_since_scraped") or 0.0
            rows.append({
                "name":        org,
                "score":       h.get("score", 0),
                "grade":       h.get("grade", "N/A"),
                "open_p1":     h.get("open_p1", 0),
                "open_p2":     h.get("open_p2", 0),
                "hours_stale": round(float(hours_stale), 1),
                "alert":       (
                    h.get("open_p1", 0) > 0
                    or h.get("score", 100) < 40
                    or float(hours_stale) > stale_hours
                ),
            })
        except Exception:
            rows.append({
                "name": org, "score": 0, "grade": "?",
                "open_p1": 0, "open_p2": 0, "hours_stale": 0.0, "alert": False,
            })
    return rows


def _execute_agent_tool(
    name: str,
    args: dict,
    cb_url: str, bucket: str, username: str, password: str,
    use_tls: bool, scope: str, collection: str,
    default_customer: str = "",
    ctx: dict | None = None,
) -> str:
    """Execute an agent tool call and return a string result for the LLM.

    ctx carries LLM/embedding/cookie config for tools that need to call out:
      provider, model, api_key, base_url,
      emb_provider, emb_model, emb_api_key, emb_base_url, emb_dims,
      cookie
    """
    ctx = ctx or {}
    # ── v1.5.0: session log tracks tools called this conversation ────────────
    _slog: dict = ctx.setdefault("_session_log", {})

    # ── 3.5: record customer access for profile tracking ─────────────────────
    _CUSTOMER_SCOPED_TOOLS = {
        "query_tickets", "count_tickets", "get_ticket", "vector_search",
        "get_customer_health_score", "check_sla_compliance", "get_digest",
        "generate_customer_report", "generate_health_report", "rescrape_customer_tickets",
        "scrape_customer_tickets", "get_cluster_health", "check_data_freshness", "smart_refresh",
    }
    if name in _CUSTOMER_SCOPED_TOOLS:
        _tracked_org = (
            args.get("organization") or args.get("customer") or
            args.get("org") or default_customer or ""
        ).strip()
        if _tracked_org and cb_url and username:
            import threading as _thr_profile
            _thr_profile.Thread(
                target=_record_customer_access,
                args=(_tracked_org, cb_url, bucket, username, password,
                      use_tls, ctx.get("profile_user", "default")),
                daemon=True,
            ).start()

    if name == "query_tickets":
        limit = min(int(args.get("limit") or 50), 200)
        filters = _agent_filters_from_args(args)
        # Auto-scope to loaded customer when LLM omits the filter
        if default_customer and not filters.get("organization"):
            filters["organization"] = default_customer
        tickets = tool_query_tickets(
            filters, cb_url, bucket, username, password,
            use_tls, scope, collection, limit=limit,
        )
        if not tickets:
            # BEFORE v1.5.0: just returned "No tickets found."
            # AFTER v1.5.0: hint at scrape_customer_tickets when data is missing for a scoped customer
            _org = filters.get("organization") or default_customer
            if _org:
                _ingest_hint = (
                    f" No local tickets for '{_org}'. "
                    "Call scrape_customer_tickets to pull them from Supportal first."
                )
            else:
                _ingest_hint = ""
            return f"No tickets found matching the given filters.{_ingest_hint}"
        _now_epoch = time.time()
        lines = [
            "| Ticket ID | Organization | Subject | Status | Priority | Created | Last Reply | Data Age | CBSEs | Jira |",
            "|-----------|-------------|---------|--------|----------|---------|------------|----------|-------|------|",
        ]
        for t in tickets:
            cbses = ", ".join(t.get("cbses") or []) or "—"
            jiras = ", ".join(t.get("jira_issues") or []) or "—"
            subj = (t.get("subject") or "")[:55].replace("|", "/")
            _lsa = t.get("last_scraped_at") or 0
            _age_h = (_now_epoch - _lsa) / 3600 if _lsa else None
            _age_str = f"{_age_h:.0f}h ago" if _age_h is not None else "unknown"
            _last_reply = (
                t.get("last_comment_at") or t.get("updated") or t.get("updated_at") or ""
            )[:10] or "—"
            lines.append(
                f"| {t.get('ticket_id','')} | {t.get('organization','')} | {subj} "
                f"| {t.get('status','')} | {t.get('priority','')} "
                f"| {(t.get('created') or '')[:10]} | {_last_reply} | {_age_str} | {cbses} | {jiras} |"
            )
        return "\n".join(lines) + f"\n\n**Total: {len(tickets)} tickets**"

    elif name == "count_tickets":
        filters = _agent_filters_from_args(args)
        if default_customer and not filters.get("organization"):
            filters["organization"] = default_customer
        tickets = tool_query_tickets(
            filters, cb_url, bucket, username, password,
            use_tls, scope, collection, limit=5000,
        )
        return str(len(tickets))

    elif name == "get_ticket":
        ticket_id = str(args.get("ticket_id") or "").strip()
        if not ticket_id:
            return "Error: ticket_id is required."
        doc_key = f"ticket::{ticket_id}"
        tickets = fetch_tickets_by_keys(
            [doc_key], cb_url, bucket, username, password, use_tls, scope, collection,
        )
        if not tickets:
            # Live fallback — fetch directly from Supportal and save to CB so future lookups work
            try:
                _live_sess = _make_api_session("")
                _live = fetch_ticket_api(ticket_id, _live_sess)
                if _live and _live.get("ticket_id"):
                    _live["last_scraped_at"] = int(time.time())
                    _live["type"]            = "ticket"
                    _live["cb_version"]      = extract_ticket_version(_live)
                    _live["feature_area"]    = classify_ticket_feature(_live)
                    _live["ticket_origin"]   = classify_ticket_origin(_live)
                    try:
                        from couchbase.cluster import Cluster as _GtCl
                        from couchbase.options import ClusterOptions as _GtCO
                        from couchbase.auth import PasswordAuthenticator as _GtPA
                        _gt_cl = _GtCl(_cb_conn_str(cb_url, use_tls), _GtCO(_GtPA(username, password)))
                        _gt_cl.wait_until_ready(timedelta(seconds=10))
                        _gt_cl.bucket(bucket).scope(scope).collection(collection).upsert(doc_key, _live)
                        _gt_cl.close()
                    except Exception:
                        pass
                    tickets = [_live]
                else:
                    return f"Ticket {ticket_id} not found in local DB or Supportal."
            except Exception as _lfe:
                return f"Ticket {ticket_id} not found locally. Live fetch also failed: {_lfe}"
        t = tickets[0]
        _tid = t.get("ticket_id", ticket_id)
        _lsa = t.get("last_scraped_at") or 0
        _age_h = (time.time() - _lsa) / 3600 if _lsa else None
        _age_str = f"{_age_h:.1f} hours ago" if _age_h is not None else "unknown"
        _supportal_url = _SUPPORTAL_TICKET_URL.format(ticket_id=_tid)
        _last_reply = (t.get("last_comment_at") or t.get("updated") or t.get("updated_at") or "")[:19] or "—"
        parts = [
            f"**Ticket {_tid}** — {t.get('subject','')}",
            f"Organization: {t.get('organization','')}",
            f"Status: {t.get('status','')} | Priority: {t.get('priority','')}",
            f"Created: {t.get('created','')} | Last Reply: {_last_reply} | Closed: {t.get('closed','')}",
            f"Requester: {t.get('requester','')}",
            f"**Data freshness:** last scraped {_age_str}",
            f"**Live verification:** {_supportal_url}",
        ]
        cbses = t.get("cbses") or []
        if cbses:
            parts.append(f"CBSEs: {', '.join(cbses)}")
        jiras = t.get("jira_issues") or []
        if jiras:
            parts.append(f"Jira Issues: {', '.join(jiras)}")
        _score = t.get("score") or {}
        clusters = _score.get("cluster_names") or []
        if clusters:
            parts.append(f"Clusters: {', '.join(clusters)}")

        # ── Snapshot topology ────────────────────────────────────────────────
        topo = t.get("snapshot_topology") or {}
        if isinstance(topo, str):
            try:
                topo = json.loads(topo)
            except Exception:
                topo = {}
        if isinstance(topo, dict) and topo:
            topo_lines = []
            if topo.get("cluster_name"):
                topo_lines.append(f"  Cluster Name:    {topo['cluster_name']}")
            if topo.get("cluster_uuid"):
                topo_lines.append(f"  Cluster UUID:    {topo['cluster_uuid']}")
            if topo.get("cb_version"):
                topo_lines.append(f"  CB Version:      {topo['cb_version']}")
            if topo.get("total_nodes"):
                topo_lines.append(f"  Nodes:           {topo['total_nodes']}")
            svc_parts = []
            for svc, key in [("KV/Data", "data_nodes"), ("Index", "index_nodes"),
                              ("Query", "query_nodes"), ("Search", "fts_nodes"),
                              ("Eventing", "eventing_nodes"), ("Analytics", "analytics_nodes")]:
                n = topo.get(key)
                if n:
                    svc_parts.append(f"{svc}×{n}")
            if svc_parts:
                topo_lines.append(f"  Services:        {', '.join(svc_parts)}")
            if topo.get("bucket_count"):
                topo_lines.append(f"  Buckets:         {topo['bucket_count']}")
            _bn = topo.get("bucket_names") or []
            if isinstance(_bn, list) and _bn:
                topo_lines.append(f"  Bucket Names:    {', '.join(_bn[:10])}")
            if topo.get("cpus_per_node"):
                topo_lines.append(f"  CPUs/Node:       {topo['cpus_per_node']}")
            if topo.get("ram_used_per_node_mib") and topo.get("ram_per_node_mib"):
                topo_lines.append(f"  RAM/Node:        {topo['ram_used_per_node_mib']}/{topo['ram_per_node_mib']} MiB used/total")
            elif topo.get("ram_per_node_mib"):
                topo_lines.append(f"  RAM/Node:        {topo['ram_per_node_mib']} MiB")
            if topo.get("auto_failover_seconds") is not None:
                topo_lines.append(f"  Auto-failover:   {topo['auto_failover_seconds']}s")
            if topo.get("n2n_encryption") is not None:
                topo_lines.append(f"  N2N Encryption:  {topo['n2n_encryption']}")
            if topo.get("data_quota_mib"):
                topo_lines.append(f"  Data Quota:      {topo['data_quota_mib']} MiB")
            if topo.get("global_index_count") is not None:
                topo_lines.append(f"  GSI Indexes:     {topo['global_index_count']}")
            if topo.get("fts_index_count") is not None:
                topo_lines.append(f"  FTS Indexes:     {topo['fts_index_count']}")
            if topo.get("eventing_function_count") is not None:
                topo_lines.append(f"  Eventing Fns:    {topo['eventing_function_count']}")
            bad  = topo.get("bad_items",  topo.get("bad_count",  0)) or 0
            warn = topo.get("warn_items", topo.get("warn_count", 0)) or 0
            if bad or warn:
                topo_lines.append(f"  Health:          bad={bad}  warn={warn}")
            if topo.get("os_name"):
                topo_lines.append(f"  OS:              {topo['os_name']}")
            if not topo.get("cpus_per_node"):
                _org_hint = t.get("organization") or ""
                _hint_arg = f'organization="{_org_hint}"' if _org_hint else "organization=<org>"
                topo_lines.append(f"  CPUs/Node:       [absent from this snapshot — MUST call get_cluster_health({_hint_arg}) to check other snapshots for this cluster]")
            if topo_lines:
                parts.append("\n**Cluster Topology (snapshot):**\n" + "\n".join(topo_lines))
        elif not topo:
            snap_ids = t.get("snap_ids") or []
            if snap_ids:
                parts.append(f"\n*Snapshot IDs linked: {len(snap_ids)} — topology not yet enriched.*")

        summary = (t.get("summary_text") or _score.get("interaction_summary") or "").strip()
        if summary:
            parts.append(f"\n**Summary:**\n{summary}")
        desc = (t.get("description") or "")[:2000]
        if desc:
            parts.append(f"\n**Description:**\n{desc}")
        comments = t.get("comments") or []
        if comments:
            parts.append(f"\n**Comments ({len(comments)}):**")
            for c in comments[:5]:
                author = c.get("author", {})
                author_name = author.get("name", "") if isinstance(author, dict) else str(author)
                body = (c.get("body") or c.get("plain_body") or "")[:500]
                parts.append(f"  [{author_name}]: {body}")
        return "\n".join(parts)

    elif name == "check_data_freshness":
        ticket_ids = args.get("ticket_ids") or []
        if not ticket_ids:
            return "No ticket IDs provided."
        if not _CB_AVAILABLE:
            return "Couchbase not available."
        try:
            from couchbase.cluster import Cluster  # type: ignore
            from couchbase.options import ClusterOptions  # type: ignore
            from couchbase.auth import PasswordAuthenticator  # type: ignore
            from couchbase.options import QueryOptions  # type: ignore
            conn_str = _cb_conn_str(cb_url, use_tls)
            cluster = Cluster(conn_str, ClusterOptions(PasswordAuthenticator(username, password)))
            cluster.wait_until_ready(timedelta(seconds=10))
            keyspace = f"`{bucket}`.`{scope}`.`{collection}`"
            phs = ", ".join(f"${i+1}" for i in range(len(ticket_ids)))
            rows = list(cluster.query(
                f"SELECT ticket_id, status, last_scraped_at "
                f"FROM {keyspace} WHERE ticket_id IN [{phs}]",
                QueryOptions(positional_parameters=[str(t) for t in ticket_ids],
                             timeout=timedelta(seconds=15)),
            ))
            cluster.close()
        except Exception as exc:
            return f"Freshness check failed: {exc}"

        _now = time.time()
        lines = [
            "| Ticket ID | Status (local) | Last Scraped | Age | Verify Live |",
            "|-----------|---------------|-------------|-----|------------|",
        ]
        for r in rows:
            _lsa = r.get("last_scraped_at") or 0
            _age_h = (_now - _lsa) / 3600 if _lsa else None
            _age_str = f"{_age_h:.1f}h" if _age_h is not None else "unknown"
            _stale = "⚠️ STALE" if (_age_h or 0) > 4 else "✓ fresh"
            _url = _SUPPORTAL_TICKET_URL.format(ticket_id=r.get("ticket_id", ""))
            lines.append(
                f"| {r.get('ticket_id','')} | {r.get('status','')} "
                f"| {_age_str} ago | {_stale} | {_url} |"
            )
        result = "\n".join(lines)
        stale_count = sum(1 for r in rows if ((_now - (r.get("last_scraped_at") or 0)) / 3600) > 4)
        if stale_count:
            result += (
                f"\n\n⚠️ {stale_count}/{len(rows)} tickets have data older than 4 hours. "
                f"Use rescrape_ticket to refresh individual tickets, or click the Verify Live "
                f"links above to check current status on Supportal directly."
            )
        else:
            result += f"\n\n✓ All {len(rows)} tickets have fresh data (scraped within 4 hours)."
        return result

    elif name == "rescrape_customer_tickets":
        cust        = (args.get("customer") or default_customer or "").strip()
        stale_hours = float(args.get("stale_hours") if args.get("stale_hours") is not None else 4.0)
        max_tix     = min(int(args.get("max_tickets") or 50), 2000)
        status_filt = (args.get("status") or "").strip().lower() or None

        cookie = _get_profile_cookie()  # optional as of v2.6.2; retained for re-enablement

        # ── Step 1: Gather existing CB tickets for this customer ──────────────
        _rs_filters: dict = {}
        if cust and cust.lower() != "all customers":
            _rs_filters["organization"] = cust
        if status_filt:
            _rs_filters["status"] = status_filt

        candidates = tool_query_tickets(
            _rs_filters, cb_url, bucket, username, password,
            use_tls, scope, collection, limit=max_tix * 8,
        )

        # The candidate list above is capped (stale-refresh pool only); the
        # new-vs-cached comparison needs the org's COMPLETE local ticket-id set,
        # otherwise cached tickets beyond the cap get misclassified as "new".
        cb_ids: set[str] = set()
        try:
            _idc = Cluster(
                _cb_conn_str(cb_url, use_tls),
                ClusterOptions(PasswordAuthenticator(username, password)),
            )
            _idc.wait_until_ready(timedelta(seconds=10))
            _id_where = ["ticket_id IS NOT MISSING",
                         "(`_deleted` IS MISSING OR `_deleted` = false)"]
            _id_params: list = []
            if cust and cust.lower() != "all customers":
                _id_params.append(f"%{cust.lower()}%")
                _id_where.append(f"LOWER(TOSTRING(organization)) LIKE ${len(_id_params)}")
            _id_rows = list(_idc.query(
                f"SELECT RAW ticket_id FROM `{bucket}`.`{scope}`.`{collection}` "
                f"WHERE {' AND '.join(_id_where)}",
                QueryOptions(positional_parameters=_id_params,
                             timeout=timedelta(seconds=30)),
            ))
            _idc.close()
            cb_ids = {str(r or "").strip() for r in _id_rows if r}
        except Exception:
            cb_ids = set()
        if not cb_ids:
            # Fallback: capped candidate set (better than treating everything as new)
            cb_ids = {str(t.get("ticket_id") or "").strip() for t in candidates if t.get("ticket_id")}

        # ── Step 2: Fetch full Supportal listing to discover new tickets ──────
        new_stubs: list[dict] = []
        _listing_note = ""
        if cust and cust.lower() != "all customers":
            try:
                _list_sess = _make_api_session(cookie)
                _listing = _get_customer_ticket_listing_api(cust, _list_sess)
                supportal_ids = {str(r.get("id") or "").strip() for r in _listing if r.get("id")}
                new_ids = supportal_ids - cb_ids
                if new_ids:
                    # Apply status filter if requested
                    if status_filt:
                        _id_map = {str(r.get("id") or ""): r for r in _listing}
                        new_ids = {
                            nid for nid in new_ids
                            if (_id_map.get(nid, {}).get("status") or "").lower() == status_filt
                        }
                    # Newest ticket ID first — sorted() on strings is lexical
                    # (oldest-first) and a max_tix cap then silently keeps only
                    # the oldest tickets instead of the newest (e.g. a first-time
                    # scrape of an org with hundreds of "new" tickets kept only
                    # 2019-era IDs, making an active account look dormant).
                    for nid in sorted(new_ids, key=lambda x: int(x) if x.isdigit() else 0, reverse=True):
                        new_stubs.append({
                            "ticket_id": nid,
                            "organization": cust,
                            "last_scraped_at": 0,  # force scrape
                        })
                    _listing_note = f"{len(new_stubs)} new"
            except Exception as _le:
                _listing_note = f"(listing unavailable: {_le})"

        # ── Step 3: Apply stale filter to existing candidates ─────────────────
        now_epoch    = time.time()
        stale_cutoff = now_epoch - stale_hours * 3600
        stale_existing = (
            [t for t in candidates if (t.get("last_scraped_at") or 0) < stale_cutoff]
            if stale_hours > 0 else list(candidates)
        )

        # Merge: new stubs first, then stale existing (deduped), cap at max_tix
        _seen_ids: set[str] = set()
        to_scrape: list[dict] = []
        for t in new_stubs + stale_existing:
            tid = str(t.get("ticket_id") or "").strip()
            if tid and tid not in _seen_ids:
                _seen_ids.add(tid)
                to_scrape.append(t)
            if len(to_scrape) >= max_tix:
                break

        new_count = min(len(new_stubs), len(to_scrape))  # new stubs are always first in to_scrape

        if not to_scrape:
            if not candidates:
                return f"No tickets found in Couchbase or Supportal for '{cust or 'all customers'}'."
            return (
                f"All {len(candidates)} tickets for {cust or 'all customers'} "
                f"were scraped within the last {stale_hours:.0f} hours — nothing to update."
            )

        _job = _make_scrape_job(cust or "all customers", "rescrape")
        _job["new_count"] = new_count
        threading.Thread(
            target=_run_rescrape_job_bg,
            args=(
                _job, cust or "all customers", to_scrape, cookie,
                {"url": cb_url, "bucket": bucket, "username": username, "password": password,
                 "use_tls": use_tls, "scope": scope, "collection": collection},
                {"provider": ctx.get("emb_provider",""), "model": ctx.get("emb_model",""),
                 "api_key": ctx.get("emb_api_key",""), "base_url": ctx.get("emb_base_url",""),
                 "dims": ctx.get("emb_dims", 0),
                 "embed_parallel": ctx.get("embed_parallel", 1),
                 "score_provider": ctx.get("provider",""), "score_model": ctx.get("model",""),
                 "score_api_key": ctx.get("api_key",""), "score_base_url": ctx.get("base_url","")},
            ),
            daemon=True,
        ).start()
        if ctx is not None:
            ctx.setdefault("_started_jobs", []).append(_job["job_id"])

        _new_label = f"{new_count} new + " if new_count else ""
        _stale_label = f"{len(to_scrape) - new_count} stale" if new_count else f"{len(to_scrape)} stale"
        return (
            f"Started rescrape job **{_job['job_id']}** for '{cust or 'all customers'}' "
            f"({_new_label}{_stale_label} tickets{f', stale > {stale_hours:.0f}h' if stale_hours > 0 else ''}). "
            f"Running in the background at ~3 req/s. "
            f"**I cannot notify you when it finishes — you must ask me.** "
            f"Ask 'what is the scrape status?' after a minute or two to check progress."
        )

    elif name == "smart_refresh":
        cust    = (args.get("organization") or default_customer or "").strip()
        max_new = min(int(args.get("max_new") or 25), 100)

        if not cust:
            return "No customer specified and no customer is currently scoped."

        # ── Step 1: Pull rich CB signals for this org ─────────────────────────
        # (includes status, solved, priority, last_scraped_at, enrichment gaps)
        try:
            from couchbase.cluster import Cluster as _SrCl
            from couchbase.options import ClusterOptions as _SrCO, QueryOptions as _SrQO
            from couchbase.auth import PasswordAuthenticator as _SrPA
            _idc = _SrCl(_cb_conn_str(cb_url, use_tls), _SrCO(_SrPA(username, password)))
            _idc.wait_until_ready(timedelta(seconds=10))
            _now_ep = time.time()
            _cb_rows = list(_idc.query(
                f"SELECT META(t).id AS doc_id, t.status, t.solved, t.priority, "
                f"t.last_scraped_at, t.requester, t.`_deleted`, "
                f"(t.score IS NOT MISSING AND t.score IS NOT NULL) AS has_score, "
                f"(t.embedding IS NOT MISSING AND t.embedding IS NOT NULL) AS has_embedding, "
                f"t.sfdc_matched "
                f"FROM `{bucket}`.`{scope}`.`{collection}` AS t "
                f"WHERE META(t).id LIKE 'ticket::%' "
                f"AND LOWER(TOSTRING(t.organization)) LIKE $1 "
                f"AND (t.`_deleted` IS MISSING OR t.`_deleted` = false)",
                _SrQO(positional_parameters=[f"%{cust.lower()}%"], timeout=timedelta(seconds=30)),
            ))
            _idc.close()
        except Exception as _e:
            return f"smart_refresh: CB query failed — {_e}"

        cb_signals: dict[str, dict] = {}
        for _row in _cb_rows:
            _tid = str(_row.get("doc_id", "")).split("::")[-1]
            if _tid:
                _lsa = _row.get("last_scraped_at") or 0
                cb_signals[_tid] = {
                    "status":        (_row.get("status") or "").lower().strip(),
                    "solved":        (_row.get("solved") or "").strip(),
                    "priority":      (_row.get("priority") or "").lower().strip(),
                    "last_scraped_at": _lsa,
                    "age_hours":     (_now_ep - _lsa) / 3600 if _lsa else None,
                    "is_stub":       not _row.get("requester"),
                    "has_score":     bool(_row.get("has_score")),
                    "has_embedding": bool(_row.get("has_embedding")),
                    "sfdc_matched":  _row.get("sfdc_matched"),
                    "_deleted":      bool(_row.get("_deleted")),
                }

        # ── Step 2: Fetch Supportal listing ───────────────────────────────────
        try:
            _sess = _make_api_session("")
            _listing = _get_customer_ticket_listing_api(cust, _sess)
        except Exception as _e:
            return f"smart_refresh: Supportal listing failed — {_e}"

        # Normalise listing items into the same shape _filter_changed_tickets expects
        _listing_summaries: list[dict] = []
        for _item in _listing:
            _itid = str(_item.get("id") or "").strip()
            if not _itid or (_item.get("status") or "").lower() == "deleted":
                continue
            _listing_summaries.append({
                "ticket_id": _itid,
                "status":    (_item.get("status") or "").lower().strip(),
                "priority":  (_item.get("Priority") or _item.get("priority") or "").lower().strip(),
                "subject":   _item.get("subject") or _item.get("raw_subject") or "",
                "created":   _item.get("created_at") or "",
                "solved":    _item.get("solved_at") or "",
                "organization": cust,
                "last_scraped_at": 0,
            })

        if not _listing_summaries:
            return f"smart_refresh: Supportal returned no tickets for '{cust}'."

        # ── Step 3: Multi-signal diff ─────────────────────────────────────────
        # Signals: new ID, status change, solved-date change, priority change,
        # stub (requester missing), stale-open (open/pending > 4h since last scrape).
        # Note: Supportal listing has no updated_at; stale-open is the best proxy.
        to_scrape, n_new, n_changed, n_skipped = _filter_changed_tickets(
            _listing_summaries, cb_signals, max_tickets=max_new, stale_open_hours=4.0,
        )

        # ── Step 4: Enrichment-gap report (CB-only, no rescrape needed) ───────
        _enrich_gaps = [
            tid for tid, sig in cb_signals.items()
            if not sig.get("has_score") or not sig.get("has_embedding")
            or sig.get("sfdc_matched") is None
        ]
        _gap_score   = sum(1 for sig in cb_signals.values() if not sig.get("has_score"))
        _gap_embed   = sum(1 for sig in cb_signals.values() if not sig.get("has_embedding"))
        _gap_sfdc    = sum(1 for sig in cb_signals.values() if sig.get("sfdc_matched") is None)

        # ── Step 5: Kick off background scrape + enrich job if needed ─────────
        _job_note = ""
        if to_scrape:
            _job = _make_scrape_job(cust, "smart_refresh")
            threading.Thread(
                target=_run_rescrape_job_bg,
                args=(
                    _job, cust, to_scrape, "",
                    {"url": cb_url, "bucket": bucket, "username": username, "password": password,
                     "use_tls": use_tls, "scope": scope, "collection": collection},
                    {"provider": ctx.get("emb_provider", ""), "model": ctx.get("emb_model", ""),
                     "api_key": ctx.get("emb_api_key", ""), "base_url": ctx.get("emb_base_url", ""),
                     "dims": ctx.get("emb_dims", 0),
                     "embed_parallel": ctx.get("embed_parallel", 1),
                     "score_provider": ctx.get("provider", ""), "score_model": ctx.get("model", ""),
                     "score_api_key": ctx.get("api_key", ""), "score_base_url": ctx.get("base_url", "")},
                ),
                daemon=True,
            ).start()
            if ctx is not None:
                ctx.setdefault("_started_jobs", []).append(_job["job_id"])
            _job_note = f"Background job **{_job['job_id']}** started (scrape → embed → score → SFDC)."

        # ── Step 6: Build human-readable diff report ──────────────────────────
        _supportal_total = len(_listing_summaries)
        _cb_total        = len(cb_signals)

        _scrape_lines = []
        for _s in to_scrape[:20]:   # cap display at 20 lines
            _reason = _s.get("_change_reason", "new")
            _scrape_lines.append(
                f"- #{_s['ticket_id']} [{_s.get('priority','—').upper()}] "
                f"{_s.get('status','—')} — {_s.get('subject','')[:55]} "
                f"_({_reason})_"
            )
        _more = f"\n  …and {len(to_scrape) - 20} more" if len(to_scrape) > 20 else ""

        _enrich_note = ""
        if _enrich_gaps:
            _enrich_note = (
                f"\n\n**Enrichment gaps** (existing tickets missing pipeline output):\n"
                f"- Missing score: {_gap_score}\n"
                f"- Missing embedding: {_gap_embed}\n"
                f"- SFDC never correlated: {_gap_sfdc}\n"
                f"Use `batch_score_tickets` or `rescrape_customer_tickets` to fill these."
            )

        if not to_scrape and not _enrich_gaps:
            return (
                f"✓ **{cust}** is fully up to date.\n"
                f"Supportal: {_supportal_total} tickets | CB: {_cb_total} tickets | "
                f"{n_skipped} unchanged."
            )

        _header = (
            f"**{cust}** diff — Supportal: {_supportal_total} | CB: {_cb_total}\n\n"
            f"**Changes found:** {n_new} new, {n_changed} updated, {n_skipped} unchanged\n"
        )
        _scrape_section = ""
        if to_scrape:
            _scrape_section = (
                f"\n**Queued for scrape + enrich ({len(to_scrape)}):**\n"
                + "\n".join(_scrape_lines) + _more + "\n\n" + _job_note
            )

        return _header + _scrape_section + _enrich_note

    elif name == "rescrape_ticket":
        ticket_id = str(args.get("ticket_id") or "").strip()
        if not ticket_id:
            # Local models often hallucinate "ticket_ids" (list) — tolerate it and take first element
            _ids = args.get("ticket_ids")
            if isinstance(_ids, list) and _ids:
                ticket_id = str(_ids[0]).strip()
                _remaining = _ids[1:]
            elif _ids:
                ticket_id = str(_ids).strip()
                _remaining = []
            else:
                _remaining = []
        else:
            _remaining = []
        if not ticket_id:
            return "Error: ticket_id is required. Pass a single numeric ticket ID, e.g. {\"ticket_id\": \"12345\"}. For bulk refresh use rescrape_customer_tickets instead."

        # Cookie is optional as of v2.6.2 — Supportal endpoints are open.
        # Retained in case auth is re-added; pass empty string to make an unauthenticated session.
        cookie = ""
        try:
            _settings = _load_settings_file()
            _active = _settings.get("__last__", "")
            _prof = _settings.get(_active, {}) if _active else {}
            cookie = _prof.get("cookie", "")
        except Exception as _pe:
            print(f"[rescrape_ticket] profile read failed: {_pe}")

        # Scrape fresh data from Supportal via REST API
        try:
            _sess = _make_api_session(cookie)
            fresh = fetch_ticket_api(ticket_id, _sess)
        except Exception as exc:
            _url = _SUPPORTAL_TICKET_URL.format(ticket_id=ticket_id)
            return (
                f"Scrape failed ({exc}).\n"
                f"Verify manually: {_url}"
            )

        if not fresh or not fresh.get("ticket_id"):
            return f"Scrape returned no data for ticket {ticket_id}."

        if fresh.get("_deleted"):
            # Persist the deletion marker so future incremental scrapes skip it
            doc_key = f"ticket::{ticket_id}"
            try:
                from couchbase.cluster import Cluster  # type: ignore
                from couchbase.options import ClusterOptions  # type: ignore
                from couchbase.auth import PasswordAuthenticator  # type: ignore
                conn_str = _cb_conn_str(cb_url, use_tls)
                cluster = Cluster(conn_str, ClusterOptions(PasswordAuthenticator(username, password)))
                cluster.wait_until_ready(timedelta(seconds=10))
                col = cluster.bucket(bucket).scope(scope).collection(collection)
                try:
                    existing = col.get(doc_key).content_as[dict]
                    existing["_deleted"] = True
                    existing["last_scraped_at"] = int(time.time())
                    col.upsert(doc_key, existing)
                except Exception:
                    col.upsert(doc_key, {
                        "ticket_id": ticket_id, "_deleted": True, "type": "ticket",
                        "last_scraped_at": int(time.time()),
                    })
                cluster.close()
            except Exception as _de:
                print(f"[rescrape_ticket] failed to persist deletion marker: {_de}")
            return f"Ticket {ticket_id} has been deleted on Supportal — marked in Couchbase, will be skipped in future scrapes."

        fresh["last_scraped_at"] = int(time.time())
        fresh["type"] = "ticket"
        # Recompute analytics classification fields from fresh data
        fresh["cb_version"]    = extract_ticket_version(fresh)
        fresh["feature_area"]  = classify_ticket_feature(fresh)
        fresh["ticket_origin"] = classify_ticket_origin(fresh)
        doc_key = f"ticket::{ticket_id}"

        # ── Inline snapshot topology enrichment ──────────────────────────────
        # Attempt before the CB write so topology lands in the same upsert.
        topo_enriched = False
        _snaps_raw = fresh.get("snapshots")
        _snaps_str = _snaps_raw if isinstance(_snaps_raw, str) else ""
        snap_ids_found = _SNAP_ID_RE.findall(_snaps_str)
        if not snap_ids_found:
            snap_ids_found = _UUID_RE.findall(_snaps_str)
        if snap_ids_found:
            fresh["snap_ids"] = list(dict.fromkeys(snap_ids_found))  # dedup, preserve order
            best_snap = _highest_snap_id(snap_ids_found)
            try:
                topo = fetch_snapshot_topology(best_snap, cookie=cookie)
                if topo:
                    fresh["snapshot_topology"] = topo
                    topo_enriched = True
                    print(f"[rescrape_ticket] topology enriched from snap {best_snap[:16]}…")
            except Exception as _te:
                print(f"[rescrape_ticket] topology fetch failed for {best_snap}: {_te}")

        # ── Merge onto existing CB doc (preserve score, embedding, etc.) ─────
        _saved = False
        try:
            from couchbase.cluster import Cluster  # type: ignore
            from couchbase.options import ClusterOptions  # type: ignore
            from couchbase.auth import PasswordAuthenticator  # type: ignore
            conn_str = _cb_conn_str(cb_url, use_tls)
            cluster = Cluster(conn_str, ClusterOptions(PasswordAuthenticator(username, password)))
            cluster.wait_until_ready(timedelta(seconds=10))
            col = cluster.bucket(bucket).scope(scope).collection(collection)
            try:
                existing = col.get(doc_key).content_as[dict]
                # Start from existing; let fresh override, but skip None/empty fresh values
                # so immutable fields (created, organization, requester) are never wiped.
                merged = {**existing}
                for _k, _v in fresh.items():
                    if _v is not None and _v != "" and _v != [] and _v != {}:
                        merged[_k] = _v
                    elif _k not in merged:
                        merged[_k] = _v  # key is new — write even if null
            except Exception:
                merged = fresh  # doc doesn't exist yet — insert as-is
            col.upsert(doc_key, merged)
            cluster.close()
            fresh = merged  # use merged for summary output
            _saved = True
        except Exception as exc:
            print(f"[rescrape_ticket] CB upsert failed: {exc}")

        # ── Embed + score the refreshed ticket ───────────────────────────────
        _pipeline_notes: list[str] = []
        if _saved:
            emb_p = (ctx.get("emb_provider") or "").lower().strip()
            emb_m = ctx.get("emb_model", "")
            emb_k = ctx.get("emb_api_key", "")
            emb_u = ctx.get("emb_base_url", "")
            emb_d = int(ctx.get("emb_dims") or 0)
            if emb_p and emb_m and emb_d:
                if emb_p == "lmstudio":
                    _lms_base = (emb_u or "http://localhost:1234").rstrip("/v1").rstrip("/")
                    _lms_emb_id = lmstudio_ensure_model_loaded(_lms_base, emb_m, timeout_s=45, model_type="embeddings")
                    if _lms_emb_id:
                        emb_m = _lms_emb_id
                try:
                    _done_emb, _errs_emb = embed_all_tickets(
                        [fresh], cb_url, bucket, username, password,
                        use_tls, scope, collection,
                        emb_p, emb_m, emb_k, emb_u, emb_d,
                        lambda msg, pct: None,
                    )
                    _pipeline_notes.append(f"Embedded ✓" if _done_emb else "Embed skipped")
                except Exception as _ee:
                    _pipeline_notes.append(f"Embed failed: {_ee}")
            s_prov = (ctx.get("provider") or "").lower().strip()
            s_mod  = ctx.get("model", "")
            s_key  = ctx.get("api_key", "")
            s_url  = ctx.get("base_url", "")
            if s_prov and s_mod:
                try:
                    _scores = score_tickets_batch(
                        [fresh], s_prov, s_mod, s_key, s_url,
                        cb_url, bucket, username, password, use_tls, scope, collection,
                        save_to_cb=True,
                    )
                    _pipeline_notes.append(f"Scored ✓" if _scores else "Score skipped")
                except Exception as _se:
                    _pipeline_notes.append(f"Score failed: {_se}")

        _url = _SUPPORTAL_TICKET_URL.format(ticket_id=ticket_id)
        _pipeline_str = " | " + " | ".join(_pipeline_notes) if _pipeline_notes else ""
        summary_lines = [
            f"**Ticket {ticket_id} re-scraped from Supportal** {'(saved to CB ✓)' if _saved else '(CB save failed ✗)'}{_pipeline_str}",
            f"Status: {fresh.get('status','')} | Priority: {fresh.get('priority','')}",
            f"Subject: {fresh.get('subject','')}",
            f"Created: {fresh.get('created','')} | Closed: {fresh.get('closed','')}",
            f"Scraped at: {time.strftime('%Y-%m-%d %H:%M:%S UTC', time.gmtime())}",
            f"Supportal URL: {_url}",
        ]
        cbses = fresh.get("cbses") or []
        if cbses:
            summary_lines.append(f"CBSEs: {', '.join(cbses)}")
        jiras = fresh.get("jira_issues") or []
        if jiras:
            summary_lines.append(f"Jira Issues: {', '.join(jiras)}")
        topo = fresh.get("snapshot_topology") or {}
        if isinstance(topo, dict) and topo:
            topo_note = []
            if topo.get("cluster_name"):  topo_note.append(f"cluster={topo['cluster_name']}")
            if topo.get("cb_version"):    topo_note.append(f"CB={topo['cb_version']}")
            if topo.get("total_nodes"):   topo_note.append(f"nodes={topo['total_nodes']}")
            bad  = topo.get("bad_items",  topo.get("bad_count",  0)) or 0
            warn = topo.get("warn_items", topo.get("warn_count", 0)) or 0
            if bad or warn:               topo_note.append(f"bad={bad} warn={warn}")
            if topo_note:
                src = " (freshly enriched)" if topo_enriched else " (from prior enrichment)"
                summary_lines.append(f"Topology{src}: {', '.join(topo_note)}")
        result = "\n".join(summary_lines)
        if _remaining:
            result += (
                f"\n\n**Note:** you passed {len(_remaining) + 1} ticket IDs but rescrape_ticket "
                f"processes one at a time. Call rescrape_ticket again for each remaining ID: "
                + ", ".join(str(x) for x in _remaining)
                + ". Or call rescrape_customer_tickets to refresh all stale tickets at once."
            )
        return result

    elif name == "list_supportal_customers":
        sort_by = (args.get("sort_by") or "name").lower()
        limit   = min(int(args.get("limit") or 200), 500)
        cookie  = _get_profile_cookie()  # unused as of v2.6.2; retained for re-enablement
        try:
            order = {"snapshots": "snaps DESC", "tickets": "tickets DESC"}.get(sort_by, "cu_name ASC")
            statement = f"""
SELECT cu.`name` AS cu_name,
       COUNT(DISTINCT META(sn).id) AS snaps,
       COUNT(DISTINCT t_id)        AS tickets
FROM snapshot sn
UNNEST sn.`zendesk` AS t_id
JOIN cluster cl ON META(cl).id = ("Cluster::" || sn.`uuid`)
JOIN customer cu ON META(cu).id = ("Customer::" || cl.`customer`)
GROUP BY cu.`name`
ORDER BY {order}
LIMIT {limit}
""".strip()
            rows = query_supportal_analytics(statement, cookie)
        except Exception as exc:
            return f"Supportal Analytics error: {exc}"
        if not rows:
            return "No customers returned from Supportal Analytics."
        lines = [f"**{len(rows)} customers in Supportal Analytics:**", ""]
        for r in rows:
            name_  = r.get("cu_name", "")
            snaps  = r.get("snaps", 0)
            tix    = r.get("tickets", 0)
            lines.append(f"- **{name_}** — {snaps} snapshot{'s' if snaps != 1 else ''}, {tix} linked ticket{'s' if tix != 1 else ''}")
        return "\n".join(lines)

    elif name == "query_supportal":
        statement  = (args.get("statement") or "").strip()
        limit_rows = min(int(args.get("limit_rows") or 100), 500)
        if not statement:
            return "Error: statement is required."
        cookie = _get_profile_cookie()  # unused as of v2.6.2; retained for re-enablement
        try:
            rows = query_supportal_analytics(statement, cookie)
        except Exception as exc:
            return f"Supportal Analytics query error: {exc}"
        if not rows:
            return "Query returned no results."
        rows = rows[:limit_rows]
        # Format as markdown table
        if not isinstance(rows[0], dict):
            return "\n".join(str(r) for r in rows)
        cols = list(rows[0].keys())
        header = "| " + " | ".join(cols) + " |"
        sep    = "| " + " | ".join("---" for _ in cols) + " |"
        body   = "\n".join(
            "| " + " | ".join(str(r.get(c, "")) for c in cols) + " |"
            for r in rows
        )
        return f"{header}\n{sep}\n{body}\n\n**{len(rows)} row(s) returned**"

    elif name == "list_organizations":
        min_tickets = max(1, int(args.get("min_tickets") or 1))
        if not _CB_AVAILABLE:
            return "Couchbase not available."
        try:
            from couchbase.cluster import Cluster  # type: ignore
            from couchbase.options import ClusterOptions  # type: ignore
            from couchbase.auth import PasswordAuthenticator  # type: ignore
            from couchbase.options import QueryOptions  # type: ignore
            conn_str = _cb_conn_str(cb_url, use_tls)
            cluster = Cluster(conn_str, ClusterOptions(PasswordAuthenticator(username, password)))
            cluster.wait_until_ready(timedelta(seconds=10))
            keyspace = f"`{bucket}`.`{scope}`.`{collection}`"
            rows = list(cluster.query(
                f"SELECT organization, COUNT(*) AS ticket_count "
                f"FROM {keyspace} "
                f"WHERE organization IS NOT MISSING AND organization != '' "
                f"GROUP BY organization "
                f"HAVING COUNT(*) >= {min_tickets} "
                f"ORDER BY organization ASC",
                QueryOptions(timeout=timedelta(seconds=30)),
            ))
            cluster.close()
        except Exception as exc:
            return f"list_organizations failed: {exc}"
        if not rows:
            return "No organizations found in the database."
        lines = [f"**{len(rows)} organizations in Couchbase:**", ""]
        for r in rows:
            lines.append(f"- **{r['organization']}** ({r['ticket_count']} ticket{'s' if r['ticket_count'] != 1 else ''})")
        return "\n".join(lines)

    elif name == "search_customer_names":
        _query = (args.get("query") or "").strip()
        _limit = min(int(args.get("limit") or 10), 20)
        if not _query:
            return "Error: query is required."
        _cookie = ctx.get("cookie") or _get_profile_cookie()
        hits = resolve_customer_name(
            _query,
            cookie=_cookie,
            cb_url=cb_url, cb_bucket=bucket, cb_user=username,
            cb_pass=password, cb_tls=use_tls, cb_scope=scope,
            cb_collection=collection, limit=_limit,
        )
        if not hits:
            return (
                f"No customers found matching '{_query}'. "
                "Try a shorter fragment, check spelling, or use list_supportal_customers to browse all."
            )
        lines = [f"**{len(hits)} customer(s) matching '{_query}':**\n"]
        for i, h in enumerate(hits, 1):
            lines.append(f"{i}. **{h['display_name']}** *(source: {h['source']})*")
        lines.append(
            "\nAsk the user which one they mean, then use that exact name as "
            "`customer=` in all subsequent tool calls for this conversation."
        )
        return "\n".join(lines)

    elif name == "get_briefing":
        _top_n    = min(int(args.get("top_n") or 5), 10)
        _pu       = ctx.get("profile_user", "default")
        _stale_h  = 12.0
        try:
            _profile = _load_customer_profile(cb_url, bucket, username, password, use_tls, _pu)
            _stale_h = float((_profile.get("alert_thresholds") or {}).get("stale_hours") or 12)
            _top = [c for c in (_profile.get("top_customers") or [])
                    if c.get("is_valid", True)][:_top_n]
        except Exception:
            _top = []
        if not _top:
            return (
                "No customer profile found yet — I don't know which accounts you care about. "
                "Start by asking about specific customers; I'll learn your top accounts over time. "
                "Or use `list_organizations` to see what's in the local database."
            )
        _snap = _get_briefing_data(
            _top, cb_url, bucket, username, password,
            use_tls, scope, collection, stale_hours=_stale_h,
        )
        alerts = [r for r in _snap if r["alert"]]
        ok     = [r for r in _snap if not r["alert"]]
        lines  = [f"## Account Briefing — {len(_snap)} accounts\n"]
        if alerts:
            lines.append("### ⚠ Needs Attention")
            for r in alerts:
                flags = []
                if r["open_p1"] > 0:
                    flags.append(f"{r['open_p1']} open P1")
                if r["score"] < 40:
                    flags.append(f"health {r['score']}/100")
                if r["hours_stale"] > _stale_h:
                    flags.append(f"data {r['hours_stale']}h old")
                lines.append(f"- **{r['name']}** — {', '.join(flags)}")
        if ok:
            lines.append("\n### ✅ Looking Good")
            for r in ok:
                lines.append(
                    f"- **{r['name']}** — score {r['score']} ({r['grade']})"
                    + (f", {r['open_p2']} open P2" if r["open_p2"] > 0 else "")
                    + (f", data {r['hours_stale']}h old" if r["hours_stale"] > 0 else "")
                )
        lines.append(
            f"\n*Briefing covers your top {len(_snap)} accounts by recent activity. "
            "Ask me to drill into any account for details.*"
        )
        return "\n".join(lines)

    elif name == "generate_chart":
        opt = _build_agent_echart_option(args)
        # Return a fenced ```echart block — _render_chat renders it as a live ui.echart
        return "```echart\n" + json.dumps(opt, ensure_ascii=False) + "\n```"

    elif name == "generate_table":
        title       = args.get("title") or "Table"
        columns     = args.get("columns") or []
        rows        = args.get("rows") or []
        description = args.get("description") or ""
        payload = {"title": title, "columns": columns, "rows": rows}
        if description:
            payload["description"] = description
        return "```table\n" + json.dumps(payload, ensure_ascii=False) + "\n```"

    elif name == "vector_search":
        query = (args.get("query") or "").strip()
        if not query:
            return "Error: query is required."
        limit = min(int(args.get("limit") or 10), 30)
        emb_p   = ctx.get("emb_provider", "ollama")
        emb_m   = ctx.get("emb_model", "nomic-embed-text")
        emb_k   = ctx.get("emb_api_key", "")
        emb_u   = ctx.get("emb_base_url", "http://localhost:11434")
        emb_d   = int(ctx.get("emb_dims") or 1024)
        try:
            qvec = embed_text(query, emb_p, emb_m, emb_k, emb_u, emb_d)
        except Exception as exc:
            return f"Embedding failed — check embedding provider config: {exc}"
        try:
            keys = vector_search_cb(qvec, cb_url, bucket, username, password, use_tls, scope, collection, top_k=limit)
        except Exception as exc:
            return f"Vector search failed: {exc}"
        if not keys:
            return "No semantically similar tickets found."
        ticket_keys = [k for k in keys if k.startswith("ticket::")]
        tickets = fetch_tickets_by_keys(ticket_keys, cb_url, bucket, username, password, use_tls, scope, collection)
        if not tickets:
            return "Vector search returned results but could not fetch ticket documents."
        lines = [
            f"**Semantic search results for:** *{query}*\n",
            "| Ticket ID | Organization | Subject | Status | Priority | CB Version |",
            "|-----------|-------------|---------|--------|----------|------------|",
        ]
        for t in tickets:
            topo = t.get("snapshot_topology") or {}
            cbv = topo.get("cb_version") or ""
            subj = (t.get("subject") or "")[:55].replace("|", "/")
            lines.append(
                f"| {t.get('ticket_id','')} | {t.get('organization','')} | {subj} "
                f"| {t.get('status','')} | {t.get('priority','')} | {cbv} |"
            )
        return "\n".join(lines) + f"\n\n**{len(tickets)} results**"

    elif name == "get_cluster_health":
        org = (args.get("organization") or default_customer or "").strip()
        if not org:
            return "Error: organization is required."
        # Load snapshots for this org from CB
        try:
            from couchbase.cluster import Cluster as _Cl
            from couchbase.options import ClusterOptions as _CO
            from couchbase.auth import PasswordAuthenticator as _PA
            from couchbase.options import QueryOptions as _QO
            _conn = _cb_conn_str(cb_url, use_tls)
            _cluster = _Cl(_conn, _CO(_PA(username, password)))
            _cluster.wait_until_ready(timedelta(seconds=15))
            _snap_ks = f"`{bucket}`.`{scope}`.`snapshots`"
            _rows = list(_cluster.query(
                f"SELECT s.* FROM {_snap_ks} AS s "
                f"WHERE LOWER(s.organization) LIKE $org "
                f"ORDER BY s.date DESC LIMIT 500",
                _QO(named_parameters={"org": f"%{org.lower()}%"}, timeout=timedelta(seconds=30)),
            ))
            _cluster.close()
        except Exception as exc:
            return f"Failed to load snapshots from Couchbase: {exc}"
        if not _rows:
            # BEFORE v1.5.0: returned a "use fetch_snapshots first" message and stopped.
            # AFTER v1.5.0: auto-triggers sync_snapshots if a cookie is available, then re-queries.
            _auto_cookie = ctx.get("cookie") or _get_profile_cookie()  # unused as of v2.6.2
            _auto_prefix = f"No snapshots in Couchbase for '{org}' — auto-syncing now.\n\n"
            _sync_result = _execute_agent_tool(
                "sync_snapshots", {"organization": org, "max_stubs": 10},
                cb_url, bucket, username, password, use_tls, scope, collection,
                default_customer=default_customer, ctx=ctx,
            )
            _auto_prefix += _sync_result + "\n\n"
            # Re-query after sync
            try:
                from couchbase.cluster import Cluster as _Cl2
                from couchbase.options import ClusterOptions as _CO2
                from couchbase.auth import PasswordAuthenticator as _PA2
                from couchbase.options import QueryOptions as _QO2
                _conn2 = _cb_conn_str(cb_url, use_tls)
                _cl2 = _Cl2(_conn2, _CO2(_PA2(username, password)))
                _cl2.wait_until_ready(timedelta(seconds=15))
                _rows = list(_cl2.query(
                    f"SELECT s.* FROM `{bucket}`.`{scope}`.`snapshots` AS s "
                    f"WHERE LOWER(s.organization) LIKE $org ORDER BY s.date DESC LIMIT 500",
                    _QO2(named_parameters={"org": f"%{org.lower()}%"}, timeout=timedelta(seconds=30)),
                ))
                _cl2.close()
            except Exception:
                pass
            if not _rows:
                return _auto_prefix + f"Still no snapshot data for '{org}' after sync. The customer may not exist in Supportal or has no snapshots."
        else:
            _auto_prefix = ""
        tickets = tool_query_tickets(
            {"organization": org}, cb_url, bucket, username, password, use_tls, scope, collection, limit=500,
        )
        health = build_cluster_health_data(_rows, tickets)
        ci = health.get("cluster_index") or {}
        if not ci:
            return f"Snapshot data found ({len(_rows)} snapshots) but cluster index could not be built."
        lines = [f"**Cluster health for {org}** — {len(_rows)} snapshots, {len(ci)} clusters\n"]
        lines.append("| Cluster | CB Version | Nodes | CPUs/node | RAM/node | Last Seen | Bad | Warn | Status |")
        lines.append("|---------|------------|-------|-----------|----------|-----------|-----|------|--------|")
        for cid, c in sorted(ci.items(), key=lambda x: x[1].get("last_seen") or "", reverse=True):
            name  = c.get("cluster_name") or cid[:12]
            ver   = (c.get("version_history") or [""])[-1] or "unknown"
            nodes = c.get("node_count_last") or c.get("node_count") or "?"
            cpus  = c.get("cpus_per_node") or "?"
            ram   = f"{c['ram_per_node_mib']} MiB" if c.get("ram_per_node_mib") else "?"
            last  = (c.get("last_seen") or "")[:10]
            bad   = c.get("avg_bad", 0)
            warn  = c.get("avg_warn", 0)
            status = "Deprecated" if c.get("is_deprecated") else ("Active" if c.get("is_active") else "Inactive")
            lines.append(f"| {name} | {ver} | {nodes} | {cpus} | {ram} | {last} | {bad} | {warn} | {status} |")
        return _auto_prefix + "\n".join(lines)

    elif name == "cluster_hw_chart":
        _org = (args.get("organization") or default_customer or "").strip()
        if not _org:
            return "Error: organization is required for cluster_hw_chart."
        _sf = (args.get("status_filter") or "open_or_pending").lower()
        _height = int(args.get("height") or 0)
        try:
            from couchbase.cluster import Cluster as _Cl
            from couchbase.options import ClusterOptions as _CO, QueryOptions as _QO
            from couchbase.auth import PasswordAuthenticator as _PA
            _conn = _cb_conn_str(cb_url, use_tls)
            _cl2 = _Cl(_conn, _CO(_PA(username, password)))
            _cl2.wait_until_ready(timedelta(seconds=15))

            _status_vals = {
                "open":            ["open", "new"],
                "pending":         ["pending"],
                "open_or_pending": ["open", "new", "pending"],
                "all":             [],
            }.get(_sf, ["open", "new", "pending"])
            _status_clause = (
                "AND LOWER(t.status) IN [" + ",".join(f'"{s}"' for s in _status_vals) + "] "
                if _status_vals else ""
            )

            # Hardware is embedded directly in snapshot_topology on each ticket.
            # Pull cluster name, hw specs, and CBSEs in one query.
            _tkt_sql = (
                f"SELECT t.ticket_id, t.cbses, "
                f"t.snapshot_topology.cluster_name, "
                f"t.snapshot_topology.total_nodes, "
                f"t.snapshot_topology.cpus_per_node, "
                f"t.snapshot_topology.ram_per_node_mib "
                f"FROM `{bucket}`.`{scope}`.`{collection}` t "
                f"WHERE t.type='ticket' AND LOWER(t.organization) LIKE $org "
                f"{_status_clause}"
                f"AND t.snapshot_topology IS NOT NULL "
                f"AND t.snapshot_topology.cluster_name IS NOT NULL "
                f"LIMIT 500"
            )
            _tkt_rows = list(_cl2.query(_tkt_sql, _QO(
                named_parameters={"org": f"%{_org.lower()}%"}, timeout=timedelta(seconds=30)
            )))
            _cl2.close()

            if not _tkt_rows:
                # Diagnose: count tickets without topology
                return (
                    f"No {_sf} tickets with embedded snapshot topology found for '{_org}'. "
                    "Tickets may be missing snapshot data — ask to rescrape or sync_snapshots."
                )

            # Deduplicate by cluster_name; collect CBSEs per cluster
            _seen: set = set()
            _hw: list[dict] = []
            _cluster_cbses: dict[str, set] = {}
            for _r in _tkt_rows:
                _cn = (_r.get("cluster_name") or "").strip()
                _cbses = [c for c in (_r.get("cbses") or []) if c]
                if _cn:
                    _cluster_cbses.setdefault(_cn, set()).update(_cbses)
                if not _cn or _cn in _seen:
                    continue
                _seen.add(_cn)
                _hw.append({
                    "cluster": _cn,
                    "nodes":   int(_r.get("total_nodes") or 0),
                    "cpus":    int(_r.get("cpus_per_node") or 0),
                    "ram_gib": round((_r.get("ram_per_node_mib") or 0) / 1024, 1),
                })

            # Attach de-duped CBSEs
            for _h in _hw:
                _h["cbses"] = sorted(_cluster_cbses.get(_h["cluster"], set()))

            if not _hw:
                return f"Tickets found but no cluster topology could be extracted for '{_org}'."

            # ── 3. Build chart ────────────────────────────────────────────────────
            # Label: cluster name + ● if has CBSEs
            _labels = []
            for _h in _hw:
                _cbse_list = _h["cbses"]
                _lbl = _h["cluster"]
                if _cbse_list:
                    # Normalise IDs and append indicator
                    _ids = [c if c.upper().startswith("CBSE-") else f"CBSE-{c}" for c in _cbse_list[:3]]
                    _lbl += f" ● ({', '.join(_ids)})"
                _labels.append(_lbl)

            _auto_h = max(300, min(700, 80 + len(_hw) * 26))
            _opt = _build_agent_echart_option({
                "chart_type":   "horizontal_bar",
                "title":        f"Cluster Hardware — Open/Pending Tickets ({_org})",
                "labels":       _labels,
                "series": [
                    {"name": "Nodes",      "data": [_h["nodes"]   for _h in _hw]},
                    {"name": "CPUs/node",  "data": [_h["cpus"]    for _h in _hw]},
                    {"name": "RAM GiB/node", "data": [_h["ram_gib"] for _h in _hw]},
                ],
                "height":       _height or _auto_h,
                "description":  "● = cluster has one or more CBSE-linked tickets. "
                                "Values: physical node count, CPU cores per node, RAM per node in GiB.",
                "color_scheme": "couchbase",
            })
            return "```echart\n" + json.dumps(_opt, ensure_ascii=False) + "\n```"

        except Exception as exc:
            import traceback as _tb2; _tb2.print_exc()
            return f"cluster_hw_chart error: {exc}"

    elif name == "query_local_snapshots":
        org   = (args.get("organization") or "").strip()
        days  = int(args.get("days") or 30)
        limit = min(int(args.get("limit") or 50), 500)
        try:
            import time as _time
            from couchbase.cluster import Cluster as _Cl
            from couchbase.options import ClusterOptions as _CO, QueryOptions as _QO
            from couchbase.auth import PasswordAuthenticator as _PA
            _conn = _cb_conn_str(cb_url, use_tls)
            _cluster = _Cl(_conn, _CO(_PA(username, password)))
            _cluster.wait_until_ready(timedelta(seconds=15))
            _snap_ks = f"`{bucket}`.`{scope}`.`snapshots`"
            _cutoff  = _time.time() - days * 86400
            _org_clause = "AND LOWER(s.organization) LIKE $org " if org else ""
            _sql = (
                f"SELECT s.organization, s.cluster_name, s.cb_version, "
                f"s.node_count, s.cpus_per_node, s.ram_per_node_mib, "
                f"s.topology.disk_total_per_node_mib, s.topology.disk_used_per_node_mib, "
                f"s.topology.ram_used_per_node_mib, "
                f"s.bad_count, s.warn_count, s.date, s.last_scraped_at "
                f"FROM {_snap_ks} AS s "
                f"WHERE s.last_scraped_at >= $cutoff "
                f"{_org_clause}"
                f"ORDER BY s.last_scraped_at DESC LIMIT {limit}"
            )
            _params: dict = {"cutoff": _cutoff}
            if org:
                _params["org"] = f"%{org.lower()}%"
            _rows = list(_cluster.query(_sql, _QO(named_parameters=_params, timeout=timedelta(seconds=30))))
            _cluster.close()
        except Exception as exc:
            return f"query_local_snapshots failed: {exc}"
        if not _rows:
            _org_hint = f" for '{org}'" if org else ""
            return f"No snapshots found{_org_hint} in the last {days} days in local Couchbase."
        _header = (
            f"**{len(_rows)} snapshots{(' for ' + org) if org else ''} — last {days} days (local CB)**\n\n"
            "| Organization | Cluster | CB Version | Nodes | CPUs/node | RAM MiB | RAM Used MiB | Disk Total MiB | Disk Used MiB | Bad | Warn | Scraped |\n"
            "|---|---|---|---|---|---|---|---|---|---|---|---|"
        )
        _lines = [_header]
        import datetime as _dt2
        for r in _rows:
            _scraped = r.get("last_scraped_at")
            _scraped_str = _dt2.datetime.fromtimestamp(_scraped).strftime("%Y-%m-%d") if _scraped else "?"
            _lines.append(
                f"| {r.get('organization','?')} | {r.get('cluster_name','?')} | {r.get('cb_version','?')} "
                f"| {r.get('node_count','?')} | {r.get('cpus_per_node','?')} | {r.get('ram_per_node_mib','?')} "
                f"| {r.get('ram_used_per_node_mib','?')} | {r.get('disk_total_per_node_mib','?')} "
                f"| {r.get('disk_used_per_node_mib','?')} "
                f"| {r.get('bad_count','?')} | {r.get('warn_count','?')} | {_scraped_str} |"
            )
        return "\n".join(_lines)

    elif name == "analyze_snapshot":
        snap_id  = (args.get("snap_id") or "").strip()
        notes    = (args.get("analysis_notes") or "").strip()
        save     = bool(args.get("save_notes", False))
        if not snap_id:
            return "Error: snap_id is required."
        try:
            topo = fetch_snapshot_topology(snap_id, cookie=None)
        except Exception as exc:
            return f"Failed to fetch snapshot {snap_id}: {exc}"
        if not topo:
            return f"Snapshot {snap_id} returned no topology data."

        bad  = topo.get("bad_items")  or []
        warn = topo.get("warn_items") or []
        lines = [
            f"**Snapshot {snap_id}**",
            f"Cluster: {topo.get('cluster_name','')} | Version: {topo.get('cb_version','')}",
            f"Nodes: {topo.get('total_nodes','')} | CPUs/node: {topo.get('cpus_per_node','')} | RAM/node: {topo.get('ram_per_node_mib','')} MiB",
            f"Bad items ({len(bad)}): {', '.join(bad) if bad else 'none'}",
            f"Warn items ({len(warn)}): {', '.join(warn) if warn else 'none'}",
        ]
        buckets = topo.get("buckets") or []
        if buckets:
            lines.append(f"Buckets: {', '.join(b.get('name','') for b in buckets if b.get('name'))}")
        if notes:
            lines.append(f"\n**Analysis notes:** {notes}")

        # Save notes + topology back to the snapshot doc in CB if requested
        if save and snap_id:
            try:
                from couchbase.cluster import Cluster as _AsCl
                from couchbase.options import ClusterOptions as _AsCO
                from couchbase.auth import PasswordAuthenticator as _AsPA
                _as_cl = _AsCl(_cb_conn_str(cb_url, use_tls), _AsCO(_AsPA(username, password)))
                _as_cl.wait_until_ready(timedelta(seconds=10))
                _as_snap_col = _as_cl.bucket(bucket).scope(scope).collection("snapshots")
                _snap_key = f"snapshot::{snap_id}"
                try:
                    _snap_doc = _as_snap_col.get(_snap_key).content_as[dict]
                except Exception:
                    _snap_doc = {"snap_id": snap_id, "type": "snapshot"}
                _snap_doc["topology"]        = topo
                _snap_doc["analysis_notes"]  = notes
                _snap_doc["analyzed_at"]     = int(time.time())
                _as_snap_col.upsert(_snap_key, _snap_doc)
                _as_cl.close()
                lines.append("_(Analysis notes saved to snapshot record.)_")
            except Exception as exc:
                lines.append(f"_(Note: could not save to CB — {exc})_")

        return "\n".join(lines)

    elif name == "fetch_snapshots":
        org   = (args.get("organization") or default_customer or "").strip()
        limit = min(int(args.get("limit") or 100), 500)
        if not org:
            return "Error: organization is required."
        cookie = ctx.get("cookie") or _get_profile_cookie()  # unused as of v2.6.2; retained for re-enablement
        try:
            stubs = fetch_snapshots_via_analytics(org, cookie, limit=limit)
        except Exception as exc:
            return f"fetch_snapshots failed: {exc}"
        if not stubs:
            return f"No snapshots found for '{org}' in the Analytics API."
        # Save stubs to CB
        _saved = 0
        try:
            from couchbase.cluster import Cluster as _Cl
            from couchbase.options import ClusterOptions as _CO
            from couchbase.auth import PasswordAuthenticator as _PA
            _conn = _cb_conn_str(cb_url, use_tls)
            _cluster = _Cl(_conn, _CO(_PA(username, password)))
            _cluster.wait_until_ready(timedelta(seconds=15))
            _col = _cluster.bucket(bucket).scope(scope).collection("snapshots")
            for s in stubs:
                try:
                    _col.upsert(f"snapshot::{s['snap_id']}", s)
                    _saved += 1
                except Exception:
                    pass
            _cluster.close()
        except Exception as exc:
            return (
                f"Fetched {len(stubs)} snapshot stubs from Analytics API but CB save failed: {exc}. "
                "Check your Couchbase connection."
            )
        return (
            f"Fetched and saved **{_saved} snapshot stubs** for '{org}'. "
            f"Topology detail is not yet loaded — call backfill_snapshot_topology to enrich them."
        )

    elif name == "backfill_snapshot_topology":
        org       = (args.get("organization") or default_customer or "").strip()
        max_stubs = min(int(args.get("max_stubs") or 10), 25)
        if not org:
            return "Error: organization is required."
        cookie = ctx.get("cookie") or _get_profile_cookie()
        if not cookie:
            return "No session cookie available — paste a cookie in the Configuration tab first."
        # Load incomplete stubs from CB
        try:
            from couchbase.cluster import Cluster as _Cl
            from couchbase.options import ClusterOptions as _CO
            from couchbase.auth import PasswordAuthenticator as _PA
            from couchbase.options import QueryOptions as _QO
            _conn = _cb_conn_str(cb_url, use_tls)
            _cluster = _Cl(_conn, _CO(_PA(username, password)))
            _cluster.wait_until_ready(timedelta(seconds=15))
            _snap_ks = f"`{bucket}`.`{scope}`.`snapshots`"
            _incomplete = list(_cluster.query(
                f"SELECT s.* FROM {_snap_ks} AS s "
                f"WHERE LOWER(s.organization) LIKE $org "
                f"AND (s.cb_version IS MISSING OR s.cb_version = '' OR s.cb_version IS NULL) "
                f"ORDER BY s.date DESC LIMIT {max_stubs}",
                _QO(named_parameters={"org": f"%{org.lower()}%"}, timeout=timedelta(seconds=30)),
            ))
            _cluster.close()
        except Exception as exc:
            return f"Failed to load incomplete snapshots from CB: {exc}"
        if not _incomplete:
            return f"No incomplete snapshot stubs found for '{org}' — topology already looks complete, or no snapshots exist. Try fetch_snapshots first."
        def _noop_prog(msg, pct): pass
        try:
            enriched = scrape_snapshots_from_stubs(_incomplete, cookie, max_detail_workers=4, progress_cb=_noop_prog)
        except Exception as exc:
            return f"Topology backfill failed: {exc}"
        # Save enriched docs back to CB
        _saved = 0
        try:
            from couchbase.cluster import Cluster as _Cl2
            from couchbase.options import ClusterOptions as _CO2
            from couchbase.auth import PasswordAuthenticator as _PA2
            _conn2 = _cb_conn_str(cb_url, use_tls)
            _cl2 = _Cl2(_conn2, _CO2(_PA2(username, password)))
            _cl2.wait_until_ready(timedelta(seconds=15))
            _col2 = _cl2.bucket(bucket).scope(scope).collection("snapshots")
            for s in enriched:
                try:
                    _col2.upsert(f"snapshot::{s['snap_id']}", s)
                    _saved += 1
                except Exception:
                    pass
            _cl2.close()
        except Exception as exc:
            return f"Enriched {len(enriched)} stubs but CB save failed: {exc}"
        bad_total  = sum(s.get("bad_count", 0) for s in enriched)
        warn_total = sum(s.get("warn_count", 0) for s in enriched)
        return (
            f"Topology backfill complete — **{_saved}/{len(_incomplete)} stubs enriched**. "
            f"Total bad items: {bad_total}, warn items: {warn_total}. "
            f"Call get_cluster_health to see the full summary."
        )

    elif name == "get_scrape_status":
        _filter_jid = (args.get("job_id") or "").strip().lower()
        if not _SCRAPE_JOBS:
            return "No scrape jobs have been started in this session."
        jobs_to_show = (
            [_SCRAPE_JOBS[_filter_jid]] if _filter_jid and _filter_jid in _SCRAPE_JOBS
            else list(reversed(list(_SCRAPE_JOBS.values())))[:10]
        )
        if _filter_jid and _filter_jid not in _SCRAPE_JOBS:
            return f"Job '{_filter_jid}' not found. Recent job IDs: {', '.join(list(_SCRAPE_JOBS)[-5:])}."
        _now = time.time()
        lines: list[str] = []
        running = [j for j in jobs_to_show if j["status"] == "running"]
        done    = [j for j in jobs_to_show if j["status"] != "running"]
        if running:
            lines.append("## Active Jobs\n")
            for j in running:
                elapsed = int(_now - j["started_at"])
                proc    = j.get("processed") or 0
                total   = j.get("total")
                pct_str = f" ({proc/total:.0%})" if total else ""
                lines.append(
                    f"**Job {j['job_id']}** — {j['org']} ({j['mode']}) — **RUNNING**\n"
                    f"  Phase: {j['phase'] or '—'}\n"
                    f"  Progress: {proc}/{total or '?'} tickets{pct_str}\n"
                    f"  Elapsed: {elapsed}s\n"
                    f"  Last: {j['last_message']}\n"
                )
        if done:
            lines.append("## Recent Jobs\n")
            for j in done:
                duration = int((j.get("finished_at") or _now) - j["started_at"])
                ago      = int(_now - (j.get("finished_at") or _now))
                icon     = "✅" if j["status"] == "done" else "❌"
                lines.append(
                    f"**Job {j['job_id']}** — {j['org']} ({j['mode']}) — {icon} {j['status'].upper()}\n"
                    f"  {j['last_message']}\n"
                    f"  Duration: {duration}s | Finished: {ago}s ago\n"
                )
        return "\n".join(lines).strip()

    elif name == "cancel_scrape_job":
        _cjid = (args.get("job_id") or "").strip().lower()
        if not _cjid:
            return "Error: job_id is required."
        # Signal the running thread to stop cleanly
        _cev = _JOB_CANCEL_EVENTS.get(_cjid)
        if _cev:
            _cev.set()
        # Update in-memory record immediately so the monitor sees it
        _cjob = _SCRAPE_JOBS.get(_cjid)
        if _cjob:
            if _cjob.get("status") == "running":
                _cjob["status"]      = "cancelled"
                _cjob["phase"]       = None
                _cjob["finished_at"] = time.time()
                _cjob["last_message"] = (
                    f"Cancelled by user at ticket "
                    f"{_cjob.get('processed',0)}/{_cjob.get('total','?')}. "
                    f"To resume: rescrape with stale_hours=1 — the "
                    f"{_cjob.get('saved',0)} already-refreshed tickets will be skipped."
                )
                # Write cancelled state to CB so it persists across restarts
                _persist_job_state(
                    _cjob, cb_url, bucket, username, password, use_tls, scope, collection
                )
                proc = _cjob.get("processed", 0)
                total_ = _cjob.get("total", "?")
                saved_ = _cjob.get("saved", 0)
                return (
                    f"Job **{_cjid}** cancelled at {proc}/{total_} tickets ({saved_} saved). "
                    f"To resume from where it stopped, run: "
                    f"*rescrape {_cjob.get('org','')} with stale_hours=1* — "
                    f"already-refreshed tickets have fresh timestamps and will be skipped automatically."
                )
            else:
                return f"Job **{_cjid}** is not running (status: {_cjob['status']})."
        # Job not in memory — look it up in CB and mark cancelled there
        if cb_url and username:
            try:
                from couchbase.cluster import Cluster as _CKCl
                from couchbase.options import ClusterOptions as _CKCO
                from couchbase.auth import PasswordAuthenticator as _CKPA
                _ckconn = _cb_conn_str(cb_url, use_tls)
                _ckc    = _CKCl(_ckconn, _CKCO(_CKPA(username, password)))
                _ckc.wait_until_ready(timedelta(seconds=5))
                _ckcol  = _ckc.bucket(bucket).scope(scope).collection(collection)
                _ckdoc  = _ckcol.get(f"scrape_job::{_cjid}").content_as[dict]
                if _ckdoc.get("status") == "running":
                    _ckdoc["status"]      = "cancelled"
                    _ckdoc["phase"]       = None
                    _ckdoc["finished_at"] = time.time()
                    _ckdoc["last_message"] = "Cancelled by user (post-restart)."
                    _ckcol.upsert(f"scrape_job::{_cjid}", {**_ckdoc, "type": "scrape_job"})
                    _ckc.close()
                    return f"Job **{_cjid}** marked cancelled in Couchbase (thread was already dead after restart)."
                _ckc.close()
                return f"Job **{_cjid}** status in CB: {_ckdoc.get('status')} — nothing to cancel."
            except Exception as _cke:
                return f"Job **{_cjid}** not found in memory or Couchbase: {_cke}"
        return f"Job '{_cjid}' not found in this session."

    elif name == "backfill_last_comment_at":
        _bf_org = (args.get("organization") or "").strip()
        def _noop_prog(msg, pct): pass
        try:
            _bf_updated, _bf_errs = backfill_last_comment_at(
                cb_url, bucket, username, password, use_tls, scope, collection,
                _noop_prog, org_filter=_bf_org,
            )
        except Exception as exc:
            return f"Backfill failed: {exc}"
        scope_desc = f" for '{_bf_org}'" if _bf_org else " across all tickets"
        return (
            f"Backfill complete{scope_desc} — **{_bf_updated} tickets updated**, {_bf_errs} errors. "
            f"The 'Last Reply' column in query_tickets will now show conversation timestamps."
        )

    elif name == "scrape_customer_tickets":
        org         = (args.get("organization") or default_customer or "").strip()
        max_tickets = min(int(args.get("max_tickets") or 25), 50)
        if not org:
            return "Error: organization is required."
        cookie = ctx.get("cookie") or _get_profile_cookie()
        if not cookie:
            return "No session cookie available — paste a cookie in the Configuration tab first."
        _job = _make_scrape_job(org, "scrape")
        threading.Thread(
            target=_run_scrape_job_bg,
            args=(
                _job, org, cookie, max_tickets,
                {"url": cb_url, "bucket": bucket, "username": username, "password": password,
                 "use_tls": use_tls, "scope": scope, "collection": collection},
                {"provider": ctx.get("emb_provider",""), "model": ctx.get("emb_model",""),
                 "api_key": ctx.get("emb_api_key",""), "base_url": ctx.get("emb_base_url",""),
                 "dims": ctx.get("emb_dims", 0),
                 "embed_parallel": ctx.get("embed_parallel", 1)},
                {"provider": ctx.get("provider",""), "model": ctx.get("model",""),
                 "api_key": ctx.get("api_key",""), "base_url": ctx.get("base_url","")},
            ),
            daemon=True,
        ).start()
        if ctx is not None:
            ctx.setdefault("_started_jobs", []).append(_job["job_id"])
        return (
            f"Started scrape job **{_job['job_id']}** for '{org}' (up to {max_tickets} tickets). "
            f"The pipeline runs in the background — scraping, saving, embedding, and scoring. "
            f"**I cannot notify you when it finishes — you must ask me.** "
            f"Ask 'what is the scrape status?' after a minute or two to check progress."
        )

    elif name == "score_ticket":
        ticket_id = str(args.get("ticket_id") or "").strip()
        if not ticket_id:
            return "Error: ticket_id is required."
        tickets = fetch_tickets_by_keys(
            [f"ticket::{ticket_id}"], cb_url, bucket, username, password, use_tls, scope, collection,
        )
        if not tickets:
            return f"Ticket {ticket_id} not found in Couchbase."
        t = tickets[0]
        provider = ctx.get("provider", "claude")
        model    = ctx.get("model", "claude-sonnet-4-6")
        api_key  = ctx.get("api_key", "")
        base_url = ctx.get("base_url", "")
        try:
            scored = score_tickets_batch([t], provider, model, api_key, base_url)
        except Exception as exc:
            return f"Scoring failed: {exc}"
        if not scored:
            return f"Scoring returned no results for ticket {ticket_id}."
        s = scored[0]
        # Save scores back to CB
        try:
            from couchbase.cluster import Cluster as _Cl
            from couchbase.options import ClusterOptions as _CO
            from couchbase.auth import PasswordAuthenticator as _PA
            _conn = _cb_conn_str(cb_url, use_tls)
            _cluster = _Cl(_conn, _CO(_PA(username, password)))
            _cluster.wait_until_ready(timedelta(seconds=15))
            _col = _cluster.bucket(bucket).scope(scope).collection(collection)
            try:
                existing = _col.get(f"ticket::{ticket_id}").content_as[dict]
                existing.update({k: v for k, v in s.items() if v is not None})
                _col.upsert(f"ticket::{ticket_id}", existing)
            except Exception:
                pass
            _cluster.close()
        except Exception:
            pass
        lines = [f"**Score results for ticket {ticket_id}**\n"]
        for field in ("stars", "temperature", "complexity", "resolution_quality",
                      "response_timeliness", "communication_clarity"):
            val = s.get(field)
            if val is not None:
                lines.append(f"- **{field}**: {val}")
        summary = s.get("interaction_summary") or s.get("summary") or ""
        if summary:
            lines.append(f"\n**Summary:** {summary[:500]}")
        return "\n".join(lines)

    # ── v1.5.0 new tools ──────────────────────────────────────────────────────

    elif name == "sync_snapshots":
        # BEFORE v1.5.0: agent had to call fetch_snapshots then backfill_snapshot_topology
        #   separately — often only completing step 1.
        # AFTER v1.5.0: single tool that does both atomically.
        org       = (args.get("organization") or default_customer or "").strip()
        max_stubs = min(int(args.get("max_stubs") or 10), 25)
        if not org:
            return "Error: organization is required."
        fetch_r = _execute_agent_tool(
            "fetch_snapshots", {"organization": org, "limit": 200},
            cb_url, bucket, username, password, use_tls, scope, collection,
            default_customer=default_customer, ctx=ctx,
        )
        backfill_r = _execute_agent_tool(
            "backfill_snapshot_topology", {"organization": org, "max_stubs": max_stubs},
            cb_url, bucket, username, password, use_tls, scope, collection,
            default_customer=default_customer, ctx=ctx,
        )
        return (
            f"**Snapshot sync complete for '{org}'**\n\n"
            f"**Step 1 — Stub fetch:** {fetch_r}\n\n"
            f"**Step 2 — Topology backfill:** {backfill_r}"
        )

    elif name == "batch_score_tickets":
        # BEFORE v1.5.0: score_ticket processed one ticket per call, burning 5-turn limit.
        # AFTER v1.5.0: scores up to 10 tickets per call, returns a summary table.
        raw_ids     = args.get("ticket_ids") or []
        org         = (args.get("organization") or default_customer or "").strip()
        limit       = min(int(args.get("limit") or 10), 50)
        unscored_only = bool(args.get("unscored_only", True))
        status_filt = (args.get("status") or "").strip().lower() or None
        provider = ctx.get("provider", "claude")
        model    = ctx.get("model", "claude-sonnet-4-6")
        api_key  = ctx.get("api_key", "")
        base_url = ctx.get("base_url", "")

        ticket_ids = [str(t).strip() for t in raw_ids if t]
        if not ticket_ids:
            if not org:
                return "Error: provide ticket_ids list or organization to run batch scoring."
            _filt: dict = {"organization": org}
            if status_filt:
                _filt["status"] = status_filt
            candidates = tool_query_tickets(
                _filt, cb_url, bucket, username, password, use_tls, scope, collection, limit=limit * 6,
            )
            if unscored_only:
                candidates = [t for t in candidates if not (t.get("score") or {}).get("stars")]
            ticket_ids = [str(t.get("ticket_id")) for t in candidates[:limit] if t.get("ticket_id")]

        ticket_ids = ticket_ids[:limit]
        if not ticket_ids:
            return f"No tickets to score{' (all already scored)' if unscored_only else ''}."

        tickets = fetch_tickets_by_keys(
            [f"ticket::{tid}" for tid in ticket_ids],
            cb_url, bucket, username, password, use_tls, scope, collection,
        )
        if not tickets:
            return f"Could not fetch tickets from Couchbase."

        try:
            scored = score_tickets_batch(tickets, provider, model, api_key, base_url)
        except Exception as exc:
            return f"Scoring failed: {exc}"
        if not scored:
            return "Scoring returned no results."

        _saved = 0
        try:
            from couchbase.cluster import Cluster as _ScCl
            from couchbase.options import ClusterOptions as _ScCO
            from couchbase.auth import PasswordAuthenticator as _ScPA
            _sc_conn = _cb_conn_str(cb_url, use_tls)
            _sc_cl = _ScCl(_sc_conn, _ScCO(_ScPA(username, password)))
            _sc_cl.wait_until_ready(timedelta(seconds=15))
            _sc_col = _sc_cl.bucket(bucket).scope(scope).collection(collection)
            for s in scored:
                _stid = str(s.get("ticket_id") or "").strip()
                if not _stid:
                    continue
                try:
                    _ex = _sc_col.get(f"ticket::{_stid}").content_as[dict]
                    # Score fields belong nested under doc["score"] (like every
                    # other scoring path) — a flat update leaves reports reading
                    # the stale nested score while new values sit top-level.
                    _sc_data = {k: v for k, v in s.items() if v is not None and k != "ticket_id"}
                    _ex["score"] = {
                        **(_ex.get("score") or {}),
                        **_sc_data,
                        "scored_at": int(time.time()),
                    }
                    _sc_col.upsert(f"ticket::{_stid}", _ex)
                    _saved += 1
                except Exception:
                    pass
            _sc_cl.close()
        except Exception:
            pass

        lines = [f"**Batch scoring — {len(scored)} tickets scored, {_saved} saved**\n"]
        lines.append("| Ticket ID | Stars | Temperature | Complexity |")
        lines.append("|-----------|-------|-------------|------------|")
        for s in scored:
            lines.append(
                f"| {s.get('ticket_id','')} "
                f"| {s.get('stars','?')} "
                f"| {s.get('temperature','?')} "
                f"| {s.get('complexity','?')} |"
            )
        return "\n".join(lines)

    elif name == "batch_rescrape_tickets":
        # BEFORE v1.5.0: rescrape_ticket took one ID per call, agent called it N times
        #   burning the 5-turn limit and often stopping after 2-3 tickets.
        # AFTER v1.5.0: re-fetches up to 20 tickets per call, returns per-ticket results.
        raw_ids    = args.get("ticket_ids") or []
        limit      = min(int(args.get("limit") or 10), 20)
        ticket_ids = [str(t).strip() for t in raw_ids if t][:limit]
        if not ticket_ids:
            return "Error: ticket_ids list is required and must not be empty."

        cookie = ctx.get("cookie") or _get_profile_cookie()
        if not cookie:
            return "No session cookie available — paste a cookie in the Configuration tab first."

        try:
            from couchbase.cluster import Cluster as _BrCl
            from couchbase.options import ClusterOptions as _BrCO
            from couchbase.auth import PasswordAuthenticator as _BrPA
            _br_conn = _cb_conn_str(cb_url, use_tls)
            _br_cl = _BrCl(_br_conn, _BrCO(_BrPA(username, password)))
            _br_cl.wait_until_ready(timedelta(seconds=10))
            _br_col = _br_cl.bucket(bucket).scope(scope).collection(collection)
        except Exception as exc:
            return f"Couchbase connection failed: {exc}"

        _ok = _skipped = _errors = 0
        _result_rows: list[str] = []
        for tid in ticket_ids:
            try:
                _sess = _make_api_session(cookie)
                fresh = fetch_ticket_api(tid, _sess)
                if not fresh or not fresh.get("ticket_id"):
                    _skipped += 1
                    _result_rows.append(f"| {tid} | ⚠ skipped (no data returned) |")
                    continue
                fresh["last_scraped_at"] = int(time.time())
                fresh["type"]            = "ticket"
                fresh["cb_version"]      = extract_ticket_version(fresh)
                fresh["feature_area"]    = classify_ticket_feature(fresh)
                fresh["ticket_origin"]   = classify_ticket_origin(fresh)
                _doc_key = f"ticket::{tid}"
                try:
                    _ex = _br_col.get(_doc_key).content_as[dict]
                    _merged = {**_ex}
                    for _k, _v in fresh.items():
                        if _v not in (None, "", [], {}):
                            _merged[_k] = _v
                        elif _k not in _merged:
                            _merged[_k] = _v
                    _br_col.upsert(_doc_key, _merged)
                except Exception:
                    _br_col.upsert(_doc_key, fresh)
                _ok += 1
                _result_rows.append(
                    f"| {tid} | ✓ {fresh.get('status','')} / {fresh.get('priority','')} |"
                )
            except Exception as exc:
                _errors += 1
                _result_rows.append(f"| {tid} | ✗ {str(exc)[:60]} |")
            time.sleep(0.3)

        try:
            _br_cl.close()
        except Exception:
            pass

        lines = [f"**Batch rescrape: {_ok} updated, {_skipped} skipped, {_errors} errors**\n"]
        lines.append("| Ticket ID | Result |")
        lines.append("|-----------|--------|")
        lines.extend(_result_rows)
        return "\n".join(lines)

    # ── v1.6.0: feature set tools ─────────────────────────────────────────────

    elif name == "get_customer_health_score":
        _org = (args.get("organization") or default_customer or "").strip()
        if not _org:
            return "Error: organization is required."
        try:
            h = _compute_health_score(_org, cb_url, bucket, username, password,
                                       use_tls, scope, collection)
            lines = [
                f"## Health Score: {_org}",
                f"**{h['score']}/100** — {h['grade']}\n",
                f"| Dimension | Value |",
                f"|---|---|",
                f"| Open P1 | {h['open_p1']} |",
                f"| Open P2 | {h['open_p2']} |",
                f"| Total tickets | {h['total_tickets']} |",
                f"| Escalation rate | {h['escalation_rate_pct']}% |",
                f"| Avg resolution | {h['avg_resolution_days']} days |",
                f"| Data age | {h['hours_since_scraped']}h |",
                f"| Cluster bad ratio | {h['cluster_bad_ratio']} |",
            ]
            return "\n".join(lines)
        except Exception as exc:
            return f"Health score error: {exc}"

    elif name == "check_sla_compliance":
        _org = (args.get("organization") or default_customer or "").strip()
        if not _org:
            return "Error: organization is required."
        try:
            s = _compute_sla_compliance(
                _org, cb_url, bucket, username, password, use_tls, scope, collection,
                date_from=args.get("date_from") or "",
                date_to=args.get("date_to") or "",
            )
            lines = [
                f"## SLA Compliance: {_org}",
                f"**Overall: {s['overall_compliance_pct']}%** ({s['tickets_analyzed']} tickets analyzed)\n",
                "| Priority | Compliance | Met | Breached | Avg Hours | SLA Threshold |",
                "|---|---|---|---|---|---|",
            ]
            for p, d in s.get("by_priority", {}).items():
                lines.append(
                    f"| {p.capitalize()} | {d['compliance_pct']}% | {d['met']} | {d['breached']} "
                    f"| {d['avg_resolution_hours']}h | {d['sla_threshold_hours']}h |"
                )
            return "\n".join(lines)
        except Exception as exc:
            return f"SLA compliance error: {exc}"

    elif name == "get_portfolio_status":
        limit           = min(int(args.get("limit") or 20), 50)
        incl_cluster    = bool(args.get("include_cluster"))
        snap_collection = ctx.get("snap_collection", "snapshots")
        try:
            from couchbase.cluster import Cluster  # type: ignore
            from couchbase.options import ClusterOptions  # type: ignore
            from couchbase.auth import PasswordAuthenticator        # type: ignore
            conn = _cb_conn_str(cb_url, use_tls)
            cl_  = Cluster(conn, ClusterOptions(PasswordAuthenticator(username, password)))
            cl_.wait_until_ready(timedelta(seconds=10))
            orgs_rows = list(cl_.query(
                f"SELECT DISTINCT t.organization FROM `{bucket}`.`{scope}`.`{collection}` t "
                f"WHERE t.type='ticket' AND t.organization IS NOT NULL LIMIT {limit * 2}",
            ))
            cl_.close()
            orgs = [r.get("organization") for r in orgs_rows if r.get("organization")][:limit]
            results = []
            _errs: list[str] = []
            for org in orgs:
                try:
                    if incl_cluster:
                        h = _compute_health_score_with_cluster(
                            org, cb_url, bucket, username, password,
                            use_tls, scope, collection, snap_collection)
                    else:
                        h = _compute_health_score(org, cb_url, bucket, username, password,
                                                   use_tls, scope, collection)
                    results.append(h)
                except Exception as _e:
                    _errs.append(f"{org}: {_e}")
            if not results:
                err_detail = "; ".join(_errs[:3]) if _errs else "no organizations found in ticket collection"
                return f"Portfolio status: no data returned. {err_detail}"
            results.sort(key=lambda x: x["score"])
            if incl_cluster:
                lines = [
                    "## Portfolio Status (ranked by urgency)\n",
                    "| Customer | Score | Grade | Open P1 | Open P2 | Esc% | Clusters | Bad Ratio | Data Age |",
                    "|---|---|---|---|---|---|---|---|---|",
                ]
                for h in results:
                    lines.append(
                        f"| {h['organization']} | {h['score']} | {h['grade']} "
                        f"| {h['open_p1']} | {h['open_p2']} | {h['escalation_rate_pct']}% "
                        f"| {h.get('cluster_count', '—')} | {h.get('cluster_bad_ratio', '—')} "
                        f"| {h['hours_since_scraped']}h |"
                    )
            else:
                lines = [
                    "## Portfolio Status (ranked by urgency)\n",
                    "| Customer | Score | Grade | Open P1 | Open P2 | Esc% | Data Age |",
                    "|---|---|---|---|---|---|---|",
                ]
                for h in results:
                    lines.append(
                        f"| {h['organization']} | {h['score']} | {h['grade']} "
                        f"| {h['open_p1']} | {h['open_p2']} "
                        f"| {h['escalation_rate_pct']}% | {h['hours_since_scraped']}h |"
                    )
            return "\n".join(lines)
        except Exception as exc:
            return f"Portfolio error: {exc}"

    elif name == "query_fleet_tickets":
        _group_by      = args.get("group_by", "organization")
        _status_filter = args.get("status_filter", "open")
        _limit         = min(int(args.get("limit") or 30), 100)
        try:
            rows = _query_fleet_tickets(
                cb_url, bucket, username, password, use_tls, scope, collection,
                group_by=_group_by, status_filter=_status_filter, limit=_limit,
            )
            if not rows:
                return f"No tickets found (group_by={_group_by}, status={_status_filter})."
            _status_label = {"open": "open", "solved": "resolved", "all": "all"}.get(_status_filter, _status_filter)
            def _fmt_cbse_id(raw):
                s = str(raw or "?").strip()
                return s if (s == "?" or s.upper().startswith("CBSE-")) else f"CBSE-{s}"

            if _group_by == "cbse":
                lines = [
                    f"## Fleet Tickets by CBSE ({_status_label})\n",
                    "| CBSE | Tickets | Orgs Affected |",
                    "|---|---|---|",
                ]
                for r in rows:
                    lines.append(f"| {_fmt_cbse_id(r.get('label'))} | {r.get('ticket_count',0)} | {r.get('org_count',0)} |")
            else:
                lines = [
                    f"## Fleet Tickets by {_group_by.replace('_',' ').title()} ({_status_label})\n",
                    "| Label | Tickets | P1 | P2 |",
                    "|---|---|---|---|",
                ]
                for r in rows:
                    lines.append(
                        f"| {r.get('label','?')} | {r.get('ticket_count',0)} "
                        f"| {r.get('p1_count',0)} | {r.get('p2_count',0)} |"
                    )
            lines.append(f"\n*{len(rows)} rows · grouped by {_group_by} · filter: {_status_filter}*")
            return "\n".join(lines)
        except Exception as exc:
            return f"Fleet query error: {exc}"

    elif name == "list_at_risk_clusters":
        _bad_t  = int(args.get("bad_threshold")  or 0)
        _warn_t = int(args.get("warn_threshold") or 3)
        _lim    = min(int(args.get("limit") or 25), 100)
        snap_collection = ctx.get("snap_collection", "snapshots")
        try:
            rows = _list_at_risk_clusters(
                cb_url, bucket, username, password, use_tls, scope,
                snap_collection=snap_collection, ticket_collection=collection,
                bad_threshold=_bad_t, warn_threshold=_warn_t, limit=_lim,
            )
            if not rows:
                return "No at-risk clusters found matching the thresholds — fleet looks clean."
            lines = [
                f"## At-Risk Clusters (bad > {_bad_t} OR warn > {_warn_t}, no open ticket)\n",
                "| Cluster | Org | CB Version | Bad Items | Warn Items | Risk Score |",
                "|---|---|---|---|---|---|",
            ]
            for r in rows:
                lines.append(
                    f"| {r.get('cluster_name','?')} | {r.get('organization','?')} "
                    f"| {r.get('cb_version','?')} | {r.get('bad_items',0)} "
                    f"| {r.get('warn_items',0)} | **{r.get('risk_score',0)}** |"
                )
            lines.append(f"\n*{len(rows)} clusters · risk score = bad×3 + warn*")
            return "\n".join(lines)
        except Exception as exc:
            return f"At-risk cluster query error: {exc}"

    elif name == "fleet_version_distribution":
        snap_collection = ctx.get("snap_collection", "snapshots")
        try:
            rows = _fleet_version_distribution(
                cb_url, bucket, username, password, use_tls, scope, snap_collection,
            )
            if not rows:
                return "No snapshot version data found."
            lines = [
                "## Fleet CB Version Distribution\n",
                "| CB Version | Clusters | Orgs |",
                "|---|---|---|",
            ]
            for r in rows:
                lines.append(f"| {r.get('version','?')} | {r.get('cluster_count',0)} | {r.get('org_count',0)} |")
            total_c = sum(r.get("cluster_count", 0) for r in rows)
            lines.append(f"\n*{len(rows)} distinct versions across {total_c} clusters*")
            return "\n".join(lines)
        except Exception as exc:
            return f"Version distribution error: {exc}"

    elif name == "fleet_cbse_impact":
        _lim = min(int(args.get("limit") or 20), 50)
        try:
            rows = _fleet_cbse_impact(
                cb_url, bucket, username, password, use_tls, scope, collection, limit=_lim,
            )
            if not rows:
                return "No CBSE data found in ticket records."

            def _fmt_cbse(raw):
                s = str(raw or "?").strip()
                if s == "?":
                    return s
                return s if s.upper().startswith("CBSE-") else f"CBSE-{s}"

            lines = [
                "## Fleet CBSE Blast Radius (ranked by orgs affected)\n",
                "| CBSE | Orgs Affected | Tickets |",
                "|---|---|---|",
            ]
            for r in rows:
                lines.append(
                    f"| {_fmt_cbse(r.get('cbse'))} | {r.get('org_count',0)} | {r.get('ticket_count',0)} |"
                )
            lines.append(f"\n*{len(rows)} CBSEs found across all tickets*")
            return "\n".join(lines)
        except Exception as exc:
            return f"CBSE impact query error: {exc}"

    elif name == "record_feedback":
        try:
            from supportal.cb_helpers import save_feedback
            _corr = None
            if args.get("correction_field") or args.get("correction_old") or args.get("correction_new"):
                _corr = {"field": args.get("correction_field", ""),
                         "old": args.get("correction_old", ""),
                         "new": args.get("correction_new", "")}
            _fb_key = save_feedback(
                cb_url, bucket, username, password, use_tls, scope,
                source="chat",
                kind="correction" if _corr else "rating",
                subject_kind=str(args.get("subject_kind") or ""),
                subject_ref=str(args.get("subject_ref") or ""),
                verdict=str(args.get("verdict") or ""),
                details=str(args.get("details") or ""),
                correction=_corr,
                organization=default_customer or "",
            )
            return f"Feedback recorded ({_fb_key}). It will feed future eval and training data."
        except Exception as exc:
            return f"Failed to record feedback: {exc}"

    elif name == "tag_ticket":
        tid  = str(args.get("ticket_id") or "").strip()
        tags = args.get("tags") or []
        repl = bool(args.get("replace"))
        if not tid:
            return "Error: ticket_id is required."
        if not tags:
            return "Error: tags list is required."
        return _tag_ticket_in_cb(tid, tags, cb_url, bucket, username, password,
                                  use_tls, scope, collection, replace=repl)

    elif name == "get_digest":
        _org    = (args.get("organization") or default_customer or "").strip()
        _hours  = max(1, min(168, int(args.get("since_hours") or 24)))
        try:
            d = _get_digest(_org, cb_url, bucket, username, password, use_tls, scope, collection, _hours)
            _new = d["new_tickets"]
            _res = d["resolved_tickets"]
            _stl = d["stale_open_tickets"]
            lines = [
                f"## What's New{' — ' + _org if _org else ''} (last {_hours}h)\n",
                f"**{len(_new)} new** · **{len(_res)} resolved** · **{len(_stl)} stale open**\n",
            ]
            if _new:
                lines.append("### New Tickets")
                for t in _new[:15]:
                    lines.append(f"- [{t.get('ticket_id','')}] **{t.get('priority','')}** {(t.get('subject') or '')[:70]}")
            if _res:
                lines.append("\n### Resolved")
                for t in _res[:10]:
                    lines.append(f"- [{t.get('ticket_id','')}] {(t.get('subject') or '')[:70]}")
            if _stl:
                lines.append("\n### Stale Open (not refreshed in window)")
                for t in _stl[:10]:
                    lines.append(f"- [{t.get('ticket_id','')}] {t.get('priority','')} — {(t.get('subject') or '')[:60]}")
            return "\n".join(lines)
        except Exception as exc:
            return f"Digest error: {exc}"

    elif name == "save_query":
        _name = (args.get("name") or "").strip()
        _q    = (args.get("question") or "").strip()
        _org  = (args.get("organization") or default_customer or "").strip()
        if not _name or not _q:
            return "Error: name and question are required."
        try:
            key = _save_query_to_cb(_name, _q, _org, cb_url, bucket, username, password,
                                     use_tls, scope, collection)
            return f"Query saved as **{_name}** (key: `{key}`). Run it anytime with list_saved_queries."
        except Exception as exc:
            return f"Save query error: {exc}"

    elif name == "list_saved_queries":
        _org = (args.get("organization") or default_customer or "").strip()
        try:
            rows = _list_saved_queries(cb_url, bucket, username, password,
                                        use_tls, scope, collection, org=_org)
            if not rows:
                return "No saved queries found."
            lines = ["## Saved Queries\n",
                     "| Name | Question | Customer | Saved |",
                     "|---|---|---|---|"]
            for r in rows:
                lines.append(f"| **{r.get('name','')}** | {(r.get('question') or '')[:60]} | {r.get('organization','—')} | {(r.get('created_at') or '')[:10]} |")
            lines.append("\nTo run a saved query, just paste or type the question above.")
            return "\n".join(lines)
        except Exception as exc:
            return f"List saved queries error: {exc}"

    elif name == "generate_health_report":
        _org = (args.get("organization") or default_customer or "").strip()
        if not _org:
            return "Error: organization is required."
        try:
            from apps.mcp.server import generate_health_report as _gen_html_report
            _result = _gen_html_report(
                organization=_org,
                ae_name=args.get("ae_name", ""),
                tse_name=args.get("tse_name", ""),
                pse_name=args.get("pse_name", ""),
                date_from=args.get("date_from", ""),
                date_to=args.get("date_to", ""),
                annotations=args.get("annotations", ""),
            )
            import json as _json
            try:
                _r = _json.loads(_result)
                if "error" in _r:
                    return f"Report generation error: {_r['error']}"
                _aid = _r.get("asset_id", "")
                _fname = _r.get("filename", "")
                _count = _r.get("ticket_count", "")
                _date = _r.get("report_date", "")
                return (
                    f"Health report generated for **{_org}** ({_count} tickets analyzed, {_date}). "
                    f"The branded HTML report has been saved to the Reports tab. "
                    f"Asset ID: `{_aid}` · Filename: `{_fname}`"
                )
            except Exception:
                return _result
        except Exception as exc:
            return f"Report generation error: {exc}"

    elif name == "generate_customer_report":
        _org = (args.get("organization") or default_customer or "").strip()
        if not _org:
            return "Error: organization is required."
        try:
            _report_md = _generate_customer_report(
                _org, cb_url, bucket, username, password,
                use_tls, scope, collection,
            )
            try:
                import threading as _thr
                _thr.Thread(
                    target=_save_asset_to_cb,
                    args=(cb_url, bucket, username, password, use_tls, scope,
                          "report", f"{_org} Report", _report_md,
                          ctx.get("session_id", ""), _org, f"{_org.lower().replace(' ','_')}_report.md"),
                    daemon=True,
                ).start()
            except Exception:
                pass
            return _report_md
        except Exception as exc:
            return f"Report generation error: {exc}"

    elif name == "save_artifact":
        _title    = (args.get("title") or "untitled").strip()
        _atype    = args.get("asset_type", "report")
        _content  = args.get("content", "")
        _filename = args.get("filename", "")
        _org      = default_customer or ""
        if not _content:
            return "Error: content is required."
        try:
            aid = _save_asset_to_cb(
                cb_url, bucket, username, password, use_tls, scope,
                _atype, _title, _content,
                session_id=ctx.get("session_id", ""),
                org=_org, filename=_filename,
            )
            return f"Asset saved: **{_title}** (ID: `{aid}`). View it in the **Assets** tab."
        except Exception as exc:
            return f"Failed to save asset: {exc}"

    elif name == "get_current_time":
        import datetime as _dt_mod
        _tz_name = (args.get("timezone") or "").strip()
        try:
            if _tz_name:
                import zoneinfo as _zi
                _tz = _zi.ZoneInfo(_tz_name)
                _now = _dt_mod.datetime.now(_tz)
            else:
                _now = _dt_mod.datetime.now(_dt_mod.timezone.utc).astimezone()
        except Exception:
            _now = _dt_mod.datetime.now(_dt_mod.timezone.utc).astimezone()
        _iso_week = _now.isocalendar()[1]
        _quarter = (_now.month - 1) // 3 + 1
        return (
            f"Current date/time: {_now.strftime('%Y-%m-%d %H:%M:%S %Z')}\n"
            f"Day of week: {_now.strftime('%A')}\n"
            f"ISO week: {_iso_week}\n"
            f"Quarter: Q{_quarter} {_now.year}\n"
            f"UTC offset: {_now.strftime('%z')}"
        )

    else:
        return f"Unknown tool: {name}"





def enrich_ticket_apps_via_analytics(
    customer_name: str,
    cookie: str | None,
    cb_url: str,
    bucket: str,
    username: str,
    password: str,
    use_tls: bool,
    scope: str,
    collection: str,
    llm_provider: str,
    llm_model: str,
    llm_api_key: str,
    llm_base_url: str,
    progress_cb: Callable[[str, float], None] = lambda m, p: None,
) -> tuple[int, int]:
    """
    For tickets that are missing app labels (no entry in the cluster→app map),
    use the LLM to infer the Couchbase product name from the cluster hostname,
    then write analytics_app_labels into ticket["score"] and re-upsert to CB.

    Returns (enriched_count, error_count).
    """
    if not _CB_AVAILABLE:
        raise RuntimeError("Couchbase SDK not installed")

    conn_str = _cb_conn_str(cb_url, use_tls)
    cluster  = Cluster(conn_str, ClusterOptions(PasswordAuthenticator(username, password)))
    cluster.wait_until_ready(timedelta(seconds=15))
    col      = cluster.bucket(bucket).scope(scope).collection(collection)

    progress_cb("Loading tickets from Couchbase…", 0.05)
    try:
        ks = f"`{bucket}`.`{scope}`.`{collection}`"
        _cf = customer_name.strip().lower().replace("%", "\\%").replace("_", "\\_")
        q   = (
            f"SELECT META().id AS _key, t.* "
            f"FROM {ks} t "
            f"WHERE t.type = 'ticket' "
            f"AND LOWER(t.organization) LIKE {json.dumps(_cf + '%')}"
        )
        rows = list(cluster.query(q))
    except Exception as exc:
        cluster.close()
        raise RuntimeError(f"CB query failed: {exc}") from exc

    progress_cb(f"Loaded {len(rows)} tickets for '{customer_name}'.", 0.15)

    c2a = _get_cluster_to_app()

    # Collect tickets lacking app labels + their unmapped cluster names
    to_enrich: list[dict] = []
    all_unmapped: set[str] = set()
    for row in rows:
        ticket = dict(row)
        ticket.pop("_key", None)
        cids = _ticket_cluster_ids(ticket)
        has_label = any(c2a.get(cid) for cid in cids)
        if not has_label:
            # Also check stored analytics_app_labels
            _sc = ticket.get("score") or {}
            if _sc.get("analytics_app_labels"):
                continue  # already enriched
            to_enrich.append(ticket)
            for cid in cids:
                if cid and not c2a.get(cid):
                    all_unmapped.add(cid)

    if not to_enrich:
        progress_cb("All tickets already have app labels — nothing to enrich.", 1.0)
        cluster.close()
        return 0, 0

    progress_cb(
        f"{len(to_enrich)} ticket(s) missing app labels; "
        f"{len(all_unmapped)} unique unmapped cluster name(s).",
        0.2,
    )

    # ── LLM: infer app name from cluster hostname ────────────────────────────
    cname_to_app: dict[str, str] = {}
    if all_unmapped and llm_provider and llm_model:
        names_list = "\n".join(f"- {n}" for n in sorted(all_unmapped))
        prompt = (
            "You are a Couchbase product expert. Given a list of Couchbase cluster "
            "hostnames or names, identify which Couchbase product each cluster is "
            "running. Reply with a JSON object mapping each cluster name to the "
            "product label. Use these exact labels when applicable: "
            "\"Couchbase Server\", \"Sync Gateway\", \"App Services\", \"Mobile\", "
            "\"Analytics\", \"Eventing\", \"Search\", \"MLE\", \"Columnar\". "
            "If you cannot determine the product, use \"Unknown\". "
            "Reply ONLY with valid JSON, no explanation.\n\n"
            f"Cluster names:\n{names_list}"
        )
        try:
            progress_cb("Calling LLM to infer app labels from cluster names…", 0.35)
            raw = call_llm(
                [{"role": "user", "content": prompt}],
                llm_provider, llm_model, llm_api_key, llm_base_url,
                max_tokens=512,
            )
            # Strip markdown fences if present
            raw = raw.strip()
            if raw.startswith("```"):
                raw = "\n".join(raw.split("\n")[1:])
            if raw.endswith("```"):
                raw = raw.rsplit("```", 1)[0]
            parsed = json.loads(raw.strip())
            for cname, label in parsed.items():
                if label and label.strip().lower() != "unknown":
                    cname_to_app[cname.strip()] = label.strip()
            progress_cb(
                f"LLM mapped {len(cname_to_app)} cluster name(s) to app labels.",
                0.50,
            )
        except Exception as exc:
            progress_cb(f"LLM inference warning: {exc} — will skip unmapped clusters.", 0.50)

    # ── Apply labels and upsert ──────────────────────────────────────────────
    enriched = errors = 0
    total = len(to_enrich)
    for i, ticket in enumerate(to_enrich):
        pct = 0.5 + 0.5 * (i / total)
        tid = ticket.get("ticket_id") or ticket.get("id", "?")
        doc_key = f"ticket::{tid}"

        cids   = _ticket_cluster_ids(ticket)
        labels = sorted({
            cname_to_app[cid]
            for cid in cids
            if cname_to_app.get(cid)
        })
        if not labels:
            continue  # LLM had no mapping for this ticket's clusters either

        sc = dict(ticket.get("score") or {})
        sc["analytics_app_labels"] = labels
        ticket = {**ticket, "score": sc}

        try:
            col.upsert(doc_key, ticket)
            enriched += 1
            progress_cb(
                f"[{i+1}/{total}] #{tid} → {', '.join(labels)}",
                pct,
            )
        except Exception as exc:
            errors += 1
            progress_cb(f"[{i+1}/{total}] CB upsert error for {tid}: {exc}", pct)

    cluster.close()
    progress_cb(
        f"App label enrichment complete: {enriched} ticket(s) updated"
        + (f", {errors} error(s)" if errors else "") + ".",
        1.0,
    )
    return enriched, errors


# ─────────────────── Snapshot Couchbase storage ───────────────────────────────

def load_snapshots_to_couchbase(
    snapshots: list[dict],
    cb_url: str,
    bucket: str,
    username: str,
    password: str,
    use_tls: bool,
    scope: str,
    snap_collection: str,
    progress_cb: Callable[[str, float], None],
) -> tuple[int, int]:
    """Upsert snapshot docs into a dedicated Couchbase collection (key: snapshot::{snap_id})."""
    conn_str = _cb_conn_str(cb_url, use_tls)
    cluster = Cluster(conn_str, ClusterOptions(PasswordAuthenticator(username, password)))
    cluster.wait_until_ready(timedelta(seconds=10))
    col = cluster.bucket(bucket).scope(scope).collection(snap_collection)
    upserted = errors = 0
    total = len(snapshots)
    _now = int(time.time())
    try:
        for i, snap in enumerate(snapshots):
            key = f"snapshot::{snap['snap_id']}"
            try:
                doc = snap.copy()
                doc["last_scraped_at"] = _now
                col.upsert(key, doc)
                upserted += 1
            except Exception as exc:
                errors += 1
                print(f"[CB-SNAP] Upsert error {key}: {exc}")
            progress_cb(f"Saved {i+1}/{total}", (i + 1) / total)
    finally:
        cluster.close()
    return upserted, errors


def load_snapshots_from_couchbase(
    cb_url: str,
    bucket: str,
    username: str,
    password: str,
    use_tls: bool,
    scope: str,
    snap_collection: str,
    customer_filter: str,
    progress_cb: Callable[[str, float], None],
) -> list[dict]:
    """Load snapshot documents from Couchbase, optionally filtered by organization."""
    conn_str = _cb_conn_str(cb_url, use_tls)
    cluster = Cluster(conn_str, ClusterOptions(PasswordAuthenticator(username, password)))
    cluster.wait_until_ready(timedelta(seconds=10))
    try:
        filt = (customer_filter or "").strip().lower()
        if filt:
            q = (
                f"SELECT s.* FROM `{bucket}`.`{scope}`.`{snap_collection}` AS s "
                f"WHERE LOWER(s.organization) LIKE $1 ORDER BY s.date DESC"
            )
            result = cluster.query(q, QueryOptions(positional_parameters=[f"%{filt}%"]))
        else:
            q = (
                f"SELECT s.* FROM `{bucket}`.`{scope}`.`{snap_collection}` AS s "
                f"ORDER BY s.date DESC"
            )
            result = cluster.query(q)
        rows = list(result)
        progress_cb(f"Loaded {len(rows)} snapshots", 1.0)
        return rows
    except CouchbaseException:
        return []
    finally:
        cluster.close()


def list_orgs_from_cb(
    cb_url: str,
    bucket: str,
    username: str,
    password: str,
    use_tls: bool,
    scope: str,
    collection: str,
) -> list[str]:
    """Return a sorted list of distinct organization names from the tickets collection."""
    if not _CB_AVAILABLE:
        raise RuntimeError("couchbase SDK not installed")
    conn_str = _cb_conn_str(cb_url, use_tls)
    cluster  = Cluster(conn_str, ClusterOptions(PasswordAuthenticator(username, password)))
    cluster.wait_until_ready(timedelta(seconds=15))
    keyspace = f"`{bucket}`.`{scope}`.`{collection}`"
    q = f"SELECT DISTINCT RAW organization FROM {keyspace} WHERE organization IS NOT MISSING"
    rows = [r for r in cluster.query(q) if r]
    cluster.close()
    return sorted(rows)


def query_customer_directory_from_cb(
    cb_url: str,
    bucket: str,
    username: str,
    password: str,
    use_tls: bool,
    scope: str,
    snap_collection: str,
    ticket_collection: str,
) -> list[dict]:
    """
    Returns one row per organization aggregating snapshot + ticket counts.
    Row keys: organization, customer_url, total_snapshots, total_clusters,
              active_clusters, last_scraped_at, total_tickets.
    """
    conn_str = _cb_conn_str(cb_url, use_tls)
    cluster = Cluster(conn_str, ClusterOptions(PasswordAuthenticator(username, password)))
    cluster.wait_until_ready(timedelta(seconds=10))
    cutoff = (
        datetime.datetime.now(datetime.timezone.utc) - datetime.timedelta(days=90)
    ).isoformat().replace("+00:00", "Z")
    try:
        # Per-org snapshot summary
        snap_q = f"""
            SELECT
                s.organization,
                FIRST cu FOR cu IN ARRAY_AGG(s.customer_url) WHEN cu IS NOT MISSING AND cu IS NOT NULL END
                    AS customer_url,
                COUNT(*) AS total_snapshots,
                COUNT(DISTINCT s.cluster_id) AS total_clusters,
                COUNT(DISTINCT CASE WHEN s.date >= $cutoff THEN s.cluster_id END) AS active_clusters,
                COUNT(DISTINCT CASE WHEN s.date < $cutoff THEN s.cluster_id END) AS stale_clusters,
                MAX(s.scraped_at) AS last_scraped_at
            FROM `{bucket}`.`{scope}`.`{snap_collection}` AS s
            WHERE s.organization IS NOT MISSING
            GROUP BY s.organization
            ORDER BY s.organization
        """
        snap_rows = {
            r["organization"]: r
            for r in cluster.query(snap_q, QueryOptions(named_parameters={"cutoff": cutoff}))
        }

        # Per-org ticket counts + customer_url from ticket docs (fallback when no snapshots)
        ticket_q = f"""
            SELECT
                t.organization,
                COUNT(*) AS total_tickets,
                FIRST cu FOR cu IN ARRAY_AGG(t.customer_url) WHEN cu IS NOT MISSING AND cu IS NOT NULL END
                    AS customer_url
            FROM `{bucket}`.`{scope}`.`{ticket_collection}` AS t
            WHERE t.organization IS NOT MISSING
            GROUP BY t.organization
        """
        try:
            for r in cluster.query(ticket_q):
                org = r.get("organization", "")
                if org in snap_rows:
                    snap_rows[org]["total_tickets"] = r.get("total_tickets", 0)
                    # Fill customer_url from tickets if snapshots didn't have one
                    if not snap_rows[org].get("customer_url") and r.get("customer_url"):
                        snap_rows[org]["customer_url"] = r["customer_url"]
                else:
                    snap_rows[org] = {
                        "organization": org,
                        "customer_url": r.get("customer_url"),
                        "total_snapshots": 0,
                        "total_clusters": 0,
                        "active_clusters": 0,
                        "last_scraped_at": None,
                        "total_tickets": r.get("total_tickets", 0),
                    }
        except Exception:
            pass  # tickets collection may not exist — snapshots-only is fine

        for row in snap_rows.values():
            row.setdefault("total_tickets", 0)
        return sorted(snap_rows.values(), key=lambda r: (r.get("organization") or "").lower())
    finally:
        cluster.close()


def ensure_cb_indexes(
    cb_url: str,
    bucket: str,
    username: str,
    password: str,
    use_tls: bool,
    scope: str,
    snap_collection: str,
    ticket_collection: str,
    progress_cb: Callable[[str, float], None],
) -> list[str]:
    """
    Create GSI indexes on the tickets and snapshots collections if they do not
    already exist.  Uses IF NOT EXISTS so it is safe to call repeatedly.

    Returns a list of result strings (one per index DDL attempted).
    """
    if not _CB_AVAILABLE:
        raise RuntimeError("couchbase SDK not installed")

    conn_str = _cb_conn_str(cb_url, use_tls)
    cluster  = Cluster(conn_str, ClusterOptions(PasswordAuthenticator(username, password)))
    cluster.wait_until_ready(timedelta(seconds=15))

    ks_t = f"`{bucket}`.`{scope}`.`{ticket_collection}`"
    ks_s = f"`{bucket}`.`{scope}`.`{snap_collection}`"

    ddls = [
        # ── Tickets collection ────────────────────────────────────────────
        (
            "idx_tickets_org_customer",
            f"CREATE INDEX IF NOT EXISTS `idx_tickets_org_customer` "
            f"ON {ks_t} (organization, customer_url, ticket_id) "
            f"WHERE organization IS NOT MISSING",
        ),
        (
            "idx_tickets_snap_ids",
            f"CREATE INDEX IF NOT EXISTS `idx_tickets_snap_ids` "
            f"ON {ks_t} (DISTINCT ARRAY sid FOR sid IN snap_ids END) "
            f"WHERE snap_ids IS NOT MISSING",
        ),
        (
            "idx_tickets_org_date",
            f"CREATE INDEX IF NOT EXISTS `idx_tickets_org_date` "
            f"ON {ks_t} (organization, created DESC) "
            f"WHERE organization IS NOT MISSING AND ticket_id IS NOT MISSING",
        ),
        (
            "idx_tickets_org_status_priority",
            f"CREATE INDEX IF NOT EXISTS `idx_tickets_org_status_priority` "
            f"ON {ks_t} (organization, status, priority, created DESC) "
            f"WHERE organization IS NOT MISSING AND ticket_id IS NOT MISSING",
        ),
        # ── Snapshots collection ──────────────────────────────────────────
        (
            "idx_snaps_org_covering",
            f"CREATE INDEX IF NOT EXISTS `idx_snaps_org_covering` "
            f"ON {ks_s} (organization, customer_url, cluster_id, date, scraped_at, bad_count, warn_count) "
            f"WHERE organization IS NOT MISSING",
        ),
        (
            "idx_snaps_cluster_date",
            f"CREATE INDEX IF NOT EXISTS `idx_snaps_cluster_date` "
            f"ON {ks_s} (cluster_id, date, bad_count, warn_count, node_count)",
        ),
        (
            "idx_snaps_ticket_ids",
            f"CREATE INDEX IF NOT EXISTS `idx_snaps_ticket_ids` "
            f"ON {ks_s} (DISTINCT ARRAY tid FOR tid IN ticket_ids END) "
            f"WHERE ticket_ids IS NOT MISSING",
        ),
        (
            "idx_snaps_org_date",
            f"CREATE INDEX IF NOT EXISTS `idx_snaps_org_date` "
            f"ON {ks_s} (organization, date) "
            f"WHERE organization IS NOT MISSING",
        ),
    ]

    results: list[str] = []
    total = len(ddls)
    try:
        for i, (name, ddl) in enumerate(ddls):
            progress_cb(f"Creating index {name} …", i / total)
            try:
                list(cluster.query(ddl, QueryOptions(timeout=timedelta(seconds=120))))
                results.append(f"OK: {name}")
            except Exception as exc:
                msg = str(exc)
                # "already exists" is non-fatal
                if "already exist" in msg.lower() or "duplicate" in msg.lower():
                    results.append(f"EXISTS: {name}")
                else:
                    results.append(f"ERROR {name}: {msg}")
        progress_cb(f"Done — {len(ddls)} indexes processed.", 1.0)
    finally:
        cluster.close()
    return results


def migrate_ticket_snapshot_links(
    cb_url: str,
    bucket: str,
    username: str,
    password: str,
    use_tls: bool,
    scope: str,
    ticket_collection: str,
    snap_collection: str,
    progress_cb: Callable[[str, float], None],
) -> tuple[int, int, int]:
    """
    One-shot migration that backfills cross-reference fields without re-scraping:

    1. For each ticket doc missing snap_ids: extract IDs from the raw `snapshots`
       text field using _SNAP_ID_RE and set snap_ids + snapshot_summary (from
       snapshot_topology if present).
    2. For each snapshot doc: query which tickets reference its snap_id and patch
       ticket_ids via SQL++ ARRAY_APPEND (deduped).

    Returns (tickets_updated, snaps_updated, errors).
    """
    if not _CB_AVAILABLE:
        raise RuntimeError("couchbase SDK not installed")

    conn_str = _cb_conn_str(cb_url, use_tls)
    cluster  = Cluster(conn_str, ClusterOptions(PasswordAuthenticator(username, password)))
    cluster.wait_until_ready(timedelta(seconds=15))
    ks_t = f"`{bucket}`.`{scope}`.`{ticket_collection}`"
    ks_s = f"`{bucket}`.`{scope}`.`{snap_collection}`"
    col_t = cluster.bucket(bucket).scope(scope).collection(ticket_collection)
    col_s = cluster.bucket(bucket).scope(scope).collection(snap_collection)

    tickets_updated = snaps_updated = errors = 0
    try:
        # ── Pass 1: backfill snap_ids + snapshot_summary on ticket docs ────
        progress_cb("Querying tickets …", 0.05)
        ticket_rows = list(cluster.query(
            f"SELECT META(t).id AS _key, t.snapshots, t.snap_ids, t.snapshot_topology, t.ticket_id "
            f"FROM {ks_t} AS t WHERE t.snapshots IS NOT MISSING",
            QueryOptions(timeout=timedelta(seconds=120)),
        ))
        total_t = len(ticket_rows)
        progress_cb(f"{total_t} tickets with snapshots text found.", 0.10)

        for i, row in enumerate(ticket_rows):
            doc_key = row.get("_key")
            if not doc_key:
                continue
            # Already has snap_ids — skip
            if row.get("snap_ids"):
                continue
            raw = row.get("snapshots") or ""
            all_ids = _SNAP_ID_RE.findall(raw)
            if not all_ids:
                continue
            deduped: list[str] = []
            for sid in all_ids:
                if sid not in deduped:
                    deduped.append(sid)
            patch: dict = {"snap_ids": deduped}
            # Build snapshot_summary from snapshot_topology if present and not already set
            topo = row.get("snapshot_topology")
            if isinstance(topo, dict) and topo:
                highest = _highest_snap_id(deduped)
                patch["snapshot_summary"] = {
                    "snap_id":      highest,
                    "cluster_name": topo.get("cluster_name") or "",
                    "cb_version":   topo.get("cb_version") or "",
                    "bad_count":    topo.get("bad_count", 0),
                    "warn_count":   topo.get("warn_count", 0),
                    "node_count":   topo.get("total_nodes", 0),
                }
            try:
                from couchbase.subdocument import upsert as _SD_upsert, MutateInOptions
                mutations = [_SD_upsert(k, v) for k, v in patch.items()]
                col_t.mutate_in(doc_key, mutations)
                tickets_updated += 1
            except Exception as exc:
                errors += 1
                print(f"[migrate] ticket {doc_key}: {exc}")
            if (i + 1) % 50 == 0:
                progress_cb(f"Tickets: {i+1}/{total_t} processed …", 0.10 + 0.40 * (i / total_t))

        progress_cb(f"Ticket pass done — {tickets_updated} updated.", 0.50)

        # ── Pass 2: backfill ticket_ids on snapshot docs ───────────────────
        # Query all ticket docs that have snap_ids and cross-reference to snaps
        snap_to_tickets: dict[str, list[str]] = {}
        progress_cb("Building snap→ticket map from tickets collection …", 0.55)
        xref_rows = list(cluster.query(
            f"SELECT t.ticket_id, ARRAY_AGG(sid) AS snap_ids "
            f"FROM {ks_t} AS t "
            f"UNNEST t.snap_ids AS sid "
            f"WHERE t.snap_ids IS NOT MISSING AND t.ticket_id IS NOT MISSING "
            f"GROUP BY t.ticket_id",
            QueryOptions(timeout=timedelta(seconds=120)),
        ))
        for xrow in xref_rows:
            tid = str(xrow.get("ticket_id", ""))
            if not tid:
                continue
            for sid in xrow.get("snap_ids") or []:
                snap_to_tickets.setdefault(sid, [])
                if tid not in snap_to_tickets[sid]:
                    snap_to_tickets[sid].append(tid)

        total_s = len(snap_to_tickets)
        progress_cb(f"{total_s} unique snap IDs referenced by tickets.", 0.60)

        for j, (snap_id, ticket_ids) in enumerate(snap_to_tickets.items()):
            snap_key = f"snapshot::{snap_id}"
            try:
                # Read current ticket_ids then merge (upsert avoids duplicates)
                result = col_s.get(snap_key)
                existing: list = result.content_as[dict].get("ticket_ids") or []
                merged = existing[:]
                for tid in ticket_ids:
                    if tid not in merged:
                        merged.append(tid)
                if merged != existing:
                    from couchbase.subdocument import upsert as _SD_upsert
                    col_s.mutate_in(snap_key, [_SD_upsert("ticket_ids", merged)])
                    snaps_updated += 1
            except Exception:
                # Snapshot doc may not exist yet (only in tickets, not yet scraped)
                pass
            if (j + 1) % 100 == 0:
                progress_cb(f"Snapshots: {j+1}/{total_s} processed …", 0.60 + 0.38 * (j / total_s))

        progress_cb(
            f"Migration complete — {tickets_updated} tickets updated, "
            f"{snaps_updated} snapshots linked, {errors} errors.",
            1.0,
        )
    finally:
        cluster.close()
    return tickets_updated, snaps_updated, errors


# ─────────────────── Cluster index + health analytics ────────────────────────

def build_cluster_index(snapshots: list[dict]) -> dict[str, dict]:
    """
    Derive a per-cluster summary from a flat list of snapshot dicts.
    Snapshots should be sorted by date descending so node_count_last is the
    most recent value.
    """
    index: dict[str, dict] = {}
    for snap in snapshots:
        cid = snap.get("cluster_id") or ""
        if not cid:
            continue
        if cid not in index:
            index[cid] = {
                "cluster_id":        cid,
                "cluster_uuid":      snap.get("cluster_uuid") or "",
                "capella_cluster_id": snap.get("capella_cluster_id") or "",
                "organization":      snap.get("organization") or "",
                "cluster_names":     [],
                "snapshot_ids":      [],
                "ticket_ids":        [],
                "dates":             [],
                "node_count_min":    None,
                "node_count_max":    None,
                "node_count_last":   None,
                "cpus_per_node":     None,
                "ram_per_node_mib":  None,
                "os_name":           None,
                "version_history":   [],
                "bucket_names_seen": set(),
                "bad_counts":        [],
                "warn_counts":       [],
                "_bad_name_counts":  {},
                "_warn_name_counts": {},
            }
        ci = index[cid]

        _cn = snap.get("cluster_name") or ""
        _cn_val = (_cn[0] if isinstance(_cn, list) else _cn)
        name = str(_cn_val).strip() if _cn_val else ""
        if name and name not in ci["cluster_names"]:
            ci["cluster_names"].append(name)

        ci["snapshot_ids"].append(snap["snap_id"])
        for tid in snap.get("ticket_ids") or []:
            if tid not in ci["ticket_ids"]:
                ci["ticket_ids"].append(tid)

        if snap.get("date"):
            ci["dates"].append(snap["date"])

        nc = snap.get("node_count") or 0
        if nc > 0:
            ci["node_count_min"] = min(ci["node_count_min"] or nc, nc)
            ci["node_count_max"] = max(ci["node_count_max"] or nc, nc)
            if ci["node_count_last"] is None:   # first (=newest) wins when sorted desc
                ci["node_count_last"] = nc

        # Promote hardware fields from top-level snap field or nested topology
        _topo_d = snap.get("topology") or {}
        for _hw_field in ("cpus_per_node", "ram_per_node_mib", "os_name"):
            if ci[_hw_field] is None:
                _val = snap.get(_hw_field) or _topo_d.get(_hw_field)
                if _val:
                    ci[_hw_field] = _val

        _ver = snap.get("cb_version") or ""
        _ver_val = (_ver[0] if isinstance(_ver, list) else _ver)
        ver = str(_ver_val).strip() if _ver_val else ""
        if ver and ver not in ci["version_history"]:
            ci["version_history"].append(ver)

        for bn in snap.get("bucket_names") or []:
            ci["bucket_names_seen"].add(bn)

        ci["bad_counts"].append(snap.get("bad_count", 0))
        ci["warn_counts"].append(snap.get("warn_count", 0))
        for _name in snap.get("bad_items") or []:
            ci["_bad_name_counts"][_name] = ci["_bad_name_counts"].get(_name, 0) + 1
        for _name in snap.get("warn_items") or []:
            ci["_warn_name_counts"][_name] = ci["_warn_name_counts"].get(_name, 0) + 1

    now_iso = datetime.datetime.now(datetime.timezone.utc).isoformat()
    cutoff  = (datetime.datetime.now(datetime.timezone.utc) - datetime.timedelta(days=90)).isoformat()
    for ci in index.values():
        ci["bucket_names_seen"] = sorted(ci["bucket_names_seen"])
        ci["cluster_name"]      = ci["cluster_names"][-1] if ci["cluster_names"] else ""
        ci["first_seen"]        = min(ci["dates"]) if ci["dates"] else None
        ci["last_seen"]         = max(ci["dates"]) if ci["dates"] else None
        ci["snapshot_count"]    = len(ci["snapshot_ids"])
        ci["is_active"]         = bool(ci["last_seen"] and ci["last_seen"] >= cutoff)
        bc = ci["bad_counts"]
        wc = ci["warn_counts"]
        ci["avg_bad"]  = round(sum(bc) / len(bc), 2) if bc else 0
        ci["avg_warn"] = round(sum(wc) / len(wc), 2) if wc else 0
        ci["total_bad"]  = sum(bc)
        ci["total_warn"] = sum(wc)
        # Top recurring checker names (sorted by frequency desc, capped at 20)
        ci["top_bad_items"]  = sorted(ci["_bad_name_counts"],  key=ci["_bad_name_counts"].get,  reverse=True)[:20]
        ci["top_warn_items"] = sorted(ci["_warn_name_counts"], key=ci["_warn_name_counts"].get, reverse=True)[:20]
        del ci["_bad_name_counts"], ci["_warn_name_counts"]
    return index


def build_cluster_health_data(snapshots: list[dict], tickets: list[dict]) -> dict:
    """
    Assemble all analytics needed for the Cluster Health tab.
    Snapshots are expected newest-first (as returned by CB query).
    """
    # Sort ascending for timeline charts
    snaps_asc = sorted(snapshots, key=lambda s: s.get("date") or "")

    cluster_index = build_cluster_index(snapshots)   # uses desc order for last-seen

    # Per-cluster timeline series
    by_cluster: dict[str, list[dict]] = {}
    for snap in snaps_asc:
        cid = snap.get("cluster_id") or "unknown"
        by_cluster.setdefault(cid, []).append({
            "date":       snap.get("date") or "",
            "bad_count":  snap.get("bad_count", 0),
            "warn_count": snap.get("warn_count", 0),
            "node_count": snap.get("node_count", 0),
            "cb_version": snap.get("cb_version") or "",
            "snap_id":    snap["snap_id"],
        })

    # Issue heatmap: month × cluster → total issues
    heatmap: dict[str, dict[str, int]] = {}
    for snap in snaps_asc:
        month = (snap.get("date") or "")[:7] or "unknown"
        cid   = snap.get("cluster_id") or "unknown"
        _hm = heatmap.setdefault(month, {})
        _hm[cid] = _hm.get(cid, 0) + snap.get("bad_count", 0) + snap.get("warn_count", 0)

    # Ticket → cluster correlation via snapshot topology
    ticket_by_cid: dict[str, list[dict]] = {}
    for t in tickets:
        topo = t.get("snapshot_topology") or {}
        if isinstance(topo, str):
            try:
                topo = json.loads(topo)
            except Exception:
                topo = {}
        snap_info = extract_cluster_snapshot_info(t)
        for raw_cid in snap_info.get("cluster_ids") or []:
            short = raw_cid.split("::")[0].replace("-", "")
            ticket_by_cid.setdefault(short, []).append(t)

    # ── Deprecation detection ───────────────────────────────────────────────
    # A stale cluster is "Deprecated" (replaced rather than just not recently
    # snapshotted) when the same org has a DIFFERENT cluster_id whose earliest
    # snapshot date is on or after this cluster's last snapshot date AND whose
    # highest CB major version is greater.  This detects the pattern:
    #   org → cluster A (6.x) stops → cluster B (7.x) appears shortly after
    # which indicates B is a fresh deployment that replaced A, not an in-place
    # upgrade (those preserve the cluster UUID and show the version jump in the
    # same cluster entry's version_history).

    def _max_major(version_history: list[str]) -> int:
        best = 0
        for v in version_history:
            m = re.match(r"(\d+)\.", v or "")
            if m:
                best = max(best, int(m.group(1)))
        return best

    # Group by org
    org_clusters: dict[str, list[dict]] = {}
    for ci in cluster_index.values():
        org_clusters.setdefault(ci.get("organization") or "", []).append(ci)

    for org_cis in org_clusters.values():
        stale_cis  = [c for c in org_cis if not c.get("is_active")]
        active_cis = [c for c in org_cis if c.get("is_active")]
        # Also include stale clusters with a later first-seen date as potential successors
        all_cis    = org_cis  # successors can themselves be stale
        for sc in stale_cis:
            sc_last  = sc.get("last_seen") or ""
            sc_major = _max_major(sc.get("version_history") or [])
            sc["is_deprecated"] = False
            if sc_major == 0:
                continue
            for ac in all_cis:
                if ac["cluster_id"] == sc["cluster_id"]:
                    continue
                ac_first  = ac.get("first_seen") or ""
                ac_major  = _max_major(ac.get("version_history") or [])
                # Successor: higher major version AND started no earlier than this cluster stopped
                if ac_major > sc_major and ac_first >= sc_last:
                    sc["is_deprecated"] = True
                    break
        # Active clusters are never deprecated
        for ac in active_cis:
            ac["is_deprecated"] = False

    active     = sum(1 for c in cluster_index.values() if c.get("is_active"))
    deprecated = sum(1 for c in cluster_index.values() if not c.get("is_active") and c.get("is_deprecated"))
    stale      = len(cluster_index) - active - deprecated

    return {
        "cluster_index":      cluster_index,
        "by_cluster":         by_cluster,
        "heatmap":            heatmap,
        "ticket_by_cid":      ticket_by_cid,
        "total_snapshots":    len(snapshots),
        "total_clusters":     len(cluster_index),
        "active_clusters":    active,
        "stale_clusters":     stale,
        "deprecated_clusters": deprecated,
        "all_months":      sorted(heatmap.keys()),
        "all_cluster_ids": sorted(cluster_index.keys()),
    }


# ══════════════════════════════════════════════════════════════════════════════
_AUTO_MSG_PREFIX = "this is an automated message"

_PROACTIVE_SUBJECT_PREFIX = "proactive ticket:"
_PROACTIVE_BODY_FRAGMENT   = "creating a proactive ticket on your behalf"


def _is_automated_comment(comment: dict) -> bool:
    body = (comment.get("body") or "").strip().lower()
    return body.startswith(_AUTO_MSG_PREFIX)


def _is_proactive_ticket(ticket: dict) -> bool:
    """Return True if this ticket was opened proactively by Couchbase on the
    customer's behalf (e.g. triggered by internal monitoring).

    Signals (either is sufficient):
      1. Subject starts with "Proactive ticket:" (case-insensitive).
      2. The oldest comment body contains the standard proactive-ticket opening
         phrase "creating a proactive ticket on your behalf".
    """
    subject = (ticket.get("subject") or "").strip().lower()
    if subject.startswith(_PROACTIVE_SUBJECT_PREFIX):
        return True

    comments_raw = ticket.get("comments")
    if comments_raw:
        try:
            comments = json.loads(comments_raw) if isinstance(comments_raw, str) else comments_raw
            if isinstance(comments, list) and comments:
                # Check oldest comment (stored newest-first, so oldest is last)
                oldest_body = (comments[-1].get("body") or "").lower()
                if _PROACTIVE_BODY_FRAGMENT in oldest_body:
                    return True
        except Exception:
            pass
    return False


def _extract_ticket_fields(raw: str | dict) -> dict:
    """Parse the ticket_fields blob and return a flat dict of non-empty values."""
    try:
        tf = json.loads(raw) if isinstance(raw, str) else (raw or {})
        return {k: v for k, v in tf.items() if v and str(v).strip()}
    except Exception:
        return {}


# ticket_fields keys worth surfacing to the scoring prompt (normalized)
_SCORING_TICKET_FIELDS = [
    "Resolution_type", "Resolution", "Root_Cause", "Component",
    "Environment_Current_Impact", "Current_Impact", "Escalation_Level",
    "Business_Impact",
]

# Tag prefixes/substrings that add scoring signal (others are administrative noise)
_USEFUL_TAG_FRAGMENTS = [
    "unresolved", "no_reply", "automation_solved", "customer_abandon",
    "root_cause", "escalat", "prod_impact", "blocker", "component__",
    "highest_not_p1", "p1", "p2", "p3",
]


def _summarize_ticket_comments(
    ticket: dict,
    provider: str,
    model: str,
    api_key: str,
    base_url: str,
) -> str:
    """Compress a ticket's full comment thread into a ~200-char narrative summary.

    Called as a fallback when the full ticket content exceeds the model's context
    window.  Uses the same model/provider already in use for scoring so no extra
    configuration is needed.  Returns an empty string on any error so the caller
    can fall back to hard truncation.
    """
    comments_raw = ticket.get("comments")
    if not comments_raw:
        return ""
    try:
        comments = json.loads(comments_raw) if isinstance(comments_raw, str) else comments_raw
        if not comments:
            return ""
        # Build a compact thread — cap at 40 comments, 400 chars each
        lines = []
        for c in comments[:40]:
            author = c.get("author", "?")
            body = (c.get("body") or "").strip()[:400]
            if body and not _is_automated_comment(c):
                lines.append(f"{author}: {body}")
        if not lines:
            return ""
        thread = "\n".join(lines)
        subject = ticket.get("subject", "")
        prompt = (
            f"Support ticket: {subject}\n\nComment thread (newest first):\n{thread}\n\n"
            "In 1-2 sentences, summarize: what was the core problem, how it progressed, "
            "and the final outcome or current state. Be factual and concise."
        )
        summary = call_llm(
            [{"role": "user", "content": prompt}],
            provider, model, api_key, base_url,
            max_tokens=100,
        )
        return (summary or "").strip()[:300]
    except Exception:
        return ""


def build_scoring_input(
    ticket: dict,
    desc_limit: int = 400,
    comment_limit: int = 300,
    comment_override: str | None = None,
    include_extras: bool = True,
) -> str:
    """Build a compact ticket representation for the scoring prompt.

    desc_limit / comment_limit — character caps for description and latest comment.
    comment_override — pre-computed summary string; replaces live comment parsing
                       (used after a context-length retry with summarization).
    include_extras — when False, skips outcome fields, tags, and topology
                     (used for the last-resort minimal retry).
    """
    escs  = ticket.get("escalations") or "none"
    proactive = _is_proactive_ticket(ticket)

    # Dates
    _created_str = (ticket.get("created") or ticket.get("created_at") or "").strip()[:10]
    _solved_raw  = (ticket.get("solved") or ticket.get("solved_at")
                    or ticket.get("closed_at") or "").strip()
    _solved_str  = _solved_raw[:10] if _solved_raw else ""
    if not _solved_str and (ticket.get("status") or "").lower() in ("closed", "solved"):
        _solved_str = (ticket.get("updated") or "").strip()[:10]

    # App impact from cluster→app map; fallback to analytics-enriched labels
    _c2a = _get_cluster_to_app()
    _cids = _ticket_cluster_ids(ticket)
    _app_labels = sorted({_c2a[c].upper() for c in _cids if _c2a.get(c)})
    if not _app_labels:
        _score_fmt = ticket.get("score") or {}
        _analytics_lbl = _score_fmt.get("analytics_app_labels") or []
        if _analytics_lbl:
            _app_labels = sorted(str(l).upper() for l in _analytics_lbl)
    _app_str = ", ".join(_app_labels) if _app_labels else ""

    _cbses_val = ticket.get("cbses") or []
    _cbse_str  = ", ".join(_cbses_val) if isinstance(_cbses_val, list) else str(_cbses_val)
    _jira_val  = ticket.get("jira_issues") or []
    _jira_str  = ", ".join(_jira_val) if isinstance(_jira_val, list) else str(_jira_val)

    parts = [
        f"ID: {ticket.get('ticket_id','?')} | "
        f"Priority: {ticket.get('priority','?')} | "
        f"Status: {ticket.get('status','?')} | "
        f"Origin: {'Proactive (Couchbase-initiated)' if proactive else 'Customer-initiated'} | "
        f"Comments: {ticket.get('comment_count', 0)} | "
        f"Escalations: {escs}",
        f"Requester: {ticket.get('requester') or '?'} | "
        f"Created: {_created_str or '?'} | Closed: {_solved_str or 'open'}",
        f"Subject: {(ticket.get('subject') or '(no subject)')[:200]}",
    ]
    if _cbse_str:
        parts.append(f"CBSEs: {_cbse_str}")
    if _jira_str:
        parts.append(f"Jira Issues: {_jira_str}")
    if _app_str:
        parts.append(f"Application: {_app_str}")
    if _cids:
        parts.append(f"Clusters: {', '.join(_cids[:5])}")
    if ticket.get("description"):
        parts.append(f"Description: {ticket['description'][:desc_limit]}")

    if include_extras:
        # Key outcome fields from ticket_fields (Resolution type, Root Cause, etc.)
        tf = _extract_ticket_fields(ticket.get("ticket_fields", {}))
        outcome_parts = []
        for key in _SCORING_TICKET_FIELDS:
            val = tf.get(key, "").strip()
            if val:
                outcome_parts.append(f"{key}: {val}")
        if outcome_parts:
            outcome_str = " | ".join(outcome_parts)
            parts.append(f"Outcome fields: {outcome_str[:desc_limit]}")

        # Informative tags (skip pure administrative/routing tags)
        tags_raw = (ticket.get("tags") or "").strip()
        if tags_raw:
            useful = [
                t for t in tags_raw.split()
                if any(frag in t.lower() for frag in _USEFUL_TAG_FRAGMENTS)
            ][:20]  # cap tag count
            if useful:
                parts.append("Signals: " + ", ".join(useful))

    # Comment section — use pre-computed summary if provided (context-length retry path),
    # otherwise extract the most recent non-automated human comment.
    if comment_override is not None:
        if comment_override:
            parts.append(f"Comment summary: {comment_override}")
    else:
        comments_raw = ticket.get("comments")
        if comments_raw:
            try:
                comments = json.loads(comments_raw) if isinstance(comments_raw, str) else comments_raw
                if comments:
                    # comments stored newest-first; find most recent non-automated one
                    human = next(
                        (c for c in comments if not _is_automated_comment(c)), comments[0]
                    )
                    body = (human.get("body") or "").strip()[:comment_limit]
                    if body:
                        parts.append(f"Most recent comment ({human.get('author','')}): {body}")
            except Exception:
                pass

    if include_extras:
        # Include snapshot cluster topology if enriched
        topo = ticket.get("snapshot_topology")
        if topo:
            svc_parts = []
            for svc, field in [("data", "data_nodes"), ("query", "query_nodes"),
                               ("index", "index_nodes"), ("fts", "fts_nodes"),
                               ("eventing", "eventing_nodes"), ("analytics", "analytics_nodes")]:
                n = topo.get(field, 0)
                if n:
                    svc_parts.append(f"{svc}={n}")
            node_detail = f"({', '.join(svc_parts)})" if svc_parts else ""
            bkts = ", ".join((topo.get("bucket_names") or [])[:10]) or "?"
            parts.append(
                f"Cluster topology [{topo.get('cluster_name','?')}]: "
                f"CB {topo.get('cb_version','?')} | "
                f"{topo.get('total_nodes','?')} nodes {node_detail} | "
                f"RAM/node: {topo.get('ram_per_node_mib','?')} MiB | "
                f"CPUs/node: {topo.get('cpus_per_node','?')} | "
                f"Buckets: {bkts} | "
                f"Checker issues: {topo.get('bad_count',0)} BAD, {topo.get('warn_count',0)} WARN"
            )
    return "\n  ".join(parts)


def _strip_think_blocks(text: str) -> str:
    """Remove <think>...</think> reasoning blocks (Qwen3 / DeepSeek-R1).

    Handles three cases:
    1. Complete <think>...</think> — stripped entirely.
    2. Unclosed <think> (token-limit truncation) — everything from <think> onward
       is removed so the JSON before it (if any) survives.
    3. JSON found only inside the think block — falls back to returning original
       text so the caller can still attempt array extraction from raw content.
    """
    # Case 1: complete blocks
    stripped = re.sub(r"<think>.*?</think>", "", text, flags=re.DOTALL).strip()
    # Case 2: unclosed <think> tag — drop everything from the opening tag onward
    if "<think>" in stripped:
        stripped = stripped[: stripped.index("<think>")].strip()
    # If stripping destroyed the only JSON array, fall back to raw text
    if "[" not in stripped and "[" in text:
        stripped = re.sub(r"<think>|</think>", "", text, flags=re.DOTALL).strip()
    return stripped


def _extract_json_array(text: str) -> list:
    """
    Extract a JSON array from an LLM response that may contain prose or fences.
    Falls back through progressively more lenient strategies to handle malformed
    output from small models (missing commas/colons, trailing commas, etc.).
    """
    text = _strip_think_blocks(text)
    # Strip markdown fences
    text = re.sub(r"^```(?:json)?\s*", "", text, flags=re.MULTILINE)
    text = re.sub(r"\s*```\s*$", "", text, flags=re.MULTILINE)
    text = text.strip()

    # Isolate the outermost [...] block
    start = text.find("[")
    end   = text.rfind("]")
    if start == -1 or end == -1:
        raise ValueError(f"No JSON array found in response:\n{text[:500]}")
    raw = text[start : end + 1]

    # 1. Strict parse
    try:
        result = json.loads(raw)
        if not result:
            raise ValueError("Model returned an empty JSON array []")
        return result
    except json.JSONDecodeError:
        pass

    # 2. Common small-model fixes: trailing commas before } or ]
    fixed = re.sub(r",\s*([}\]])", r"\1", raw)
    try:
        result = json.loads(fixed)
        if not result:
            raise ValueError("Model returned an empty JSON array []")
        return result
    except json.JSONDecodeError:
        pass

    # 3. Single quotes → double quotes (naive but catches simple cases)
    try:
        fixed2 = fixed.replace("'", '"')
        result = json.loads(fixed2)
        if not result:
            raise ValueError("Model returned an empty JSON array []")
        return result
    except json.JSONDecodeError:
        pass

    # 4. Last resort: extract individual {...} objects and parse each one
    objects = []
    depth = 0
    buf   = []
    for ch in raw:
        if ch == "{":
            depth += 1
            buf.append(ch)
        elif ch == "}":
            buf.append(ch)
            depth -= 1
            if depth == 0 and buf:
                candidate = "".join(buf).strip()
                try:
                    obj = json.loads(candidate)
                    objects.append(obj)
                except json.JSONDecodeError:
                    # try trailing-comma fix on individual object
                    try:
                        obj = json.loads(re.sub(r",\s*([}\]])", r"\1", candidate))
                        objects.append(obj)
                    except json.JSONDecodeError:
                        pass
                buf = []
        elif depth > 0:
            buf.append(ch)

    if objects:
        return objects

    raise ValueError(f"Could not parse JSON array from response:\n{raw[:500]}")


def score_tickets_batch(
    batch: list[dict],
    provider: str,
    model: str,
    api_key: str,
    base_url: str,
    cb_url: str = "",
    bucket: str = "",
    username: str = "",
    password: str = "",
    use_tls: bool = False,
    scope: str = "transcripts",
    collection: str = "tickets",
    num_ctx: int | None = None,
    no_think: bool = False,
    save_to_cb: bool = False,
) -> list[dict]:
    """Send one batch to the LLM and return parsed score objects.

    If the model returns prose instead of JSON (common with smaller local models),
    a second 'repair' call asks it to convert its own output to the required array.
    """
    # Auto-force no_think for Ollama thinking models regardless of the UI toggle —
    # covers the case where settings loaded without triggering the model-change event.
    _effective_no_think = no_think or (
        provider == "ollama" and _model_has_thinking_by_name(model)
    )

    def _build_block(
        desc_limit: int = 400,
        comment_limit: int = 300,
        comment_overrides: dict | None = None,
        include_extras: bool = True,
    ) -> str:
        overrides = comment_overrides or {}
        return "\n\n".join(
            f"=== TICKET_ID: {t.get('ticket_id','?')} ===\n  "
            f"{build_scoring_input(t, desc_limit=desc_limit, comment_limit=comment_limit, comment_override=overrides.get(str(t.get('ticket_id',''))), include_extras=include_extras)}"
            for t in batch
        )

    def _make_messages(block: str) -> list[dict]:
        content = block + ("\n\n/no_think" if _effective_no_think else "")
        return [
            {"role": "system", "content": SCORING_SYSTEM_PROMPT},
            {"role": "user",   "content": content},
        ]

    # Scale output budget: ~1200 tokens per ticket (schema + enriched interaction_summary
    # now includes requester, dates, app, clusters — larger than original).
    # min 2048, max 16384.  LMStudio/Ollama reserves max_tokens from the context
    # window before reading the prompt, so this must cover the full JSON array.
    _score_max_tokens = max(2048, min(16384, len(batch) * 1200))
    _llm = lambda msgs, max_tok=_score_max_tokens: call_llm(
        msgs, provider, model, api_key, base_url,
        max_tokens=max_tok, num_ctx=num_ctx, no_think=_effective_no_think,
    )

    def _is_context_error(exc: Exception) -> bool:
        s = str(exc).lower()
        return "context length" in s or "n_keep" in s or "n_ctx" in s

    def _looks_truncated(text: str) -> bool:
        """Output was cut off: empty, bare code fence opener, or unclosed JSON array."""
        t = text.strip()
        if not t:
            return True
        # Strip markdown fence opener — if nothing remains, it's effectively empty
        inner = re.sub(r"^```(?:json)?\s*\n?", "", t).strip()
        if not inner:
            return True
        # JSON array started but never closed
        if inner.startswith("[") and not inner.rstrip().endswith("]"):
            return True
        return False

    # Proactive size cap: if the block is already huge before sending, step down
    # immediately rather than paying for a round-trip failure first.
    # 5 000 chars ≈ 1 250 tokens of user content, leaving ample room for output.
    _PROACTIVE_CHAR_LIMIT = 5000
    ticket_block = _build_block()
    if len(ticket_block) > _PROACTIVE_CHAR_LIMIT:
        ticket_block = _build_block(desc_limit=200, comment_limit=150)

    messages = _make_messages(ticket_block)

    def _call_with_empty_as_ctx_error(msgs, max_tok=_score_max_tokens):
        """Call LLM; treat empty response or truncated JSON as a context overflow."""
        try:
            result = _llm(msgs, max_tok)
        except Exception as exc:
            if _is_context_error(exc):
                raise RuntimeError("context_length") from exc
            raise
        if not result or not result.strip() or _looks_truncated(result):
            raise RuntimeError("context_length")
        return result

    try:
        raw = _call_with_empty_as_ctx_error(messages)
    except RuntimeError as ctx_exc:
        if "context_length" not in str(ctx_exc):
            raise

        # ── Tier 1 retry: tighten desc/comments ──────────────────────────────
        _LARGE_TICKET_CHARS = 1200
        summaries: dict[str, str] = {}
        for t in batch:
            tid = str(t.get("ticket_id", ""))
            rendered = build_scoring_input(t)
            if len(rendered) > _LARGE_TICKET_CHARS:
                summaries[tid] = _summarize_ticket_comments(
                    t, provider, model, api_key, base_url
                )
        ticket_block = _build_block(desc_limit=200, comment_overrides=summaries)
        messages = _make_messages(ticket_block)
        try:
            raw = _call_with_empty_as_ctx_error(messages)
        except RuntimeError as ctx_exc2:
            if "context_length" not in str(ctx_exc2):
                raise

            # ── Tier 2 retry: absolute minimum — header + subject only ────────
            ticket_block = _build_block(
                desc_limit=80,
                comment_overrides={str(t.get("ticket_id", "")): "" for t in batch},
                include_extras=False,
            )
            messages = _make_messages(ticket_block)
            raw = _llm(messages)  # propagate if still failing
    except Exception as orig_exc:
        if not _is_context_error(orig_exc):
            raise
        # Original exception was a real context error (not our sentinel)
        ticket_block = _build_block(desc_limit=80,
            comment_overrides={str(t.get("ticket_id", "")): "" for t in batch},
            include_extras=False)
        messages = _make_messages(ticket_block)
        raw = _llm(messages)

    try:
        return _extract_json_array(raw)
    except ValueError as first_err:
        # Only attempt repair if the response isn't truncated — appending a
        # truncated assistant turn makes the repair prompt even longer and fails.
        if _looks_truncated(raw):
            snippet = raw[:300].replace("\n", " ")
            raise ValueError(
                f"Initial: {first_err}. Output truncated (no repair attempted). "
                f"Model snippet: {snippet!r}"
            ) from first_err
        # Model returned prose/markdown — ask it to repair its own output
        repair_messages = messages + [
            {"role": "assistant", "content": raw},
            {"role": "user", "content": (
                "Your response was not a JSON array. "
                "Convert your analysis above into ONLY a valid JSON array "
                "matching the schema exactly — no prose, no markdown, no explanation. "
                "Start your response with [ and end with ]."
            )},
        ]
        raw2 = _llm(repair_messages)
        try:
            return _extract_json_array(raw2)
        except ValueError as second_err:
            snippet = (raw2 or raw)[:300].replace("\n", " ")
            raise ValueError(
                f"Initial: {first_err}. Repair also failed: {second_err}. "
                f"Model snippet: {snippet!r}"
            ) from second_err


def score_all_tickets(
    tickets: list[dict],
    provider: str,
    model: str,
    api_key: str,
    base_url: str,
    batch_size: int,
    progress_cb: Callable[[str, float], None],
    cancel_event: threading.Event | None = None,
    num_ctx: int | None = None,
    no_think: bool = False,
    max_workers: int = 1,
) -> dict[str, dict]:
    """
    Score all tickets in batches.  Returns a dict keyed by ticket_id.
    Errors on individual batches are logged but do not abort the run.
    max_workers > 1 dispatches multiple batches to the LLM concurrently.
    """
    import concurrent.futures
    import traceback as _tb

    # Filter out non-ticket documents (Couchbase design docs, metadata, etc.)
    tickets = [t for t in tickets if isinstance(t, dict) and t.get("ticket_id")]

    # Skip scraping-failure stubs: HTTP error pages stored as ticket subjects
    # (scraper received a 4xx/5xx response and stored the error title as subject).
    # These have no description and no comments — the LLM would invent scores.
    _HTTP_ERR_SUBJECTS = frozenset({
        "404 page not found", "403 forbidden", "401 unauthorized",
        "500 internal server error", "502 bad gateway", "503 service unavailable",
        "access denied",
    })
    def _is_scrape_stub(t: dict) -> bool:
        subj = (t.get("subject") or "").strip().lower()
        if subj in _HTTP_ERR_SUBJECTS:
            return True
        # Also skip truly empty tickets (no subject, no description, no comments)
        has_desc = bool((t.get("description") or "").strip())
        has_cmts = bool(t.get("comments"))
        return not subj and not has_desc and not has_cmts

    scoreable = [t for t in tickets if not _is_scrape_stub(t)]
    skipped   = len(tickets) - len(scoreable)
    if skipped:
        progress_cb(f"Skipping {skipped} empty/stub ticket(s) — no scoreable content.", 0.0)
    tickets = scoreable

    total   = len(tickets)
    results: dict[str, dict] = {}
    results_lock = threading.Lock()
    batches = [tickets[i : i + batch_size] for i in range(0, total, batch_size)]
    n_batches = len(batches)
    effective = max(1, min(max_workers, n_batches))

    def _run_batch(b_idx: int, batch: list[dict]) -> None:
        if cancel_event and cancel_event.is_set():
            return
        pct = b_idx / n_batches
        progress_cb(
            f"Scoring batch {b_idx + 1}/{n_batches} ({len(results)}/{total} done)…",
            pct,
        )
        try:
            scored = score_tickets_batch(
                batch, provider, model, api_key, base_url,
                num_ctx=num_ctx, no_think=no_think,
            )
            batch_by_id = {str(t.get("ticket_id", "")): t for t in batch}
            if not scored:
                progress_cb(
                    f"[SCORE] Batch {b_idx + 1} returned empty — LLM may have refused or returned unparseable output",
                    pct,
                )
            for s in scored:
                if not isinstance(s, dict):
                    if isinstance(s, str):
                        try:
                            s = json.loads(s)
                        except Exception:
                            continue
                    else:
                        continue
                tid = str(s.get("ticket_id", ""))
                if not tid or tid not in batch_by_id:
                    if tid:
                        progress_cb(
                            f"[SCORE] Batch {b_idx + 1}: skipping ID {tid!r} — not in batch",
                            pct,
                        )
                    continue
                src_ticket = batch_by_id.get(tid, {})
                try:
                    s.update(extract_cluster_snapshot_info(src_ticket))
                except Exception as merge_exc:
                    progress_cb(f"[SCORE] Ticket {tid} cluster-merge error: {merge_exc}", pct)
                # Write interaction_summary back to the ticket dict so build_embed_text
                # can include it without needing the separate scores dict.
                summary = s.get("interaction_summary", "")
                if summary:
                    src_ticket["interaction_summary"] = summary
                with results_lock:
                    results[tid] = s
        except Exception as exc:
            progress_cb(
                f"[SCORE] Batch {b_idx + 1} FAILED [{provider}/{model}]: {exc} — "
                f"{_tb.format_exc().splitlines()[-1]}",
                b_idx / n_batches,
            )

    with concurrent.futures.ThreadPoolExecutor(max_workers=effective) as pool:
        futures = [pool.submit(_run_batch, i, b) for i, b in enumerate(batches)]
        concurrent.futures.wait(futures)

    progress_cb(f"Scoring complete — {len(results)}/{total} tickets scored.", 1.0)
    return results


def backfill_analytics_fields(
    cb_url: str,
    bucket: str,
    username: str,
    password: str,
    use_tls: bool,
    scope: str,
    collection: str,
    progress_cb: Callable[[str, float], None],
) -> tuple[int, int]:
    """
    Read every ticket doc from Couchbase, compute cb_version / feature_area /
    ticket_origin if missing (or overwrite if already present), and upsert back.
    Does NOT touch the embedding field.
    Returns (updated, errors).
    """
    if not _CB_AVAILABLE:
        raise RuntimeError("couchbase SDK not installed")

    conn_str = _cb_conn_str(cb_url, use_tls)
    cluster  = Cluster(conn_str, ClusterOptions(PasswordAuthenticator(username, password)))
    cluster.wait_until_ready(timedelta(seconds=15))
    scope_obj  = cluster.bucket(bucket).scope(scope)
    col        = scope_obj.collection(collection)

    # Fetch all ticket doc keys via N1QL
    progress_cb("Fetching ticket list from Couchbase …", 0.0)
    fqn = f"`{bucket}`.`{scope}`.`{collection}`"
    rows = list(scope_obj.query(
        f"SELECT META().id AS doc_key, * FROM {fqn} WHERE META().id LIKE 'ticket::%'",
        QueryOptions(timeout=timedelta(seconds=120)),
    ))

    total   = len(rows)
    updated = errors = 0

    for i, row in enumerate(rows, 1):
        doc_key = row.get("doc_key") or row.get(collection, {}).get("doc_key")
        # N1QL wraps the doc under the collection name
        ticket  = row.get(collection) or {k: v for k, v in row.items() if k != "doc_key"}
        if not doc_key or not ticket:
            errors += 1
            continue
        try:
            ticket["cb_version"]    = extract_ticket_version(ticket)
            ticket["feature_area"]  = classify_ticket_feature(ticket)
            ticket["ticket_origin"] = classify_ticket_origin(ticket)
            col.upsert(doc_key, ticket)
            updated += 1
        except Exception as exc:
            errors += 1
            progress_cb(f"Error on {doc_key}: {exc}", i / total)
            continue
        if i % 50 == 0 or i == total:
            progress_cb(f"Backfilled {i}/{total} …", i / total)

    cluster.close()
    return updated, errors


def backfill_missing_cbse_fields(
    cb_url: str,
    bucket: str,
    username: str,
    password: str,
    use_tls: bool,
    scope: str,
    collection: str,
    progress_cb: Callable[[str, float], None],
) -> tuple[int, int]:
    """
    Find tickets where `cbses` or `jira_issues` is MISSING (never written, not null).
    For each ticket, derive cbses from ticket_fields.CBSE if present; otherwise write
    null so the field is no longer MISSING. Returns (updated, errors).
    """
    if not _CB_AVAILABLE:
        raise RuntimeError("couchbase SDK not installed")

    conn_str = _cb_conn_str(cb_url, use_tls)
    cluster  = Cluster(conn_str, ClusterOptions(PasswordAuthenticator(username, password)))
    cluster.wait_until_ready(timedelta(seconds=15))
    scope_obj = cluster.bucket(bucket).scope(scope)
    col       = scope_obj.collection(collection)
    fqn       = f"`{bucket}`.`{scope}`.`{collection}`"

    progress_cb("Querying for tickets with MISSING cbses/jira_issues …", 0.0)
    rows = list(scope_obj.query(
        f"SELECT META().id AS doc_key, * FROM {fqn} "
        f"WHERE META().id LIKE 'ticket::%' "
        f"AND (cbses IS MISSING OR jira_issues IS MISSING)",
        QueryOptions(timeout=timedelta(seconds=120)),
    ))

    total   = len(rows)
    updated = errors = 0

    if total == 0:
        progress_cb("No tickets with MISSING cbses/jira_issues found.", 1.0)
        cluster.close()
        return 0, 0

    progress_cb(f"Found {total} tickets to backfill …", 0.0)

    for i, row in enumerate(rows, 1):
        doc_key = row.get("doc_key") or row.get(collection, {}).get("doc_key")
        ticket  = row.get(collection) or {k: v for k, v in row.items() if k != "doc_key"}
        if not doc_key or not ticket:
            errors += 1
            continue
        try:
            if "cbses" not in ticket:
                tf       = ticket.get("ticket_fields") or {}
                cbse_raw = tf.get("CBSE") or tf.get("cbse") or ""
                if cbse_raw and str(cbse_raw).strip():
                    cbses = [
                        c.strip().upper()
                        for c in re.split(r"[,\s]+", str(cbse_raw))
                        if c.strip()
                    ]
                    ticket["cbses"] = cbses if cbses else None
                else:
                    ticket["cbses"] = None

            if "jira_issues" not in ticket:
                ticket["jira_issues"] = None

            col.upsert(doc_key, ticket)
            updated += 1
        except Exception as exc:
            errors += 1
            progress_cb(f"Error on {doc_key}: {exc}", i / total)
            continue
        if i % 50 == 0 or i == total:
            progress_cb(f"Backfilled {i}/{total} …", i / total)

    cluster.close()
    return updated, errors


def backfill_last_comment_at(
    cb_url: str,
    bucket: str,
    username: str,
    password: str,
    use_tls: bool,
    scope: str,
    collection: str,
    progress_cb: Callable[[str, float], None],
    org_filter: str = "",
) -> tuple[int, int]:
    """
    Derive last_comment_at from the stored comments array for tickets that are
    missing the field. Falls back to ticket_fields Updated/Last_Updated entries
    for tickets with no comments. Returns (updated, errors).
    """
    if not _CB_AVAILABLE:
        raise RuntimeError("couchbase SDK not installed")

    conn_str  = _cb_conn_str(cb_url, use_tls)
    cluster   = Cluster(conn_str, ClusterOptions(PasswordAuthenticator(username, password)))
    cluster.wait_until_ready(timedelta(seconds=15))
    scope_obj = cluster.bucket(bucket).scope(scope)
    col       = scope_obj.collection(collection)
    fqn       = f"`{bucket}`.`{scope}`.`{collection}`"

    org_clause = ""
    params: list = []
    if org_filter.strip():
        org_clause = " AND LOWER(TOSTRING(t.organization)) LIKE $1"
        params.append(f"%{org_filter.strip().lower()}%")

    progress_cb("Querying for tickets missing last_comment_at …", 0.0)
    rows = list(scope_obj.query(
        f"SELECT META().id AS doc_key, t.* FROM {fqn} AS t "
        f"WHERE META().id LIKE 'ticket::%' "
        f"AND t.last_comment_at IS MISSING{org_clause}",
        QueryOptions(positional_parameters=params, timeout=timedelta(seconds=120)),
    ))

    total   = len(rows)
    updated = errors = 0

    if total == 0:
        progress_cb("All tickets already have last_comment_at — nothing to do.", 1.0)
        cluster.close()
        return 0, 0

    progress_cb(f"Found {total} tickets to backfill …", 0.0)

    for i, row in enumerate(rows, 1):
        doc_key = row.get("doc_key")
        ticket  = {k: v for k, v in row.items() if k != "doc_key"}
        if not doc_key or not ticket.get("ticket_id"):
            errors += 1
            continue
        try:
            last_comment_at: str | None = None
            for c in reversed(ticket.get("comments") or []):
                ts = c.get("timestamp") if isinstance(c, dict) else None
                if ts:
                    last_comment_at = ts
                    break
            if not last_comment_at:
                tf = ticket.get("ticket_fields") or {}
                for _k in ("Updated", "Last_Updated", "Last_Comment", "Last_Reply", "updated"):
                    if tf.get(_k):
                        last_comment_at = tf[_k]
                        break
            ticket["last_comment_at"] = last_comment_at
            col.upsert(doc_key, ticket)
            updated += 1
        except Exception as exc:
            errors += 1
            progress_cb(f"Error on {doc_key}: {exc}", i / total)
            continue
        if i % 100 == 0 or i == total:
            progress_cb(f"Backfilled {i}/{total} …", i / total)

    cluster.close()
    return updated, errors


def embed_snapshots_from_cb(
    cb_url: str,
    bucket: str,
    username: str,
    password: str,
    use_tls: bool,
    scope: str,
    snap_collection: str,
    embed_provider: str,
    embed_model: str,
    embed_api_key: str,
    embed_base_url: str,
    vector_dims: int,
    progress_cb: Callable[[str, float], None],
    max_workers: int = 1,
) -> tuple[int, int]:
    """
    Read every snapshot doc from Couchbase, embed it, and write back only the
    `embedding` field via subdocument mutation.  Never loads snapshot data into
    UI state — safe to run against the full collection.
    Returns (done, errors).
    """
    import concurrent.futures
    from couchbase.subdocument import upsert as _SD_upsert

    if not _CB_AVAILABLE:
        raise RuntimeError("couchbase SDK not installed")

    conn_str = _cb_conn_str(cb_url, use_tls)
    progress_cb("Connecting to Couchbase …", 0.0)
    cluster = Cluster(conn_str, ClusterOptions(PasswordAuthenticator(username, password)))
    cluster.wait_until_ready(timedelta(seconds=15))
    scope_obj = cluster.bucket(bucket).scope(scope)
    col       = scope_obj.collection(snap_collection)

    fqn = f"`{bucket}`.`{scope}`.`{snap_collection}`"
    progress_cb("Fetching snapshot list …", 0.0)
    rows = list(scope_obj.query(
        f"SELECT META().id AS doc_key, * FROM {fqn} "
        f"WHERE META().id LIKE 'snapshot::%'",
        QueryOptions(timeout=timedelta(seconds=120)),
    ))

    total      = len(rows)
    done_count = errors = 0
    lock       = threading.Lock()

    progress_cb(f"Embedding {total} snapshots …", 0.0)

    def _embed_one(row: dict) -> tuple[str, list[float] | None, str | None]:
        doc_key = row.get("doc_key")
        snap    = row.get(snap_collection) or {k: v for k, v in row.items() if k != "doc_key"}
        if not doc_key or not snap:
            return doc_key or "?", None, "missing doc_key or body"
        try:
            text = build_snapshot_embed_text(snap)
            vec  = embed_text(text, embed_provider, embed_model, embed_api_key,
                              embed_base_url, dims=vector_dims)
            if vector_dims and len(vec) > vector_dims:
                vec = vec[:vector_dims]
                norm = sum(x * x for x in vec) ** 0.5
                if norm > 0:
                    vec = [x / norm for x in vec]
            return doc_key, vec, None
        except Exception as exc:
            return doc_key, None, str(exc)

    with concurrent.futures.ThreadPoolExecutor(max_workers=max(1, max_workers)) as pool:
        futs = {pool.submit(_embed_one, r): r for r in rows}
        for fut in concurrent.futures.as_completed(futs):
            doc_key, vec, err = fut.result()
            if err or not vec:
                with lock:
                    errors += 1
                progress_cb(f"Skipped {doc_key}: {err}", done_count / max(total, 1))
                continue
            try:
                col.mutate_in(doc_key, [_SD_upsert("embedding", vec)])
            except Exception as exc:
                with lock:
                    errors += 1
                progress_cb(f"Write error {doc_key}: {exc}", done_count / max(total, 1))
                continue
            with lock:
                done_count += 1
                if done_count % 25 == 0 or done_count == total:
                    progress_cb(f"Embedded {done_count}/{total} snapshots …", done_count / total)

    cluster.close()
    return done_count, errors


# ─────────────────────────── Phase 2a: Ticket Summaries ──────────────────────

# (moved to supportal/ package, imported at top of file)


def _build_summary_prompt(ticket: dict) -> str:
    topo = ticket.get("snapshot_topology") or {}
    cluster_block = ""
    if topo:
        bad  = topo.get("bad_items")  or []
        warn = topo.get("warn_items") or []
        cluster_block = (
            f"\n## Cluster Snapshot\n"
            f"- Cluster: {topo.get('cluster_name','?')} | CB {topo.get('cb_version','?')} "
            f"| {topo.get('total_nodes','?')} nodes\n"
            f"- Bad ({topo.get('bad_count',0)}): {', '.join(bad[:8]) or 'none'}\n"
            f"- Warn ({topo.get('warn_count',0)}): {', '.join(str(w) for w in warn[:8]) or 'none'}\n"
        )
    cbses  = ticket.get("cbses")  or []
    jiras  = ticket.get("jira_issues") or []
    cbse_block = ""
    if cbses or jiras:
        cbse_block = "\n## References\n"
        if cbses:
            cbse_block += f"- CBSEs: {', '.join(cbses)}\n"
        if jiras:
            cbse_block += f"- Jira: {', '.join(jiras)}\n"

    import re as _re

    def _clean(text: str) -> str:
        text = _re.sub(r"<[^>]+>", " ", text)
        return _re.sub(r"\s{2,}", " ", text).strip()

    # Prefer comments thread over description — description is often just the
    # final closure message; comments contain the full conversation.
    comments_raw = ticket.get("comments")
    if comments_raw:
        try:
            comments = json.loads(comments_raw) if isinstance(comments_raw, str) else comments_raw
            comments = sorted(comments, key=lambda c: c.get("timestamp") or "")
            lines = []
            budget = 3500
            for c in comments:
                body = _clean(c.get("body") or "").strip()
                if not body:
                    continue
                line = f"[{c.get('timestamp','')}] {c.get('author','')}: {body[:600]}"
                if len("\n".join(lines)) + len(line) > budget:
                    break
                lines.append(line)
            desc_clean = "\n".join(lines)
        except Exception:
            desc_clean = _clean(ticket.get("description") or "")[:2500]
    else:
        desc_clean = _clean(ticket.get("description") or "")[:2500]

    return _SUMMARY_PROMPT_TMPL.format(
        ticket_id   = ticket.get("ticket_id", "?"),
        organization= ticket.get("organization", "?"),
        subject     = ticket.get("subject", "?"),
        priority    = ticket.get("priority", "?"),
        status      = ticket.get("status", "?"),
        requester   = ticket.get("requester", "?"),
        created_at  = ticket.get("created_at", "?"),
        description = desc_clean,
        cluster_block = cluster_block,
        cbse_block    = cbse_block,
    )


def _parse_summary_tags(text: str) -> dict:
    """Extract CLUSTER/CB_VERSION/HEALTH/RESOLUTION tagged lines from LLM output."""
    result = {"cluster": None, "cb_version": None, "health": "unknown", "resolution": None}
    for line in text.splitlines():
        line = line.strip()
        for key, field in (
            ("CLUSTER:", "cluster"),
            ("CB_VERSION:", "cb_version"),
            ("HEALTH:", "health"),
            ("RESOLUTION:", "resolution"),
        ):
            if line.upper().startswith(key):
                val = line[len(key):].strip()
                if val and val.lower() != "unknown":
                    result[field] = val
                elif val.lower() == "unknown":
                    result[field] = None
    return result


def summarize_ticket(
    ticket: dict,
    provider: str,
    model: str,
    api_key: str,
    base_url: str,
    max_tokens: int = 512,
    max_retries: int = 3,
) -> dict:
    """
    Generate a summary document for a single ticket.
    Returns a dict ready to upsert to the summary collection.
    Retries up to max_retries times on transient connection errors.
    """
    prompt   = _build_summary_prompt(ticket)
    messages = [
        {"role": "system",  "content": _SUMMARY_SYSTEM},
        {"role": "user",    "content": prompt},
    ]
    last_exc: Exception | None = None
    for attempt in range(max(1, max_retries)):
        try:
            raw = call_llm(messages, provider, model, api_key, base_url,
                           max_tokens=max_tokens, no_think=True)
            break
        except Exception as exc:
            last_exc = exc
            err_str = str(exc).lower()
            transient = any(k in err_str for k in (
                "connection", "timeout", "reset", "eof", "broken pipe",
                "remotedisconnected", "503", "502", "429",
            ))
            if not transient or attempt >= max_retries - 1:
                raise
            time.sleep(2 ** attempt)  # 1s, 2s, 4s …
    else:
        raise last_exc  # type: ignore[misc]

    # Split prose from tagged block
    tag_start = -1
    for i, line in enumerate(raw.splitlines()):
        if line.strip().upper().startswith("CLUSTER:"):
            tag_start = i
            break
    if tag_start >= 0:
        prose = "\n".join(raw.splitlines()[:tag_start]).strip()
        tags  = _parse_summary_tags("\n".join(raw.splitlines()[tag_start:]))
    else:
        prose = raw.strip()
        tags  = _parse_summary_tags(raw)

    topo = ticket.get("snapshot_topology") or {}
    return {
        "type":                 "ticket_summary",
        "ticket_id":            str(ticket.get("ticket_id", "")),
        "organization":         ticket.get("organization"),
        "subject":              ticket.get("subject"),
        "status":               ticket.get("status"),
        "priority":             ticket.get("priority"),
        "created_at":           ticket.get("created_at"),
        "summary_text":         prose,
        "cluster_name":         tags["cluster"] or topo.get("cluster_name"),
        "cb_version":           tags["cb_version"] or topo.get("cb_version"),
        "health":               tags["health"] or "unknown",
        "resolution":           tags["resolution"],
        "cbses":                ticket.get("cbses") or [],
        "jira_issues":          ticket.get("jira_issues") or [],
        "source_last_scraped_at": ticket.get("last_scraped_at"),
        "generated_at":         int(time.time()),
        "model":                f"{provider}/{model}",
    }


def summarize_tickets_from_cb(
    cb_url: str,
    bucket: str,
    username: str,
    password: str,
    use_tls: bool,
    scope: str,
    collection: str,
    summary_collection: str,
    provider: str,
    model: str,
    api_key: str,
    base_url: str,
    progress_cb: Callable[[str, float], None],
    customer_filter: str = "",
    force: bool = False,
    max_workers: int = 1,
) -> tuple[int, int]:
    """
    Read tickets from CB, generate summaries via LLM, write to summary collection.
    Skips tickets that already have a summary unless force=True.
    Returns (done, errors).
    """
    import concurrent.futures

    if not _CB_AVAILABLE:
        raise RuntimeError("couchbase SDK not installed")

    conn_str = _cb_conn_str(cb_url, use_tls)
    progress_cb("Connecting to Couchbase …", 0.0)
    cluster   = Cluster(conn_str, ClusterOptions(PasswordAuthenticator(username, password)))
    cluster.wait_until_ready(timedelta(seconds=15))
    scope_obj = cluster.bucket(bucket).scope(scope)
    src_col   = scope_obj.collection(collection)
    sum_col   = scope_obj.collection(summary_collection)

    fqn = f"`{bucket}`.`{scope}`.`{collection}`"
    where_parts = ["ticket_id IS NOT MISSING"]
    params: list = []
    if customer_filter.strip():
        where_parts.append("LOWER(organization) LIKE $1")
        params.append(f"%{customer_filter.strip().lower()}%")
    where = " AND ".join(where_parts)

    progress_cb("Fetching ticket list …", 0.0)
    rows = list(scope_obj.query(
        f"SELECT META().id AS doc_key, ticket_id, organization, subject, status, priority, "
        f"created_at, last_scraped_at, requester, cbses, jira_issues, snapshot_topology, "
        f"SUBSTR(description, 0, 3000) AS description, comments "
        f"FROM {fqn} WHERE {where}",
        QueryOptions(positional_parameters=params, timeout=timedelta(seconds=120)),
    ))

    if not force:
        # Filter to tickets without an existing summary
        progress_cb("Checking existing summaries …", 0.0)
        sum_fqn = f"`{bucket}`.`{scope}`.`{summary_collection}`"
        existing_q = list(scope_obj.query(
            f"SELECT META().id AS doc_key FROM {sum_fqn} WHERE type = 'ticket_summary'",
            QueryOptions(timeout=timedelta(seconds=60)),
        ))
        existing_keys = {r.get("doc_key") for r in existing_q}
        rows = [r for r in rows
                if f"summary::{r.get('ticket_id','')}" not in existing_keys]

    total      = len(rows)
    done_count = errors = 0
    lock       = threading.Lock()

    progress_cb(f"Summarizing {total} tickets …", 0.0)

    def _summarize_one(row: dict) -> tuple[str, dict | None, str | None]:
        doc_key    = row.get("doc_key", "")
        sum_key    = f"summary::{row.get('ticket_id','')}"
        ticket     = {k: v for k, v in row.items() if k != "doc_key"}
        try:
            summary = summarize_ticket(ticket, provider, model, api_key, base_url)
            return sum_key, summary, None
        except Exception as exc:
            return sum_key, None, str(exc)

    with concurrent.futures.ThreadPoolExecutor(max_workers=max(1, max_workers)) as pool:
        futs = {pool.submit(_summarize_one, r): r for r in rows}
        for fut in concurrent.futures.as_completed(futs):
            sum_key, summary, err = fut.result()
            if err or not summary:
                with lock:
                    errors += 1
                progress_cb(f"Skipped {sum_key}: {err}", done_count / max(total, 1))
                continue
            try:
                sum_col.upsert(sum_key, summary)
            except Exception as exc:
                with lock:
                    errors += 1
                progress_cb(f"Write error {sum_key}: {exc}", done_count / max(total, 1))
                continue
            with lock:
                done_count += 1
                if done_count % 10 == 0 or done_count == total:
                    progress_cb(
                        f"Summarized {done_count}/{total} …", done_count / total
                    )

    cluster.close()
    return done_count, errors


def fetch_ticket_summary(
    ticket_id: str,
    cb_url: str,
    bucket: str,
    username: str,
    password: str,
    use_tls: bool,
    scope: str,
    summary_collection: str,
) -> dict | None:
    """Fetch a single summary doc by ticket ID. Returns None if not found."""
    if not _CB_AVAILABLE:
        return None
    try:
        conn_str = _cb_conn_str(cb_url, use_tls)
        cluster  = Cluster(conn_str, ClusterOptions(PasswordAuthenticator(username, password)))
        cluster.wait_until_ready(timedelta(seconds=10))
        col  = cluster.bucket(bucket).scope(scope).collection(summary_collection)
        doc  = col.get(f"summary::{ticket_id}").content_as[dict]
        cluster.close()
        return doc
    except Exception:
        return None


def _org_slug(org: str) -> str:
    """Convert an org name to a lowercase underscore-separated slug for use as a doc key."""
    import re as _re
    return _re.sub(r"[^a-z0-9]+", "_", org.lower()).strip("_")


def rescore_all_customers_cb(
    cb_url: str,
    bucket: str,
    username: str,
    password: str,
    use_tls: bool,
    scope: str,
    collection: str,
    llm_provider: str,
    llm_model: str,
    llm_api_key: str,
    llm_base_url: str,
    batch_size: int,
    progress_cb: Callable[[str, float], None],
    cancel_event: threading.Event | None = None,
    max_workers: int = 1,
) -> tuple[int, int]:
    """
    Load every ticket doc from Couchbase, group by organization, then score and
    persist one customer at a time.  Returns (total_scored, total_errors).
    """
    import datetime
    if not _CB_AVAILABLE:
        raise RuntimeError("couchbase SDK not installed")

    conn_str  = _cb_conn_str(cb_url, use_tls)
    cluster   = Cluster(conn_str, ClusterOptions(PasswordAuthenticator(username, password)))
    cluster.wait_until_ready(timedelta(seconds=15))
    scope_obj = cluster.bucket(bucket).scope(scope)
    col       = scope_obj.collection(collection)

    fqn = f"`{bucket}`.`{scope}`.`{collection}`"
    progress_cb("Fetching all ticket docs from Couchbase …", 0.0)
    rows = list(scope_obj.query(
        f"SELECT META().id AS doc_key, * FROM {fqn} WHERE META().id LIKE 'ticket::%'",
        QueryOptions(timeout=timedelta(seconds=180)),
    ))
    progress_cb(f"Loaded {len(rows)} ticket docs — grouping by customer …", 0.02)

    # Group rows by organization
    customers: dict[str, list[tuple[str, dict]]] = {}
    for row in rows:
        doc_key = row.get("doc_key") or ""
        ticket  = row.get(collection) or {k: v for k, v in row.items() if k != "doc_key"}
        if not doc_key or not ticket:
            continue
        org = (ticket.get("organization") or "Unknown").strip() or "Unknown"
        customers.setdefault(org, []).append((doc_key, ticket))

    org_list   = sorted(customers.keys())
    total_orgs = len(org_list)
    scored_total = errors_total = 0
    error_log: list[str] = []
    scored_at    = datetime.datetime.now(datetime.UTC).strftime("%Y-%m-%dT%H:%M:%SZ")

    for org_idx, org in enumerate(org_list):
        if cancel_event and cancel_event.is_set():
            progress_cb(f"Cancelled — {scored_total} scored across {org_idx} customers.", org_idx / total_orgs)
            break
        pairs   = customers[org]
        tickets = [t for _, t in pairs]
        key_map = {str(t.get("ticket_id", "")): dk for dk, t in pairs}
        org_pct_base = org_idx / total_orgs

        def _org_prog(msg: str, pct: float):
            overall = org_pct_base + pct / total_orgs
            progress_cb(
                f"[{org_idx + 1}/{total_orgs}] {org} — {msg}",
                min(overall, 0.99),
            )

        _org_prog(f"scoring {len(tickets)} tickets …", 0.0)

        # Score in batches, with optional parallelism
        batches = [tickets[i: i + batch_size] for i in range(0, len(tickets), batch_size)]
        scores: dict[str, dict] = {}
        scores_lock = threading.Lock()
        effective_workers = max(1, min(max_workers, len(batches)))
        completed_batches = [0]

        def _run_org_batch(b_idx: int, batch: list[dict]) -> None:
            if cancel_event and cancel_event.is_set():
                return
            try:
                scored = score_tickets_batch(
                    batch, llm_provider, llm_model, llm_api_key, llm_base_url,
                )
                with scores_lock:
                    for s in scored:
                        tid = str(s.get("ticket_id", ""))
                        if tid:
                            scores[tid] = s
                            # Write summary back to ticket dict in-place
                            summary = s.get("interaction_summary", "")
                            if summary:
                                batch_map = {str(t.get("ticket_id", "")): t for t in batch}
                                if tid in batch_map:
                                    batch_map[tid]["interaction_summary"] = summary
            except Exception as exc:
                with scores_lock:
                    errors_total_ref[0] += len(batch)
                    err_msg = f"[{org}] batch {b_idx + 1}: {exc}"
                    error_log.append(err_msg)
            finally:
                with scores_lock:
                    completed_batches[0] += 1
                    _org_prog(
                        f"batch {completed_batches[0]}/{len(batches)} …",
                        completed_batches[0] / max(len(batches), 1) * 0.7,
                    )

        errors_total_ref = [errors_total]
        import concurrent.futures as _cf_rescore
        with _cf_rescore.ThreadPoolExecutor(max_workers=effective_workers) as pool:
            futs = [pool.submit(_run_org_batch, i, b) for i, b in enumerate(batches)]
            _cf_rescore.wait(futs)
        errors_total = errors_total_ref[0]

        # Persist scores back to CB — merge into existing score to preserve
        # cluster_names, cluster_ids, snapshot_count, last_snapshot_id, etc.
        for p_idx, (tid, score_data) in enumerate(scores.items()):
            doc_key = key_map.get(tid) or f"ticket::{tid}"
            try:
                result  = col.get(doc_key)
                doc     = result.content_as[dict]
                existing_score = doc.get("score") or {}
                doc["score"] = {**existing_score, **score_data, "scored_at": scored_at}
                col.upsert(doc_key, doc)
                scored_total += 1
            except Exception as exc:
                errors_total += 1
                err_msg = f"[{org}] save error ticket {tid}: {exc}"
                error_log.append(err_msg)
                _org_prog(f"save error {tid}: {exc}", 0.7 + p_idx / max(len(scores), 1) * 0.3)

        org_errs = sum(1 for e in error_log if e.startswith(f"[{org}]"))
        _org_prog(f"done — {len(scores)} scored.", 1.0)

        # Update per-org inventory doc (best-effort)
        try:
            inv_slug = _org_slug(org)
            inv_doc = {
                "type": "customer_inventory",
                "organization": org,
                "updated_at": datetime.datetime.now(datetime.timezone.utc).isoformat().replace("+00:00", "Z"),
                "pipeline": {
                    "score": {
                        "at": scored_at,
                        "done": len(scores),
                        "total": len(tickets),
                        "errors": org_errs,
                    }
                },
            }
            col.upsert(f"inventory::{inv_slug}", inv_doc)
        except Exception:
            pass

    progress_cb(
        f"Bulk rescore complete — {scored_total} scored, {errors_total} errors across {total_orgs} customers.",
        1.0,
    )
    cluster.close()
    return scored_total, errors_total, error_log


def recover_score_cluster_fields_cb(
    cb_url: str, bucket: str, username: str, password: str,
    use_tls: bool, scope: str, collection: str,
    progress_cb: Callable[[str, float], None],
) -> tuple[int, int]:
    """Restore score.cluster_names/cluster_ids/snapshot_count/last_snapshot_id
    for tickets where those fields were wiped by a broken rescore run.

    Reads raw ticket fields (snapshots, ticket_fields, description, comments,
    snapshot_topology) and calls extract_cluster_snapshot_info() — no LLM.
    Only updates tickets that have a score but are missing cluster_names.
    Returns (recovered, errors).
    """
    if not _CB_AVAILABLE:
        raise RuntimeError("couchbase SDK not installed")
    conn_str = _cb_conn_str(cb_url, use_tls)
    cl  = Cluster(conn_str, ClusterOptions(PasswordAuthenticator(username, password)))
    cl.wait_until_ready(timedelta(seconds=15))
    col = cl.bucket(bucket).scope(scope).collection(collection)
    keyspace = f"`{bucket}`.`{scope}`.`{collection}`"

    # Fetch all tickets with a score but missing cluster_names
    rows = list(cl.query(
        f"SELECT META(t).id AS doc_id FROM {keyspace} AS t "
        f"WHERE t.score IS NOT MISSING "
        f"AND (t.score.cluster_names IS MISSING OR t.score.cluster_names IS NULL)",
        QueryOptions(timeout=timedelta(seconds=60)),
    ))
    total = len(rows)
    progress_cb(f"Found {total} tickets with missing cluster fields …", 0.0)

    recovered = errors = 0
    for i, row in enumerate(rows):
        doc_key = row.get("doc_id", "")
        if not doc_key:
            continue
        try:
            result = col.get(doc_key)
            doc    = result.content_as[dict]
            # Recompute cluster fields from raw ticket data (deterministic, no LLM)
            cluster_info = extract_cluster_snapshot_info(doc)
            # Also supplement with text-scan via _ticket_cluster_ids (reads topo + text)
            for cname in _ticket_cluster_ids(doc):
                if cname and cname not in cluster_info["cluster_names"]:
                    cluster_info["cluster_names"].append(cname)
            existing_score = doc.get("score") or {}
            existing_score.update({
                "cluster_names":    cluster_info["cluster_names"],
                "cluster_ids":      cluster_info["cluster_ids"],
                "snapshot_count":   cluster_info["snapshot_count"],
                "last_snapshot_id": cluster_info["last_snapshot_id"],
            })
            doc["score"] = existing_score
            col.upsert(doc_key, doc)
            recovered += 1
        except Exception as exc:
            errors += 1
            print(f"[recover_cluster_fields] {doc_key}: {exc}")
        if i % 100 == 0:
            progress_cb(f"Recovered {recovered}/{total} …", i / max(total, 1))

    cl.close()
    progress_cb(f"Done — {recovered} recovered, {errors} errors.", 1.0)
    return recovered, errors


def upsert_inventory_doc(
    org: str,
    pipeline_steps: dict,
    cb_url: str,
    bucket: str,
    username: str,
    password: str,
    use_tls: bool,
    scope: str,
    collection: str,
) -> None:
    """Write/merge a customer_inventory doc to Couchbase (best-effort, does not raise)."""
    import datetime as _dt
    if not _CB_AVAILABLE or not cb_url:
        return
    try:
        conn_str = _cb_conn_str(cb_url, use_tls)
        cluster  = Cluster(conn_str, ClusterOptions(PasswordAuthenticator(username, password)))
        cluster.wait_until_ready(timedelta(seconds=15))
        col = cluster.bucket(bucket).scope(scope).collection(collection)
        key = f"inventory::{_org_slug(org)}"
        try:
            inv_doc = col.get(key).content_as[dict]
        except Exception:
            inv_doc = {"type": "customer_inventory", "organization": org, "pipeline": {}}
        inv_doc["updated_at"] = (
            _dt.datetime.now(_dt.timezone.utc).isoformat().replace("+00:00", "Z")
        )
        inv_doc.setdefault("pipeline", {}).update(pipeline_steps)
        col.upsert(key, inv_doc)
        cluster.close()
    except Exception:
        pass


def persist_scores_to_cb(
    scores: dict[str, dict],
    cb_url: str,
    bucket: str,
    username: str,
    password: str,
    use_tls: bool,
    scope: str,
    collection: str,
    progress_cb: Callable[[str, float], None],
    tickets: list[dict] | None = None,
) -> tuple[int, int]:
    """
    Merge score data into each existing ticket document in Couchbase.
    Adds a top-level `score` field with all LLM-generated metrics + a `scored_at` timestamp.
    If a ticket document is missing and `tickets` is provided, the full ticket is saved first
    before writing the score.
    Returns (saved, errors).
    """
    import datetime
    if not _CB_AVAILABLE:
        raise RuntimeError("couchbase SDK not installed")

    conn_str = _cb_conn_str(cb_url, use_tls)
    cluster  = Cluster(conn_str, ClusterOptions(PasswordAuthenticator(username, password)))
    cluster.wait_until_ready(timedelta(seconds=15))
    col = cluster.bucket(bucket).scope(scope).collection(collection)

    # Build a fast lookup from ticket_id → ticket dict if provided
    _ticket_lookup: dict[str, dict] = {}
    if tickets:
        for t in tickets:
            _tid = str(t.get("ticket_id", ""))
            if _tid:
                _ticket_lookup[_tid] = t

    total = len(scores)
    saved = errors = 0
    scored_at = datetime.datetime.now(datetime.UTC).strftime("%Y-%m-%dT%H:%M:%SZ")

    for i, (tid, score_data) in enumerate(scores.items(), 1):
        doc_key = f"ticket::{tid}"
        try:
            result = col.get(doc_key)
            doc    = result.content_as[dict]
            existing_score = doc.get("score") or {}
            doc["score"] = {**existing_score, **score_data, "scored_at": scored_at}
            col.upsert(doc_key, doc)
            saved += 1
        except CouchbaseException as exc:
            if "document_not_found" in str(exc) or "KEY_ENOENT" in str(exc):
                # Ticket missing from CB — save it first if we have the data, then apply score
                base_doc = dict(_ticket_lookup.get(str(tid), {"ticket_id": tid, "_stub": True}))
                existing_score = base_doc.get("score") or {}
                base_doc["score"] = {**existing_score, **score_data, "scored_at": scored_at}
                try:
                    col.upsert(doc_key, base_doc)
                    recovered = "_stub" not in base_doc
                    progress_cb(
                        f"Ticket {tid} was missing — {'re-saved from session data' if recovered else 'stub created'} with score.",
                        i / total,
                    )
                    saved += 1
                except Exception as inner_exc:
                    errors += 1
                    progress_cb(f"Error saving score for {tid}: {inner_exc}", i / total)
            else:
                errors += 1
                progress_cb(f"Error saving score for {tid}: {exc}", i / total)
            continue
        if i % 25 == 0 or i == total:
            progress_cb(f"Saved {i}/{total} scores …", i / total)

    cluster.close()
    return saved, errors


def load_scores_from_cb(
    ticket_ids: list[str],
    cb_url: str,
    bucket: str,
    username: str,
    password: str,
    use_tls: bool,
    scope: str,
    collection: str,
    progress_cb: Callable[[str, float], None],
) -> dict[str, dict]:
    """
    Read the `score` field from each ticket document in Couchbase.
    Returns a dict keyed by ticket_id for any docs that have a score field.
    """
    if not _CB_AVAILABLE:
        raise RuntimeError("couchbase SDK not installed")

    conn_str = _cb_conn_str(cb_url, use_tls)
    cluster  = Cluster(conn_str, ClusterOptions(PasswordAuthenticator(username, password)))
    cluster.wait_until_ready(timedelta(seconds=15))
    col = cluster.bucket(bucket).scope(scope).collection(collection)

    total   = len(ticket_ids)
    results: dict[str, dict] = {}

    for i, tid in enumerate(ticket_ids, 1):
        doc_key = f"ticket::{tid}"
        try:
            result = col.get(doc_key)
            doc    = result.content_as[dict]
            if "score" in doc:
                results[tid] = doc["score"]
        except Exception:
            pass
        if i % 100 == 0 or i == total:
            progress_cb(f"Loading scores {i}/{total} …", i / total)

    cluster.close()
    return results


# ── Analytics helpers ─────────────────────────────────────────────────────────



_EOL_FIELD_PAIRS = [
    # (normalized current field,                    normalized EOL variant)
    ("Couchbase_Server",                            "Couchbase_Server_EOL"),
    ("Couchbase_Lite",                              "Couchbase_Lite_EOL"),
    ("Couchbase_Analytics",                         "Couchbase_Analytics_EOL"),
    ("Couchbase_Analytics_SDK",                     "Couchbase_Analytics_SDK_EOL"),
    ("Couchbase_Sync_Gateway",                      "Couchbase_Sync_Gateway_EOL"),
    ("Couchbase_Autonomous_Operator",               "Couchbase_Autonomous_Operator_EOL"),
    ("Couchbase_Connector",                         "Couchbase_Connector_EOL"),
    ("Couchbase_Edge_Server",                       "Couchbase_Edge_Server_EOL"),
    ("Couchbase_Server_SDK_or_Connector",           "Couchbase_Server_SDK_or_Connector_EOL"),
]

# Values that look like versions but aren't — Zendesk dropdown placeholders
_NON_VERSION_VALS = {
    "unknown", "n/a", "-", "",
    "end of life",
    "end of life (please specify in the ticket details)",
    "end of life (please specify)",
}


def _is_usable_version(v: str) -> bool:
    """True if v looks like an actual version number (contains a digit, not a placeholder)."""
    return bool(v and re.search(r"\d", v) and v.lower().strip() not in _NON_VERSION_VALS)


def extract_ticket_version(ticket: dict) -> str:
    """
    Return the best available version string from a ticket.

    Priority order:
      1. Ticket custom fields — check each product's current field first, then
         its (EOL) variant (for tickets where the version is End-of-Life and the
         agent filled in the actual version number in the EOL field instead).
         "End of Life" dropdown values are ignored — only real version numbers
         (containing at least one digit) are accepted.
      2. Tags (cb_7_2, server__cb_7_6_0, etc.)
      3. snapshot_topology.cb_version (cluster version at time of snapshot)
      4. 'Unknown'

    Returns the full version string (e.g. '7.6.3') rather than truncating to
    major.minor, so the document preserves precision for filtering/analytics.
    """
    fields = _parse_ticket_fields(ticket)

    for current_field, eol_field in _EOL_FIELD_PAIRS:
        v = (fields.get(current_field) or "").strip()
        if _is_usable_version(v):
            return v
        # Main field was empty or an EOL-placeholder — try the EOL variant
        v_eol = (fields.get(eol_field) or "").strip()
        if _is_usable_version(v_eol):
            return v_eol

    # Tags fallback: cb_7_2, server__cb_7_6_0_enterprise, etc.
    tags = (ticket.get("tags") or "").lower()
    m = re.search(r'cb[_-](\d+)[_-](\d+)(?:[_-](\d+))?', tags)
    if m:
        parts = [m.group(1), m.group(2)]
        if m.group(3):
            parts.append(m.group(3))
        return ".".join(parts)

    # Snapshot topology fallback — covers tickets where customer left fields blank
    topo = ticket.get("snapshot_topology")
    if topo:
        if isinstance(topo, str):
            try:
                topo = json.loads(topo)
            except Exception:
                topo = {}
        if isinstance(topo, dict):
            snap_ver = _topo_str(topo.get("cb_version"))
            if snap_ver:
                # Strip build suffix: "7.2.3-6705-enterprise" → "7.2.3"
                m2 = re.match(r"(\d+\.\d+(?:\.\d+)?)", snap_ver)
                return m2.group(1) if m2 else snap_ver

    return "Unknown"


def classify_ticket_feature(ticket: dict) -> str:
    """
    Classify which Couchbase feature area the ticket relates to.
    Uses Component field (parsed as Product::Layer::Sub hierarchy), SDK/SGW
    flags, tags, and subject as signals.

    Capella tickets are sub-classified using the Component hierarchy and
    component__ tag patterns rather than lumped into one generic bucket.
    """
    fields  = _parse_ticket_fields(ticket)
    comp    = (fields.get("Component") or "").lower()
    sdk     = (fields.get("SDK_Related") or "").lower()
    sgw     = (fields.get("SGW_Related") or "").lower()
    tags    = (ticket.get("tags") or "").lower()
    subject = (ticket.get("subject") or "").lower()

    # Detect Capella platform
    is_capella = (
        "capella" in comp
        or "product_type__capella" in tags
        or "capella" in subject
    )

    # SDK check applies regardless of platform
    if sdk == "yes" or "sdk_related__yes" in tags or "sdk" in comp:
        return "SDK"

    if is_capella:
        # Parse sub-component from Component field: "Capella::dataplane::couchbase Server"
        # → sub = "dataplane couchbase server"
        sub = ""
        if "::" in comp:
            parts = [p.strip() for p in comp.split("::")]
            sub = " ".join(parts[1:])   # everything after "capella"
        # Also extract from tag pattern: component__capella__dataplane__couchbase_server
        if not sub:
            m = re.search(r'component__capella__([\w]+(?:__[\w]+)*)', tags)
            if m:
                sub = m.group(1).replace("__", " ")

        if "app services" in sub or "appservices" in sub or sgw == "yes":
            return "Capella — App Services"
        if "control plane" in sub or "controlplane" in sub or "app plane" in sub:
            return "Capella — Control Plane"
        if "analytics" in sub or "columnar" in sub:
            return "Capella — Analytics"
        if "search" in sub or "fts" in sub:
            return "Capella — Search"
        if "query" in sub or "n1ql" in sub:
            return "Capella — Query"
        if "index" in sub:
            return "Capella — Indexing"
        if "eventing" in sub:
            return "Capella — Eventing"
        if "xdcr" in sub:
            return "Capella — XDCR"
        # "dataplane couchbase server" or just "dataplane" → Server layer
        if "couchbase server" in sub or "dataplane" in sub or "data plane" in sub:
            return "Capella — Server"
        return "Capella / Cloud Platform"

    # Non-Capella service checks
    if (sgw == "yes" or "sgw_related" in tags or "sync_gateway" in comp
            or "sync gateway" in subject or "app services" in subject):
        return "Sync Gateway / App Services"
    if ("analytics" in comp or "columnar" in comp
            or "analytics" in subject or "component__analytics" in tags):
        return "Analytics / Columnar"
    if "xdcr" in comp or "xdcr" in tags or "xdcr" in subject:
        return "XDCR"
    if ("search" in comp or "fts" in comp
            or "component__search" in tags or "full-text" in subject or "fts" in subject):
        return "Full-Text Search"
    if ("query" in comp or "n1ql" in comp
            or "component__query" in tags or "component__n1ql__" in tags
            or "component__server__query" in tags
            or "n1ql" in subject or "sql++" in subject):
        return "Query / N1QL / SQL++"
    if ("index" in comp or "component__index" in tags
            or "component__server__secondary_index" in tags or "index" in subject):
        return "Indexing"
    if ("kv" in comp or "key_value" in comp or "bucket" in comp
            or "component__server__couchbase_bucket" in tags
            or "component__server__service__data" in tags or "key value" in subject):
        return "Data / KV"
    if "eventing" in comp or "eventing" in tags or "eventing" in subject:
        return "Eventing"
    if "component__client__" in tags or "component__framework_and_library__" in tags:
        return "SDK"
    if "component__sync_gateway__" in tags:
        return "Sync Gateway / App Services"
    if "component__server__tools__cbbackup" in tags or "component__server__tools__cbrestore" in tags:
        return "Backup / Restore"
    if ("component__server__tools__" in tags or "component__tools__cbtools" in tags):
        return "CLI Tools"
    if "component__server__cluster_manager" in tags:
        return "Cluster Management"
    if "component__server__rebalance" in tags:
        return "Rebalance / Cluster Ops"
    if "component__server__installer" in tags:
        return "Installation / Upgrade"
    if "component__server__certificates" in tags or "component__server__ldap" in tags:
        return "Security / Certificates"
    if "component__server__ui" in tags:
        return "Web UI"
    if "component__server__views" in tags:
        return "Views"
    if "component__server__storage_engine__" in tags:
        return "Storage Engine"
    if "component__server__rest_api" in tags:
        return "REST API"
    if "component__kubernetes_operator__" in tags:
        return "Kubernetes / Operator"
    if "component__connector__" in tags:
        return "Connectors"
    if "component__lite__" in tags:
        return "Couchbase Lite"
    if "component__documentation" in tags:
        return "Documentation"
    if "component__other__sizing" in tags:
        return "Sizing / Capacity Planning"
    if "component__other__networking" in tags:
        return "Networking"
    return "Other / General"


def _parse_created(tickets: list[dict]) -> list[str]:
    """Return YYYY-MM strings for each ticket that has a parseable created date."""
    import datetime
    months = []
    for t in tickets:
        raw = (t.get("created") or "").strip()
        if not raw:
            continue
        dt = None
        # Fast path: ISO 8601 (Zendesk/Couchbase standard)
        try:
            dt = datetime.datetime.fromisoformat(raw[:19])
        except ValueError:
            for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%B %d, %Y", "%b %d, %Y"):
                try:
                    dt = datetime.datetime.strptime(raw[:10], fmt)
                    break
                except ValueError:
                    continue
        if dt:
            months.append(f"{dt.year}-{dt.month:02d}")
    return months


def _has_carr_tag(ticket: dict) -> bool:
    """Return True if the Zendesk 'carr' revenue-tier tag is present."""
    return "carr" in (ticket.get("tags") or "").lower().split()


def build_customer_profile(
    tickets: list[dict],
    scores: dict[str, dict],
    org: str,
) -> dict:
    """
    Build a rich single-customer profile dict for deep-dive charts.
    Returns a flat dict consumed by the customer profile UI section.
    """
    import datetime
    from collections import Counter, defaultdict

    def _parse_dt(raw: str):
        if not raw:
            return None
        raw = raw.strip()
        # Fast path: ISO 8601 datetime (Zendesk/Couchbase standard: YYYY-MM-DDTHH:MM:SS[Z/±tz])
        try:
            return datetime.datetime.fromisoformat(raw[:19])
        except ValueError:
            pass
        # Fallback: date-only and locale formats
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%B %d, %Y", "%b %d, %Y"):
            try:
                return datetime.datetime.strptime(raw[:10], fmt)
            except ValueError:
                continue
        return None

    org_tickets = [t for t in tickets
                   if (t.get("organization") or "").strip() == org]
    if not org_tickets:
        return {}

    # ── CARR detection ────────────────────────────────────────────────────────
    is_carr = any(_has_carr_tag(t) for t in org_tickets)
    carr_first_year = None
    if is_carr:
        carr_years = []
        for t in org_tickets:
            if _has_carr_tag(t):
                dt = _parse_dt((t.get("created") or "").strip())
                if dt:
                    carr_years.append(dt.year)
        if carr_years:
            carr_first_year = min(carr_years)

    # ── Dates ─────────────────────────────────────────────────────────────────
    dated = []
    for t in org_tickets:
        dt = _parse_dt((t.get("created") or "").strip())
        if dt:
            dated.append((dt, t))

    dates = [d for d, _ in dated]
    first_ticket = min(dates).strftime("%Y-%m-%d") if dates else None
    last_ticket  = max(dates).strftime("%Y-%m-%d") if dates else None

    # ── Volume by month / year ────────────────────────────────────────────────
    month_counts: Counter = Counter(d.strftime("%Y-%m") for d in dates)
    month_keys = sorted(month_counts.keys())
    year_counts: Counter  = Counter(str(d.year) for d in dates)
    year_keys  = sorted(year_counts.keys())

    # ── Priority per month (stacked) ──────────────────────────────────────────
    pri_by_month: dict[str, Counter] = defaultdict(Counter)
    for dt, t in dated:
        p = (t.get("priority") or "Unknown").capitalize()
        pri_by_month[dt.strftime("%Y-%m")][p] += 1
    all_priorities = sorted({p for pc in pri_by_month.values() for p in pc})

    # ── Unique clusters ───────────────────────────────────────────────────────
    clusters: set[str] = set()
    for t in org_tickets:
        topo = t.get("snapshot_topology")
        if isinstance(topo, dict):
            name = _topo_str(topo.get("cluster_name"))
            uuid = _topo_str(topo.get("cluster_uuid"))
            if name:
                clusters.add(name)
            elif uuid:
                clusters.add(uuid)

    # ── Average resolution time (solved/closed tickets only) ──────────────────
    resolution_days: list[float] = []
    for t in org_tickets:
        if (t.get("status") or "").lower() not in ("solved", "closed"):
            continue
        c_dt = _parse_dt((t.get("created") or "").strip())
        # Prefer explicit solved date; fall back to updated
        closed_raw = (t.get("solved") or t.get("updated") or "").strip()
        u_dt = _parse_dt(closed_raw)
        if c_dt and u_dt:
            d = (u_dt - c_dt).days
            if 0 <= d <= 730:          # cap at 2 years — sanity guard
                resolution_days.append(d)
    avg_resolution_days = (
        round(sum(resolution_days) / len(resolution_days), 1)
        if resolution_days else None
    )

    # ── Satisfaction trend per month ──────────────────────────────────────────
    stars_by_month: dict[str, list[float]] = defaultdict(list)
    for dt, t in dated:
        tid = str(t.get("ticket_id", ""))
        sc  = scores.get(tid, {})
        if sc.get("stars"):
            stars_by_month[dt.strftime("%Y-%m")].append(float(sc["stars"]))
    stars_trend_keys   = [k for k in month_keys if k in stars_by_month]
    stars_trend_values = [
        round(sum(stars_by_month[k]) / len(stars_by_month[k]), 2)
        for k in stars_trend_keys
    ]

    # ── Feature area breakdown ────────────────────────────────────────────────
    feature_counts: Counter = Counter(
        classify_ticket_feature(t) for t in org_tickets
    )

    # ── CB version distribution ───────────────────────────────────────────────
    version_counts: Counter = Counter()
    _p_eol_count   = 0
    _p_admin_count = 0
    _p_blank_count = 0
    for t in org_tickets:
        ver = extract_ticket_version(t)
        if ver and ver != "Unknown":
            version_counts[ver] += 1
        else:
            _tf_p = _parse_ticket_fields(t)
            _cs_p = (_tf_p.get("Couchbase_Server") or "").strip().lower()
            if "Couchbase_Server" not in _tf_p:
                _p_admin_count += 1
            elif "end of life" in _cs_p:
                _p_eol_count += 1
            else:
                _p_blank_count += 1

    _prof_ver_breakdown: list[tuple[str, int, str]] = [
        (v, c, "version") for v, c in version_counts.most_common(12)
    ]
    if _p_eol_count:
        _prof_ver_breakdown.append(("EOL / End of Life", _p_eol_count, "eol"))
    if _p_admin_count:
        _prof_ver_breakdown.append(("Admin / No Product Fields", _p_admin_count, "admin"))
    if _p_blank_count:
        _prof_ver_breakdown.append(("Version Not Specified", _p_blank_count, "blank"))

    return {
        "org":                  org,
        "ticket_count":         len(org_tickets),
        "is_carr":              is_carr,
        "carr_first_year":      carr_first_year,
        "first_ticket":         first_ticket,
        "last_ticket":          last_ticket,
        "unique_clusters":      len(clusters),
        "cluster_list":         sorted(clusters)[:20],
        "avg_resolution_days":  avg_resolution_days,
        "escalation_count":     sum(1 for t in org_tickets if t.get("escalations")),
        "proactive_count":      sum(1 for t in org_tickets if _is_proactive_ticket(t)),
        # Volume
        "year_keys":            year_keys,
        "year_values":          [year_counts[y] for y in year_keys],
        "month_keys":           month_keys,
        "month_values":         [month_counts[k] for k in month_keys],
        # Priority stacked by month
        "pri_month_keys":       month_keys,
        "all_priorities":       all_priorities,
        "pri_by_month":         {p: [pri_by_month[k].get(p, 0) for k in month_keys]
                                 for p in all_priorities},
        # Stars trend
        "stars_trend_keys":     stars_trend_keys,
        "stars_trend_values":   stars_trend_values,
        # Composition
        "feature_labels":       [f for f, _ in feature_counts.most_common()],
        "feature_values":       [c for _, c in feature_counts.most_common()],
        "version_labels":          [v for v, _ in version_counts.most_common(12)],
        "version_values":          [c for _, c in version_counts.most_common(12)],
        "version_breakdown":       _prof_ver_breakdown,
        "version_eol_count":       _p_eol_count,
        "version_admin_count":     _p_admin_count,
        "version_blank_count":     _p_blank_count,
    }


# ── Common legal suffixes / noise words stripped during org-name normalisation ──
_ORG_SUFFIX_RE = re.compile(
    r"\b("
    r"inc(?:orporated)?|llc|l\.l\.c|ltd|limited|corp(?:oration)?|co(?:mpany)?|"
    r"plc|gmbh|ag|sa|b\.?v|n\.?v|pty|pvt|s\.?a\.?s|s\.?r\.?l|"
    r"group|holding(?:s)?|international|intl|global|worldwide|"
    r"enterprise(?:s)?|venture(?:s)?|partner(?:s)?|associates?|"
    r"solution(?:s)?|service(?:s)?|system(?:s)?|technolog(?:y|ies)|tech|"
    r"software|consulting|consultant(?:s)?|lab(?:s|oratories?)?|"
    r"the"
    r")\b",
    re.IGNORECASE,
)


def _normalize_org(name: str) -> str:
    """
    Reduce an org name to a canonical key for deduplication.
    Steps: lowercase → strip punctuation → remove legal suffixes → collapse space.
    """
    s = name.strip().lower()
    s = re.sub(r"[^\w\s]", " ", s)        # punctuation → space
    s = _ORG_SUFFIX_RE.sub(" ", s)         # remove legal/noise words
    s = re.sub(r"\s+", " ", s).strip()
    return s


def build_org_name_map(
    tickets: list[dict],
    enabled: bool = True,
    threshold: float = 0.90,
) -> dict[str, str]:
    """
    Return a {raw_org_name: canonical_org_name} mapping built from *tickets*.

    Parameters
    ----------
    enabled   : when False the identity map is returned (no consolidation).
    threshold : fuzzy-match cutoff 0.0–1.0 (default 0.90 = 90 %).
                Higher values only merge near-identical names; lower values
                are more aggressive.  0.85 is a reasonable minimum.

    Algorithm
    ---------
    1. Normalize every raw name (case, punctuation, legal suffixes).
    2. Group raw names by normalized key — each group shares one canonical name
       (the most frequently occurring raw form).
    3. Fuzzy-merge groups whose normalized keys are ≥ *threshold* similar using
       difflib.SequenceMatcher (no extra dependencies).
    4. Final canonical = most-frequent raw name in the merged super-group.
    """
    from collections import Counter
    import difflib

    raw_counts: Counter = Counter()
    for t in tickets:
        org = (t.get("organization") or "").strip()
        if org:
            raw_counts[org] += 1

    if not raw_counts:
        return {}

    if not enabled:
        return {r: r for r in raw_counts}

    # Step 1 & 2: group by normalized key
    norm_to_raws: dict[str, list[str]] = {}
    for raw in raw_counts:
        nk = _normalize_org(raw)
        if nk:
            norm_to_raws.setdefault(nk, []).append(raw)

    # Step 3: fuzzy-merge normalized keys that are very similar
    norm_keys    = list(norm_to_raws.keys())
    merged_into: dict[str, str] = {}   # norm_key → canonical_norm_key

    for nk in norm_keys:
        if nk in merged_into:
            continue
        close = difflib.get_close_matches(nk, norm_keys, n=10, cutoff=threshold)
        if len(close) <= 1:
            continue
        total = {c: sum(raw_counts[r] for r in norm_to_raws.get(c, [])) for c in close}
        best  = max(total, key=total.__getitem__)
        for c in close:
            merged_into[c] = best

    # Step 4: build the final {raw: canonical_raw} mapping
    mapping: dict[str, str] = {}
    for nk, raws in norm_to_raws.items():
        target_nk   = merged_into.get(nk, nk)
        target_raws = norm_to_raws.get(target_nk, raws)
        canonical   = max(target_raws, key=lambda r: raw_counts[r])
        for raw in raws:
            mapping[raw] = canonical

    for raw in raw_counts:
        mapping.setdefault(raw, raw)

    return mapping


def _apply_org_map(org: str, org_map: dict[str, str]) -> str:
    """Return the canonical org name, falling back to the raw value."""
    return org_map.get(org.strip(), org.strip()) if org else org


def build_customer_analytics(tickets: list[dict], scores: dict[str, dict]) -> dict[str, dict]:
    """
    Aggregate per-customer (organization) metrics from tickets + scores.
    Returns dict keyed by org name with volume, avg scored dimensions, etc.
    """
    from collections import defaultdict
    _oc = _load_settings_file().get("__org_consolidation__", {})
    org_map = build_org_name_map(
        tickets,
        enabled=bool(_oc.get("enabled", True)),
        threshold=int(_oc.get("threshold", 90)) / 100.0,
    )
    data: dict[str, dict] = defaultdict(lambda: {
        "ticket_count": 0, "escalations": 0,
        "stars": [], "complexity": [], "resolution_quality": [],
        "response_timeliness": [], "communication_clarity": [],
        "temperatures": [],
    })

    for t in tickets:
        org = _apply_org_map((t.get("organization") or "Unknown").strip() or "Unknown", org_map)
        d   = data[org]
        d["ticket_count"] += 1
        if t.get("escalations"):
            d["escalations"] += 1
        tid = str(t.get("ticket_id", ""))
        s   = scores.get(tid, {})
        if s.get("stars"):                 d["stars"].append(float(s["stars"]))
        if s.get("complexity"):            d["complexity"].append(float(s["complexity"]))
        if s.get("resolution_quality"):    d["resolution_quality"].append(float(s["resolution_quality"]))
        if s.get("response_timeliness"):   d["response_timeliness"].append(float(s["response_timeliness"]))
        if s.get("communication_clarity"): d["communication_clarity"].append(float(s["communication_clarity"]))
        if s.get("temperature") and isinstance(s["temperature"], str): d["temperatures"].append(s["temperature"].lower())

    def _avg(lst): return round(sum(lst) / len(lst), 2) if lst else 0.0

    result = {}
    for org, d in data.items():
        temps = d["temperatures"]
        hot_pct = round(100 * temps.count("hot")  / len(temps), 1) if temps else 0.0
        result[org] = {
            "ticket_count":             d["ticket_count"],
            "escalations":              d["escalations"],
            "avg_stars":                _avg(d["stars"]),
            "avg_complexity":           _avg(d["complexity"]),
            "avg_resolution_quality":   _avg(d["resolution_quality"]),
            "avg_response_timeliness":  _avg(d["response_timeliness"]),
            "avg_communication_clarity":_avg(d["communication_clarity"]),
            "hot_pct":                  hot_pct,
            "scored_count":             len(d["stars"]),
        }
    return result


def classify_ticket_origin(ticket: dict) -> str:
    """
    Classify how a ticket was opened based on subject prefix, comment body, and tags.
    Returns one of: 'Proactive/Automated', 'Agent-Initiated', 'Customer-Initiated'

    Delegates to _is_proactive_ticket() for consistency with the scoring prompt.
    Note: 'carr' is a revenue-tier tag (CARR), NOT a proactive-origin signal.
    """
    if _is_proactive_ticket(ticket):
        return "Proactive/Automated"
    tags = (ticket.get("tags") or "").lower()
    if "req_by_agent" in tags or "opened_by_agent" in tags:
        return "Agent-Initiated"
    return "Customer-Initiated"


def build_cluster_timeline(tickets: list[dict], cluster_key: str) -> list[dict]:
    """
    Return time-ordered topology snapshots for the cluster identified by
    cluster_key (matched against cluster_name or cluster_uuid).

    Each entry in the returned list corresponds to one ticket and contains the
    topology metrics captured at the time that ticket was created/enriched.
    """
    import datetime

    _DATE_FMTS = (
        "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d", "%m/%d/%Y", "%B %d, %Y", "%b %d, %Y",
    )

    def _parse_date(raw: str):
        for fmt in _DATE_FMTS:
            try:
                return datetime.datetime.strptime(raw[:19], fmt)
            except ValueError:
                continue
        return None

    key_lower = (cluster_key or "").strip().lower()
    points: list[dict] = []

    for ticket in tickets:
        topo = ticket.get("snapshot_topology")
        if not isinstance(topo, dict) or not topo:
            continue

        # Match on cluster_name, capella_cluster_id, or ticket_fields.Cluster_ID.
        # cluster_uuid (internal CB hex UUID) is kept for dedup but not shown to users.
        tf = _parse_ticket_fields(ticket)
        identifiers = [
            _topo_str(topo.get("cluster_name")).lower(),
            _topo_str(topo.get("capella_cluster_id")).lower(),
            _topo_str(topo.get("cluster_uuid")).lower(),
            _topo_str(tf.get("Cluster_ID")).lower(),
        ]
        if key_lower not in identifiers:
            continue

        raw_date = (ticket.get("created") or ticket.get("updated") or "").strip()
        dt = _parse_date(raw_date)
        if dt is None:
            continue

        # SDK version from ticket fields (Zendesk custom field)
        sdk_ver = _topo_str(tf.get("Couchbase_Server_SDK_or_Connector"))
        if not sdk_ver:
            sdk_ver = _topo_str(tf.get("Couchbase_Analytics_SDK"))

        points.append({
            "dt":                dt,
            "ts_ms":             int(dt.timestamp() * 1000),
            "date_label":        dt.strftime("%Y-%m-%d"),
            "ticket_id":         ticket.get("ticket_id", ""),
            "subject":           (ticket.get("subject") or "")[:80],
            "node_count":        topo.get("total_nodes"),
            "bucket_count":      topo.get("bucket_count"),
            "cb_version":        topo.get("cb_version"),
            "sdk_version":       sdk_ver or None,
            "bad_count":         topo.get("bad_count", 0),
            "warn_count":        topo.get("warn_count", 0),
            "auto_failover_sec": topo.get("auto_failover_seconds"),
            "ram_mib":           topo.get("ram_per_node_mib"),
            "orchestrator":      topo.get("orchestrator"),
            "ldap_enabled":      topo.get("ldap_enabled"),
            "data_nodes":        topo.get("data_nodes", 0),
            "query_nodes":       topo.get("query_nodes", 0),
            "index_nodes":       topo.get("index_nodes", 0),
            "fts_nodes":         topo.get("fts_nodes", 0),
            "eventing_nodes":    topo.get("eventing_nodes", 0),
            "analytics_nodes":   topo.get("analytics_nodes", 0),
        })

    points.sort(key=lambda p: p["dt"])
    return points


def _cluster_timeline_charts(points: list[dict]) -> list[dict]:
    """
    Convert a sorted list of timeline points into a list of ECharts option
    dicts ready to be passed to ui.echart().  Each dict has a '_height' key
    (popped at render time) so the caller can set the container height.

    All x-axes use type="time" so points are spaced by actual date gap, not
    evenly as category labels.  Series data is [[ts_ms, value], ...] pairs.
    Returns an empty list when there is no plottable data.
    """
    if not points:
        return []

    _zoom = [{"type": "inside"}, {"type": "slider", "bottom": 30, "height": 20}]
    _xaxis_time = {
        "type": "time",
        "name": "Date",
        "nameLocation": "end",
        "axisLabel": {"formatter": "{yyyy}-{MM}-{dd}", "rotate": 30},
    }
    _tooltip_time = {
        "trigger": "axis",
        "axisPointer": {"type": "cross", "snap": True},
    }

    charts = []

    # ── 1. CB Version & SDK version over time ─────────────────────────────────
    # Scatter chart: y-axis is version string category, x-axis is time.
    # Each row in the scatter data is [ts_ms, category_index].
    _ver_set = sorted(set(p["cb_version"] for p in points if p["cb_version"]))
    _sdk_set  = sorted(set(p["sdk_version"] for p in points if p["sdk_version"]))
    _has_sdk  = bool(_sdk_set)

    if _ver_set:
        # Build combined category list: CB versions first, then SDK versions with prefix
        _cb_labels  = _ver_set
        _sdk_labels = [f"SDK {v}" for v in _sdk_set] if _has_sdk else []
        _all_labels = _cb_labels + _sdk_labels
        _label_idx  = {lbl: i for i, lbl in enumerate(_all_labels)}

        _cb_scatter = [
            [p["ts_ms"], _label_idx[p["cb_version"]], p["date_label"], str(p["ticket_id"]), p["cb_version"]]
            for p in points if p["cb_version"] and p["cb_version"] in _label_idx
        ]
        _sdk_scatter = [
            [p["ts_ms"], _label_idx[f"SDK {p['sdk_version']}"], p["date_label"], str(p["ticket_id"]), p["sdk_version"]]
            for p in points if p["sdk_version"] and f"SDK {p['sdk_version']}" in _label_idx
        ] if _has_sdk else []

        _dims = ["ts", "y_idx", "Date", "Ticket", "Version"]
        _ver_series = [{
            "name": "CB Server",
            "type": "scatter",
            "symbolSize": 14,
            "color": "#1565C0",
            "dimensions": _dims,
            "encode": {"x": "ts", "y": "y_idx", "tooltip": ["Date", "Ticket", "Version"]},
            "data": _cb_scatter,
        }]
        if _sdk_scatter:
            _ver_series.append({
                "name": "SDK",
                "type": "scatter",
                "symbolSize": 10,
                "symbol": "diamond",
                "color": "#E65100",
                "dimensions": _dims,
                "encode": {"x": "ts", "y": "y_idx", "tooltip": ["Date", "Ticket", "Version"]},
                "data": _sdk_scatter,
            })

        charts.append({
            "_height": max(200, len(_all_labels) * 32 + 120),
            "title":   {"text": "CB Server & SDK Version History"},
            "tooltip": {"trigger": "item"},
            "legend":  {"bottom": 0} if _has_sdk else {},
            "dataZoom": _zoom,
            "grid":    {"bottom": 70, "left": 120, "right": 30},
            "xAxis":   dict(_xaxis_time),
            "yAxis":   {
                "type": "category",
                "data": _all_labels,
                "axisLabel": {"fontSize": 11},
                "splitLine": {"show": True, "lineStyle": {"type": "dashed", "opacity": 0.4}},
            },
            "series": _ver_series,
        })

    # ── 2. Node count over time ────────────────────────────────────────────────
    node_vals = [p["node_count"] for p in points]
    if any(v is not None for v in node_vals):
        svc_fields = [
            ("Data", "data_nodes"), ("Query", "query_nodes"),
            ("Index", "index_nodes"), ("Search", "fts_nodes"),
            ("Eventing", "eventing_nodes"), ("Analytics", "analytics_nodes"),
        ]
        has_mds = any(p.get(f) for p in points for _, f in svc_fields)
        if has_mds:
            charts.append({
                "_height": 300,
                "title":    {"text": "Node Count Over Time (by Service)"},
                "tooltip":  {"trigger": "axis", "axisPointer": {"type": "shadow"}},
                "legend":   {"bottom": 0},
                "dataZoom": _zoom,
                "grid":     {"bottom": 70},
                "xAxis":    dict(_xaxis_time),
                "yAxis":    {"type": "value", "name": "Nodes", "minInterval": 1},
                "series":   [
                    {
                        "name": label,
                        "type": "bar",
                        "stack": "total",
                        "data": [[p["ts_ms"], p.get(field, 0) or 0] for p in points],
                    }
                    for label, field in svc_fields
                    if any(p.get(field, 0) for p in points)
                ],
            })
        else:
            charts.append({
                "_height": 280,
                "title":    {"text": "Node Count Over Time"},
                "tooltip":  _tooltip_time,
                "dataZoom": _zoom,
                "grid":     {"bottom": 70},
                "xAxis":    dict(_xaxis_time),
                "yAxis":    {"type": "value", "name": "Nodes", "minInterval": 1},
                "color":    ["#1E88E5"],
                "series":   [{"name": "Total Nodes", "type": "line", "smooth": True,
                              "data": [[p["ts_ms"], p["node_count"]] for p in points]}],
            })

    # ── 3. Bucket count over time ──────────────────────────────────────────────
    bucket_vals = [p["bucket_count"] for p in points]
    if any(v is not None for v in bucket_vals):
        charts.append({
            "_height": 260,
            "title":    {"text": "Bucket Count Over Time"},
            "tooltip":  _tooltip_time,
            "dataZoom": _zoom,
            "grid":     {"bottom": 70},
            "xAxis":    dict(_xaxis_time),
            "yAxis":    {"type": "value", "name": "Buckets", "minInterval": 1},
            "color":    ["#43A047"],
            "series":   [{"name": "Buckets", "type": "line", "smooth": True,
                          "data": [[p["ts_ms"], p["bucket_count"]] for p in points]}],
        })

    # ── 4. Checker health (BAD + WARN) over time ───────────────────────────────
    if any(p["bad_count"] or p["warn_count"] for p in points):
        charts.append({
            "_height": 280,
            "title":    {"text": "Checker Health Over Time", "subtext": "Lower is better"},
            "tooltip":  _tooltip_time,
            "legend":   {"bottom": 0},
            "dataZoom": _zoom,
            "grid":     {"bottom": 70},
            "xAxis":    dict(_xaxis_time),
            "yAxis":    {"type": "value", "name": "Issue Count", "minInterval": 1},
            "color":    ["#E53935", "#FB8C00"],
            "series":   [
                {"name": "BAD checks",  "type": "line", "smooth": True,
                 "data": [[p["ts_ms"], p["bad_count"]]  for p in points]},
                {"name": "WARN checks", "type": "line", "smooth": True,
                 "data": [[p["ts_ms"], p["warn_count"]] for p in points]},
            ],
        })

    # ── 5. Auto-failover setting over time ────────────────────────────────────
    af_vals = [p["auto_failover_sec"] for p in points]
    if any(v is not None for v in af_vals):
        charts.append({
            "_height": 260,
            "title":    {"text": "Auto-Failover Threshold Over Time"},
            "tooltip":  _tooltip_time,
            "dataZoom": _zoom,
            "grid":     {"bottom": 70},
            "xAxis":    dict(_xaxis_time),
            "yAxis":    {"type": "value", "name": "Seconds", "minInterval": 1},
            "color":    ["#8E24AA"],
            "series":   [{"name": "Auto-failover (s)", "type": "line", "step": "start",
                          "data": [[p["ts_ms"], p["auto_failover_sec"]] for p in points]}],
        })

    # ── 6. RAM per node over time ─────────────────────────────────────────────
    ram_vals = [p["ram_mib"] for p in points]
    if any(v is not None for v in ram_vals):
        charts.append({
            "_height": 260,
            "title":    {"text": "RAM per Node Over Time"},
            "tooltip":  _tooltip_time,
            "dataZoom": _zoom,
            "grid":     {"bottom": 70},
            "xAxis":    dict(_xaxis_time),
            "yAxis":    {"type": "value", "name": "MiB", "minInterval": 1},
            "color":    ["#FB8C00"],
            "series":   [{"name": "RAM (MiB)", "type": "line", "smooth": True,
                          "data": [[p["ts_ms"], p["ram_mib"]] for p in points]}],
        })

    return charts


def build_analytics_data(tickets: list[dict], scores: dict[str, dict]) -> dict:
    """
    Compute all chart series from tickets + scores dict.
    Returns a nested dict consumed by the chart-rendering functions.
    """
    from collections import Counter
    import datetime

    _sf = _topo_str  # local alias for readability inside this function

    # Frequency over time — total and per origin
    month_by_origin: dict[str, Counter] = {
        "Customer-Initiated":   Counter(),
        "Agent-Initiated":      Counter(),
        "Proactive/Automated":  Counter(),
    }
    all_months = []
    for t in tickets:
        parsed = _parse_created([t])
        if parsed:
            m = parsed[0]
            all_months.append(m)
            month_by_origin[classify_ticket_origin(t)][m] += 1
    month_freq = Counter(all_months)
    month_keys = sorted(month_freq.keys())
    year_freq: Counter = Counter(m[:4] for m in all_months)
    year_keys = sorted(year_freq.keys())

    # Priority
    priority_counts = Counter(
        (t.get("priority") or "unknown").capitalize() for t in tickets
    )

    # Status
    status_counts = Counter(
        (t.get("status") or "unknown").capitalize() for t in tickets
    )

    # Comment count buckets
    buckets    = {"1": 0, "2-5": 0, "6-10": 0, "11-20": 0, "21+": 0}
    for t in tickets:
        c = int(t.get("comment_count") or 0)
        if c <= 1:   buckets["1"]     += 1
        elif c <= 5:  buckets["2-5"]   += 1
        elif c <= 10: buckets["6-10"]  += 1
        elif c <= 20: buckets["11-20"] += 1
        else:         buckets["21+"]   += 1

    # Escalation rate
    with_esc    = sum(1 for t in tickets if t.get("escalations"))
    without_esc = len(tickets) - with_esc

    # Ticket origin
    origin_counts = Counter(classify_ticket_origin(t) for t in tickets)

    # Proactive signal breakdown — which signal triggered classification
    proactive_by_subject = 0
    proactive_by_comment = 0
    proactive_tickets_sample: list[dict] = []
    for t in tickets:
        subj = (t.get("subject") or "").strip().lower()
        if subj.startswith(_PROACTIVE_SUBJECT_PREFIX):
            proactive_by_subject += 1
            proactive_tickets_sample.append({"signal": "subject", "id": t.get("ticket_id"), "subject": t.get("subject", "")[:80], "org": t.get("organization", "")})
            continue
        comments_raw = t.get("comments")
        if comments_raw:
            try:
                clist = json.loads(comments_raw) if isinstance(comments_raw, str) else comments_raw
                if isinstance(clist, list) and clist:
                    oldest_body = (clist[-1].get("body") or "").lower()
                    if _PROACTIVE_BODY_FRAGMENT in oldest_body:
                        proactive_by_comment += 1
                        proactive_tickets_sample.append({"signal": "comment", "id": t.get("ticket_id"), "subject": t.get("subject", "")[:80], "org": t.get("organization", "")})
            except Exception:
                pass

    # Version distribution + unique cluster UUIDs per version
    version_counts: Counter = Counter()
    _uuid_re = re.compile(r"[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}", re.IGNORECASE)
    all_unique_cluster_uuids: set = set()
    clusters_by_version: dict[str, set] = {}
    _snap_info_cache: dict[str, dict] = {}  # ticket_id → extract_cluster_snapshot_info result

    _ver_eol_count   = 0  # Couchbase_Server = "End of Life*"
    _ver_admin_count = 0  # No product version fields at all (Account Admin schema)
    _ver_blank_count = 0  # Couchbase_Server key present but blank (SE didn't fill in)

    for t in tickets:
        ver = extract_ticket_version(t)
        # Parse ticket_fields once per ticket (used for version fallback + cluster UUID)
        tf = _parse_ticket_fields(t)
        if ver and ver != "Unknown":
            version_counts[ver] += 1
        else:
            _cs = (tf.get("Couchbase_Server") or "").strip().lower()
            if "Couchbase_Server" not in tf:
                _ver_admin_count += 1
            elif "end of life" in _cs:
                _ver_eol_count += 1
            else:
                _ver_blank_count += 1

        # Collect cluster UUIDs from fast structured sources only.
        # Source 3 (extract_cluster_snapshot_info over description+comments) is omitted here —
        # it is O(n * comment_size) and produces many false positives for analytics charts.
        ticket_uuids: set = set()

        # Source 1: ticket_fields["Cluster ID"]
        cf_uuid = (tf.get("Cluster_ID") or "").strip()
        if _uuid_re.match(cf_uuid):
            ticket_uuids.add(cf_uuid.lower())

        # Source 2: snapshot_topology identifiers
        topo_raw = t.get("snapshot_topology")
        if topo_raw:
            topo_d = topo_raw if isinstance(topo_raw, dict) else {}
            if isinstance(topo_raw, str):
                try:
                    topo_d = json.loads(topo_raw)
                except Exception:
                    topo_d = {}
            for _uid_key in ("capella_cluster_id", "cluster_uuid"):
                _uid_val = (topo_d.get(_uid_key) or "").strip()
                if _uuid_re.match(_uid_val):
                    ticket_uuids.add(_uid_val.lower())

        all_unique_cluster_uuids.update(ticket_uuids)

        # Lightweight snapshot count — just count non-empty lines in the snapshots field
        _snaps_raw = t.get("snapshots") or ""
        if isinstance(_snaps_raw, list):
            _snaps_raw = "\n".join(str(s) for s in _snaps_raw)
        _snap_count = sum(1 for ln in str(_snaps_raw).splitlines() if ln.strip())
        _snap_info_cache[str(t.get("ticket_id", ""))] = {"snapshot_count": _snap_count}

        # Map each UUID to the ticket's version (skip if no real version)
        if ver and ver != "Unknown":
            clusters_by_version.setdefault(ver, set()).update(ticket_uuids)

    version_counts.pop("Unknown", None)
    version_items = sorted(version_counts.items())   # sort by version string

    # Build a combined breakdown list for the color-coded chart:
    #   (label, count, category)  — category drives bar colour + checkbox visibility
    _version_breakdown: list[tuple[str, int, str]] = [
        (v, c, "version") for v, c in version_items
    ]
    if _ver_eol_count:
        _version_breakdown.append(("EOL / End of Life", _ver_eol_count, "eol"))
    if _ver_admin_count:
        _version_breakdown.append(("Admin / No Product Fields", _ver_admin_count, "admin"))
    if _ver_blank_count:
        _version_breakdown.append(("Version Not Specified", _ver_blank_count, "blank"))

    # Build parallel lists for chart: versions with ≥1 known cluster UUID
    _ver_with_clusters = sorted(clusters_by_version.keys())
    unique_cluster_total = len(all_unique_cluster_uuids)

    # Feature area distribution
    feature_counts = Counter(classify_ticket_feature(t) for t in tickets)
    feature_items  = sorted(feature_counts.items(), key=lambda x: -x[1])  # by frequency

    # Scored metrics (only if scores available)
    stars_counts      = Counter()
    temp_counts       = Counter()
    complexity_counts = Counter()
    rq_counts         = Counter()
    rt_counts         = Counter()
    cc_counts         = Counter()

    for s in scores.values():
        if s.get("stars"):        stars_counts[str(s["stars"])]                 += 1
        if s.get("temperature") and isinstance(s["temperature"], str):  temp_counts[s["temperature"].capitalize()] += 1
        if s.get("complexity"):   complexity_counts[str(s["complexity"])]        += 1
        if s.get("resolution_quality"):   rq_counts[str(s["resolution_quality"])] += 1
        if s.get("response_timeliness"):  rt_counts[str(s["response_timeliness"])] += 1
        if s.get("communication_clarity"): cc_counts[str(s["communication_clarity"])] += 1

    # Cluster & snapshot metrics — always derived from the full ticket dicts so
    # that snapshots/cluster fields are available regardless of scoring state.
    # Build a tid→ticket lookup for quick cross-reference.
    cluster_id_counts:   Counter = Counter()
    cluster_name_counts: Counter = Counter()
    snapshot_bucket_counts: Counter = Counter({"0": 0, "1": 0, "2-5": 0, "6-10": 0, "11+": 0})
    tickets_with_snapshots = 0

    for ticket in tickets:
        tid = str(ticket.get("ticket_id", ""))
        score = scores.get(tid, {})

        # Snapshot count: score field first, then lightweight cache from first loop.
        s_count = score.get("snapshot_count") if score else None
        if s_count is None:
            s_count = _snap_info_cache.get(tid, {}).get("snapshot_count", 0)

        # For analytics charts use ONLY authoritative sources — snapshot_topology
        # (from the nutshell API) and the structured ticket_fields "Cluster ID".
        # The heuristic text extraction in extract_cluster_snapshot_info scans
        # comments/descriptions and produces thousands of false positives
        # (e.g. "production cluster", "test cluster") that pollute the chart.
        topo = ticket.get("snapshot_topology") if isinstance(ticket.get("snapshot_topology"), dict) else {}
        if isinstance(ticket.get("snapshot_topology"), str):
            try:
                topo = json.loads(ticket["snapshot_topology"])
            except Exception:
                topo = {}

        if topo and s_count == 0:
            s_count = 1  # enriched topology counts as at least one snapshot

        # Cluster name: only from snapshot_topology.cluster_name
        # Skip UUID-shaped values — those are IDs that leaked into the name field
        cn = _sf(topo.get("cluster_name"))
        if cn and not re.match(r"^[0-9a-f]{8}-[0-9a-f]{4}-", cn, re.IGNORECASE):
            cluster_name_counts[cn] += 1

        # Cluster ID: prefer Capella UUID (user-facing, matches ticket_fields.Cluster_ID)
        # Don't count cluster_uuid (CB internal 32-hex UUID, no dashes) — users don't
        # recognise it and it doesn't match Zendesk Cluster_ID format.
        tf = _parse_ticket_fields(ticket)
        cid_field   = _sf(tf.get("Cluster_ID"))
        capella_cid = _sf(topo.get("capella_cluster_id"))
        # De-duplicate: ticket_fields.Cluster_ID and capella_cluster_id are often
        # the same value; only count a given UUID once per ticket.
        _seen_cids: set = set()
        for cid in filter(None, [cid_field, capella_cid]):
            if cid not in _seen_cids:
                _seen_cids.add(cid)
                cluster_id_counts[cid] += 1

        sc = int(s_count or 0)
        if sc > 0:
            tickets_with_snapshots += 1
        if sc == 0:    snapshot_bucket_counts["0"]    += 1
        elif sc == 1:  snapshot_bucket_counts["1"]    += 1
        elif sc <= 5:  snapshot_bucket_counts["2-5"]  += 1
        elif sc <= 10: snapshot_bucket_counts["6-10"] += 1
        else:          snapshot_bucket_counts["11+"]  += 1

    # ── Enriched topology metrics (tickets that have snapshot_topology) ──────────
    _af_order   = ["Disabled", "10s", "30s", "60s", "120s", "Other"]
    _ram_order  = ["≤16 GiB", "32 GiB", "64 GiB", "128 GiB", ">128 GiB"]
    _node_order = ["1", "2", "3", "4", "5", "6-10", "11+"]

    af_counts:          Counter = Counter({k: 0 for k in _af_order})
    ldap_counts:        Counter = Counter({"Enabled": 0, "Disabled": 0, "Unknown": 0})
    ram_counts:         Counter = Counter({k: 0 for k in _ram_order})
    node_dist_counts:   Counter = Counter({k: 0 for k in _node_order})
    bucket_dist_counts: Counter = Counter()
    topo_version_counts: Counter = Counter()
    orchestrator_counts: Counter = Counter()
    enriched_ticket_count = 0

    for ticket in tickets:
        topo = ticket.get("snapshot_topology")
        if not isinstance(topo, dict) or not topo:
            continue
        enriched_ticket_count += 1

        # Auto-failover
        af = topo.get("auto_failover_seconds")
        if af is None:
            af_counts["Disabled"] += 1
        elif af <= 0:
            af_counts["Disabled"] += 1
        elif af <= 10:
            af_counts["10s"] += 1
        elif af <= 30:
            af_counts["30s"] += 1
        elif af <= 60:
            af_counts["60s"] += 1
        elif af <= 120:
            af_counts["120s"] += 1
        else:
            af_counts["Other"] += 1

        # LDAP
        ldap = topo.get("ldap_enabled")
        if ldap is True:
            ldap_counts["Enabled"] += 1
        elif ldap is False:
            ldap_counts["Disabled"] += 1
        else:
            ldap_counts["Unknown"] += 1

        # RAM per node tier
        ram = topo.get("ram_per_node_mib")
        if ram is not None:
            if ram <= 16384:
                ram_counts["≤16 GiB"] += 1
            elif ram <= 32768:
                ram_counts["32 GiB"] += 1
            elif ram <= 65536:
                ram_counts["64 GiB"] += 1
            elif ram <= 131072:
                ram_counts["128 GiB"] += 1
            else:
                ram_counts[">128 GiB"] += 1

        # Node count distribution
        nc = topo.get("total_nodes")
        if nc is not None:
            nc = int(nc)
            if nc <= 5:
                node_dist_counts[str(nc)] += 1
            elif nc <= 10:
                node_dist_counts["6-10"] += 1
            else:
                node_dist_counts["11+"] += 1

        # Bucket count distribution
        bc = topo.get("bucket_count")
        if bc is not None:
            bucket_dist_counts[str(int(bc))] += 1

        # Orchestrator hotspot (top 10)
        orch = _sf(topo.get("orchestrator"))
        if orch:
            orchestrator_counts[orch] += 1

    # CB version distribution — all tickets (ticket fields primary, snapshot fallback)
    for ticket in tickets:
        cv = extract_ticket_version(ticket)
        if cv and cv.lower() not in ("unknown", "n/a", "-", ""):
            topo_version_counts[cv] += 1

    # ── CBSE document metrics ────────────────────────────────────────────────────
    _cbse_re = re.compile(r"cbse[-_]?\d+", re.IGNORECASE)
    cbse_total          = 0
    cbse_by_year:  Counter = Counter()
    cbse_by_month: Counter = Counter()

    for ticket in tickets:
        tf = _parse_ticket_fields(ticket)
        raw_cbse = _sf(tf.get("CBSE"))
        if not raw_cbse or not _cbse_re.search(raw_cbse):
            continue
        cbse_total += 1
        # Derive year and month from ticket created date
        raw_date = (ticket.get("created") or "").strip()
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%B %d, %Y", "%b %d, %Y",
                    "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S"):
            try:
                import datetime as _dt
                _d = _dt.datetime.strptime(raw_date[:len(fmt) + 2], fmt)
                cbse_by_year[str(_d.year)] += 1
                cbse_by_month[f"{_d.year}-{_d.month:02d}"] += 1
                break
            except ValueError:
                continue

    cbse_year_items  = sorted(cbse_by_year.items())
    cbse_month_keys  = sorted(cbse_by_month.keys())
    total_tickets    = max(len(tickets), 1)
    cbse_avg_per_ticket = round(cbse_total / total_tickets, 3)

    top_cluster_ids        = cluster_id_counts.most_common(20)
    top_cluster_names      = cluster_name_counts.most_common(20)
    top_orchestrators      = orchestrator_counts.most_common(10)
    top_bucket_dist        = sorted(bucket_dist_counts.items(), key=lambda x: int(x[0]) if x[0].isdigit() else 99)
    topo_version_items     = sorted(topo_version_counts.items())
    snap_bucket_order      = ["0", "1", "2-5", "6-10", "11+"]

    # Build cluster ID display labels (first 8 chars before any dash, or 12 chars)
    def _cid_label(cid: str) -> str:
        parts = cid.split("-")
        return parts[0] if len(parts) > 1 else cid[:12]

    return {
        "year_keys":           year_keys,
        "year_values":         [year_freq[y] for y in year_keys],
        "month_keys":          month_keys,
        "month_values":        [month_freq[k] for k in month_keys],
        "month_customer":      [month_by_origin["Customer-Initiated"][k]  for k in month_keys],
        "month_agent":         [month_by_origin["Agent-Initiated"][k]     for k in month_keys],
        "month_proactive":     [month_by_origin["Proactive/Automated"][k] for k in month_keys],
        "priority_labels":   list(priority_counts.keys()),
        "priority_values":   list(priority_counts.values()),
        "status_labels":     list(status_counts.keys()),
        "status_values":     list(status_counts.values()),
        "comment_labels":    list(buckets.keys()),
        "comment_values":    list(buckets.values()),
        "esc_labels":        ["With escalation", "No escalation"],
        "esc_values":        [with_esc, without_esc],
        "origin_labels":     ["Customer-Initiated", "Agent-Initiated", "Proactive/Automated"],
        "origin_values":     [
            origin_counts["Customer-Initiated"],
            origin_counts["Agent-Initiated"],
            origin_counts["Proactive/Automated"],
        ],
        "proactive_by_subject":       proactive_by_subject,
        "proactive_by_comment":       proactive_by_comment,
        "proactive_tickets_sample":   proactive_tickets_sample[:50],
        "version_labels":          [v for v, _ in version_items],
        "version_values":          [c for _, c in version_items],
        "version_breakdown":       _version_breakdown,
        "version_eol_count":       _ver_eol_count,
        "version_admin_count":     _ver_admin_count,
        "version_blank_count":     _ver_blank_count,
        "unique_cluster_total":          unique_cluster_total,
        "clusters_by_version_labels":    _ver_with_clusters,
        "clusters_by_version_values":    [len(clusters_by_version[v]) for v in _ver_with_clusters],
        "tickets_by_version_for_clusters": [version_counts.get(v, 0) for v in _ver_with_clusters],
        "feature_labels":    [f for f, _ in feature_items],
        "feature_values":    [c for _, c in feature_items],
        "stars_labels":      [str(i) for i in range(1, 6)],
        "stars_values":      [stars_counts[str(i)] for i in range(1, 6)],
        "temp_labels":       ["Cold", "Warm", "Hot"],
        "temp_values":       [temp_counts["Cold"], temp_counts["Warm"], temp_counts["Hot"]],
        "complexity_labels": [str(i) for i in range(1, 6)],
        "complexity_values": [complexity_counts[str(i)] for i in range(1, 6)],
        "dim_categories":    ["Resolution Quality", "Response Timeliness", "Communication Clarity"],
        "dim_avg":           [
            round(sum(int(k)*v for k,v in rq_counts.items()) / max(sum(rq_counts.values()),1), 2),
            round(sum(int(k)*v for k,v in rt_counts.items()) / max(sum(rt_counts.values()),1), 2),
            round(sum(int(k)*v for k,v in cc_counts.items()) / max(sum(cc_counts.values()),1), 2),
        ],
        # Cluster & snapshot
        "cluster_id_labels":   [_cid_label(k) for k, _ in top_cluster_ids],
        "cluster_id_full":     [k for k, _ in top_cluster_ids],
        "cluster_id_values":   [v for _, v in top_cluster_ids],
        "cluster_name_labels": [k for k, _ in top_cluster_names],
        "cluster_name_values": [v for _, v in top_cluster_names],
        "snap_bucket_labels":  snap_bucket_order,
        "snap_bucket_values":  [snapshot_bucket_counts[k] for k in snap_bucket_order],
        "tickets_with_snapshots": tickets_with_snapshots,
        # Enriched topology metrics (from #nutshell-alternative HTML)
        "enriched_ticket_count":    enriched_ticket_count,
        "af_labels":                _af_order,
        "af_values":                [af_counts[k] for k in _af_order],
        "ldap_labels":              ["Enabled", "Disabled", "Unknown"],
        "ldap_values":              [ldap_counts["Enabled"], ldap_counts["Disabled"], ldap_counts["Unknown"]],
        "ram_labels":               _ram_order,
        "ram_values":               [ram_counts[k] for k in _ram_order],
        "node_dist_labels":         _node_order,
        "node_dist_values":         [node_dist_counts[k] for k in _node_order],
        "bucket_dist_labels":       [k for k, _ in top_bucket_dist],
        "bucket_dist_values":       [v for _, v in top_bucket_dist],
        "topo_version_labels":      [v for v, _ in topo_version_items],
        "topo_version_values":      [c for _, c in topo_version_items],
        "orchestrator_labels":      [k for k, _ in top_orchestrators],
        "orchestrator_values":      [v for _, v in top_orchestrators],
        # CBSE document metrics
        "cbse_total":               cbse_total,
        "cbse_avg_per_ticket":      cbse_avg_per_ticket,
        "cbse_year_labels":         [y for y, _ in cbse_year_items],
        "cbse_year_values":         [c for _, c in cbse_year_items],
        "cbse_month_keys":          cbse_month_keys,
        "cbse_month_values":        [cbse_by_month[k] for k in cbse_month_keys],
    }


# ─────────────────────────── Entry point ──────────────────────────────────────

if __name__ == "__main__":
    # Raise Socket.IO / Engine.IO max WebSocket payload from 1 MB → 16 MB
    # so that PDF export (SVG collection via run_javascript) doesn't hit the limit.
    import nicegui.core as _ngcore
    _ngcore.sio.eio.max_http_buffer_size = 16 * 1024 * 1024

    ui.run(
        title=f"Strabo v{__version__}",
        port=int(os.environ.get("STRABO_PORT", 8765)),
        reload=False,   # reload=True would destroy _browser_state mid-session
        show=os.environ.get("STRABO_OPEN_BROWSER", "1") == "1",
        favicon="🔍",
    )
